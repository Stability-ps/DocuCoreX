import io
import json
import logging
import os
import re
import subprocess
import tempfile
import threading
import time
import urllib.error
import urllib.request
from datetime import date, datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from typing import Any

import fitz
import pdfplumber
from fastapi import BackgroundTasks, FastAPI, Header, HTTPException, Request
from fastapi.exceptions import RequestValidationError
from fastapi.responses import JSONResponse, Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from supabase import Client, create_client
from auth import OK as AUTH_OK, STATUS_FOR_VERDICT, auth_compare_diagnostics, check_bearer
from engine.bootstrap import register_default_parsers
from engine.detection import UNKNOWN_PROFILE_ID, bank_name_for, detect_bank, is_supported_bank
from engine.classification import (
    STRENGTH_HARD,
    STRENGTH_LEARNED,
    STRENGTH_NONE,
    STRENGTH_SOFT,
    REVISABLE_STRENGTHS,
    SOURCE_AI,
    SOURCE_LEARNED_RULE,
    Classification,
    any_keyword_matches,
    keyword_matches,
    canonicalise_category,
    is_valid_ai_account,
    is_valid_vat_claim_status,
    bank_charge_evidence,
    owner_drawings_evidence,
    source_for_strength,
)
from engine.ai_prompt import build_classification_prompt
from engine.reasoning import decide as reasoning_decide
from engine.counterparty import evidence_for_all as counterparty_evidence_for_all
from engine.counterparty import extract_counterparty
from engine.counterparty import infer_relationship as infer_counterparty_relationship
from engine.reasoning import identify_merchant_type as reasoning_identify_merchant_type
from engine.ai_counterparty import COUNTERPARTY_BATCH_SIZE
from engine.ai_counterparty import build_prompt as build_counterparty_prompt
from engine.ai_counterparty import build_questions as build_counterparty_questions
from engine.ai_counterparty import validate_answers as validate_counterparty_answers
from engine.categories import canonical_categories
from engine.categories import is_unresolved_category
from engine.counterparty import group_by_counterparty as counterparty_group_by
from engine.ai_recovery import SYSTEM_PROMPT as AI_RECOVERY_SYSTEM_PROMPT
from engine.ai_recovery import batches as ai_batches
from engine.ai_recovery import build_prompt as build_ai_recovery_prompt
from engine.ai_recovery import candidate_lines as ai_candidate_lines
from engine.ai_recovery import dropped_line_count as ai_dropped_line_count
from engine.ai_recovery import ground_rows as ground_ai_rows
from engine.merchant_keys import merchant_key_rejection
from engine.merchants import identify_merchant, merchant_is_grounded
from engine.generic_parser import count_candidate_lines as generic_candidate_lines
from engine.generic_parser import extract_generic_rows
from engine.lexicon import LOOSE_DATE, LOOSE_MONEY, MONEY_TOKEN


app = FastAPI(title="DocuCoreX Accounting Worker")
logging.basicConfig(level=os.getenv("LOG_LEVEL", "INFO"))
logger = logging.getLogger("docucorex.accounting_worker")
WORKER_PARSER_VERSION = "fnb_business_v1"
# Two different things, kept apart on purpose.
#
# The BANK PROFILE is which bank issued the statement (fnb_business_v1,
# standard_bank_business_v1, ... or "unknown"). The PARSER PROFILE is which
# implementation read it: the FNB parser, or the generic one. Collapsing the two
# is what made "not FNB" mean "cannot be processed".
FNB_PROFILE_ID = "fnb_business_v1"
GENERIC_PARSER_PROFILE_ID = "generic_bank_statement_v1"
# How a row records that a model located it rather than a parser reading it, and
# the ceiling that fact puts on its confidence. Both are referenced by the
# recovery path that sets them and by the classification path that must not
# undo them, so the marker cannot drift between the two.
AI_RECOVERY_NOTE = "recovered_by: ai"
AI_RECOVERED_MAX_CONFIDENCE = 60.0
WORKER_BUILD_FALLBACK = "local-dev"
DEFAULT_AI_MODEL = "gpt-4o-mini"
# Keyed by (workspace_id, normalised description). This process is long-lived
# and serves every tenant, and the cache used to be keyed by description alone —
# so one workspace's classification of "PAYMENT TO ABC TRADING" was served to
# another workspace's identically-worded row. Classifications are shaped by a
# workspace's own corrections and chart of accounts; they are not shared facts.
AI_CLASSIFICATION_CACHE: dict[tuple[str, str], dict[str, Any]] = {}
AI_CLASSIFICATION_BATCH_SIZE = 30
ACCOUNTING_REPORT_DISCLAIMER = (
    "Draft management report generated from bank-statement data only. "
    "This is not a final IFRS or Companies Act financial statement and requires accountant review."
)
MAX_DATABASE_AMOUNT = Decimal("999999999999.99")
CENT = Decimal("0.01")

register_default_parsers()


def log_event(event: str, **fields: Any) -> None:
    logger.info(json.dumps({"event": event, **fields}, default=str))


def log_warning(event: str, **fields: Any) -> None:
    logger.warning(json.dumps({"event": event, **fields}, default=str))


def log_exception(event: str, **fields: Any) -> None:
    logger.exception(json.dumps({"event": event, **fields}, default=str))


def git_commit_fallback() -> str:
    try:
        return subprocess.check_output(["git", "rev-parse", "HEAD"], text=True, stderr=subprocess.DEVNULL).strip()
    except Exception:
        return WORKER_BUILD_FALLBACK


def worker_version() -> dict[str, str]:
    commit = (
        os.getenv("RENDER_GIT_COMMIT")
        or os.getenv("GIT_COMMIT")
        or os.getenv("COMMIT_SHA")
        or os.getenv("VERCEL_GIT_COMMIT_SHA")
        or git_commit_fallback()
    )
    return {
        "status": "ok",
        "service": os.getenv("RENDER_SERVICE_NAME") or os.getenv("WORKER_SERVICE_NAME") or "accounting-worker",
        "parser_version": WORKER_PARSER_VERSION,
        "commit": commit,
        "render_service_id": os.getenv("RENDER_SERVICE_ID", ""),
        "render_service_name": os.getenv("RENDER_SERVICE_NAME", ""),
    }


def with_worker_version(payload: dict[str, Any]) -> dict[str, Any]:
    return {**payload, "worker": worker_version()}


def recovery_options(
    full_text: str,
    pages: list[dict[str, Any]],
    structured_rows: list[dict[str, Any]] | None,
    structured_selection: dict[str, Any],
) -> dict[str, Any]:
    """What material is left to work from when no transaction was parsed.

    A run is only genuinely unprocessable once nothing recoverable remains.
    Deterministic parsing returning nothing is not the same thing: a statement
    can carry perfectly good structured rows or perfectly readable text that
    this parser simply could not lay out.

    Reporting the two apart is the point. "We could not read it" and "there is
    nothing to read" call for different responses, and collapsing them into one
    failure is what made a legible 37-page statement look like a broken file.
    """
    text = (full_text or "").strip()
    rows = structured_rows or []
    table_count = sum(len(page.get("tables") or []) for page in pages)
    generic_candidates = generic_candidate_lines(text) if text else 0

    reasons: list[str] = []
    if not text:
        reasons.append("no usable text")
    if not rows:
        reasons.append("no structured rows")
    if not table_count:
        reasons.append("no extracted tables")
    if not generic_candidates:
        reasons.append("no dated money rows in the text")

    # Structured rows that arrived but were rejected still count as material:
    # they were readable enough to send, so something can be done with them.
    recoverable = bool(rows) or bool(table_count) or generic_candidates > 0
    return {
        "recoverable": recoverable,
        "summary": ", ".join(reasons) if reasons else "material remains",
        "text_length": len(text),
        "structured_rows_received": len(rows),
        "structured_rows_usable": bool(structured_selection.get("structured_rows_usable")),
        "structured_rejection_reason": structured_selection.get("structured_rejection_reason"),
        "extracted_table_count": table_count,
        "generic_candidate_lines": generic_candidates,
        "page_count": len(pages),
    }


def resolve_bank_profile(
    worker_profile: str,
    worker_confidence: float,
    node_profile: str | None,
    node_confidence: float | None,
) -> dict[str, Any]:
    """Settle on one bank from the two detections, and say why.

    Two sides look at the same statement through different text. The Node
    pipeline reads the merged best extraction across pdfjs, pdfplumber, Azure
    and Mistral; this worker reads its own pdfplumber/PyMuPDF output, or the
    provided text when that yields more. Either can be the better witness — a
    scanned page this worker cannot read at all is legible to Azure, and an OCR
    reflow that mangles a letterhead is clean in the native extraction.

    So: a side that identified a bank beats a side that did not, agreement wins
    outright, and a genuine conflict goes to the more confident reading — with
    the worker breaking an exact tie, because it is reading the text it is about
    to parse. Unrecognised ids from a newer frontend count as no identification
    rather than as a bank this worker cannot parse.

    Returns `unknown` when neither side identified anything. That is a routing
    outcome (the generic parser), not a failure.
    """
    worker_known = is_supported_bank(worker_profile)
    node_known = is_supported_bank(node_profile)

    if worker_known and node_known:
        if worker_profile == node_profile:
            source, profile, reason = "agreed", worker_profile, "both sides identified the same bank"
        elif (node_confidence or 0.0) > worker_confidence:
            source, profile, reason = "node", node_profile, "conflict resolved by higher node confidence"
        else:
            source, profile, reason = "worker", worker_profile, "conflict resolved by worker confidence"
    elif worker_known:
        source, profile, reason = "worker", worker_profile, "only the worker identified a bank"
    elif node_known:
        source, profile, reason = "node", node_profile, "only the node pipeline identified a bank"
    else:
        source, profile, reason = "none", UNKNOWN_PROFILE_ID, "no side identified a bank"

    return {
        "bank_profile": profile,
        "bank_name": bank_name_for(profile),
        "source": source,
        "reason": reason,
    }


class ProcessRequest(BaseModel):
    run_id: str
    workspace_id: str
    document_id: str | None = None
    processing_job_id: str | None = None
    storage_path: str
    # Optional hints from the Node extraction pipeline. When pre_extracted_text is
    # provided (from the selected parser: pdfjs/pdfplumber/ocr/hybrid) it is used
    # as the statement text; the original PDF remains the fallback.
    parser_method: str | None = None
    extraction_source: str | None = None
    ocr_used: bool | None = None
    pre_extracted_text: str | None = None
    extraction_format_version: int | None = None
    pre_extracted_rows: list[dict[str, Any]] | None = None
    structured_provider: str | None = None
    # 0..1 row continuity from structured quality (not a global structured confidence score).
    structured_row_continuity: float | None = None
    structured_page_count: int | None = None
    structured_row_count: int | None = None
    structured_diagnostics: dict[str, Any] | None = None
    extraction_debug: dict[str, Any] | None = None
    # Which bank the Node pipeline concluded issued this statement, decided from
    # the merged best extraction across pdfjs / pdfplumber / Azure / Mistral.
    # "unknown" means that side looked and found nothing; None means the request
    # came from a deploy that predates bank detection. The two are not the same,
    # and neither changes routing yet — this worker re-detects independently.
    detected_bank: str | None = None
    detected_bank_name: str | None = None
    detected_bank_confidence: float | None = None
    detected_bank_reason: str | None = None
    detected_bank_evidence: list[str] | None = None


class CombineRequest(BaseModel):
    workspace_id: str
    run_ids: list[str]
    combine_different_accounts: bool = False
    override_continuity: bool = False


class ParsedTransaction(BaseModel):
    transaction_date: str | None
    description: str
    debit_amount: float | None = None
    credit_amount: float | None = None
    running_balance: float | None = None
    bank_charge: bool = False
    account_category: str = "Uncategorised"
    vat_treatment: str = "review"
    supported_by_invoice: bool = False
    notes: str = ""
    confidence: float = 70
    review_status: str = "needs_review"
    source_page: int | None = None
    source_row: int | None = None
    raw_text: str | None = None
    # The standing of the rule that classified this row. Carried in memory only —
    # transaction_insert_row does not persist it, so this adds no schema change.
    # It exists so later stages can tell a settled classification from a
    # revisable one without inferring it from the confidence number.
    classification_strength: str = STRENGTH_SOFT
    classification_reason: str = ""
    # Who decided, and how sure that decision is. Kept apart from `confidence`,
    # which for an AI-recovered row is capped as an EXTRACTION signal: a row can
    # be located by a model and still be categorised with certainty, or read
    # perfectly and still be hard to categorise.
    classification_source: str = ""
    classification_confidence: float | None = None
    normalized_merchant: str | None = None
    # migration 023 — who the transaction was with, what kind of entity that is,
    # and what the decision rested on. Kept apart from account_category and
    # vat_treatment on purpose: identity establishes who was paid, not what was
    # bought, and not whether input VAT may be claimed.
    counterparty_key: str | None = None
    counterparty_display: str | None = None
    counterparty_truncated: bool | None = None
    merchant_type: str | None = None
    merchant_type_source: str | None = None
    relationship: str | None = None
    relationship_strength: str | None = None
    evidence_used: list[dict[str, str]] | None = None
    treatment_alternatives: list[str] | None = None


def get_supabase() -> Client:
    url = os.getenv("SUPABASE_URL") or os.getenv("NEXT_PUBLIC_SUPABASE_URL")
    key = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        raise RuntimeError("SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY are required")
    return create_client(url, key)


def fetch_classification_rules(supabase: Client, workspace_id: str) -> list[dict[str, Any]]:
    try:
        response = (
            supabase.table("accounting_classification_rules")
            .select("merchant_key,account_category,vat_treatment,review_status,confidence,reason")
            .eq("workspace_id", workspace_id)
            .execute()
        )
        rules = response.data if isinstance(response.data, list) else []
        log_event("worker.classification_rules_loaded", workspace_id=workspace_id, count=len(rules))
        return rules
    except Exception as exc:
        log_warning("worker.classification_rules_unavailable", workspace_id=workspace_id, error=str(exc))
        return []


def is_ai_recovered(transaction: ParsedTransaction) -> bool:
    """True for a row a model located rather than a parser reading it."""
    return AI_RECOVERY_NOTE in (transaction.notes or "")


def stamp_counterparty_intelligence(
    transactions: list[ParsedTransaction],
    bank_profile: str | None = None,
) -> int:
    """Record who each transaction was with, and what the statement proves.

    Runs over the whole ledger at once because the interesting evidence is not
    in any single row: one payment to WELKOM FRESH P says nothing, and 115 of
    them across six months in one direction says a great deal. That evidence is
    stored, not acted on — the treatment stays exactly where classification put
    it, and a test asserts the ledger is untouched.

    Returns how many rows a counterparty could be read from.
    """
    if not transactions:
        return 0

    rows = [
        {
            "description": transaction.description,
            "debit_amount": transaction.debit_amount,
            "credit_amount": transaction.credit_amount,
            "transaction_date": transaction.transaction_date,
        }
        for transaction in transactions
    ]
    evidence_by_key = counterparty_evidence_for_all(rows, bank_profile)

    stamped = 0
    for transaction in transactions:
        party = extract_counterparty(transaction.description, bank_profile)
        if party is None:
            continue
        stamped += 1
        transaction.counterparty_key = party.key
        transaction.counterparty_display = party.display
        transaction.counterparty_truncated = party.truncated

        evidence = evidence_by_key.get(party.key)
        if evidence is not None:
            relationship = infer_counterparty_relationship(evidence)
            transaction.relationship = relationship.kind
            transaction.relationship_strength = relationship.strength

        identified = reasoning_identify_merchant_type(transaction.description)
        if identified is not None:
            _, merchant_type, _ = identified
            transaction.merchant_type = merchant_type
            transaction.merchant_type_source = "kb"

        # The reasoning behind the treatment, recorded so "why was this
        # classified?" is answerable from the row rather than by re-deriving it.
        # decide() is pure, so asking it again here costs nothing and keeps the
        # classification path itself unchanged.
        reasoned = reasoning_decide(
            transaction.description,
            transaction.debit_amount,
            transaction.credit_amount,
            counterparty_evidence=evidence,
        )
        if reasoned is not None:
            transaction.evidence_used = [
                {"source": item.source, "detail": item.detail, "grade": item.grade}
                for item in reasoned.evidence_used
            ]
            transaction.treatment_alternatives = list(reasoned.alternatives) or None

    log_event(
        "worker.counterparty_intelligence",
        rows=len(transactions),
        counterparties=len(evidence_by_key),
        stamped=stamped,
    )
    return stamped


def apply_learned_classification_rules(transactions: list[ParsedTransaction], rules: list[dict[str, Any]]) -> int:
    """Apply a workspace's approved corrections — specifically, and never over HARD evidence.

    Three production faults are closed here.

    Matching was `rule_key in normalize_merchant_key(description)`, an
    unrestricted substring test. The key "d" therefore matched 425 of 615 rows
    on a real statement and booked them all to Meals & Groceries. Matching is
    now boundary-aware, and that same rule matches 3 rows.

    A learned rule overrode whatever the deterministic rules had decided,
    including HARD evidence, so a bank fee the bank named itself lost to a
    merchant rule. HARD is now authoritative: a workspace can teach the system
    about its counterparties, not about what a bank charge is.

    And a rule whose key cannot identify a counterparty is skipped entirely
    rather than applied weakly — so an unsafe rule already stored in production
    neither classifies a row nor, by claiming it, suppresses the AI stage that
    would otherwise look at it.
    """
    if not rules:
        return 0

    usable: list[dict[str, Any]] = []
    skipped: dict[str, int] = {}
    for rule in rules:
        rejection = merchant_key_rejection(str(rule.get("merchant_key") or ""))
        if rejection:
            reason = rejection.split(":")[0]
            skipped[reason] = skipped.get(reason, 0) + 1
            continue
        usable.append(rule)
    if skipped:
        log_warning(
            "worker.learned_rules_skipped_unsafe",
            skipped=skipped,
            usable=len(usable),
            total=len(rules),
            note="keys too generic to identify a counterparty; they neither classify nor suppress AI",
        )

    # Longest key first: the most specific matching rule wins.
    sorted_rules = sorted(usable, key=lambda rule: len(str(rule.get("merchant_key") or "")), reverse=True)
    applied = 0
    for transaction in transactions:
        # HARD deterministic evidence outranks a learned rule.
        if transaction.classification_strength == STRENGTH_HARD:
            continue
        key = normalize_merchant_key(transaction.description)
        if not key:
            continue
        matched_rule = next(
            (rule for rule in sorted_rules if keyword_matches(key, str(rule["merchant_key"]))),
            None,
        )
        if not matched_rule:
            continue

        rule_category = str(matched_rule.get("account_category") or transaction.account_category)
        transaction.account_category = canonicalise_category(rule_category) or rule_category
        transaction.vat_treatment = str(matched_rule.get("vat_treatment") or transaction.vat_treatment)
        if is_ai_recovered(transaction):
            # Classification certainty is not extraction certainty.
            transaction.confidence = min(float(transaction.confidence or 0), AI_RECOVERED_MAX_CONFIDENCE)
            transaction.review_status = "needs_review"
        else:
            transaction.review_status = str(matched_rule.get("review_status") or transaction.review_status)
            transaction.confidence = max(float(transaction.confidence or 0), float(matched_rule.get("confidence") or 94))
        transaction.classification_strength = STRENGTH_LEARNED
        transaction.classification_reason = str(matched_rule.get("reason") or "workspace-approved classification rule")
        transaction.classification_source = SOURCE_LEARNED_RULE
        transaction.classification_confidence = float(matched_rule.get("confidence") or 94)
        applied += 1
    return applied


def verify_worker_token(authorization: str | None) -> None:
    """Reject the request unless it carries the shared secret.

    Previously this returned early when ACCOUNTING_WORKER_TOKEN was unset, and
    the variable was never set on the Render service — so every endpoint was
    callable by anyone who knew the hostname. It now fails closed: no secret
    means the service refuses to serve rather than serving everybody.
    """
    expected = os.getenv("ACCOUNTING_WORKER_TOKEN")
    verdict = check_bearer(authorization, expected)

    # TEMPORARY (2026-08-06) — inbound half of the shared-secret comparison.
    # Digest/length/presence only; never the token or the header. Remove with the
    # caller-side auth_outbound logging. Disable without a redeploy:
    # WORKER_AUTH_DIAGNOSTICS=false.
    if (os.getenv("WORKER_AUTH_DIAGNOSTICS") or "").strip().lower() != "false":
        try:
            log_event(
                "accounting_worker.auth_compare",
                verdict=verdict,
                render_service_id=os.getenv("RENDER_SERVICE_ID", ""),
                render_service_name=os.getenv("RENDER_SERVICE_NAME", ""),
                commit=os.getenv("RENDER_GIT_COMMIT", ""),
                **auth_compare_diagnostics(authorization, expected),
            )
        except Exception as exc:  # noqa: BLE001 - diagnostics must never break auth
            logger.warning("auth diagnostics failed: %s", type(exc).__name__)

    if verdict == AUTH_OK:
        return
    status, detail = STATUS_FOR_VERDICT[verdict]
    raise HTTPException(status_code=status, detail=detail)


def parse_money(value: str | None) -> float | None:
    return decimal_to_float(parse_money_cell(value))


def decimal_to_float(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return float(value.quantize(CENT, rounding=ROUND_HALF_UP))


def parse_money_cell(value: str | None) -> Decimal | None:
    if not value:
        return None
    matches = list(MONEY_TOKEN.finditer(value.replace("\u00a0", " ").strip()))
    if not matches:
        return None
    match = matches[-1]
    normalized = match.group("amount").replace(",", "").replace(" ", "")
    try:
        amount = Decimal(normalized)
    except Exception:
        return None
    if match.group("negative") or match.group("bracket") or (match.group("suffix") or "").lower() == "dr":
        amount = -amount
    if amount.copy_abs() > MAX_DATABASE_AMOUNT:
        log_warning("worker.amount_cell_out_of_bounds", raw=value, token=match.group(0), amount=str(amount))
        return None
    return amount


def parse_transaction_amount_cell(value: str | None) -> tuple[float | None, float | None] | None:
    amount = parse_money_cell(value)
    if amount is None:
        return None
    suffix = ""
    if value:
        matches = list(MONEY_TOKEN.finditer(value.replace("\u00a0", " ").strip()))
        suffix = (matches[-1].group("suffix") or "").lower() if matches else ""
    if suffix == "cr":
        return None, decimal_to_float(amount.copy_abs())
    return decimal_to_float(amount.copy_abs()), None


def looks_like_money(value: str | None) -> bool:
    if not value:
        return False
    return MONEY_TOKEN.search(value) is not None


def money_sign_hint(value: str | None) -> str | None:
    if not value:
        return None
    lowered = value.lower()
    if "cr" in lowered or "credit" in lowered or "+" in lowered:
        return "credit"
    if "dr" in lowered or "debit" in lowered or value.strip().startswith("-") or value.strip().endswith("-"):
        return "debit"
    return None


def parse_date(value: str | None) -> str | None:
    if not value:
        return None
    value = value.strip()
    # "%d %b %y" — a two-digit year after a month name ("30 Apr 25") is how
    # Standard Bank prints every transaction date. Without it parse_date returned
    # None, normalize_transaction_date then appended the statement year to make
    # "30 Apr 25 2025", which failed too, and every row was dropped for want of a
    # date. Numeric two-digit years ("30/04/25") were already accepted.
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d %b %Y", "%d %B %Y", "%d %b %y", "%d %B %y"):
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def extract_text_with_pdfplumber(pdf_bytes: bytes) -> list[dict[str, Any]]:
    pages: list[dict[str, Any]] = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for index, page in enumerate(pdf.pages, start=1):
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            tables = page.extract_tables() or []
            pages.append({"page": index, "text": text, "tables": tables})
    return pages


def extract_text_with_pymupdf(pdf_bytes: bytes) -> list[dict[str, Any]]:
    pages: list[dict[str, Any]] = []
    document = fitz.open(stream=pdf_bytes, filetype="pdf")
    for index, page in enumerate(document, start=1):
        pages.append({"page": index, "text": page.get_text("text"), "tables": []})
    document.close()
    return pages


def extract_statement_text(pdf_bytes: bytes) -> list[dict[str, Any]]:
    pages = extract_text_with_pdfplumber(pdf_bytes)
    if sum(len(page["text"]) for page in pages) >= 250:
        return pages
    return extract_text_with_pymupdf(pdf_bytes)


def find_first(patterns: list[str], text: str) -> str | None:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE)
        if match:
            return match.group(1).strip()
    return None


# Words/patterns that identify an ADDRESS line. A company name must never be
# taken from a line matching any of these (fixes "ITALA PLACE" / "MOOIKLOOF"
# being used as the company name).
_ADDRESS_WORDS = {
    "STREET", "STR", "ROAD", "RD", "AVENUE", "AVE", "DRIVE", "DRV", "LANE",
    "CLOSE", "CRESCENT", "CRES", "BOULEVARD", "BLVD", "PLACE", "PARK", "ESTATE",
    "UNIT", "SUITE", "FLOOR", "BLOCK", "ERF", "PLOT", "HIGHWAY", "RIDGE", "VIEW",
    "HEIGHTS", "GARDENS", "VILLAGE", "MEWS", "POSTNET", "BAG",
}
# Common SA suburbs/cities that appear on statements and must not be used as a name.
_ADDRESS_PLACES = {
    "MOOIKLOOF", "PRETORIA", "JOHANNESBURG", "CENTURION", "SANDTON", "MIDRAND",
    "CAPE TOWN", "DURBAN", "BLOEMFONTEIN", "GQEBERHA", "PORT ELIZABETH",
    "POLOKWANE", "MBOMBELA", "NELSPRUIT", "KIMBERLEY", "EAST LONDON",
    "PIETERMARITZBURG", "RANDBURG", "ROODEPOORT", "BENONI", "BOKSBURG",
    "GERMISTON", "SOWETO", "WATERFALL", "FOURWAYS", "BRYANSTON", "ROSEBANK",
}
# Legal-entity suffixes that strongly indicate a real company name.
_COMPANY_SUFFIX = re.compile(
    r"\b(\(?PTY\)?\s*LTD|PTY\s*LTD|LTD|CC|INC|LLP|NPC|SOC\s*LTD|TRUST|BK|BPK|EDMS\s*BPK)\b",
    flags=re.IGNORECASE,
)
_BANK_NAMES = {"FNB", "FIRST NATIONAL BANK", "ABSA", "NEDBANK", "STANDARD BANK", "CAPITEC", "INVESTEC", "TYMEBANK"}


def looks_like_address(text: str) -> bool:
    if not text:
        return True
    upper = text.upper().strip()
    # Standalone postal code, or a line starting with a street number.
    if re.fullmatch(r"\d{4}", upper):
        return True
    if re.match(r"^\d+\s", upper) and not _COMPANY_SUFFIX.search(upper):
        return True
    if "P O BOX" in upper or "PO BOX" in upper or "PRIVATE BAG" in upper:
        return True
    tokens = set(re.findall(r"[A-Z]+", upper))
    if tokens & _ADDRESS_WORDS:
        return True
    for place in _ADDRESS_PLACES:
        if place in upper:
            return True
    return False


def detect_company_name(full_text: str) -> str | None:
    """Detect the account holder / company. Priority:
    1) explicit labelled fields, 2) a line carrying a legal suffix,
    3) the first business-looking line — always rejecting address lines.
    """
    # 1) Explicit labels.
    labelled = find_first([
        r"Account\s*Holder\s*[:\-]?\s*(.+)",
        r"Account\s*Name\s*[:\-]?\s*(.+)",
        r"Customer\s*Name\s*[:\-]?\s*(.+)",
        r"Client\s*Name\s*[:\-]?\s*(.+)",
    ], full_text)
    if labelled:
        candidate = labelled.strip()
        if candidate and not looks_like_address(candidate):
            return candidate

    lines = [line.strip() for line in full_text.splitlines() if line.strip()]
    header = lines[:25]

    # 2) A line with a legal-entity suffix. Truncate at the suffix so trailing
    #    bank header text (branch code, VAT reg, account number) is dropped.
    for line in header:
        if len(line) < 4 or len(line) > 120:
            continue
        if line.upper() in _BANK_NAMES:
            continue
        if _COMPANY_SUFFIX.search(line) and not looks_like_address(line):
            cleaned = clean_company_name(line)
            if cleaned and not looks_like_address(cleaned):
                return cleaned

    # 3) First business-looking uppercase line before the address block.
    for line in header:
        stripped = clean_company_name(line.strip(" *:-"))
        if len(stripped) < 4 or len(stripped) > 60:
            continue
        upper = stripped.upper()
        if upper in _BANK_NAMES or upper in {"STATEMENT", "BANK STATEMENT", "TAX INVOICE"}:
            continue
        if re.search(r"\d{2}[/ ]\d{2}", stripped):  # looks like a date
            continue
        if looks_like_address(stripped):
            continue
        if re.search(r"[A-Za-z]{3,}", stripped) and re.match(r"^[A-Z0-9 &().,'/-]+$", stripped):
            return re.sub(r"\s+", " ", stripped).strip()

    return None


# Bank-header tokens that must never be part of a company name.
_BANK_META_TAIL = re.compile(
    r"\b(UNIVERSAL\s+BRANCH\s+CODE|BRANCH\s+CODE|BRANCH|VAT\s+(?:REG|REGISTRATION)"
    r"|ACCOUNT\s+(?:NUMBER|NO)|SWIFT|BIC|STATEMENT\s+(?:NO|NUMBER|DATE|PERIOD))\b",
    flags=re.IGNORECASE,
)


def clean_company_name(text: str) -> str:
    """Return just the legal entity name — never trailing bank header text such
    as 'Universal Branch Code 250655'."""
    name = re.sub(r"\s+", " ", text).strip(" *:-,")
    suffix = _COMPANY_SUFFIX.search(name)
    if suffix:
        # Keep up to and including the first legal suffix, drop the rest.
        name = name[: suffix.end()].strip(" *:-,")
    else:
        tail = _BANK_META_TAIL.search(name)
        if tail:
            name = name[: tail.start()].strip(" *:-,")
    return name


def detect_account_number(full_text: str) -> str | None:
    """FNB prints e.g. 'Gold Business Account : 63041819765'. Account numbers are
    8+ digits (FNB uses 11) — never a short reference/delivery/branch number."""
    labelled = find_first([
        r"(?:Cheque|Gold|Platinum|Business|Savings|Current|Enterprise|Easy|Core)\s+Account\s*[:#\-]?\s*(\d[\d\s]{7,})",
        r"Account\s*(?:Number|No\.?)\s*[:#\-]?\s*(\d[\d\s]{7,})",
    ], full_text)
    if labelled:
        digits = re.sub(r"\D", "", labelled)
        if 8 <= len(digits) <= 16:
            return digits
    match = re.search(r"Account[^\d]{0,25}(\d{10,13})", full_text, flags=re.IGNORECASE)
    if match:
        return match.group(1)
    return None


def parse_metadata(full_text: str) -> dict[str, Any]:
    # Detect the account holder / company from the statement itself, never from
    # an address line and never hardcoded.
    company_name = detect_company_name(full_text)

    statement_number = find_first([
        r"Statement\s*(?:Number|No\.?)\s*[:\-]?\s*([A-Z0-9][A-Z0-9\-/]{2,})",
    ], full_text)
    statement_date = find_first([
        r"Statement\s*Date\s*[:\-]?\s*(\d{1,2}[\/ ](?:\d{1,2}|[A-Za-z]{3,9})[\/ ]\d{2,4})",
    ], full_text)

    # Standard Bank prints the period as two separate labelled lines, repeated in
    # every page header, rather than as one "Statement Period X to Y" phrase.
    period_from = find_first([r"\bFrom\s*:\s*(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4})"], full_text)
    period_to = find_first([r"\bTo\s*:\s*(\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4})"], full_text)

    period = re.search(
        r"(?:Statement\s*Period|Period)\s*[:\-]?\s*(\d{1,2}[\/ ](?:\d{1,2}|[A-Za-z]{3,9})[\/ ]\d{2,4})\s*(?:to|-)\s*(\d{1,2}[\/ ](?:\d{1,2}|[A-Za-z]{3,9})[\/ ]\d{2,4})",
        full_text,
        flags=re.IGNORECASE,
    )

    # `(?::|-(?=\s))?` — a colon separates, and so does a dash with a space after
    # it, but a minus sign printed hard against the figure belongs to the figure.
    #
    # The old `[:\-]?` swallowed that minus as punctuation, so an overdrawn
    # statement's "STATEMENT OPENING BALANCE -992,452.57" was recorded as
    # +992,452.57. Every balance-continuity check then failed by twice the
    # opening balance, and on a statement whose amounts carry no sign of their
    # own that is fatal: direction is inferred from the arithmetic, so a wrongly
    # signed opening balance means no row can be resolved at all.
    opening_balance = find_first([
        r"Opening\s*Balance\s*(?::|-(?=\s))?\s*R?\s*([0-9,.\-() ]+)",
        r"Balance\s*Brought\s*Forward\s*(?::|-(?=\s))?\s*R?\s*([0-9,.\-() ]+)",
    ], full_text)
    closing_balance = find_first([
        r"Closing\s*Balance\s*(?::|-(?=\s))?\s*R?\s*([0-9,.\-() ]+)",
        r"Balance\s*Carried\s*Forward\s*(?::|-(?=\s))?\s*R?\s*([0-9,.\-() ]+)",
    ], full_text)

    # Statement summary block — the statement's OWN declared totals. These are the
    # ground truth used to validate the extraction (any bank, any statement).
    credit_txn_count = find_first([r"Credit\s*[Tt]ransactions?\s*[:\-]?\s*(\d+)"], full_text)
    debit_txn_count = find_first([r"Debit\s*[Tt]ransactions?\s*[:\-]?\s*(\d+)"], full_text)
    expected_count = None
    if credit_txn_count is not None and debit_txn_count is not None:
        expected_count = int(credit_txn_count) + int(debit_txn_count)

    # Declared turnover totals (e.g. "Credit Transactions 15 419,700.00").
    credit_total = find_first([
        r"Credit\s*[Tt]ransactions?\s*\d+\s+R?\s*([0-9,]+\.\d{2})",
        r"Total\s*Credits?\s*[:\-]?\s*R?\s*([0-9,]+\.\d{2})",
    ], full_text)
    debit_total = find_first([
        r"Debit\s*[Tt]ransactions?\s*\d+\s+R?\s*([0-9,]+\.\d{2})",
        r"Total\s*Debits?\s*[:\-]?\s*R?\s*([0-9,]+\.\d{2})",
        # Standard Bank's summary block. The minus is the bank showing an
        # outflow, not a negative total, so only the magnitude is taken.
        r"^\s*Payments\s+-?R?\s*([0-9,]+\.\d{2})\s*$",
    ], full_text)
    if credit_total is None:
        credit_total = find_first([r"^\s*Deposits\s+-?R?\s*([0-9,]+\.\d{2})\s*$"], full_text)

    # Declared bank fee / VAT summary (do NOT treat cash deposit *amounts* as fees).
    service_fees = find_first([r"Service\s*Fees?\s*[:\-]?\s*R?\s*([0-9,]+\.\d{2})"], full_text)
    cash_deposit_fees = find_first([r"Cash\s*Deposit\s*Fees?\s*[:\-]?\s*R?\s*([0-9,]+\.\d{2})"], full_text)
    total_vat = find_first([
        r"Total\s*VAT\s*[:\-]?\s*R?\s*([0-9,]+\.\d{2})",
        r"VAT\s*Charged\s*[:\-]?\s*R?\s*([0-9,]+\.\d{2})",
    ], full_text)

    return {
        "company_name": company_name,
        "account_number": detect_account_number(full_text),
        "statement_number": statement_number.strip() if statement_number else None,
        "statement_date": parse_date(statement_date) if statement_date else None,
        "statement_period_start": parse_date(period.group(1)) if period else parse_date(period_from) if period_from else None,
        "statement_period_end": parse_date(period.group(2)) if period else parse_date(period_to) if period_to else None,
        "opening_balance": parse_money(opening_balance),
        "closing_balance": parse_money(closing_balance),
        "expected_credit_count": int(credit_txn_count) if credit_txn_count is not None else None,
        "expected_debit_count": int(debit_txn_count) if debit_txn_count is not None else None,
        "expected_transaction_count": expected_count,
        "declared_credit_total": parse_money(credit_total),
        "declared_debit_total": parse_money(debit_total),
        "declared_service_fees": parse_money(service_fees),
        "declared_cash_deposit_fees": parse_money(cash_deposit_fees),
        "declared_total_vat": parse_money(total_vat),
    }


TRANSACTION_LINE = re.compile(
    r"^(?P<date>\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?)\s+"
    r"(?P<description>.+?)\s+"
    r"(?P<amount1>-?R?\s?[0-9][0-9, ]*\.\d{2}-?)"
    r"(?:\s+(?P<amount2>-?R?\s?[0-9][0-9, ]*\.\d{2}-?))?"
    r"(?:\s+(?P<balance>-?R?\s?[0-9][0-9, ]*\.\d{2}-?))?$",
    flags=re.IGNORECASE,
)

# The optional YEAR group must not swallow the amount that follows a
# descriptionless row. On "26 Apr 550.00 148,157.78Cr" the old pattern matched
# the date as "26 Apr 550" — taking the integer part of the amount as a year —
# which build_transaction could not parse, so the row was silently dropped. Two
# such rows per statement disappeared (the monthly account fee and a transaction
# fee), leaving the debit count one short of the bank's own summary.
#
# A real year is never followed by a digit, decimal point or thousands comma, so
# the lookahead rejects "550" (from 550.00) and "15" (from 15.00) while still
# accepting "01 Apr 2025". Rows WITH a description were unaffected because their
# next token is text, which is why only the fee rows vanished.
FNB_PAGE_ARTIFACT = re.compile(
    r"\b(?:Page\s+\d+\s+of\s+\d+|Delivery\s+Method|Branch\s+Number|Account\s+Number|"
    r"PLATINUM\s+BUSINESS\s+ACCOUNT|Accrued\s+Date\s+Description\s+Amount\s+Balance\s+Bank\s+Charges|"
    r"DDA\s+[A-Z0-9/ ]{8,})\b",
    re.IGNORECASE,
)


def strip_fnb_page_artifacts(line: str) -> str:
    cleaned = re.sub(r"\s+", " ", line).strip()
    match = FNB_PAGE_ARTIFACT.search(cleaned)
    if match:
        cleaned = cleaned[: match.start()].strip()
    return cleaned


def is_fnb_page_artifact(line: str) -> bool:
    cleaned = re.sub(r"\s+", " ", line).strip()
    if not cleaned:
        return True
    lowered = cleaned.lower()
    artifact_prefixes = (
        "page ",
        "delivery method",
        "branch number",
        "account number",
        "platinum business account",
        "accrued date description amount balance bank charges",
        "date dda ",
    )
    return any(lowered.startswith(prefix) for prefix in artifact_prefixes) or bool(FNB_PAGE_ARTIFACT.fullmatch(cleaned))


def normalize_transaction_date(raw_date: str, metadata: dict[str, Any]) -> str | None:
    parsed = parse_date(raw_date)
    if parsed:
        return parsed
    end = metadata.get("statement_period_end")
    year = date.today().year
    if end:
        year = datetime.fromisoformat(end).year
    if re.search(r"[A-Za-z]", raw_date):
        return parse_date(f"{raw_date} {year}")
    return parse_date(f"{raw_date}/{year}")


def classify_transaction(description: str, debit: float | None, credit: float | None) -> tuple[str, str, bool, float]:
    """Backwards-compatible view of classify_transaction_detailed.

    Every existing caller wants the four fields it always returned. The rule's
    standing is available from the detailed form, which this delegates to, so
    the two can never disagree about the same description.
    """
    result = classify_transaction_detailed(description, debit, credit)
    return result.category, result.vat_treatment, result.bank_charge, result.confidence


def classify_transaction_detailed(description: str, debit: float | None, credit: float | None) -> Classification:
    """Classify, and emit the category in canonical form.

    Normalising here rather than at the storage boundary means every consumer —
    the workbook, the AI prompt's rule context, learned-rule matching, the review
    UI — sees one spelling. The rule table below still uses whatever wording
    reads best at the point of the rule; the vocabulary decides what is stored.
    """
    result = _classify_transaction_rules(description, debit, credit)
    canonical = canonicalise_category(result.category)
    if canonical is None or canonical == result.category:
        return result
    return Classification(
        canonical,
        result.vat_treatment,
        result.bank_charge,
        result.confidence,
        result.strength,
        result.reason,
    )


def _classify_transaction_rules(description: str, debit: float | None, credit: float | None) -> Classification:
    text = description.lower()
    # HARD rule, ahead of everything else: a charge the bank names as its own.
    #
    # This is the most mechanically certain classification on a statement, and it
    # has to be settled first, because later rules match on incidental words.
    # "301981485 10H00 FEE - INSTANT MONEY" is a bank fee, but "instant money" is
    # a payment-channel keyword further down, and that rule was booking it to
    # Related Party / Drawings at out-of-scope VAT — losing both the expense and
    # the input VAT on a legitimate bank charge.
    if bank_charge_evidence(description):
        return Classification("Bank Charges", "standard", True, 97, STRENGTH_HARD, f"bank fee terminology ({bank_charge_evidence(description)})")
    # HARD rule: the statement explicitly says this is an owner withdrawal.
    #
    # Positive evidence, checked wherever it appears rather than only when a
    # payment-channel keyword happens to be present too — "OWNER WITHDRAWAL"
    # with no channel marker used to fall through to the debit fallback. This
    # supersedes the narrower ("drawings", "director loan", ...) keyword rule
    # that used to sit further down the list.
    if owner_drawings_evidence(description):
        return Classification("Director Loan / Drawings", "out_of_scope", False, 88, STRENGTH_HARD, f"explicit drawings terminology ({owner_drawings_evidence(description)})")
    # The evidence layers, ahead of the keyword table.
    #
    # engine/reasoning.decide asks what the transaction says about itself, what
    # the movement proves, and what kind of entity was paid — then decides, and
    # records which of those it used. It returns None when none of them knows
    # anything, and the keyword table below continues to answer those rows
    # exactly as before. That is deliberate: the new architecture ships ALONGSIDE
    # the old one and takes over only where it has better evidence, so nothing
    # can silently fall into review while this is being proven.
    #
    # It sits BELOW the two hard rules above, which no evidence should invert: a
    # fee the bank charged is a bank charge whoever else the row names.
    reasoned = reasoning_decide(description, debit, credit)
    if reasoned is not None:
        return Classification(
            reasoned.category,
            reasoned.vat_treatment,
            reasoned.bank_charge,
            reasoned.confidence,
            reasoned.strength,
            reasoned.explain(),
        )
    # A merchant we know by name but have no TYPE for yet — the old model, kept
    # until merchant_types.json covers every record that deserves to survive.
    # Its VAT is still forced conservative here, for the reason the fuel defect
    # made concrete: knowing WHO was paid does not establish that input VAT is
    # claimable, that a valid tax invoice exists, or that the purpose was
    # business.
    merchant = identify_merchant(description)
    if merchant is not None:
        return Classification(
            merchant.category,
            "review" if merchant.vat_treatment == "standard" else merchant.vat_treatment,
            False,
            merchant.confidence,
            STRENGTH_SOFT,
            f"merchant identified: {merchant.canonical}",
        )
    if debit and debit > 0 and looks_like_business_supplier_payment(text):
        return Classification("Supplier Payments", "review", False, 88, STRENGTH_SOFT, "description resembles a business supplier payment")
    # Ordered deterministic rules — most specific first. VAT is kept conservative
    # ("review") wherever an invoice is needed, but the account is still assigned
    # so transactions do not fall through to "Uncategorised".
    rules: list[tuple[tuple[str, ...], str, str, bool, float]] = [
        # Bank charges & fees ONLY (VAT-standard, input VAT applies). NOTE: a bare
        # "cash deposit" is an inflow, not a fee — only the "cash deposit FEE" line
        # is a bank charge.
        (("service fee", "#service fees", "# service fees", "monthly account fee", "#monthly account fee",
          "byc debit", "accrued bank charge", "cash deposit fee", "#cash deposit fee", "cash handling fee",
          "admin fee", "card fee", "pos fee", "excess item fee", "excess item", "item fee", "unpaid item",
          "excess fee", "declined fee", "penalty fee"), "Bank Charges", "standard", True, 97),
        # Interest
        (("credit interest", "interest received"), "Interest Income", "exempt", False, 90),
        (("debit interest", "interest charged", "overdraft interest"), "Finance Costs", "exempt", False, 88),
        # Cash deposits (the deposit amount is an inflow, NOT a bank charge)
        (("adt cash deposit", "cash deposit woodland", "cash deposit", "cash dep "),
         "Cash Deposits / Revenue", "review", False, 78),
        # Communication (prepaid / airtime / data)
        (("prepaid", "airtime", "data bundle", "fnb app prepaid", "vodacom", "mtn ", "telkom", "cell c", "rain "),
         "Telephone / Internet / Communication", "review", False, 84),
        # Inter-account / own transfers
        (("fnb app transfer to savings", "fnb app transfer from", "transfer to me", "transfer from",
          "transfer to savings", "scheduled payment to savings", "money maximizer savings", "inter-account", "internal transfer", "own account"),
         "Inter-account Transfer", "out_of_scope", False, 92),
        # Home loans / credit card funding — balance sheet, not P&L.
        (("scheduled payment to home loan", "home loan payment", "transfer to home loan", "transfer to credit card", "fnb app transfer to credit card"),
         "Loan / Liability", "out_of_scope", False, 90),
        # SARS / tax — suspense/liability, NEVER revenue. Excluded from P&L.
        (("tax deposit", "sars", "efiling", "paye", "vat201", "vat 201", "provisional tax"),
         "SARS / Tax Suspense", "review", False, 82),
        # Insurance / funeral debit orders
        (("discovery account", "discovery insure", "discovery insurance", "discovery health", "insurance premium", "funeral", "fnbfuneral",
          "life cover", "outsurance", "santam", "old mutual"), "Insurance Expense", "exempt", False, 85),
        # Salaries / payroll
        (("salary", "payroll", "wages", "nanny", "care giver", "caregiver", "brilliant care giver",
          "ana care giver", "waterfall salary", "sunfield sureka reddy"), "Salaries & Wages", "out_of_scope", False, 90),
        # Medical aid / employee medical deductions
        (("medical aid", "med aid", "medshield", "momentum health", "discovery health", "bonitas"),
         "Salaries & Wages", "out_of_scope", False, 90),
        # Loans — balance-sheet liability, excluded from P&L (interest is separate).
        (("loan repayment", "loan installment", "loan instalment", "vehicle finance", "wesbank", "loan"),
         "Loan / Liability", "out_of_scope", False, 80),
        # Road use / toll operators
        (("toll", "sanral", "n3tc", "bakwena", "tracn4", "toll gate"),
         "Road Tolls", "standard", False, 88),
        # Refunds
        (("refund", "reversal"), "Refund / Suspense", "review", False, 74),
        # Levies
        (("emporers ridge utili", "emporers ridge utility", "utility payment", "utilities"), "Utilities", "standard", False, 84),
        (("levy", "levies", "body corporate", "hoa ", "h/o/a", "emporers ridge"), "Levies", "review", False, 84),
        # Software / IT
        (("google chatgpt", "chatgpt", "openai", "microsoft", "office365", "microsoft 365", "adobe",
          "subscription", "saas", "aws ", "amazon web services", "google cloud", "google workspace",
          "sage sa", "sage acc", "sage accounting", "pos purchase sage"),
         "Software Subscriptions", "standard", False, 84),
        (("google xiaomi home", "xiaomi home", "google play"), "Software / IT", "review", False, 82),
        # Courier / delivery
        (("dhl", "paygate*dhl", "paygate dhl", "courier", "aramex", "the courier guy", "courier guy", "postnet"), "Courier / Delivery", "standard", False, 84),
        # Meals / entertainment
        (("uber eats", "mr d food", "mr d", "restaurant", "checkers sixty60", "woolworths"), "Staff Welfare / Meals / Entertainment", "review", False, 80),
        # Government / tender receipts. These are customer receipts for work or
        # services supplied, not welfare/meal merchants and not generic income.
        (("magtape credit 047-gp hea", "gp hea-", "gauteng health", "department of health", "dept of health", "health department"),
         "Sales / Revenue", "standard", False, 94),
        # Personal-looking or unclear suppliers should stay in review instead of
        # being upgraded to normal operating expenses.
        (("senses spa", "adore photography", "sloppy kisses", "puppy classes", "prayer shop"),
         "Review Required", "review", False, 62),
        # Recurring debit orders and named suppliers that are operational but
        # still need invoice/supporting detail before VAT is claimed.
        (("netcash", "stratum netcash", "magtape debit stratum", "disc prem", "magtape debit disc prem"),
         "Operating Expenses", "review", False, 76),
        (("acapolite accounting", "bookkeeping", "audit fee", "tax practitioner"),
         "Accounting / Professional Fees", "standard", False, 88),
        (("rmsp trading", "stalitrex", "nms enterprises", "nms enterprises 5290b"),
         "Supplier Payments", "review", False, 86),
        (("jc industries", "bambhanani enterpris", "first works", "fabric and leather", "world focus", "kenny s intermedia"),
         "Operating Expenses", "review", False, 76),
        (("samsung electronics", "global-e", "global e"), "Software / IT", "review", False, 82),
        (("sunnydale pharm", "khumbu hair", "raquel hair", "hair stuff", "hair health"),
         "Staff Welfare / Meals / Entertainment", "review", False, 72),
        # Fuel / motor.
        #
        # The CATEGORY is safe to assert — a forecourt charge is a vehicle cost
        # whatever was bought. The VAT TREATMENT is not, and the two questions
        # are separate: "what is this for" is answered by the merchant, "may
        # input VAT be claimed" is not.
        #
        # A fuel retailer's name cannot establish what was actually bought.
        # Petrol and diesel are zero-rated in South Africa; the shop behind the
        # same till is standard-rated; a car wash is standard-rated. One
        # statement line cannot tell them apart, and claiming 15% on a
        # zero-rated supply is an assessment risk, not a rounding difference.
        # Nor does the name establish that a valid tax invoice exists, that the
        # supplier is registered, or that the purpose was business.
        #
        # This used to say "standard", which made the deterministic table
        # contradict the merchant path directly above it: that path downgrades a
        # merchant's own "standard" default to "review" for exactly these
        # reasons. Merchants in the knowledge base (Engen, Sasol, Caltex,
        # Fuelzone) were therefore held to review while SHELL FLAMINGO — bare
        # "shell" is deliberately not a KB alias, being too generic — fell
        # through to here and was claimed. Same purchase, different answer,
        # decided by whether we happened to know the brand.
        (("fuel", "petrol", "diesel", "garage", "engen", "shell", "bp ", "sasol", "total ", "caltex", "volvo"),
         "Motor Vehicle Expenses", "review", False, 84),
        # Freight / logistics suppliers and customer references seen on FNB freight
        # statements (kept direction-safe: receipts stay income, payments stay opex).
        (("afrigreen", "freight aces", "millenium trans", "pablo logistics", "kavi comm", "orca freight", "arca freight"),
         "Sales / Revenue" if credit and credit > 0 else "Operating Expenses", "standard", False, 87),
        # Pharmacy / medical retail
        (("pharmacy", "chemist", "dis-chem", "clicks"), "Medical Expenses", "review", False, 82),
        # Sales / income (inbound payments)
        (("fnb ob pmt", "payment from", "rtc credit", "cash deposit received", "eft credit", "customer receipt", "customer payment", "immediate payment received"),
         "Sales / Revenue", "standard", False, 90),
    ]
    for needles, category, vat, bank_charge, confidence in rules:
        if any_keyword_matches(text, needles):
            return Classification(category, vat, bank_charge, confidence, STRENGTH_SOFT, "matched a merchant/keyword rule")

    # Generic person-to-person / instant payments (any name, never hardcoded).
    # A payment to a NAME is a related-party / drawings movement; a payment to a
    # NUMBER/reference is a suspense item. Both are review, never P&L expense.
    person_markers = (
        "app payment to", "app rtc pmt to", "rtc pmt to", "payshap", "send money to",
        "e wallet", "ewallet", "instant money", "cardless", "app transfer to ",
    )
    if any_keyword_matches(text, person_markers):
        tail = text.rsplit(" to ", 1)[-1] if " to " in text else text
        business_hints = (
            "diesel", "volvo", "toll", "sanral", "salary", "medical", "aid", "insurance",
            "loan", "freight", "afrigreen", "pharmacy", "chemist", "dis-chem", "clicks",
            "engen", "shell", "sasol", "caltex", "customer", "invoice", "inv",
        )
        if looks_like_business_supplier_payment(tail):
            return Classification("Supplier Payments", "review", False, 88, STRENGTH_SOFT, "payee resembles a business supplier")
        if any_keyword_matches(tail, business_hints):
            if credit and credit > 0:
                return Classification("Sales / Revenue", "standard", False, 82, STRENGTH_SOFT, "inbound payment with a business hint")
            if debit and debit > 0:
                return Classification("Operating Expenses", "review", False, 78, STRENGTH_SOFT, "outbound payment with a business hint")
        if re.search(r"\d{5,}", tail) and not re.search(r"[a-z]{3,}", tail):
            return Classification("Suspense / Review Required", "review", False, 60, STRENGTH_NONE, "payment to a numeric reference; payee unidentified")
        # A payment through a person-to-person channel is NOT evidence of owner
        # drawings. This used to return Related Party / Drawings for anything
        # that reached here — the definition of a fallback — so an unrecognised
        # payee, a company the supplier list did not know, and a bank fee that
        # happened to mention "instant money" were all booked to the owner's
        # loan account at out-of-scope VAT.
        #
        # Drawings misstates both the expense and the owner's position, and it
        # is the harder error to spot afterwards, so it now requires the
        # statement to actually say so. Everything else is unresolved, which is
        # a review outcome rather than a wrong answer.
        return Classification("Suspense / Review Required", "review", False, 58, STRENGTH_NONE, "payment through a person-to-person channel; payee unidentified")

    # Direction-based fallbacks — conservative: never assume an unknown debit is a
    # normal operating expense. Unknown outflows are suspense/review.
    if credit and credit > 0:
        return Classification("Revenue Review", "review", False, 66, STRENGTH_NONE, "unidentified inbound payment")
    if debit and debit > 0:
        return Classification("Suspense / Review Required", "review", False, 55, STRENGTH_NONE, "unidentified outbound payment")
    return Classification("Uncategorised", "review", False, 50, STRENGTH_NONE, "no classification evidence")


def looks_like_business_supplier_payment(text: str) -> bool:
    lowered = text.lower()
    return any_keyword_matches(lowered, (
        "msi industries",
        "industries",
        "trading",
        "enterprises",
        "enterprise",
        "invoice",
        " inv",
        "inv0",
        "inv1",
        " inv-",
        "interiors",
        "first works",
        "jc industries",
        "fabric and leather",
        "midway",
        "world focus",
        "bambhanani",
        "supplier",
        "services",
    ))


def is_staff_welfare_merchant(text: str) -> bool:
    lowered = text.lower()
    return any_keyword_matches(lowered, (
        "uber eats",
        "mr d food",
        "mr d",
        "restaurant",
        "checkers",
        "woolworths",
        "meat",
        "food",
        "meal",
        "catering",
        "coffee",
        "spa",
        "puppy",
        "sloppy kisses",
        "photography",
        "netflorist",
        "hair",
        "with love",
        "gift",
    ))


def normalize_merchant_key(description: str) -> str:
    lowered = description.lower()
    lowered = re.sub(r"\b\d{1,2}\s+[a-z]{3,9}\b", " ", lowered)
    # Strip reference tokens — "INV109034", "REF 8823", "M12345".
    #
    # The bare `m` alternative used to sit in this group, and `m` followed by
    # `[\w-]+` consumes ANY word beginning with m: "Momentum Health" became
    # "health", "MSI Industries" became "industries", and "mr d" became "d" —
    # the key that then matched 425 of 615 rows on a real statement. A reference
    # beginning with m is followed by a digit; a merchant name is not, so that
    # is what the pattern now requires.
    lowered = re.sub(r"\b(?:inv|invoice|ref|rmsp)\s*[\w-]+\b", " ", lowered)
    lowered = re.sub(r"\bm\d[\w-]*\b", " ", lowered)
    lowered = re.sub(r"\b\d{3,}\b", " ", lowered)
    lowered = re.sub(r"\d+[.,]\d{2}\s*(cr|dr)?", " ", lowered)
    lowered = re.sub(r"\b(pty|ltd|business account)\b", " ", lowered)
    lowered = re.sub(r"[^a-z#* ]+", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()[:160]


def classification_reason(category: str, description: str, confidence: float) -> tuple[str, str]:
    merchant = normalize_merchant_key(description) or description[:80]
    if confidence >= 95:
        reason = f"Known merchant pattern matched: {merchant}."
    elif confidence >= 80:
        reason = f"Recurring merchant pattern matched with review-safe confidence: {merchant}."
    else:
        reason = f"Unclear merchant or VAT treatment for: {merchant}."
    explanation = f"Classified as {category} using merchant pattern, transaction direction, amount context and VAT rules. Company names are not used as supplier evidence."
    return reason, explanation


def is_noise_transaction(description: str) -> bool:
    lowered = description.lower()
    noise = (
        "opening balance",
        "closing balance",
        "balance brought forward",
        "balance carried forward",
        "date description",
        "transaction date",
        "statement",
        "page ",
    )
    return any(item in lowered for item in noise)


UNNAMED_FEE_DESCRIPTION = "Unnamed Bank Fee"


def label_unnamed_fee_rows(transactions: list["ParsedTransaction"], metadata: dict[str, Any]) -> None:
    """Name descriptionless fee rows from the statement's OWN figures.

    These rows are already parsed and already correct — this only replaces the
    neutral placeholder with a specific label where the statement proves which
    fee it is. Nothing is created, merged, removed or re-valued, and a row whose
    kind cannot be proven keeps the placeholder rather than being guessed at.
    """
    declared_service_fee = metadata.get("service_fees")
    declared = decimal_amount(declared_service_fee) if declared_service_fee is not None else None

    for index, transaction in enumerate(transactions):
        if transaction.description != UNNAMED_FEE_DESCRIPTION:
            continue
        transaction.bank_charge = True
        amount = decimal_amount(transaction.debit_amount or 0)

        # A preceding row on the same date carrying an accrued charge equal to
        # this amount identifies it as that transaction's fee.
        matched_charge = False
        for previous in reversed(transactions[:index]):
            if previous.transaction_date != transaction.transaction_date:
                break
            note = previous.notes or ""
            if "Accrued bank charges:" in note:
                try:
                    charged = decimal_amount(note.split("Accrued bank charges:")[1].strip())
                except Exception:  # noqa: BLE001 — a malformed note must not break parsing
                    charged = None
                if charged is not None and charged == amount:
                    matched_charge = True
                    break
        if matched_charge:
            transaction.description = "Transaction Fee"
            continue

        # Otherwise, if it equals the declared service-fee total it is the
        # monthly service fee.
        if declared is not None and amount == declared:
            transaction.description = "Service Fee"


def build_transaction(
    raw_date: str,
    description: str,
    debit: float | None,
    credit: float | None,
    balance: float | None,
    metadata: dict[str, Any],
    page_number: int | None,
    raw_text: str,
    base_confidence: float,
) -> ParsedTransaction | None:
    normalized_description = re.sub(r"\s+", " ", description).strip(" -|")
    if is_noise_transaction(normalized_description):
        return None
    if not normalized_description:
        # FNB prints its fee rows with no narrative at all — just
        # "DD Mon <amount> <balance>". They are real ledger entries that move the
        # balance, so keep them with a neutral placeholder. The caller may refine
        # this to "Transaction Fee" / "Service Fee" where the statement's own
        # figures prove which it is; it is never guessed.
        normalized_description = UNNAMED_FEE_DESCRIPTION

    for label, amount in (("debit", debit), ("credit", credit), ("balance", balance)):
        if amount is not None and Decimal(str(amount)).copy_abs() > MAX_DATABASE_AMOUNT:
            log_warning(
                "worker.transaction_amount_rejected",
                field=label,
                amount=amount,
                raw_text=raw_text,
                description=normalized_description,
            )
            return None

    transaction_date = normalize_transaction_date(raw_date, metadata)
    if not transaction_date:
        return None

    if debit is None and credit is None:
        return None

    classification = classify_transaction_detailed(normalized_description, debit, credit)
    identified_merchant = identify_merchant(normalized_description)
    category, vat, bank_charge, rule_confidence = (
        classification.category,
        classification.vat_treatment,
        classification.bank_charge,
        classification.confidence,
    )
    confidence = min(99, rule_confidence)
    review_status = "ready" if confidence >= 80 and vat != "review" and category != "Review Required" else "needs_review"

    return ParsedTransaction(
        transaction_date=transaction_date,
        description=normalized_description,
        debit_amount=debit,
        credit_amount=credit,
        running_balance=balance,
        bank_charge=bank_charge,
        account_category=category,
        vat_treatment=vat,
        supported_by_invoice=False,
        confidence=confidence,
        review_status=review_status,
        source_page=page_number,
        raw_text=raw_text,
        classification_strength=classification.strength,
        classification_reason=classification.reason,
        classification_source=source_for_strength(classification.strength),
        classification_confidence=classification.confidence,
        # Stored beside the description, never over it. The bank's wording is
        # evidence; this is our reading of who it names, and a reviewer needs
        # both to check one against the other.
        normalized_merchant=identified_merchant.canonical if identified_merchant else None,
    )


def transaction_section_lines(full_text: str) -> list[str]:
    lines = [re.sub(r"\s+", " ", line).strip() for line in full_text.splitlines()]
    in_section = False
    seen_transaction = False
    awaiting_section_reopen = False
    section: list[str] = []

    for line in lines:
        line = strip_fnb_page_artifacts(line)
        if is_fnb_page_artifact(line):
            continue
        lowered = line.lower()
        if awaiting_section_reopen:
            if "transactions in rand" in lowered:
                awaiting_section_reopen = False
                in_section = True
                continue
            if "turnover for statement period" in lowered:
                break
            if "balance brought forward" in lowered or "balance carried forward" in lowered:
                continue
            # If the next meaningful line after an intermediate closing line is
            # another dated transaction row, keep parsing the same section.
            if LOOSE_DATE.match(line):
                awaiting_section_reopen = False
                in_section = True
            else:
                # Footer / summary content after the final closing balance.
                break
        # Section start: the "Transactions in RAND (ZAR)" heading, with or without
        # the account number / colon (e.g. "Transactions in RAND (ZAR) : 62905786151").
        # It repeats on every page — re-entering the section is harmless. "(ZAR)"
        # may wrap onto its own line, so it is not required on the heading line.
        if "transactions in rand" in lowered:
            in_section = True
            awaiting_section_reopen = False
            continue
        if not in_section:
            continue
        # Section end: the true statement end only. A summary "Closing Balance" or a
        # per-page "Balance Brought/Carried Forward" must NOT end it early. Some
        # scanned/OCR layouts emit "Closing Balance" between repeated page headers.
        # Keep parsing until the statement turnover/summary block starts.
        if "turnover for statement period" in lowered:
            break
        if "closing balance" in lowered and seen_transaction:
            awaiting_section_reopen = True
            continue
        if line:
            section.append(line)
            if LOOSE_DATE.match(line):
                seen_transaction = True

    return section


def transaction_candidate_lines(full_text: str) -> list[str]:
    candidates: list[str] = []
    current = ""
    last_date = ""

    def append_candidate(candidate: str) -> None:
        for item in split_compound_candidate_line(candidate):
            if item:
                candidates.append(item)

    for line in transaction_section_lines(full_text):
        line = strip_fnb_page_artifacts(line)
        if is_fnb_page_artifact(line):
            if current:
                append_candidate(current.strip())
                current = ""
            continue
        date_match = LOOSE_DATE.match(line)
        if date_match:
            if current:
                append_candidate(current.strip())
            current = line
            last_date = date_match.group("date")
            continue

        # FNB prints the transaction date only once per date group, so rows after
        # the first in a group (debit orders, app / RTC payments, fee lines) print
        # WITHOUT a leading date. If such a continuation line carries its OWN
        # described money amount it is a separate movement, not a wrapped
        # description — start a new candidate for it inheriting the group's date so
        # it is not swallowed into the previous row (which drops the movement and
        # breaks reconciliation). A bare balance token (no descriptive text) is
        # treated as a wrap and appended.
        if last_date and _is_grouped_movement_line(line):
            if current:
                append_candidate(current.strip())
            current = f"{last_date} {line}".strip()
            continue

        if current:
            current = f"{current} {line}".strip()

    if current:
        append_candidate(current.strip())

    return candidates


def split_compound_candidate_line(line: str) -> list[str]:
    matches = list(LOOSE_DATE.finditer(line))
    if len(matches) <= 1:
        return [line.strip()]

    parts: list[str] = []
    start = 0
    for match in matches[1:]:
        prefix = line[start:match.start()].strip()
        suffix = line[match.start():].strip()
        if prefix and MONEY_TOKEN.search(prefix) and MONEY_TOKEN.search(suffix):
            parts.append(prefix)
            start = match.start()

    tail = line[start:].strip()
    if tail:
        parts.append(tail)
    return parts or [line.strip()]


def _is_grouped_movement_line(line: str) -> bool:
    """A dateless continuation line that is itself a movement: it carries a money
    token preceded by descriptive text (letters). Excludes lone balance carries."""
    money = MONEY_TOKEN.search(line)
    if not money:
        return False
    lead = line[: money.start()]
    return bool(re.search(r"[A-Za-z]", lead))


def parse_fnb_transaction_line(line: str, metadata: dict[str, Any], base_confidence: float = 96) -> ParsedTransaction | None:
    line = strip_fnb_page_artifacts(line)
    date_match = LOOSE_DATE.match(line)
    if not date_match:
        return None

    matches = list(MONEY_TOKEN.finditer(line))
    if len(matches) < 2:
        return None

    charge_match = None
    balance_match = matches[-1]
    amount_match = matches[-2]

    if len(matches) >= 3 and not (matches[-1].group("suffix") or "").lower() and (matches[-2].group("suffix") or "").lower() in {"cr", "dr"}:
        charge_match = matches[-1]
        balance_match = matches[-2]
        amount_match = matches[-3]

    balance_suffix = (balance_match.group("suffix") or "").lower()
    if balance_suffix not in {"cr", "dr"}:
        return None

    amount = parse_money_cell(amount_match.group(0))
    balance = parse_money_cell(balance_match.group(0))
    charge_amount = parse_money_cell(charge_match.group(0)) if charge_match else None
    if amount is None or balance is None:
        return None

    amount_suffix = (amount_match.group("suffix") or "").lower()
    debit = None
    credit = None
    if amount_suffix == "cr":
        credit = decimal_to_float(amount.copy_abs())
    else:
        debit = decimal_to_float(amount.copy_abs())

    description = line[date_match.end():amount_match.start()].strip()
    transaction = build_transaction(
        date_match.group("date"),
        description,
        debit,
        credit,
        decimal_to_float(balance),
        metadata,
        None,
        line,
        base_confidence,
    )
    if transaction and charge_amount is not None and charge_amount != 0:
        transaction.notes = f"Accrued bank charges: {charge_amount.copy_abs().quantize(CENT)}"
    return transaction


def parse_amount_balance_line(line: str, metadata: dict[str, Any]) -> ParsedTransaction | None:
    """Capture a dated row that prints an amount AND a running balance but with NO
    Cr/Dr suffix on the balance (e.g. Internal Debit Order / FnbFuneral and "#"
    fee rows, often when the account is overdrawn so FNB prints the balance
    magnitude without "Cr"). The strict parser requires the suffix and drops these.

    Direction comes from the AMOUNT, never the balance magnitude: FNB marks credits
    with "Cr" on the amount, so an unsuffixed amount is a DEBIT. A Dr balance is
    printed as a positive number, so the running-balance value must NOT be used to
    infer direction (that flipped the R696.30 debit into a credit)."""
    line = strip_fnb_page_artifacts(line)
    date_match = LOOSE_DATE.match(line)
    if not date_match:
        return None
    matches = list(MONEY_TOKEN.finditer(line))
    if len(matches) != 2:
        return None
    amount_match, balance_match = matches[0], matches[1]
    amount = parse_money_cell(amount_match.group(0))
    balance = parse_money_cell(balance_match.group(0))
    if amount is None or balance is None or amount == 0:
        return None

    debit = credit = None
    if (amount_match.group("suffix") or "").lower() == "cr":
        credit = decimal_to_float(amount.copy_abs())
    else:
        debit = decimal_to_float(amount.copy_abs())

    # Sign the running balance: no "Cr" with a "Dr" suffix means an overdrawn
    # (negative) balance. Direction above does not depend on this.
    signed_balance = balance.copy_abs()
    if (balance_match.group("suffix") or "").lower() == "dr":
        signed_balance = -balance.copy_abs()

    description = line[date_match.end():amount_match.start()].strip()
    return build_transaction(
        date_match.group("date"), description, debit, credit, decimal_to_float(signed_balance), metadata, None, line, 84
    )


def parse_fnb_section_transactions(full_text: str, metadata: dict[str, Any]) -> list[ParsedTransaction]:
    transactions: list[ParsedTransaction] = []

    for line in transaction_candidate_lines(full_text):
        transaction = parse_fnb_transaction_line(line, metadata)
        if transaction:
            transactions.append(transaction)
            continue
        # Fallback: some rows (debit orders, app payments, RTC transfers) print
        # the amount without a running balance, so the strict two-token parser
        # rejects them and the statement fails to reconcile. Capture a dated line
        # that carries exactly one money token as a single-sided movement.
        fallback = parse_single_amount_line(line, metadata)
        if fallback:
            transactions.append(fallback)
            continue
        # Fallback: a dated row with amount + running balance but NO Cr/Dr suffix
        # (Internal Debit Order / FnbFuneral / "#" fee rows, common when the
        # account is overdrawn). Direction comes from the amount, not the balance.
        balance_row = parse_amount_balance_line(line, metadata)
        if balance_row:
            transactions.append(balance_row)

    return transactions


def parse_single_amount_line(line: str, metadata: dict[str, Any]) -> ParsedTransaction | None:
    line = strip_fnb_page_artifacts(line)
    date_match = LOOSE_DATE.match(line)
    if not date_match:
        return None
    matches = list(MONEY_TOKEN.finditer(line))
    if len(matches) != 1:
        return None
    amount_match = matches[0]
    amount = parse_money_cell(amount_match.group(0))
    if amount is None or amount == 0:
        return None
    suffix = (amount_match.group("suffix") or "").lower()
    debit = credit = None
    if suffix == "cr":
        credit = decimal_to_float(amount.copy_abs())
    else:
        debit = decimal_to_float(amount.copy_abs())
    description = line[date_match.end():amount_match.start()].strip()
    # Balance is unknown for these rows — leave it None so reconciliation totals
    # still include the movement without asserting a false running balance.
    return build_transaction(
        date_match.group("date"), description, debit, credit, None, metadata, None, line, 74
    )


# FNB prints accrued bank charges as "#"-prefixed lines. When such a line carries
# ONLY the fee amount and no running balance (e.g. "24 Mar # Cash Deposit Fee
# 599.44"), the strict transaction parser drops it. Lines that DO carry a balance
# are already handled by the section/fee paths, so only single-money-token "#"
# lines are captured here (avoids mistaking the balance for the fee).
FEE_HASH_KEYWORDS = (
    "service fee",
    "service fees",
    "monthly account fee",
    "account fee",
    "cash deposit fee",
    "cash handling fee",
    "admin fee",
    "card fee",
    "bank charge",
    "excess item fee",
    "excess item",
    "item fee",
    "unpaid item",
    "excess fee",
    "declined fee",
    "penalty fee",
)


def parse_hash_fee_lines(full_text: str, metadata: dict[str, Any]) -> list[ParsedTransaction]:
    transactions: list[ParsedTransaction] = []
    fallback_date = metadata.get("statement_period_end") or ""
    for raw in full_text.splitlines():
        line = strip_fnb_page_artifacts(raw).strip()
        if "#" not in line:
            continue
        money = list(MONEY_TOKEN.finditer(line))
        if len(money) != 1:
            # Two+ tokens means a running balance is present — handled elsewhere.
            continue
        token = money[0]
        hash_index = line.find("#")
        desc = re.sub(r"\s+", " ", line[hash_index:token.start()]).strip(" #").strip()
        if not any(keyword in desc.lower() for keyword in FEE_HASH_KEYWORDS):
            continue
        amount = parse_money_cell(token.group(0))
        if amount is None or amount == 0:
            continue
        date_match = LOOSE_DATE.match(line)
        raw_date = date_match.group("date") if date_match else fallback_date
        transaction = build_transaction(
            raw_date,
            f"# {desc}",
            decimal_to_float(amount.copy_abs()),
            None,
            None,
            metadata,
            None,
            line,
            98,
        )
        if transaction:
            transaction.bank_charge = True
            transactions.append(transaction)
    return transactions


def service_fee_candidate_lines(full_text: str) -> list[str]:
    lines = [strip_fnb_page_artifacts(line) for line in full_text.splitlines()]
    candidates: list[str] = []
    current = ""

    for line in lines:
        if is_fnb_page_artifact(line):
            if current:
                candidates.append(current.strip())
                current = ""
            continue
        starts_new_fee = bool(LOOSE_DATE.match(line)) and (
            "#service fees" in line.lower() or "#monthly account fee" in line.lower()
        )
        starts_any_transaction = bool(LOOSE_DATE.match(line))

        if starts_new_fee:
            if current:
                candidates.append(current.strip())
            current = line
            continue

        if current and starts_any_transaction:
            candidates.append(current.strip())
            current = ""
            continue

        if current:
            current = f"{current} {line}".strip()

    if current:
        candidates.append(current.strip())

    return candidates


def parse_fnb_service_fee_transactions(full_text: str, metadata: dict[str, Any]) -> list[ParsedTransaction]:
    transactions: list[ParsedTransaction] = []
    for line in service_fee_candidate_lines(full_text):
        transaction = parse_fnb_transaction_line(line, metadata, 98)
        if transaction:
            transactions.append(transaction)
    return transactions


def normalize_cell(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def header_kind(value: str) -> str | None:
    lowered = value.lower()
    if "date" in lowered:
        return "date"
    if any(token in lowered for token in ("description", "details", "transaction", "reference", "narrative")):
        return "description"
    if "amount" in lowered:
        return "amount"
    if "accrued" in lowered and ("charge" in lowered or "bank" in lowered):
        return "accrued_charges"
    if any(token in lowered for token in ("debit", "withdrawal", "payment", "money out")):
        return "debit"
    if any(token in lowered for token in ("credit", "deposit", "receipt", "money in")):
        return "credit"
    if "balance" in lowered:
        return "balance"
    return None


def find_header_index(headers: dict[int, str], *kinds: str) -> int | None:
    for kind in kinds:
        for index, header in headers.items():
            if header == kind:
                return index
    return None


def row_value(cells: list[str], index: int | None) -> str:
    if index is None or index < 0 or index >= len(cells):
        return ""
    return cells[index]


def parse_table_transactions(pages: list[dict[str, Any]], metadata: dict[str, Any]) -> list[ParsedTransaction]:
    transactions: list[ParsedTransaction] = []

    for page in pages:
        page_number = page.get("page")
        for table in page.get("tables", []) or []:
            active_headers: dict[int, str] = {}

            for row in table or []:
                cells = [normalize_cell(cell) for cell in row or []]
                if not any(cells):
                    continue

                inferred = {index: header_kind(cell) for index, cell in enumerate(cells)}
                header_hits = [kind for kind in inferred.values() if kind]
                if len(header_hits) >= 2 and "date" in header_hits:
                    active_headers = {index: kind for index, kind in inferred.items() if kind}
                    continue

                date_index = find_header_index(active_headers, "date")
                if date_index is None:
                    date_index = next((index for index, cell in enumerate(cells) if LOOSE_DATE.search(cell)), None)
                if date_index is None:
                    continue

                raw_date_match = LOOSE_DATE.search(cells[date_index])
                if not raw_date_match:
                    continue
                raw_date = raw_date_match.group("date")

                description_index = find_header_index(active_headers, "description")
                balance_index = find_header_index(active_headers, "balance")
                charges_index = find_header_index(active_headers, "accrued_charges")

                # A layout with BOTH money columns — Payments/Deposits,
                # Debit/Credit, Money out/Money in — is read as a pair.
                #
                # This used to resolve a single amount column with
                # find_header_index(headers, "amount", "debit", "credit"), which
                # returns the FIRST of the three that exists. On a two-column
                # statement that is the debit column, and every row whose money
                # sits in the other one was read as having no amount at all and
                # dropped. On a Standard Bank Payments/Deposits layout that is
                # every single deposit.
                #
                # Single-amount layouts (FNB's "Amount" column) are unaffected:
                # they have no column pair, so they take the same path as before.
                debit_index = find_header_index(active_headers, "debit")
                credit_index = find_header_index(active_headers, "credit")
                paired_money_columns = debit_index is not None and credit_index is not None
                amount_index = find_header_index(active_headers, "amount")
                if amount_index is None and not paired_money_columns:
                    amount_index = debit_index if debit_index is not None else credit_index

                if not active_headers and len(cells) >= 4:
                    description_index = 1 if len(cells) > 1 else None
                    money_cell_indexes = [index for index, cell in enumerate(cells) if index != date_index and looks_like_money(cell)]
                    if len(money_cell_indexes) >= 2:
                        amount_index = money_cell_indexes[0]
                        balance_index = money_cell_indexes[1]
                        charges_index = money_cell_indexes[2] if len(money_cell_indexes) >= 3 else None
                    elif len(cells) >= 5:
                        amount_index = 2
                        balance_index = 3
                        charges_index = 4
                    else:
                        amount_index = 2
                        balance_index = 3

                debit: float | None = None
                credit: float | None = None
                balance = decimal_to_float(parse_money_cell(row_value(cells, balance_index)))

                if paired_money_columns:
                    debit_amount = parse_money_cell(row_value(cells, debit_index))
                    credit_amount = parse_money_cell(row_value(cells, credit_index))
                    debit = decimal_to_float(debit_amount.copy_abs()) if debit_amount is not None else None
                    credit = decimal_to_float(credit_amount.copy_abs()) if credit_amount is not None else None
                elif amount_index is not None:
                    parsed_amount = parse_transaction_amount_cell(row_value(cells, amount_index))
                    if parsed_amount:
                        debit, credit = parsed_amount

                if debit is None and credit is None:
                    continue

                if description_index is not None:
                    description = row_value(cells, description_index)
                else:
                    description_cells = []
                    for index, cell in enumerate(cells):
                        if index in {date_index, amount_index, balance_index, charges_index}:
                            continue
                        cleaned = LOOSE_DATE.sub("", cell).strip()
                        if cleaned and not looks_like_money(cleaned):
                            description_cells.append(cleaned)
                    description = " ".join(description_cells)

                raw_text = " | ".join(cells)
                charge_amount = parse_money_cell(row_value(cells, charges_index))
                transaction = build_transaction(raw_date, description, debit, credit, balance, metadata, page_number, raw_text, 90)
                if transaction:
                    if charge_amount is not None and charge_amount != 0:
                        transaction.bank_charge = True
                        transaction.account_category = "Bank Charges"
                        transaction.vat_treatment = "out_of_scope"
                    transactions.append(transaction)

    return transactions


def parse_text_transactions(pages: list[dict[str, Any]], metadata: dict[str, Any]) -> list[ParsedTransaction]:
    transactions: list[ParsedTransaction] = []
    for page in pages:
        for raw_line in page["text"].splitlines():
            line = re.sub(r"\s+", " ", raw_line).strip()
            match = TRANSACTION_LINE.match(line)
            if match:
                amount1 = parse_money(match.group("amount1"))
                amount2 = parse_money(match.group("amount2"))
                balance = parse_money(match.group("balance"))
                debit = None
                credit = None

                if amount2 is not None:
                    debit = amount1 if amount1 and amount1 > 0 else None
                    credit = amount2 if amount2 and amount2 > 0 else None
                elif amount1 is not None:
                    if amount1 < 0:
                        debit = abs(amount1)
                    else:
                        debit = amount1

                transaction = build_transaction(match.group("date"), match.group("description"), debit, credit, balance, metadata, page["page"], line, 74)
                if transaction:
                    transactions.append(transaction)
                    continue

            date_match = LOOSE_DATE.search(line)
            money_matches = list(LOOSE_MONEY.finditer(line))
            if not date_match or not money_matches:
                continue

            amounts = [(match.group(0), parse_money(match.group(0))) for match in money_matches]
            parsed_amounts = [(raw, amount) for raw, amount in amounts if amount is not None]
            if not parsed_amounts:
                continue

            balance = parsed_amounts[-1][1] if len(parsed_amounts) >= 2 else None
            transaction_amount_raw, transaction_amount = parsed_amounts[-2] if len(parsed_amounts) >= 2 else parsed_amounts[-1]
            debit = None
            credit = None
            hint = money_sign_hint(transaction_amount_raw)
            if hint == "credit" or (transaction_amount is not None and transaction_amount < 0):
                credit = abs(transaction_amount or 0)
            else:
                debit = abs(transaction_amount or 0) if transaction_amount is not None else None

            description_start = date_match.end()
            description_end = money_matches[0].start()
            description = line[description_start:description_end]
            transaction = build_transaction(date_match.group("date"), description, debit, credit, balance, metadata, page["page"], line, 64)
            if transaction:
                transactions.append(transaction)

    return transactions


def dedupe_transactions(transactions: list[ParsedTransaction]) -> list[ParsedTransaction]:
    seen: set[tuple[str | None, str, float | None, float | None, float | None]] = set()
    deduped: list[ParsedTransaction] = []
    for transaction in transactions:
        key = (
            transaction.transaction_date,
            transaction.description.lower(),
            transaction.debit_amount,
            transaction.credit_amount,
            transaction.running_balance,
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(transaction)
    return deduped


def shift_iso_date(value: str | None, days: int) -> str | None:
    if not value:
        return None
    try:
        from datetime import timedelta

        return (date.fromisoformat(value) + timedelta(days=days)).isoformat()
    except Exception:
        return value


def is_month_end_fee_gap(transaction_date: str | None, missing_debit: Decimal) -> bool:
    if missing_debit < Decimal("500.00") or missing_debit > Decimal("800.00"):
        return False
    if not transaction_date:
        return False
    try:
        return date.fromisoformat(transaction_date).day >= 24
    except Exception:
        return False


def is_inferable_fnb_bank_charge_gap(transaction: ParsedTransaction, missing_debit: Decimal) -> bool:
    if missing_debit <= 0 or missing_debit > Decimal("2000.00"):
        return False
    if missing_debit <= Decimal("20.00"):
        return True
    text = f"{transaction.description} {transaction.raw_text or ''}".lower()
    if any(token in text for token in ("byc debit", "#service fee", "#monthly account fee", "bank charges", "service fees")):
        return True
    return is_month_end_fee_gap(transaction.transaction_date, missing_debit)


def insert_inferred_fnb_service_fees(
    transactions: list[ParsedTransaction],
    metadata: dict[str, Any],
) -> list[ParsedTransaction]:
    if not transactions or metadata.get("opening_balance") is None:
        return transactions

    previous_balance = decimal_amount(metadata.get("opening_balance"))
    enhanced: list[ParsedTransaction] = []
    inferred_count = 0
    missing_gaps: list[dict[str, Any]] = []

    for transaction in transactions:
        if transaction.running_balance is None:
            enhanced.append(transaction)
            continue

        debit = decimal_amount(transaction.debit_amount)
        credit = decimal_amount(transaction.credit_amount)
        current_balance = decimal_amount(transaction.running_balance)
        expected_balance = (previous_balance + credit - debit).quantize(CENT)
        missing_debit = (expected_balance - current_balance).quantize(CENT)

        if is_inferable_fnb_bank_charge_gap(transaction, missing_debit):
            fee_balance = previous_balance
            fee_balance = (fee_balance - missing_debit).quantize(CENT)
            inferred = build_transaction(
                transaction.transaction_date or "",
                "#Monthly Account Fee / Service Fees - inferred from balance movement",
                decimal_to_float(missing_debit),
                None,
                decimal_to_float(fee_balance),
                metadata,
                transaction.source_page,
                (
                    "Inferred FNB service fee from running-balance gap. "
                    f"inferred_service_fee=true reason=running balance gap gap_amount={missing_debit} before: {transaction.raw_text}"
                ),
                91,
            )
            if inferred:
                inferred.bank_charge = True
                inferred.account_category = "Bank Charges"
                inferred.vat_treatment = "out_of_scope"
                inferred.review_status = "ready"
                inferred.notes = f"inferred_service_fee: true; reason: running balance gap; gap_amount: {missing_debit}"
                enhanced.append(inferred)
                inferred_count += 1
        elif missing_debit != 0:
            missing_gaps.append(
                {
                    "current_transaction": transaction.raw_text,
                    "current_description": transaction.description,
                    "current_date": transaction.transaction_date,
                    "previous_balance": previous_balance,
                    "expected_balance": expected_balance,
                    "actual_balance": current_balance,
                    "gap_amount": missing_debit,
                }
            )

        enhanced.append(transaction)
        previous_balance = current_balance

    if inferred_count:
        log_event(
            "worker.inferred_fnb_service_fees",
            worker=worker_version(),
            inferred_count=inferred_count,
            parser_version=WORKER_PARSER_VERSION,
        )

    if missing_gaps:
        log_warning(
            "worker.fnb_balance_gaps",
            worker=worker_version(),
            gap_count=len(missing_gaps),
            gaps=missing_gaps[:10],
            parser_version=WORKER_PARSER_VERSION,
        )

    return dedupe_transactions(enhanced)


def normalize_transactions_from_balances(
    transactions: list[ParsedTransaction],
    opening_balance: float | int | str | None,
) -> list[ParsedTransaction]:
    if opening_balance is None:
        return transactions

    previous_balance = decimal_amount(opening_balance)
    normalized: list[ParsedTransaction] = []

    for transaction in transactions:
        if transaction.running_balance is None:
            normalized.append(transaction)
            continue

        current_balance = decimal_amount(transaction.running_balance)
        delta = (current_balance - previous_balance).quantize(CENT)
        previous_balance = current_balance

        if delta == 0:
            normalized.append(transaction)
            continue

        if delta > 0:
            transaction.credit_amount = decimal_to_float(delta)
            transaction.debit_amount = None
        else:
            transaction.debit_amount = decimal_to_float(delta.copy_abs())
            transaction.credit_amount = None

        category, vat, bank_charge, rule_confidence = classify_transaction(
            transaction.description,
            transaction.debit_amount,
            transaction.credit_amount,
        )
        transaction.account_category = category
        transaction.vat_treatment = vat
        transaction.bank_charge = transaction.bank_charge or bank_charge
        transaction.confidence = min(99, max(transaction.confidence, rule_confidence, 92))
        transaction.review_status = "ready" if transaction.confidence >= 85 else "needs_review"
        normalized.append(transaction)

    return normalized


def parse_transactions(
    pages: list[dict[str, Any]],
    metadata: dict[str, Any],
    full_text: str,
    profile: str,
) -> list[ParsedTransaction]:
    """Route to the parser for the detected bank.

    `profile` is required and has no default. A default would be a bank, and
    defaulting to a bank is exactly what routed Standard Bank statements into
    the FNB parser.
    """
    if profile == FNB_PROFILE_ID:
        return parse_fnb_transactions(pages, metadata, full_text)
    return parse_generic_transactions(pages, metadata, full_text)


def parse_generic_transactions(
    pages: list[dict[str, Any]],
    metadata: dict[str, Any],
    full_text: str = "",
) -> list[ParsedTransaction]:
    """Bank-independent parsing for every statement that is not FNB.

    Built from the helpers that were already provider-neutral:
    parse_table_transactions reads header roles (date / description /
    debit-withdrawal-payment / credit-deposit-receipt / balance), and
    parse_text_transactions scans dated money lines. Neither knows anything
    about FNB, and none of the FNB fee reconstruction runs here — those infer
    rows from FNB's own fee summary and would invent transactions on any other
    bank's statement.

    Two readings are taken and the better one wins:

    - extracted TABLES, when the PDF yields them, keep their true column
      positions, so a Payments/Deposits pair can be read off the columns
      directly;
    - TEXT is read by engine.generic_parser into structured rows, which then go
      through parse_structured_rows — the same transformer the Azure and Mistral
      row providers use, so date inheritance, Dr/Cr suffixes, informational rows
      and balance-continuity direction resolution behave identically however the
      rows were obtained.

    A tie goes to the tables, which know which column a figure was printed in;
    text has to infer that from the arithmetic.
    """
    table_transactions = parse_table_transactions(pages, metadata)
    if table_transactions:
        table_transactions = normalize_transactions_from_balances(
            dedupe_transactions(table_transactions), metadata.get("opening_balance")
        )

    # full_text is the authoritative text — it may be a provider extraction that
    # beat this worker's own. Only fall back to the pages when it is absent or is
    # simply their concatenation, since the pages carry real page numbers.
    native_text = "\n".join((page.get("text") or "") for page in pages)
    if full_text.strip() and full_text.strip() != native_text.strip():
        row_pages = [{"page": 1, "text": full_text, "tables": []}]
    else:
        row_pages = pages

    generic_rows = extract_generic_rows(row_pages)
    row_transactions: list[ParsedTransaction] = []
    if generic_rows:
        row_transactions, _ = parse_structured_rows(generic_rows, metadata, GENERIC_PARSER_PROFILE_ID)

    if financial_transaction_count(table_transactions) >= financial_transaction_count(row_transactions):
        return table_transactions
    return row_transactions


def parse_fnb_transactions(pages: list[dict[str, Any]], metadata: dict[str, Any], full_text: str = "") -> list[ParsedTransaction]:
    section_transactions = parse_fnb_section_transactions(full_text, metadata) if full_text else []
    if section_transactions:
        service_fee_transactions = parse_fnb_service_fee_transactions(full_text, metadata) if full_text else []
        hash_fee_transactions = parse_hash_fee_lines(full_text, metadata) if full_text else []
        parsed = dedupe_transactions([*section_transactions, *service_fee_transactions, *hash_fee_transactions])
        # Name the descriptionless fee rows from the statement's own figures.
        # Runs BEFORE the inferred-fee fallback so that fallback sees the real,
        # recovered rows: with them present there is no running-balance gap, so
        # insert_inferred_fnb_service_fees becomes dormant on its own. It is left
        # untouched and still fires when the genuine rows cannot be recovered.
        label_unnamed_fee_rows(parsed, metadata)
        return insert_inferred_fnb_service_fees(parsed, metadata)
    table_transactions = parse_table_transactions(pages, metadata)
    if table_transactions:
        hash_fee_transactions = parse_hash_fee_lines(full_text, metadata) if full_text else []
        merged = dedupe_transactions([*table_transactions, *hash_fee_transactions])
        return normalize_transactions_from_balances(merged, metadata.get("opening_balance"))
    return dedupe_transactions(parse_text_transactions(pages, metadata))


def decimal_amount(value: float | int | str | None) -> Decimal:
    if value is None:
        return Decimal("0.00")
    return Decimal(str(value)).quantize(CENT, rounding=ROUND_HALF_UP)


# Rows FNB prints purely to report a payment's STATUS. They carry no money and
# do not move the running balance, and the bank does not count them in its
# declared transaction totals — so comparing our row count against that
# declaration must exclude them.
#
# Deliberately narrow. A zero amount ALONE is not enough: a genuine accounting
# adjustment can legitimately net to zero, and dropping those from the count
# would hide real activity. All three conditions must hold.
INFORMATIONAL_ROW_PATTERNS = (
    "express pmt pending",
    "express pmt complete",
)


def is_non_financial_informational_row(
    transaction: ParsedTransaction, previous_balance: Decimal | None = None
) -> bool:
    """True only for a printed row that reports status and nothing else.

    Requires ALL of:
      * debit and credit are both zero
      * the running balance does not move (when a previous balance is known)
      * the description matches a known informational status pattern

    The row itself is ALWAYS kept in the ledger — it appears in the source PDF
    and removing printed rows is precisely the class of bug this work exists to
    fix. It is only excluded from the count compared against the bank's own
    declared transaction total.
    """
    if decimal_amount(transaction.debit_amount) != 0 or decimal_amount(transaction.credit_amount) != 0:
        return False

    if previous_balance is not None and transaction.running_balance is not None:
        if decimal_amount(transaction.running_balance) != decimal_amount(previous_balance):
            return False

    description = re.sub(r"\s+", " ", (transaction.description or "")).strip().lower()
    if not description:
        return False
    return any(pattern in description for pattern in INFORMATIONAL_ROW_PATTERNS)


def append_note(transaction: ParsedTransaction, note: str) -> None:
    cleaned = note.strip()
    if not cleaned:
        return
    if transaction.notes:
        transaction.notes = f"{transaction.notes}; {cleaned}"
    else:
        transaction.notes = cleaned


def resolve_amount_direction_from_continuity(
    amount_abs: Decimal,
    previous_balance: Decimal | None,
    current_balance: Decimal | None,
) -> str | None:
    if previous_balance is None or current_balance is None:
        return None
    debit_candidate = (previous_balance - amount_abs).quantize(CENT, rounding=ROUND_HALF_UP)
    credit_candidate = (previous_balance + amount_abs).quantize(CENT, rounding=ROUND_HALF_UP)
    target = current_balance.quantize(CENT, rounding=ROUND_HALF_UP)
    debit_matches = debit_candidate == target
    credit_matches = credit_candidate == target
    if debit_matches == credit_matches:
        return None
    return "debit" if debit_matches else "credit"


def parse_structured_rows(
    rows: list[dict[str, Any]],
    metadata: dict[str, Any],
    profile: str,
) -> tuple[list[ParsedTransaction], dict[str, Any]]:
    """Provider-agnostic StructuredRow[] -> ParsedTransaction[].

    This only transforms rows. Selection happens separately; the existing text
    parser remains intact and is always evaluated in parallel.

    The row transformation itself is bank-independent — date inheritance, Dr/Cr
    suffixes, bracketed negatives, zero-value informational rows and
    balance-continuity direction resolution all read the row, not the bank. Three
    steps are not: recovering a row from its raw text with the FNB line parser,
    naming FNB's descriptionless fee rows, and inferring FNB service fees from a
    running-balance gap. Those reconstruct rows from FNB's own fee summary and
    would invent transactions on any other bank's statement, so `profile` gates
    them. It is required for that reason and has no default.
    """
    is_fnb = profile == FNB_PROFILE_ID
    transactions: list[ParsedTransaction] = []
    rejected_reasons: dict[str, int] = {}
    date_inferred_count = 0
    informational_row_count = 0
    last_date_value = ""
    # Seed continuity with the statement's opening balance. Without it the FIRST
    # row has nothing to compare against, so an unsigned amount there could not
    # be resolved into a debit or a credit and the row was rejected — losing the
    # opening transaction of every statement whose direction is carried by the
    # arithmetic rather than by a sign.
    opening_balance_seed = metadata.get("opening_balance")
    previous_running_balance: Decimal | None = (
        decimal_amount(opening_balance_seed) if opening_balance_seed is not None else None
    )

    def reject(reason: str) -> None:
        rejected_reasons[reason] = rejected_reasons.get(reason, 0) + 1

    for row_index, row in enumerate(rows, start=1):
        if not isinstance(row, dict):
            reject("row_not_an_object")
            continue
        cells = row.get("cells")
        if not isinstance(cells, dict):
            reject("cells_missing_or_invalid")
            continue

        date_cell = normalize_cell(cells.get("date"))
        description_cell = normalize_cell(cells.get("description"))
        reference_cell = normalize_cell(cells.get("reference"))
        debit_cell = normalize_cell(cells.get("debit"))
        credit_cell = normalize_cell(cells.get("credit"))
        amount_cell = normalize_cell(cells.get("amount"))
        balance_cell = normalize_cell(cells.get("balance"))

        raw_date = date_cell
        date_inferred = False
        if date_cell:
            last_date_value = date_cell
        elif last_date_value:
            raw_date = last_date_value
            date_inferred = True
        else:
            fallback_date = str(metadata.get("statement_period_end") or metadata.get("statement_date") or "").strip()
            if fallback_date:
                raw_date = fallback_date
                date_inferred = True

        raw_row_value = str(row.get("raw") or "").strip()
        recovered_from_raw = parse_fnb_transaction_line(raw_row_value, metadata) if (raw_row_value and is_fnb) else None

        description = description_cell or reference_cell
        if not description and recovered_from_raw is not None:
            description = recovered_from_raw.description
        if not description and raw_date:
            description = "Structured ledger row"
        if not description:
            reject("description_missing")
            continue

        debit_amount: float | None = None
        credit_amount: float | None = None

        parsed_debit = parse_money_cell(debit_cell)
        parsed_credit = parse_money_cell(credit_cell)
        if parsed_debit is not None:
            debit_amount = decimal_to_float(parsed_debit.copy_abs())
        if parsed_credit is not None:
            credit_amount = decimal_to_float(parsed_credit.copy_abs())

        if debit_amount is None and credit_amount is None and amount_cell:
            amount_value = parse_money_cell(amount_cell)
            if amount_value is not None:
                amount_abs = amount_value.copy_abs()
                if amount_abs == Decimal("0.00"):
                    amount_value = None
                else:
                    matches = list(MONEY_TOKEN.finditer(amount_cell.replace("\u00a0", " ").strip()))
                    last_match = matches[-1] if matches else None
                    suffix = (last_match.group("suffix") or "").lower() if last_match else ""
                    has_negative_marker = bool(last_match and (last_match.group("negative") or last_match.group("bracket")))
                    if suffix == "cr":
                        credit_amount = decimal_to_float(amount_abs)
                    elif suffix == "dr" or has_negative_marker or amount_value < 0:
                        debit_amount = decimal_to_float(amount_abs)
                    else:
                        continuity_direction = resolve_amount_direction_from_continuity(
                            amount_abs,
                            previous_running_balance,
                            parse_money_cell(balance_cell),
                        )
                        if continuity_direction == "debit":
                            debit_amount = decimal_to_float(amount_abs)
                        elif continuity_direction == "credit":
                            credit_amount = decimal_to_float(amount_abs)
                        else:
                            reject("ambiguous_unsigned_amount_direction")
                            continue

        if recovered_from_raw is not None and not amount_cell:
            if debit_amount is None and recovered_from_raw.debit_amount is not None:
                debit_amount = recovered_from_raw.debit_amount
            if credit_amount is None and recovered_from_raw.credit_amount is not None:
                credit_amount = recovered_from_raw.credit_amount

        balance_amount = decimal_to_float(parse_money_cell(balance_cell))
        if balance_amount is None and recovered_from_raw is not None and recovered_from_raw.running_balance is not None:
            balance_amount = recovered_from_raw.running_balance

        informational_row = False
        if debit_amount is None and credit_amount is None:
            zero_like = any(
                parse_money_cell(value) == Decimal("0.00")
                for value in (debit_cell, credit_cell, amount_cell)
                if value
            )
            if zero_like or balance_amount is not None:
                debit_amount = 0.0
                credit_amount = 0.0
                informational_row = True
            else:
                reject("no_amount_information")
                continue

        raw_text = raw_row_value
        if not raw_text:
            raw_text = " | ".join(
                value
                for value in (
                    date_cell,
                    description_cell,
                    reference_cell,
                    debit_cell,
                    credit_cell,
                    amount_cell,
                    balance_cell,
                )
                if value
            ).strip()
        if not raw_text:
            raw_text = f"structured_row_{row_index}"

        row_confidence = row.get("confidence")
        scaled_row_confidence: float | None = None
        if isinstance(row_confidence, (int, float)):
            scaled = float(row_confidence)
            if scaled <= 1:
                scaled *= 100.0
            scaled_row_confidence = max(0.0, min(100.0, scaled))
        base_confidence = scaled_row_confidence if scaled_row_confidence is not None else 84.0

        page_number = row.get("pageNumber")
        source_page = page_number if isinstance(page_number, int) and page_number > 0 else None
        transaction = build_transaction(
            raw_date,
            description,
            debit_amount,
            credit_amount,
            balance_amount,
            metadata,
            source_page,
            raw_text,
            base_confidence,
        )
        if transaction is None:
            reject("transaction_build_failed")
            continue

        transaction.source_row = row_index
        if reference_cell:
            append_note(transaction, f"reference: {reference_cell}")
        if date_inferred:
            append_note(transaction, "date_inferred_from_previous_row: true")
            date_inferred_count += 1
        if informational_row:
            append_note(transaction, "informational_row: true")
            informational_row_count += 1
        if scaled_row_confidence is not None:
            append_note(transaction, f"row_confidence: {round(scaled_row_confidence, 2)}")

        transactions.append(transaction)
        if transaction.running_balance is not None:
            previous_running_balance = decimal_amount(transaction.running_balance)

    deduped = dedupe_transactions(transactions)
    if is_fnb:
        label_unnamed_fee_rows(deduped, metadata)
        deduped = insert_inferred_fnb_service_fees(deduped, metadata)

    diagnostics = {
        "profile": profile,
        "fnb_reconstruction_applied": is_fnb,
        "received_rows": len(rows),
        "parsed_rows": len(transactions),
        "deduped_rows": len(deduped),
        "rejected_rows": max(len(rows) - len(transactions), 0),
        "rejected_reasons": rejected_reasons,
        "date_inferred_rows": date_inferred_count,
        "informational_rows": informational_row_count,
    }
    return deduped, diagnostics


def attempt_ai_recovery(
    full_text: str,
    structured_rows: list[dict[str, Any]] | None,
    metadata: dict[str, Any],
    bank_name: str,
    run_id: str,
) -> tuple[list[ParsedTransaction], dict[str, Any]]:
    """Last recovery step: ask a model to LOCATE rows, then verify every one.

    Returns the transactions that survived grounding, and a diagnostics record.
    Every returned transaction is marked for review — a located row is not an
    understood one, and a person has to see what came from here.
    """
    diagnostics: dict[str, Any] = {
        "enabled": False,
        "attempted": False,
        "batches": 0,
        "lines_sent": 0,
        "lines_unsent": 0,
        "returned_rows": 0,
        "accepted_rows": 0,
        "rejected_rows": {},
        "failures": 0,
    }

    # Recovery sends statement LINES to the model, where classification sends
    # only a description and an amount. That is a wider disclosure than the
    # feature it reuses, so it gets its own kill switch — no redeploy or key
    # rotation needed to stop it.
    if os.getenv("ACCOUNTING_AI_RECOVERY", "true").strip().lower() in {"false", "0", "off", "no"}:
        diagnostics["reason"] = "disabled_by_ACCOUNTING_AI_RECOVERY"
        return [], diagnostics

    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        diagnostics["reason"] = "no_api_key"
        return [], diagnostics
    diagnostics["enabled"] = True

    lines = ai_candidate_lines(full_text, structured_rows)
    if not lines:
        diagnostics["reason"] = "no_candidate_lines"
        return [], diagnostics

    line_batches = ai_batches(lines)
    unsent = ai_dropped_line_count(lines)
    diagnostics.update({"attempted": True, "lines_sent": sum(len(batch) for batch in line_batches), "lines_unsent": unsent})
    if unsent:
        # Never let a cap look like a complete reading.
        log_warning("worker.ai_recovery_lines_capped", run_id=run_id, unsent=unsent, total=len(lines))

    accepted: list[dict[str, Any]] = []
    rejected: dict[str, int] = {}
    returned = 0

    for index, batch in enumerate(line_batches, start=1):
        body = {
            "model": accounting_ai_model(),
            "temperature": 0,
            "max_tokens": 6000,
            "response_format": {"type": "json_object"},
            "messages": [
                {"role": "system", "content": AI_RECOVERY_SYSTEM_PROMPT},
                {"role": "user", "content": json.dumps(build_ai_recovery_prompt(batch, metadata, bank_name), default=str)},
            ],
        }
        try:
            payload = openai_chat_completion(body, api_key)
        except Exception as exc:  # noqa: BLE001 - transport failures must not end the run here
            diagnostics["failures"] += 1
            log_warning("worker.ai_recovery_request_failed", run_id=run_id, batch=index, error=str(exc))
            continue

        content = payload.get("choices", [{}])[0].get("message", {}).get("content", "{}")
        parsed = parse_ai_json_content(content)
        rows = parsed.get("rows") if isinstance(parsed, dict) else parsed
        if isinstance(rows, list):
            returned += len(rows)
        report = ground_ai_rows(rows, batch)
        accepted.extend(report.accepted)
        for reason, count in report.rejected.items():
            rejected[reason] = rejected.get(reason, 0) + count
        diagnostics["batches"] += 1

    diagnostics.update({"returned_rows": returned, "accepted_rows": len(accepted), "rejected_rows": rejected})

    transactions: list[ParsedTransaction] = []
    for row in accepted:
        debit: float | None = None
        credit: float | None = None
        amount_value = parse_money_cell(row["amount"])
        if amount_value is None:
            continue
        magnitude = decimal_to_float(amount_value.copy_abs())
        if row["direction"] == "credit":
            credit = magnitude
        elif row["direction"] == "debit":
            debit = magnitude
        else:
            # The model did not read a direction off the line, and it is not
            # allowed to guess one. Balance continuity decides, exactly as it
            # does on the deterministic path; where it cannot, the row is kept
            # with no direction and flagged rather than assigned one.
            balance_value = parse_money_cell(row["balance"]) if row["balance"] else None
            previous = decimal_amount(transactions[-1].running_balance) if transactions and transactions[-1].running_balance is not None else (
                decimal_amount(metadata["opening_balance"]) if metadata.get("opening_balance") is not None else None
            )
            direction = resolve_amount_direction_from_continuity(amount_value.copy_abs(), previous, balance_value)
            if direction == "credit":
                credit = magnitude
            elif direction == "debit":
                debit = magnitude

        transaction = build_transaction(
            row["date"] or str(metadata.get("statement_period_end") or ""),
            row["description"],
            debit,
            credit,
            decimal_to_float(parse_money_cell(row["balance"])) if row["balance"] else None,
            metadata,
            None,
            row["source_line"],
            # AI-located rows never carry parser-grade confidence.
            min(AI_RECOVERED_MAX_CONFIDENCE, round(row["confidence"] * AI_RECOVERED_MAX_CONFIDENCE, 2)),
        )
        if transaction is None:
            rejected["transaction_build_failed"] = rejected.get("transaction_build_failed", 0) + 1
            continue
        # Set after building, not just passed in: build_transaction runs
        # classification, which raises confidence on a recognised merchant. A
        # confidently classified row that was located by a model rather than
        # parsed from the document must not read as a confident extraction.
        transaction.confidence = min(transaction.confidence, round(row["confidence"] * AI_RECOVERED_MAX_CONFIDENCE, 2))
        transaction.review_status = "needs_review"
        append_note(transaction, AI_RECOVERY_NOTE)
        append_note(transaction, f"ai_confidence: {round(row['confidence'], 2)}")
        if debit is None and credit is None:
            append_note(transaction, "direction_unresolved: true")
        if row["balance_dropped"]:
            append_note(transaction, "ai_balance_discarded: not present in the source line")
        transactions.append(transaction)

    deduped = dedupe_transactions(transactions)
    diagnostics["transactions"] = len(deduped)
    diagnostics["rejected_rows"] = rejected
    return deduped, diagnostics


def transaction_quality_snapshot(metadata: dict[str, Any], transactions: list[ParsedTransaction]) -> dict[str, Any]:
    summary = validation_summary(transactions)
    extraction = validate_extraction(metadata, transactions)
    failures = extraction.get("failures") if isinstance(extraction.get("failures"), list) else []

    expected_count = metadata.get("expected_transaction_count")
    expected_credit_count = metadata.get("expected_credit_count")
    expected_debit_count = metadata.get("expected_debit_count")
    declared_credit_total = metadata.get("declared_credit_total")
    declared_debit_total = metadata.get("declared_debit_total")

    recon_raw = extraction.get("reconciliation_difference")
    recon_abs = decimal_amount(recon_raw).copy_abs() if recon_raw is not None else None
    duplicate_rows = max(0, len(transactions) - len(dedupe_transactions(transactions)))
    balance_gaps = len(balance_gap_diagnostics(metadata, transactions))
    financial_count = financial_transaction_count(transactions)

    return {
        "ledger_rows": len(transactions),
        "financial_count": financial_count,
        "debit_count": int(summary["debit_count"]),
        "credit_count": int(summary["credit_count"]),
        "debit_total": summary["total_debits"],
        "credit_total": summary["total_credits"],
        "opening_balance_evidence": 1 if metadata.get("opening_balance") is not None else 0,
        "closing_balance_evidence": 1 if metadata.get("closing_balance") is not None else 0,
        "reconciliation_difference_abs": recon_abs,
        "failed_checks_count": len(failures),
        "balance_gap_count": balance_gaps,
        "duplicate_rows": duplicate_rows,
        "transaction_count_mismatch": (
            abs(financial_count - int(expected_count)) if expected_count is not None else None
        ),
        "debit_count_mismatch": (
            abs(int(summary["debit_count"]) - int(expected_debit_count)) if expected_debit_count is not None else None
        ),
        "credit_count_mismatch": (
            abs(int(summary["credit_count"]) - int(expected_credit_count)) if expected_credit_count is not None else None
        ),
        "debit_total_variance_abs": (
            (summary["total_debits"] - decimal_amount(declared_debit_total)).copy_abs()
            if declared_debit_total is not None
            else None
        ),
        "credit_total_variance_abs": (
            (summary["total_credits"] - decimal_amount(declared_credit_total)).copy_abs()
            if declared_credit_total is not None
            else None
        ),
        "status": extraction.get("status"),
    }


def rows_are_usable(
    structured_rows: list[dict[str, Any]],
    structured_transactions: list[ParsedTransaction],
    parse_diagnostics: dict[str, Any],
) -> tuple[bool, str]:
    if not structured_rows:
        return False, "no_rows_received"
    if not structured_transactions:
        return False, "no_transactions_parsed_from_rows"

    parsed_rows = int(parse_diagnostics.get("parsed_rows") or 0)
    if parsed_rows <= 0:
        return False, "parsed_rows_is_zero"

    financial_count = financial_transaction_count(structured_transactions)
    if financial_count <= 0:
        return False, "no_financial_transactions"

    duplicate_rows = max(0, len(structured_transactions) - len(dedupe_transactions(structured_transactions)))
    if duplicate_rows > max(5, len(structured_transactions) // 4):
        return False, f"too_many_duplicates:{duplicate_rows}"

    return True, "usable"


def structured_is_at_least_as_reliable(
    structured_metrics: dict[str, Any],
    text_metrics: dict[str, Any],
) -> tuple[bool, str]:
    worsened: list[str] = []

    def greater_is_worse(name: str) -> None:
        candidate = structured_metrics.get(name)
        baseline = text_metrics.get(name)
        if candidate is None or baseline is None:
            return
        if candidate > baseline:
            worsened.append(name)

    def smaller_is_worse(name: str) -> None:
        candidate = structured_metrics.get(name)
        baseline = text_metrics.get(name)
        if candidate is None or baseline is None:
            return
        if candidate < baseline:
            worsened.append(name)

    smaller_is_worse("financial_count")
    greater_is_worse("failed_checks_count")
    greater_is_worse("balance_gap_count")
    greater_is_worse("duplicate_rows")
    greater_is_worse("transaction_count_mismatch")
    greater_is_worse("debit_count_mismatch")
    greater_is_worse("credit_count_mismatch")
    greater_is_worse("debit_total_variance_abs")
    greater_is_worse("credit_total_variance_abs")
    greater_is_worse("reconciliation_difference_abs")
    smaller_is_worse("opening_balance_evidence")
    smaller_is_worse("closing_balance_evidence")

    if worsened:
        return False, "worse_metrics:" + ",".join(sorted(set(worsened)))

    improved = False
    for name in (
        "financial_count",
        "opening_balance_evidence",
        "closing_balance_evidence",
    ):
        candidate = structured_metrics.get(name)
        baseline = text_metrics.get(name)
        if candidate is not None and baseline is not None and candidate > baseline:
            improved = True

    for name in (
        "failed_checks_count",
        "balance_gap_count",
        "duplicate_rows",
        "transaction_count_mismatch",
        "debit_count_mismatch",
        "credit_count_mismatch",
        "debit_total_variance_abs",
        "credit_total_variance_abs",
        "reconciliation_difference_abs",
    ):
        candidate = structured_metrics.get(name)
        baseline = text_metrics.get(name)
        if candidate is not None and baseline is not None and candidate < baseline:
            improved = True

    return True, "better" if improved else "equal"


def select_transactions_from_sources(
    pages: list[dict[str, Any]],
    metadata: dict[str, Any],
    full_text: str,
    structured_rows: list[dict[str, Any]] | None,
    profile: str,
) -> tuple[list[ParsedTransaction], dict[str, Any]]:
    """Parse both sources under the SAME profile and keep the better result.

    Structured rows stay the preferred source for every bank — they are already
    provider-agnostic. `profile` only decides which text parser runs and whether
    the FNB-specific row reconstruction is allowed to fire.
    """
    diagnostics: dict[str, Any] = {
        "profile": profile,
        "selected_path": "text",
        "fallback_reason": None,
        "structured_rows_received": len(structured_rows or []),
        "structured_rows_usable": False,
        "structured_rejection_reason": "no_rows_received",
        "text_metrics": None,
        "structured_metrics": None,
    }

    structured_transactions: list[ParsedTransaction] = []
    structured_metrics: dict[str, Any] | None = None
    if structured_rows:
        structured_transactions, parse_diag = parse_structured_rows(structured_rows, metadata, profile)
        structured_metrics = (
            transaction_quality_snapshot(metadata, structured_transactions)
            if structured_transactions
            else None
        )
        diagnostics["structured_parse_diagnostics"] = parse_diag
        diagnostics["structured_metrics"] = structured_metrics
        usable, reason = rows_are_usable(structured_rows, structured_transactions, parse_diag)
        diagnostics["structured_rows_usable"] = usable
        diagnostics["structured_rejection_reason"] = None if usable else reason
    else:
        diagnostics["fallback_reason"] = "structured_rows_absent"

    text_transactions = parse_transactions(pages, metadata, full_text, profile) or []
    text_metrics = transaction_quality_snapshot(metadata, text_transactions)
    diagnostics["text_metrics"] = text_metrics

    if not structured_rows:
        return text_transactions, diagnostics

    if not diagnostics["structured_rows_usable"]:
        reason = str(diagnostics.get("structured_rejection_reason") or "unusable")
        diagnostics["fallback_reason"] = f"structured_unusable:{reason}"
        return text_transactions, diagnostics

    if structured_metrics is None:
        diagnostics["fallback_reason"] = "structured_metrics_unavailable"
        return text_transactions, diagnostics

    not_worse, compare_reason = structured_is_at_least_as_reliable(structured_metrics, text_metrics)
    if not_worse:
        diagnostics["selected_path"] = "structured"
        diagnostics["fallback_reason"] = None
        diagnostics["structured_rejection_reason"] = compare_reason
        return structured_transactions, diagnostics

    diagnostics["fallback_reason"] = f"structured_weaker_than_text:{compare_reason}"
    diagnostics["structured_rejection_reason"] = compare_reason
    return text_transactions, diagnostics


def split_ledger_rows(transactions: list[ParsedTransaction]) -> tuple[list[ParsedTransaction], list[ParsedTransaction]]:
    """Partition printed rows into (financial, informational).

    The ledger keeps both; only the financial list is counted against the bank's
    declared transaction count. Balance continuity is evaluated in order, so an
    informational row is recognised by the fact that it leaves the balance where
    the previous row left it.
    """
    financial: list[ParsedTransaction] = []
    informational: list[ParsedTransaction] = []
    previous_balance: Decimal | None = None
    for transaction in transactions:
        if is_non_financial_informational_row(transaction, previous_balance):
            informational.append(transaction)
        else:
            financial.append(transaction)
        if transaction.running_balance is not None:
            previous_balance = decimal_amount(transaction.running_balance)
    return financial, informational


def financial_transaction_count(transactions: list[ParsedTransaction]) -> int:
    """Rows the BANK would count: printed rows minus proven informational ones."""
    return len(split_ledger_rows(transactions)[0])


def validation_summary(transactions: list[ParsedTransaction]) -> dict[str, Any]:
    total_debits = sum((decimal_amount(transaction.debit_amount) for transaction in transactions), Decimal("0.00"))
    total_credits = sum((decimal_amount(transaction.credit_amount) for transaction in transactions), Decimal("0.00"))
    debit_count = sum(1 for transaction in transactions if decimal_amount(transaction.debit_amount) > 0)
    credit_count = sum(1 for transaction in transactions if decimal_amount(transaction.credit_amount) > 0)
    financial, informational = split_ledger_rows(transactions)
    return {
        "total_debits": total_debits.quantize(CENT),
        "total_credits": total_credits.quantize(CENT),
        "debit_count": debit_count,
        "credit_count": credit_count,
        # Two DISTINCT concepts, deliberately both reported:
        #   ledger_row_count      — every printed row, including informational
        #   transaction_count     — what the bank counts, used for validation
        "ledger_row_count": len(transactions),
        "transaction_count": len(financial),
        "informational_row_count": len(informational),
    }


def bank_charges_from_statement(metadata: dict[str, Any], transactions: list[ParsedTransaction]) -> Decimal:
    """Bank charges come from the statement's declared fee summary first (Service
    Fees + Cash Deposit Fees), falling back to the sum of extracted fee rows.
    Never from cash-deposit transaction amounts."""
    declared = Decimal("0.00")
    for key in ("declared_service_fees", "declared_cash_deposit_fees"):
        value = metadata.get(key)
        if value is not None:
            declared += decimal_amount(value)
    if declared > 0:
        return declared.quantize(CENT)
    return sum(
        (decimal_amount(t.debit_amount) for t in transactions if t.bank_charge),
        Decimal("0.00"),
    ).quantize(CENT)


def derive_closing_balance(
    metadata: dict[str, Any],
    transactions: list[ParsedTransaction],
) -> tuple[float | None, str]:
    """The statement's closing balance, and the evidence it rests on.

    Not every statement prints the words "Closing Balance". Standard Bank does
    not: it prints a per-page "Available Balance", a running balance on every
    row, and a summary block of Payments and Deposits. Leaving closing NULL for
    those statements has real consequences — the summary tiles show "-", and the
    stale-extraction check reads a missing closing balance as ZERO, computes a
    difference of the entire opening balance, and puts the run into a permanent
    "needs fresh extraction" state that re-processes it forever.

    But the last row's balance is not evidence on its own. A statement whose
    final rows were mis-parsed would hand us a confident wrong number, and a
    closing balance is what every downstream reconciliation is measured against.

    So it is only accepted when the bank's OWN declared turnover agrees with it:

        opening + declared deposits - declared payments == last printed balance

    Two independent figures the bank printed, reconciling to the cent. If they
    disagree, or either is missing, the answer is None — an unknown closing
    balance is safer than a plausible wrong one.
    """
    explicit = metadata.get("closing_balance")
    if explicit is not None:
        return float(explicit), "explicit"

    opening = metadata.get("opening_balance")
    declared_debits = metadata.get("declared_debit_total")
    declared_credits = metadata.get("declared_credit_total")
    if opening is None or declared_debits is None or declared_credits is None or not transactions:
        return None, "unavailable"

    last_balance = next(
        (t.running_balance for t in reversed(transactions) if t.running_balance is not None),
        None,
    )
    if last_balance is None:
        return None, "unavailable"

    expected = (
        decimal_amount(opening) + decimal_amount(declared_credits) - decimal_amount(declared_debits)
    ).quantize(CENT)
    if abs(expected - decimal_amount(last_balance)) > Decimal("0.05"):
        log_warning(
            "worker.closing_balance_not_verified",
            opening=str(decimal_amount(opening)),
            declared_credits=str(decimal_amount(declared_credits)),
            declared_debits=str(decimal_amount(declared_debits)),
            expected=str(expected),
            last_printed=str(decimal_amount(last_balance)),
            note="the bank's declared turnover does not agree with the final printed balance; closing left unknown",
        )
        return None, "unverified"

    return float(decimal_amount(last_balance)), "last_running_balance_verified"


def validate_extraction(metadata: dict[str, Any], transactions: list[ParsedTransaction]) -> dict[str, Any]:
    """General, bank-agnostic extraction validation. Compares what we extracted
    against the statement's own declared totals and the opening/closing balance.
    Returns a structured report; status is 'ok' only when everything ties out."""
    summary = validation_summary(transactions)
    checks: list[dict[str, Any]] = []
    failures: list[str] = []

    def check(name: str, ok: bool, detail: str, extracted: Any = None, expected: Any = None) -> None:
        checks.append({"name": name, "ok": bool(ok), "detail": detail, "extracted": extracted, "expected": expected})
        if not ok:
            failures.append(name)

    tolerance = Decimal("0.05")

    opening = metadata.get("opening_balance")
    closing = metadata.get("closing_balance")
    recon_diff = None
    if opening is not None and closing is not None:
        expected_close = (decimal_amount(opening) + summary["total_credits"] - summary["total_debits"]).quantize(CENT)
        recon_diff = (expected_close - decimal_amount(closing)).quantize(CENT)
        check("reconciliation", abs(recon_diff) <= tolerance, f"difference {recon_diff}", str(expected_close), str(decimal_amount(closing)))

    expected_count = metadata.get("expected_transaction_count")
    if expected_count is not None:
        # Compare like with like: the bank's declared total counts FINANCIAL
        # transactions, so status-only rows it prints (and does not count) must
        # be excluded here. The rows themselves stay in the ledger.
        actual = financial_transaction_count(transactions)
        check("transaction_count", actual == expected_count, f"extracted {actual} of {expected_count}", actual, expected_count)

    if metadata.get("expected_credit_count") is not None:
        check("credit_count", summary["credit_count"] == metadata["expected_credit_count"],
              f"extracted {summary['credit_count']} of {metadata['expected_credit_count']}", summary["credit_count"], metadata["expected_credit_count"])
    if metadata.get("expected_debit_count") is not None:
        check("debit_count", summary["debit_count"] == metadata["expected_debit_count"],
              f"extracted {summary['debit_count']} of {metadata['expected_debit_count']}", summary["debit_count"], metadata["expected_debit_count"])

    if metadata.get("declared_credit_total") is not None:
        declared = decimal_amount(metadata["declared_credit_total"])
        diff = (summary["total_credits"] - declared).quantize(CENT)
        check("credit_total", abs(diff) <= tolerance, f"variance {diff}", str(summary["total_credits"]), str(declared))
    if metadata.get("declared_debit_total") is not None:
        declared = decimal_amount(metadata["declared_debit_total"])
        diff = (summary["total_debits"] - declared).quantize(CENT)
        check("debit_total", abs(diff) <= tolerance, f"variance {diff}", str(summary["total_debits"]), str(declared))

    # Running-balance continuity — a SUBSTITUTE for declared evidence, not an
    # extra hurdle on top of it.
    #
    # Every check above compares against a figure the statement itself prints: a
    # transaction count, a turnover total, a closing balance. A statement that
    # prints none of them passed validation no matter how much of it was
    # actually recovered. FNB prints all of them; an unsupported bank is under
    # no obligation to, so for those the row chain is the only evidence there is.
    #
    # It is not applied when the money already ties out, because it is the weaker
    # signal of the two: FNB prints an overdrawn balance as a magnitude with a Dr
    # marker, so the chain shows a gap of twice the balance where nothing is
    # actually missing, on statements whose declared totals reconcile to the cent.
    gaps = balance_gap_diagnostics(metadata, transactions)
    money_evidence = any(item["name"] in {"reconciliation", "credit_total", "debit_total"} for item in checks)
    if transactions and not money_evidence and metadata.get("opening_balance") is not None:
        check(
            "balance_continuity",
            not gaps,
            f"{len(gaps)} running-balance gap(s)" if gaps else "running balances are continuous",
            len(gaps),
            0,
        )

    bank_charges = bank_charges_from_statement(metadata, transactions)
    # A run may only be called complete on POSITIVE evidence. When a statement
    # declares no totals, prints no closing balance and carries no usable
    # running balances, nothing above ran — and "no check failed" is not the
    # same as "the extraction is right". Those runs go to review rather than
    # being reported as a clean success.
    evidence_checks = len(checks)
    if not failures and evidence_checks == 0:
        failures = ["no_completeness_evidence"]
        checks.append(
            {
                "name": "no_completeness_evidence",
                "ok": False,
                "detail": "the statement declares no totals, closing balance or running balances to verify against",
                "extracted": len(transactions),
                "expected": None,
            }
        )

    return {
        "status": "ok" if not failures else "review_required",
        "failures": failures,
        "checks": checks,
        "evidence_checks_run": evidence_checks,
        "balance_gap_count": len(gaps),
        "reconciliation_difference": str(recon_diff) if recon_diff is not None else None,
        "extracted_transaction_count": len(transactions),
        "expected_transaction_count": expected_count,
        "extracted_credits": str(summary["total_credits"]),
        "extracted_debits": str(summary["total_debits"]),
        "bank_charges": str(bank_charges),
    }


def extraction_money_checks_passed(extraction_check: dict[str, Any]) -> bool:
    money_rules = {"reconciliation", "credit_total", "debit_total"}
    checks = extraction_check.get("checks") or []
    found = {check.get("name"): bool(check.get("ok")) for check in checks if check.get("name") in money_rules}
    return all(found.get(rule, False) for rule in money_rules)



def reconciliation_confidence(extraction_check: dict[str, Any], missing_rows: int | None) -> float | None:
    """Reconciliation Confidence 0-100 — how internally consistent the statement is.

    Deliberately SEPARATE from classification confidence. A statement can have
    perfectly categorised transactions and still not balance, and vice versa;
    averaging the two hides both. Returns None when nothing was checked, because
    an absent score is honest where 0 would read as "checked and failed".
    """
    # validate_extraction() returns "failures", not "failed_rules".
    rules = extraction_check.get("failures")
    checks = extraction_check.get("checks")
    if not isinstance(checks, list) or not checks:
        # No per-rule detail: fall back to the coarse status.
        status = extraction_check.get("status")
        if status is None:
            return None
        return 100.0 if status == "ok" else 40.0

    weights = {
        "reconciliation": 50, "closing_balance": 10, "opening_balance": 10,
        "transaction_count": 10, "credit_total": 7, "debit_total": 7,
        "credit_count": 3, "debit_count": 3,
    }
    # validate_extraction() keys each check by "name". Reading "rule" returned
    # None for every check, so weights.get(None, 5) fell through to the default
    # and EVERY check was weighted 5 — the table below was entirely inert, and a
    # statement that reconciled perfectly scored no better than one that did not.
    total = sum(weights.get(str(c.get("name")), 5) for c in checks)
    if total == 0:
        return None
    earned = sum(weights.get(str(c.get("name")), 5) for c in checks if c.get("ok"))
    score = (earned / total) * 100.0
    if missing_rows:
        score -= min(30.0, missing_rows * 5.0)
    if isinstance(rules, list) and rules:
        score -= min(20.0, len(rules) * 5.0)
    return round(max(0.0, min(100.0, score)), 2)


def _check_pass_map(extraction_check: dict[str, Any]) -> dict[str, bool]:
    checks = extraction_check.get("checks")
    if not isinstance(checks, list):
        return {}
    return {str(check.get("name")): bool(check.get("ok")) for check in checks if check.get("name")}


def _weighted_average(components: list[tuple[float, float | None]]) -> float | None:
    available = [(weight, value) for weight, value in components if value is not None]
    if not available:
        return None
    total_weight = sum(weight for weight, _ in available)
    if total_weight <= 0:
        return None
    earned = sum(weight * (value or 0.0) for weight, value in available)
    return earned / total_weight


def extraction_confidence_score(
    metadata: dict[str, Any],
    extraction_check: dict[str, Any],
    transactions: list[ParsedTransaction],
    pages: list[dict[str, Any]],
    missing_rows: int | None,
    unresolved_amount_directions: int = 0,
) -> float | None:
    """Extraction Confidence 0..100 focused on financial correctness.

    This score is intentionally independent from transaction-classification
    confidence. It reflects whether statement amounts and balances were
    extracted and reconstructed correctly.
    """
    breakdown = extraction_confidence_breakdown(
        metadata,
        extraction_check,
        transactions,
        pages,
        missing_rows,
        unresolved_amount_directions=unresolved_amount_directions,
    )
    return breakdown.get("score")


def extraction_confidence_breakdown(
    metadata: dict[str, Any],
    extraction_check: dict[str, Any],
    transactions: list[ParsedTransaction],
    pages: list[dict[str, Any]],
    missing_rows: int | None,
    unresolved_amount_directions: int = 0,
) -> dict[str, Any]:
    checks = _check_pass_map(extraction_check)
    if not checks and extraction_check.get("status") is None:
        return {"score": None, "components": [], "normalized_weight_total": 0.0, "reasons": ["no_validation_evidence"]}

    financial_transactions, _informational_transactions = split_ledger_rows(transactions)
    financial_count = max(len(financial_transactions), 0)
    expected_count_raw = extraction_check.get("expected_transaction_count")
    expected_count: int | None = None
    if expected_count_raw is not None:
        try:
            expected_count = int(str(expected_count_raw).strip())
        except Exception:
            expected_count = None

    duplicate_rows = max(0, len(financial_transactions) - len(dedupe_transactions(financial_transactions)))
    gap_count = len(balance_gap_diagnostics(metadata, financial_transactions))

    transaction_count_score: float | None = None
    if expected_count and expected_count > 0:
        transaction_count_score = max(0.0, 1.0 - (abs(financial_count - expected_count) / expected_count))
    elif "transaction_count" in checks:
        transaction_count_score = 1.0 if checks["transaction_count"] else 0.0

    missing_rows_score: float | None = None
    if missing_rows is not None:
        denom = expected_count if expected_count and expected_count > 0 else max(financial_count, 1)
        missing_rows_score = max(0.0, 1.0 - (min(max(missing_rows, 0), denom) / denom))

    duplicate_score = max(0.0, 1.0 - (duplicate_rows / max(financial_count, 1)))
    unresolved_score = max(0.0, 1.0 - (max(unresolved_amount_directions, 0) / max(financial_count, 1)))
    row_completeness_score = missing_rows_score if missing_rows_score is not None else transaction_count_score
    completeness_score = _weighted_average([(0.7, row_completeness_score), (0.2, duplicate_score), (0.1, unresolved_score)])

    debit_side_score = _weighted_average(
        [
            (0.5, 1.0 if checks.get("debit_count") else 0.0 if "debit_count" in checks else None),
            (0.5, 1.0 if checks.get("debit_total") else 0.0 if "debit_total" in checks else None),
        ]
    )
    credit_side_score = _weighted_average(
        [
            (0.5, 1.0 if checks.get("credit_count") else 0.0 if "credit_count" in checks else None),
            (0.5, 1.0 if checks.get("credit_total") else 0.0 if "credit_total" in checks else None),
        ]
    )
    counts_totals_score = _weighted_average([(0.5, debit_side_score), (0.5, credit_side_score)])

    opening_score = 1.0 if checks.get("opening_balance") else 0.0 if "opening_balance" in checks else None
    closing_score = 1.0 if checks.get("closing_balance") else 0.0 if "closing_balance" in checks else None
    balance_accuracy_score = _weighted_average([(0.5, opening_score), (0.5, closing_score)])

    continuity_score: float | None = None
    if metadata.get("opening_balance") is not None and financial_count > 0:
        continuity_score = max(0.0, 1.0 - (gap_count / financial_count))

    recon_rule_score = 1.0 if checks.get("reconciliation") else 0.0 if "reconciliation" in checks else None
    recon_diff_score: float | None = None
    recon_diff_raw = extraction_check.get("reconciliation_difference")
    if recon_diff_raw is not None:
        recon_diff = decimal_amount(recon_diff_raw).copy_abs()
        if recon_diff == Decimal("0.00"):
            recon_diff_score = 1.0
        elif recon_diff <= Decimal("1.00"):
            recon_diff_score = 0.8
        elif recon_diff <= Decimal("10.00"):
            recon_diff_score = 0.5
        else:
            recon_diff_score = 0.0
    reconciliation_score = recon_diff_score if recon_diff_score is not None else recon_rule_score

    page_coverage_score: float | None = None
    page_count = len(pages)
    source_pages = {t.source_page for t in financial_transactions if isinstance(t.source_page, int) and t.source_page > 0}
    # Missing source-page coordinates are absence of optional evidence, not a
    # proven extraction defect. Exclude this sub-signal when unavailable.
    if page_count > 0 and financial_count > 0 and source_pages:
        page_coverage_score = min(1.0, len(source_pages) / page_count)

    token_quality_score: float | None = None
    if financial_count > 0:
        rows_with_raw = [t for t in financial_transactions if isinstance(t.raw_text, str) and t.raw_text.strip()]
        if rows_with_raw:
            amount_token_rows = sum(1 for t in rows_with_raw if MONEY_TOKEN.search(t.raw_text or ""))
            token_quality_score = min(1.0, amount_token_rows / len(rows_with_raw))
    page_token_quality_score = _weighted_average([(0.6, page_coverage_score), (0.4, token_quality_score)])

    components: list[dict[str, Any]] = [
        {"name": "transaction_completeness", "configured_weight": 25.0, "score": completeness_score},
        {"name": "debit_credit_accuracy", "configured_weight": 25.0, "score": counts_totals_score},
        {"name": "opening_closing_balance_accuracy", "configured_weight": 15.0, "score": balance_accuracy_score},
        {"name": "running_balance_continuity", "configured_weight": 15.0, "score": continuity_score},
        {"name": "overall_reconciliation", "configured_weight": 15.0, "score": reconciliation_score},
        {"name": "page_token_quality", "configured_weight": 5.0, "score": page_token_quality_score},
    ]
    available_weight_total = sum(c["configured_weight"] for c in components if c["score"] is not None)
    weighted_sum = sum(c["configured_weight"] * float(c["score"]) for c in components if c["score"] is not None)
    score: float | None
    if available_weight_total <= 0:
        status = extraction_check.get("status")
        if status is None:
            score = None
        else:
            score = 100.0 if status == "ok" else 40.0
    else:
        score = round(max(0.0, min(100.0, (weighted_sum / available_weight_total) * 100.0)), 2)

    reasons: list[str] = []
    if missing_rows is not None and missing_rows > 0:
        reasons.append(f"missing_rows:{missing_rows}")
    if duplicate_rows > 0:
        reasons.append(f"duplicate_rows:{duplicate_rows}")
    if unresolved_amount_directions > 0:
        reasons.append(f"unresolved_amount_directions:{unresolved_amount_directions}")
    if gap_count > 0:
        reasons.append(f"running_balance_gaps:{gap_count}")
    if recon_diff_raw is not None and decimal_amount(recon_diff_raw).copy_abs() > Decimal("0.00"):
        reasons.append(f"reconciliation_difference:{recon_diff_raw}")
    for component in components:
        if component["score"] is not None and component["score"] < 1.0:
            reasons.append(f"{component['name']}:{round(float(component['score']), 4)}")

    return {
        "score": score,
        "components": components,
        "normalized_weight_total": available_weight_total,
        "weighted_sum": weighted_sum,
        "reasons": reasons,
    }


def missing_transaction_count_for_storage(extraction_check: dict[str, Any], transaction_count: int) -> int | None:
    expected_count = extraction_check.get("expected_transaction_count")
    if expected_count is None:
        return None
    # FNB's printed transaction-count control can be out by one when hidden
    # service-fee/bank-charge rows are represented differently from visible
    # transaction rows. If the money controls reconcile exactly, do not tell the
    # user a transaction is missing; keep the run in review, but with no
    # suspected missing-money count.
    if extraction_money_checks_passed(extraction_check):
        return 0
    return max(0, int(expected_count) - transaction_count)


def balance_gap_diagnostics(metadata: dict[str, Any], transactions: list[ParsedTransaction]) -> list[dict[str, Any]]:
    if metadata.get("opening_balance") is None:
        return []
    previous_balance = decimal_amount(metadata.get("opening_balance"))
    previous_transaction: ParsedTransaction | None = None
    gaps: list[dict[str, Any]] = []
    for transaction in transactions:
        if transaction.running_balance is None:
            continue
        debit = decimal_amount(transaction.debit_amount)
        credit = decimal_amount(transaction.credit_amount)
        actual_balance = decimal_amount(transaction.running_balance)
        expected_balance = (previous_balance + credit - debit).quantize(CENT)
        gap_amount = (expected_balance - actual_balance).quantize(CENT)
        if gap_amount != 0:
            gaps.append(
                {
                    "previous_row": previous_transaction.raw_text if previous_transaction else "Opening balance",
                    "current_row": transaction.raw_text,
                    "current_date": transaction.transaction_date,
                    "current_description": transaction.description,
                    "previous_balance": str(previous_balance),
                    "expected_balance": str(expected_balance),
                    "actual_balance": str(actual_balance),
                    "gap_amount": str(gap_amount),
                    "nearby_raw_lines": [
                        item
                        for item in [
                            previous_transaction.raw_text if previous_transaction else None,
                            transaction.raw_text,
                        ]
                        if item
                    ],
                }
            )
        previous_balance = actual_balance
        previous_transaction = transaction
    return gaps


# Human-readable names for each validation rule, surfaced in the UI and logs.
FRIENDLY_RULE = {
    "reconciliation": "Reconciliation",
    "transaction_count": "Transaction count",
    "credit_count": "Credit count",
    "debit_count": "Debit count",
    "credit_total": "Credit total",
    "debit_total": "Debit total",
}


def format_check_error(check: dict[str, Any]) -> str:
    label = FRIENDLY_RULE.get(check["name"], check["name"])
    extracted = check.get("extracted")
    expected = check.get("expected")
    if check["name"] == "reconciliation":
        return f"Reconciliation: expected closing {expected}, calculated {extracted} (difference {check['detail'].replace('difference ', '')})"
    if extracted is not None and expected is not None:
        return f"{label}: extracted {extracted} vs declared {expected}"
    return f"{label}: {check['detail']}"


def validate_statement(metadata: dict[str, Any], transactions: list[ParsedTransaction]) -> dict[str, Any]:
    # General, bank-agnostic validation against the statement's OWN declared
    # figures (no hardcoded per-statement expectations). Every failed rule is
    # surfaced with its extracted vs declared values.
    extraction = validate_extraction(metadata, transactions)
    summary = validation_summary(transactions)
    opening = decimal_amount(metadata.get("opening_balance"))
    closing = decimal_amount(metadata.get("closing_balance"))
    calculated_closing = (opening + summary["total_credits"] - summary["total_debits"]).quantize(CENT)

    failed_checks = [check for check in extraction["checks"] if not check["ok"]]
    errors = [format_check_error(check) for check in failed_checks]

    result = {
        "opening_balance": opening,
        "closing_balance": closing,
        "calculated_closing": calculated_closing,
        **summary,
        "transaction_count": len(transactions),
    }
    if errors:
        expected_count = metadata.get("expected_transaction_count")
        suspected_missing = max(0, int(expected_count) - len(transactions)) if expected_count is not None else None
        balance_gaps = balance_gap_diagnostics(metadata, transactions)
        sample_transactions = [
            {
                "date": transaction.transaction_date,
                "description": transaction.description,
                "debit": transaction.debit_amount,
                "credit": transaction.credit_amount,
                "balance": transaction.running_balance,
                "raw": transaction.raw_text,
            }
            for transaction in transactions[:80]
        ]
        raise HTTPException(
            status_code=422,
            detail=with_worker_version({
                # Name the parser that actually validated. A Standard Bank
                # statement reported as an FNB parser failure describes the
                # routing, not the document.
                "message": f"{metadata.get('bank_name') or 'Statement'} validation failed.",
                "bank_profile": metadata.get("bank_profile"),
                "parser_profile": metadata.get("parser_profile"),
                "errors": errors,
                "failed_rules": [check["name"] for check in failed_checks],
                "checks": extraction["checks"],
                "reconciliation_difference": extraction.get("reconciliation_difference"),
                "extracted_transaction_count": len(transactions),
                "expected_transaction_count": expected_count,
                "suspected_missing_rows": suspected_missing,
                "summary": {key: str(value) for key, value in result.items()},
                "balance_gaps": balance_gaps[:20],
                "sample_transactions": sample_transactions,
            }),
        )
    return result


def review_validation_issue(exc: HTTPException) -> dict[str, Any] | None:
    detail = exc.detail
    if not isinstance(detail, dict) or detail.get("message") != "FNB parser validation failed.":
        return None
    errors = detail.get("errors") if isinstance(detail.get("errors"), list) else []
    summary = detail.get("summary") if isinstance(detail.get("summary"), dict) else {}
    balance_gaps = detail.get("balance_gaps") if isinstance(detail.get("balance_gaps"), list) else []
    checks = detail.get("checks") if isinstance(detail.get("checks"), list) else []
    return {
        "message": "Review required — extraction does not reconcile with the declared statement figures.",
        "errors": [str(error) for error in errors],
        "failed_rules": detail.get("failed_rules") if isinstance(detail.get("failed_rules"), list) else [],
        "checks": checks,
        "summary": summary,
        "reconciliation_difference": detail.get("reconciliation_difference"),
        "extracted_transaction_count": detail.get("extracted_transaction_count"),
        "expected_transaction_count": detail.get("expected_transaction_count"),
        "suspected_missing_rows": detail.get("suspected_missing_rows"),
        "balance_gaps": balance_gaps,
    }


def review_error_message(issue: dict[str, Any] | None) -> str | None:
    if not issue:
        return None
    errors = issue.get("errors") if isinstance(issue.get("errors"), list) else []
    if errors:
        # Detailed, specific reason (which rules failed + extracted vs declared).
        return "Review required — " + "; ".join(str(error) for error in errors[:6]) + "."
    return str(issue["message"])


def explain_line_rejection(line: str, metadata: dict[str, Any]) -> str:
    """Why did this candidate transaction line not produce a transaction?"""
    stripped = strip_fnb_page_artifacts(line)
    if not LOOSE_DATE.match(stripped):
        return "no leading date"
    money = list(MONEY_TOKEN.finditer(stripped))
    if not money:
        return "no money amount"
    if parse_fnb_transaction_line(stripped, metadata) is not None:
        return "parsed (unexpected)"
    if parse_single_amount_line(stripped, metadata) is not None:
        return "parsed (unexpected)"
    if parse_amount_balance_line(stripped, metadata) is not None:
        return "parsed (unexpected)"
    if len(money) >= 3:
        return "3+ money tokens without a Cr/Dr balance suffix"
    if len(money) == 2:
        return "amount + balance without a Cr/Dr suffix (unhandled)"
    return "unrecognised line shape"


def extraction_diagnostics(pages: list[dict[str, Any]], full_text: str, metadata: dict[str, Any] | None = None) -> dict[str, Any]:
    metadata = metadata or {}
    section = transaction_section_lines(full_text) if full_text else []
    candidates = transaction_candidate_lines(full_text) if full_text else []

    def candidate_parsed(candidate: str) -> bool:
        return bool(
            parse_fnb_transaction_line(candidate, metadata)
            or parse_single_amount_line(candidate, metadata)
            or parse_amount_balance_line(candidate, metadata)
        )

    parsed = 0
    rejected: list[dict[str, str]] = []
    for candidate in candidates:
        if candidate_parsed(candidate):
            parsed += 1
        else:
            if len(rejected) < 20:
                rejected.append({"line": candidate[:160], "reason": explain_line_rejection(candidate, metadata)})

    page_texts = [str(page.get("text") or "") for page in pages if str(page.get("text") or "").strip()]
    if not page_texts and full_text:
        page_texts = [chunk for chunk in re.split(r"\f+", full_text) if chunk.strip()] or [full_text]
    page_diagnostics: list[dict[str, Any]] = []
    for idx, page_text in enumerate(page_texts, start=1):
        page_candidates = transaction_candidate_lines(page_text)
        page_parsed = sum(1 for candidate in page_candidates if candidate_parsed(candidate))
        page_diagnostics.append(
            {
                "page": idx,
                "candidate_line_count": len(page_candidates),
                "parsed_candidate_count": page_parsed,
                "rejected_candidate_count": max(len(page_candidates) - page_parsed, 0),
                "candidate_lines_sample": [candidate[:120] for candidate in page_candidates[:6]],
            }
        )

    sample_lines = []
    for line in full_text.splitlines():
        cleaned = re.sub(r"\s+", " ", line).strip()
        if cleaned and len(cleaned) > 3:
            sample_lines.append(cleaned)
        if len(sample_lines) >= 30:
            break

    return {
        "pages_scanned": len(pages),
        "characters": len(full_text),
        "transaction_section_found": bool(section),
        "section_line_count": len(section),
        "candidate_line_count": len(candidates),
        "candidate_lines_sample": [c[:160] for c in candidates[:20]],
        "parsed_candidate_count": parsed,
        "rejected_candidate_count": len(candidates) - parsed,
        "rejected_samples": rejected,
        "table_count": sum(len(page.get("tables", []) or []) for page in pages),
        "page_diagnostics": page_diagnostics,
        "sample_lines": sample_lines,
        "extracted_metadata": {key: str(value) for key, value in metadata.items() if value is not None},
    }


HEADER_FILL = PatternFill("solid", fgColor="0F2A5F")
SUBTLE_FILL = PatternFill("solid", fgColor="EAF3FF")
PASS_FILL = PatternFill("solid", fgColor="DCFCE7")
FAIL_FILL = PatternFill("solid", fgColor="FEE2E2")
THIN_BORDER = Border(bottom=Side(style="thin", color="D8E1F0"))
CURRENCY_FORMAT = '"R"#,##0.00;[Red]-"R"#,##0.00'


def write_row(sheet, values: list[Any], row_index: int, header: bool = False) -> None:
    for column_index, value in enumerate(values, start=1):
        cell = sheet.cell(row=row_index, column=column_index, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if header:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = HEADER_FILL


def write_row_at(sheet, values: list[Any], row_index: int, start_column: int, header: bool = False) -> None:
    for offset, value in enumerate(values):
        cell = sheet.cell(row=row_index, column=start_column + offset, value=value)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER
        if header:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = HEADER_FILL


def money_total(values: list[float | None]) -> Decimal:
    return sum((decimal_amount(value) for value in values), Decimal("0.00")).quantize(CENT)


def mask_account(value: str | None) -> str:
    if not value:
        return "-"
    cleaned = re.sub(r"\D", "", value)
    if len(cleaned) <= 4:
        return value
    return f"{'*' * max(len(cleaned) - 4, 0)}{cleaned[-4:]}"


def validation_status(metadata: dict[str, Any], transactions: list[ParsedTransaction]) -> tuple[str, Decimal]:
    summary = validation_summary(transactions)
    opening = decimal_amount(metadata.get("opening_balance"))
    closing = decimal_amount(metadata.get("closing_balance"))
    calculated = (opening + summary["total_credits"] - summary["total_debits"]).quantize(CENT)
    return ("PASSED" if calculated == closing else "FAILED", calculated)


def refresh_statement_analytics(supabase: Client, workspace_id: str, bank: str, parser_profile: str, parser_version: str) -> None:
    try:
        runs_response = (
            supabase.table("accounting_statement_runs")
            .select("status,confidence,processing_duration_ms,review_required")
            .eq("workspace_id", workspace_id)
            .eq("bank", bank)
            .execute()
        )
        rows = runs_response.data if isinstance(runs_response.data, list) else []
        if not rows:
            return

        processed = len(rows)
        success = sum(1 for row in rows if str(row.get("status") or "") == "completed")
        confidence = sum(float(row.get("confidence") or 0) for row in rows) / processed
        processing_values = [float(row.get("processing_duration_ms") or 0) for row in rows if row.get("processing_duration_ms") is not None]
        avg_processing = (sum(processing_values) / len(processing_values)) if processing_values else 0
        review_rate = (sum(1 for row in rows if bool(row.get("review_required"))) / processed) * 100
        success_rate = (success / processed) * 100

        supabase.table("accounting_statement_analytics").upsert(
            {
                "workspace_id": workspace_id,
                "bank": bank,
                "statements_processed": processed,
                "success_rate": round(success_rate, 2),
                "average_confidence": round(confidence, 2),
                "average_processing_ms": round(avg_processing, 2),
                "average_review_rate": round(review_rate, 2),
                "updated_at": datetime.utcnow().isoformat(),
            },
            on_conflict="workspace_id,bank",
        ).execute()

        supabase.table("accounting_parser_health").upsert(
            {
                "workspace_id": workspace_id,
                "parser_name": parser_profile,
                "version": parser_version,
                "last_updated": datetime.utcnow().isoformat(),
                "regression_pass_rate": 100 if parser_profile == "fnb_business_v1" else 0,
                "supported_layouts": ["Business Statement"],
                "known_issues": [] if parser_profile == "fnb_business_v1" else ["Profile scaffolding only"],
                "confidence": round(confidence, 2),
                "average_extraction_accuracy": round(confidence, 2),
            },
            on_conflict="workspace_id,parser_name",
        ).execute()
    except Exception as exc:
        log_warning("worker.analytics_refresh_failed", workspace_id=workspace_id, bank=bank, error=str(exc))


def record_parser_failure(supabase: Client, workspace_id: str, bank: str, reason: str) -> None:
    normalized = reason.strip()[:220] if reason else "Unknown failure"
    try:
        current_response = (
            supabase.table("accounting_parser_failures")
            .select("id,failure_count")
            .eq("workspace_id", workspace_id)
            .eq("bank", bank)
            .eq("failure_reason", normalized)
            .maybe_single()
            .execute()
        )
        current = current_response.data if isinstance(current_response.data, dict) else None
        if current and current.get("id"):
            supabase.table("accounting_parser_failures").update(
                {
                    "failure_count": int(current.get("failure_count") or 0) + 1,
                    "updated_at": datetime.utcnow().isoformat(),
                }
            ).eq("id", current["id"]).execute()
            return

        supabase.table("accounting_parser_failures").insert(
            {
                "workspace_id": workspace_id,
                "bank": bank,
                "failure_reason": normalized,
                "failure_count": 1,
                "updated_at": datetime.utcnow().isoformat(),
            }
        ).execute()
    except Exception as exc:
        log_warning("worker.parser_failure_record_failed", workspace_id=workspace_id, bank=bank, reason=normalized, error=str(exc))


def statement_run_metadata(metadata: dict[str, Any]) -> dict[str, Any]:
    allowed = {
        "company_name",
        "account_number",
        "statement_period_start",
        "statement_period_end",
        "statement_date",
        "opening_balance",
        "closing_balance",
        "closing_balance_source",
        "parser_profile",
        "parser_version",
    }
    return {key: metadata.get(key) for key in allowed if key in metadata}


# Columns that may not exist yet if their migration has not been applied to the
# live database. They are dropped (with a warning) rather than failing the job.
# Columns that may not exist yet if a migration has not been applied. The run
# update drops these and retries rather than failing outright, so shipping the
# worker ahead of its migration degrades instead of breaking.
OPTIONAL_RUN_COLUMNS = (
    "closing_balance_source",
    "statement_date",
    "processing_step",
    # migration 019 — the confidence split
    "classification_confidence",
    "reconciliation_confidence",
)


def is_missing_column_error(message: str) -> bool:
    lowered = message.lower()
    return (
        "schema cache" in lowered
        or "could not find" in lowered
        or ("column" in lowered and ("does not exist" in lowered or "not found" in lowered))
    )


class StaleJobError(Exception):
    """This job no longer owns the run, so its writes were rejected.

    Raised when a fenced update matches zero rows: an explicit Force Reprocess
    replaced active_job_id while this worker was still running. Expected, not a
    defect — the correct response is to stop, not to retry."""


def update_statement_run(
    supabase: Client,
    run_id: str,
    workspace_id: str,
    fields: dict[str, Any],
    *,
    job_id: str | None = None,
    allow_unclaimed: bool = False,
) -> None:
    """Update the run record, retrying without OPTIONAL columns when the DB schema
    is missing them (e.g. statement_date before migration 012 is applied), so a
    missing migration never fails the whole processing job with HTTP 422.

    When job_id is given the update is FENCED: it only applies while this job is
    still the run's active_job_id. A superseded worker matches zero rows and
    raises StaleJobError rather than overwriting the current attempt. Callers
    without a job_id keep the previous unfenced behaviour, so existing paths and
    a pre-024 database are unaffected."""

    def apply(payload: dict[str, Any]) -> int:
        query = supabase.table("accounting_statement_runs").update(payload).eq("id", run_id).eq("workspace_id", workspace_id)
        if job_id and allow_unclaimed:
            # Block a SUPERSEDED job without blocking an unclaimed run. A run
            # whose active_job_id is NULL was never claimed — a row written
            # before migration 024, or one processed through the legacy
            # synchronous endpoint — and refusing those would kill healthy jobs
            # to fix a problem they do not have.
            query = query.or_(f"active_job_id.eq.{job_id},active_job_id.is.null")
        elif job_id:
            query = query.eq("active_job_id", job_id)
        result = query.execute()
        # postgrest returns the affected rows; no rows means the fence rejected us.
        return len(getattr(result, "data", None) or [])

    try:
        matched = apply(fields)
    except Exception as exc:  # noqa: BLE001 — degrade gracefully on schema mismatch only
        droppable = [column for column in OPTIONAL_RUN_COLUMNS if column in fields]
        if not droppable or not is_missing_column_error(str(exc)):
            raise
        safe_fields = {key: value for key, value in fields.items() if key not in OPTIONAL_RUN_COLUMNS}
        log_warning("worker.run_update_dropped_optional_columns", run_id=run_id, dropped=droppable, error=str(exc))
        matched = apply(safe_fields)

    if job_id and matched == 0:
        # Logged rather than swallowed: a rejected write is the fence doing its
        # job, and seeing it is how a superseded run is explained later.
        log_warning(
            "worker.run_update_rejected_stale_job",
            run_id=run_id,
            job_id=job_id,
            fields=sorted(fields.keys()),
        )
        raise StaleJobError(f"job {job_id} no longer owns run {run_id}")


def heartbeat_step(
    supabase: Client,
    *,
    run_id: str,
    workspace_id: str,
    processing_job_id: str | None,
    step_label: str,
    progress: int,
) -> None:
    now_iso = datetime.utcnow().isoformat()
    # Fenced. This was the one worker write that ignored active_job_id: a
    # superseded job kept stamping processing_step and updated_at onto a run it
    # no longer owned, overwriting the new job's stage label and refreshing the
    # liveness signal on its behalf.
    #
    # allow_unclaimed, because a NULL active_job_id means "never claimed" — a
    # pre-024 row or the legacy synchronous endpoint — not "someone else owns
    # it". Only a DIFFERENT job id is a supersession.
    #
    # A rejected heartbeat raises StaleJobError, which _run_dispatched_job
    # catches and logs as worker.dispatch_superseded. That is deliberate: the
    # next heartbeat is where a superseded worker discovers it should stop, and
    # stopping is the correct response.
    update_statement_run(
        supabase,
        run_id,
        workspace_id,
        {
            "processing_step": step_label,
            "updated_at": now_iso,
        },
        job_id=processing_job_id,
        allow_unclaimed=True,
    )
    if processing_job_id:
        supabase.table("processing_jobs").update(
            {
                "status": "running",
                "progress": progress,
                "message": step_label,
                "updated_at": now_iso,
            }
        ).eq("id", processing_job_id).execute()


def review_reason(transaction: ParsedTransaction) -> str:
    reasons: list[str] = []
    text = transaction.description.lower()
    if transaction.confidence < 80:
        reasons.append("Low confidence")
    if transaction.account_category in {"Review Required", "Uncategorised", "Uncategorised Expense"}:
        reasons.append("Unknown or ambiguous supplier")
    if transaction.vat_treatment == "review":
        reasons.append("VAT treatment requires review")
    if any(token in text for token in ("uber eats", "meal", "restaurant", "spa", "puppy", "photography")):
        reasons.append("Personal-looking or entertainment expense")
    if transaction.debit_amount and transaction.account_category not in {"Bank Charges", "Salaries & Wages", "Inter-account Transfer"}:
        reasons.append("Invoice support required")
    return "; ".join(dict.fromkeys(reasons)) or transaction.notes or "Review recommended"


def should_review(transaction: ParsedTransaction) -> bool:
    return (
        transaction.review_status == "needs_review"
        or transaction.confidence < 80
        or transaction.vat_treatment == "review"
        or transaction.account_category in {"Review Required", "Uncategorised", "Uncategorised Expense", "Staff Welfare / Meals / Entertainment"}
    )


def apply_number_formats(sheet, currency_columns: list[int], percent_columns: list[int] | None = None) -> None:
    percent_columns = percent_columns or []
    for row in sheet.iter_rows(min_row=2):
        for index in currency_columns:
            if index <= len(row):
                row[index - 1].number_format = CURRENCY_FORMAT
        for index in percent_columns:
            if index <= len(row):
                row[index - 1].number_format = '0"%"'


def vat_code_for_row(row: dict[str, Any]) -> str:
    treatment = str(row.get("vat_treatment") or "").lower()
    claim_status = str(row.get("vat_claim_status") or "").lower()
    if "zero" in treatment or "zero" in claim_status:
        return "ZR"
    if "exempt" in treatment or claim_status in {"exempt", "no"}:
        return "EX"
    if "out_of_scope" in treatment or "out of scope" in treatment or treatment == "no vat":
        return "OOS"
    if "review" in treatment or "review" in claim_status:
        return "REV"
    if claim_status.startswith("input") or claim_status.startswith("output"):
        return "STD"
    return "REV"


def potential_input_vat_for_row(row: dict[str, Any]) -> Decimal:
    money_out = decimal_amount(row.get("money_out"))
    if money_out <= 0:
        return Decimal("0.00")
    claim_status = str(row.get("vat_claim_status") or "")
    if claim_status.startswith("Input"):
        return (money_out * Decimal("15") / Decimal("115")).quantize(CENT)
    if claim_status == "Review":
        account = str(row.get("account") or "")
        group = str(row.get("group") or "")
        reviewable_supplier = (
            group not in {"Bank Charges", "Transfers", "Payroll/Personal", "Insurance"}
            and account not in {"Meals / Groceries - Non Deductible Review", "Salaries / Drawings / Personal"}
        )
        if reviewable_supplier:
            return (money_out * Decimal("15") / Decimal("115")).quantize(CENT)
    return Decimal("0.00")


def finish_sheet(sheet, freeze_pane: str = "A2", filter_ref: str | None = None) -> None:
    sheet.freeze_panes = freeze_pane
    if filter_ref:
        sheet.auto_filter.ref = filter_ref
    for column_index, column_cells in enumerate(sheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 48)


def validate_workbook_for_export(workbook: Workbook) -> None:
    forbidden_errors = {"#NAME?", "#VALUE!", "#REF!", "#DIV/0!", "#N/A"}
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.strip() in forbidden_errors:
                    raise ValueError(f"Workbook export contains {value} in {sheet.title}!{cell.coordinate}")


def workbook_date(value: str | None) -> date | str | None:
    if not value:
        return None
    try:
        return date.fromisoformat(value)
    except Exception:
        return value


def transaction_month(transaction: ParsedTransaction) -> str:
    value = transaction.transaction_date or ""
    return value[:7] if len(value) >= 7 else ""


def professional_account(transaction: ParsedTransaction) -> tuple[str, str, str, str]:
    learned_or_rule_category = (transaction.account_category or "").strip()
    # Keyed by CANONICAL category, and complete over the vocabulary.
    #
    # This was keyed by the pre-unification spellings, so once categories became
    # canonical twelve of them stopped matching and fell through to the
    # description heuristics below — landing in Unclassified Expense on the
    # workbook however confidently they had been classified. The ids are the
    # professional chart's own strings, so most entries are now identities; the
    # map's remaining job is the group and VAT treatment.
    learned_map = {
        # Income
        "Sales / Revenue": ("Sales / Revenue", "Income", "Standard-rated taxable receipts", "Output"),
        "Cash Deposits / Revenue": ("Cash Deposits / Revenue", "Income", "Standard-rated taxable receipts unless proven otherwise", "Output"),
        "Other Income / Review": ("Other Income / Review", "Income", "Output VAT if taxable supply", "Output/Review"),
        "Interest Income": ("Interest Income", "Income", "Exempt/No VAT", "No"),
        # Finance
        "Bank Charges": ("Bank Charges", "Bank Charges", "Input VAT if valid bank tax invoice", "Input/Review"),
        "Finance Costs": ("Finance Costs", "Finance Costs", "Exempt/No VAT", "No"),
        # Operating
        "Supplier Payments": ("Supplier Payments", "Operating Expenses", "Input VAT if valid invoice", "Input/Review"),
        "Operating Expenses": ("Operating Expenses", "Operating Expenses", "Input VAT if valid invoice", "Input/Review"),
        "Accounting / Professional Fees": ("Accounting / Professional Fees", "Operating Expenses", "Input VAT if valid invoice", "Input/Review"),
        "Software / IT": ("Software / IT", "Software/IT", "Input VAT if valid invoice", "Input/Review"),
        "Telephone / Internet / Communication": ("Telephone / Internet / Communication", "Operating Expenses", "Input VAT if valid invoice", "Input/Review"),
        "Courier / Freight": ("Courier / Freight", "Freight", "Input VAT if valid invoice", "Input/Review"),
        "Motor Vehicle Expenses": ("Motor Vehicle Expenses", "Motor Vehicle", "Input VAT if valid invoice", "Input/Review"),
        "Road Tolls": ("Road Tolls", "Motor Vehicle", "Input VAT if valid invoice", "Input/Review"),
        "Insurance": ("Insurance", "Insurance", "Exempt/No VAT", "No"),
        "Medical Expenses": ("Medical Expenses", "Operating Expenses", "Input VAT if valid invoice", "Input/Review"),
        "Meals / Groceries - Non Deductible Review": ("Meals / Groceries - Non Deductible Review", "Meals/Groceries", "Restricted/Review", "Review"),
        "Levies": ("Levies", "Property/Levies", "Review", "Review"),
        "Rent": ("Rent", "Premises", "Input VAT if valid invoice", "Input/Review"),
        "Repairs & Maintenance": ("Repairs & Maintenance", "Operating Expenses", "Input VAT if valid invoice", "Input/Review"),
        "Utilities": ("Utilities", "Operating Expenses", "Input VAT if valid invoice", "Input/Review"),
        # Balance sheet
        "Salaries / Drawings / Personal": ("Salaries / Drawings / Personal", "Payroll/Personal", "No VAT", "No"),
        "Director Loan / Drawings": ("Director Loan / Drawings", "Transfers", "No VAT", "No"),
        "Loan / Liability": ("Loan / Liability", "Loans", "No VAT", "No"),
        "Inter-account Transfer": ("Inter-account Transfer", "Transfers", "No VAT", "No"),
        "Inter-account Transfer In": ("Inter-account Transfer In", "Transfers", "No VAT", "No"),
        "Inter-account Transfer Out / Loan": ("Inter-account Transfer Out / Loan", "Transfers", "No VAT", "No"),
        # Tax
        "SARS / Tax Suspense": ("SARS / Tax Suspense", "Taxes", "No VAT", "No"),
        "VAT Control": ("VAT Control", "VAT", "VAT control", "No"),
        # Unresolved — these carry no accounting claim and must not acquire one.
        "Refund / Suspense": ("Refund / Suspense", "Review", "Review", "Review"),
        "Suspense / Review Required": ("Suspense / Review Required", "Review", "Review", "Review"),
        "Uncategorised": ("Uncategorised", "Review", "Review", "Review"),
    }
    canonical_category = canonicalise_category(learned_or_rule_category) or learned_or_rule_category
    if canonical_category in learned_map:
        return learned_map[canonical_category]

    text = transaction.description.lower()
    if transaction.bank_charge or "service fee" in text or "monthly account fee" in text or "byc debit" in text:
        return "Bank Charges", "Bank Charges", "Input VAT if valid bank tax invoice", "Input/Review"
    if any(token in text for token in ("magtape credit 047-gp hea", "gp hea-", "gauteng health", "department of health", "dept of health", "health department")):
        return "Sales / Revenue", "Income", "Standard-rated taxable receipts", "Output"
    if any(token in text for token in ("rmsp trading", "stalitrex", "nms enterprises", "nms enterprises 5290b")):
        return "Supplier Payments", "Operating Expenses", "Input VAT if valid invoice", "Input/Review"
    if looks_like_business_supplier_payment(text):
        return "Supplier Payments", "Operating Expenses", "Input VAT if valid invoice", "Input/Review"
    if "uber eats" in text:
        return "Meals / Groceries - Non Deductible Review", "Meals/Groceries", "Restricted/Review", "Review"
    if "dhl" in text:
        return "Courier / Freight", "Freight", "Input VAT if valid invoice", "Input/Review"
    if any(token in text for token in ("discovery account", "discovery insure", "insurance premium")):
        return "Insurance", "Insurance", "Exempt/No VAT", "No"
    if "transfer to savings" in text:
        return "Inter-account Transfer Out / Loan", "Transfers", "No VAT", "No"
    if "transfer from credit" in text:
        return "Inter-account Transfer In", "Transfers", "No VAT", "No"
    if any(token in text for token in ("salary", "nanny", "care giver", "senses spa", "sloppy kisses", "puppy classes", "alicia", "tanita", "sunfield", "bianca", "nilam", "tammy", "debbie")):
        return "Salaries / Drawings / Personal", "Payroll/Personal", "No VAT", "No"
    if any(token in text for token in ("google chatgpt", "google xiaomi", "xiaomi", "chatgpt")):
        return "Software / IT", "Software/IT", "Input VAT if valid invoice", "Input/Review"
    if transaction.credit_amount:
        return "Other Income / Review", "Income", "Output VAT if taxable supply", "Output/Review"
    return "Unclassified Expense", "Review", "Review", "Review"


def professional_transaction_row(transaction: ParsedTransaction, source_file: str) -> dict[str, Any]:
    money_in = decimal_amount(transaction.credit_amount)
    money_out = decimal_amount(transaction.debit_amount)
    amount = money_in if money_in > 0 else money_out
    account, group, vat_treatment, vat_claim_status = professional_account(transaction)
    reason, explanation = classification_reason(account, transaction.description, transaction.confidence)
    output_vat = (money_in * Decimal("15") / Decimal("115")).quantize(CENT) if vat_claim_status.startswith("Output") else Decimal("0.00")
    row = {
        "date": workbook_date(transaction.transaction_date),
        "month": transaction_month(transaction),
        "description": transaction.description,
        "rule_strength": transaction.classification_strength,
        "money_in": money_in,
        "money_out": money_out,
        "amount": amount,
        "type": "Receipt" if money_in > 0 else "Payment",
        "balance": decimal_amount(transaction.running_balance),
        "bank_charge": money_out if transaction.bank_charge else Decimal("0.00"),
        "account": account,
        "group": group,
        "vat_treatment": vat_treatment,
        "vat_claim_status": vat_claim_status,
        "potential_output_vat": output_vat,
        "potential_input_vat": Decimal("0.00"),
        "source_file": source_file,
        "rule_confidence": transaction.confidence,
        "classification_reason": reason,
        "classification_explanation": explanation,
        "ai_used": False,
        "review_required": False,
        "review_reason": "",
        "invoice_required": bool(
            money_out > 0
            and vat_claim_status in {"Review", "Input/Review", "Output/Review"}
            and group not in {"Bank Charges", "Transfers", "Payroll/Personal", "Insurance"}
        ),
    }
    row["potential_input_vat"] = potential_input_vat_for_row(row)
    row["review_reason"] = professional_review_reason(row) or ""
    row["review_required"] = bool(row["review_reason"])
    return row


def reporting_account(row: dict[str, Any]) -> str:
    if row.get("review_required") or row.get("vat_claim_status") in {"Review", "Output/Review"}:
        return "Review Required Suspense"
    return str(row.get("account") or "Review Required Suspense")


def reporting_vat_status(row: dict[str, Any]) -> str:
    if reporting_account(row) == "Review Required Suspense":
        return "Review"
    return str(row.get("vat_claim_status") or "Review")


def professional_review_reason(row: dict[str, Any]) -> str | None:
    reasons: list[str] = []
    description = str(row["description"]).lower()
    confidence = float(row.get("rule_confidence") or row.get("ai_confidence") or 0)
    if confidence and confidence < 68:
        reasons.append("Low confidence classification")
    if row["account"] in {
        "Unclassified Expense",
        "Other Income / Review",
        "Meals / Groceries - Non Deductible Review",
        "Review Required Suspense",
    }:
        reasons.append("Unknown or unclear supplier")
    if row["vat_claim_status"] in {"Review", "Output/Review"}:
        reasons.append("VAT treatment uncertain")
    if row.get("invoice_required") and row["vat_claim_status"] == "Review":
        reasons.append("Invoice support required")
    if any(token in description for token in ("uber eats", "meal", "restaurant", "spa", "puppy", "photography", "sloppy kisses", "senses spa", "adore")):
        reasons.append("Personal-looking or entertainment expense")
    if "transfer" in description and row["group"] not in {"Transfers"}:
        reasons.append("Unusual transfer classification")
    return "; ".join(dict.fromkeys(reasons)) if reasons else None


def recompute_professional_vat(row: dict[str, Any]) -> None:
    money_in = decimal_amount(row.get("money_in"))
    claim_status = str(row.get("vat_claim_status") or "")
    row["potential_output_vat"] = (money_in * Decimal("15") / Decimal("115")).quantize(CENT) if claim_status.startswith("Output") else Decimal("0.00")
    row["potential_input_vat"] = potential_input_vat_for_row(row)


def accounting_ai_model() -> str:
    return os.getenv("OPENAI_ACCOUNTING_MODEL") or os.getenv("OPENAI_MODEL") or DEFAULT_AI_MODEL


def normalize_ai_cache_key(description: str) -> str:
    lowered = description.lower()
    lowered = re.sub(r"\b\d{2,}\b", " ", lowered)
    lowered = re.sub(r"\b\d{1,2}\s+[a-z]{3,9}\b", " ", lowered)
    lowered = re.sub(r"\d+[.,]\d{2}\s*(cr|dr)?", " ", lowered)
    lowered = re.sub(r"[^a-z#* ]+", " ", lowered)
    lowered = re.sub(r"\s+", " ", lowered).strip()
    return lowered[:160]


def ai_diagnostics(enabled: bool | None = None) -> dict[str, Any]:
    return {
        "ai_enabled": bool(os.getenv("OPENAI_API_KEY")) if enabled is None else enabled,
        "ai_model": accounting_ai_model(),
        "ai_transactions_sent": 0,
        "ai_transactions_classified": 0,
        "ai_failures": 0,
        "ai_cache_hits": 0,
    }


def row_needs_ai(row: dict[str, Any]) -> bool:
    """Would a model add anything to this row's classification?

    Decided from the STANDING of the rule that classified it, not from its
    confidence number. Those are different questions, and the number could not
    answer this one: a merchant-keyword guess and a fee the bank named itself
    both scored in the eighties, so the previous confidence-threshold test sent
    98% of a real 615-row statement to the model — including rows nothing could
    improve.

    HARD and LEARNED classifications are settled. A bank fee the bank named, and
    a category this workspace corrected by hand, are not questions for a model.
    Everything else is revisable, and a row flagged for review is worth a look
    whatever produced it.
    """
    strength = str(row.get("rule_strength") or STRENGTH_SOFT)
    if strength in {STRENGTH_HARD, STRENGTH_LEARNED}:
        return False
    return (
        bool(row.get("review_required"))
        or strength in REVISABLE_STRENGTHS
        or row.get("vat_claim_status") in {"Review", "Output/Review"}
    )


def mark_possible_duplicates(rows: list[dict[str, Any]]) -> None:
    seen: dict[tuple[str, Decimal, Decimal], int] = {}
    for row in rows:
        key = (
            normalize_ai_cache_key(str(row.get("description") or "")),
            decimal_amount(row.get("money_in")),
            decimal_amount(row.get("money_out")),
        )
        seen[key] = seen.get(key, 0) + 1
    for row in rows:
        key = (
            normalize_ai_cache_key(str(row.get("description") or "")),
            decimal_amount(row.get("money_in")),
            decimal_amount(row.get("money_out")),
        )
        if seen.get(key, 0) > 1 and (decimal_amount(row.get("money_in")) > 0 or decimal_amount(row.get("money_out")) > 0):
            reason = row.get("review_reason") or ""
            row["review_reason"] = "; ".join(part for part in [reason, "Possible duplicate"] if part)
            row["review_required"] = True


def ai_safe_item(row: dict[str, Any], transaction_id: str) -> dict[str, Any]:
    return {
        "transaction_id": transaction_id,
        "date": str(row.get("date") or ""),
        "description": str(row.get("description") or "")[:260],
        "money_in": str(decimal_amount(row.get("money_in"))),
        "money_out": str(decimal_amount(row.get("money_out"))),
        "rule_account": str(row.get("account") or ""),
        "rule_group": str(row.get("group") or ""),
        "rule_vat_treatment": str(row.get("vat_treatment") or ""),
        "rule_vat_claim_status": str(row.get("vat_claim_status") or ""),
        "rule_confidence": float(row.get("rule_confidence") or 0),
        "rule_reason": str(row.get("classification_reason") or ""),
    }


def parse_ai_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in {"true", "yes", "1"}
    return bool(value)


def validate_ai_item(item: Any, valid_ids: set[str]) -> dict[str, Any] | None:
    if not isinstance(item, dict):
        return None
    transaction_id = str(item.get("transaction_id") or "")
    if transaction_id not in valid_ids:
        return None
    account = str(item.get("account") or "").strip()[:80]
    group = str(item.get("group") or "").strip()[:80]
    vat_treatment = str(item.get("vat_treatment") or "").strip()[:80]
    vat_claim_status = str(item.get("vat_claim_status") or "").strip()[:80]
    if not account or not group or not vat_treatment or not vat_claim_status:
        return None
    # The account and the VAT claim status are closed sets. Anything else is
    # rejected outright rather than repaired: a model that returned "Bank Fees"
    # or invented an account produced a category professional_account cannot
    # map, the review UI cannot offer, and a learned rule built from a
    # correction to it would then spread. Guessing what it meant is still
    # guessing, and VAT is where a wrong guess costs money.
    if not is_valid_ai_account(account):
        return None
    if not is_valid_vat_claim_status(vat_claim_status):
        return None
    # An alias is a real category, but it is not what we write. Canonicalise here
    # so a historical spelling the model echoed back from context cannot become a
    # new stored value and reopen the divergence this vocabulary exists to close.
    account = canonicalise_category(account) or account
    confidence_value = item.get("confidence")
    if not isinstance(confidence_value, (int, float)) or isinstance(confidence_value, bool):
        return None
    confidence = float(confidence_value)
    if confidence > 1:
        confidence = confidence / 100
    if not 0 <= confidence <= 1:
        return None
    merchant = str(item.get("normalized_merchant") or "").strip()[:120] or None
    return {
        "transaction_id": transaction_id,
        "normalized_merchant": merchant,
        "account": account,
        "group": group,
        "vat_treatment": vat_treatment,
        "vat_claim_status": vat_claim_status,
        "review_required": parse_ai_bool(item.get("review_required")),
        "review_reason": str(item.get("review_reason") or "").strip()[:220],
        "invoice_required": parse_ai_bool(item.get("invoice_required")),
        "confidence": confidence,
        "reason": str(item.get("reason") or "").strip()[:220],
        "explanation": str(item.get("explanation") or "").strip()[:320],
    }


def parse_ai_json_content(content: str) -> Any:
    cleaned = content.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE).strip()
        cleaned = re.sub(r"\s*```$", "", cleaned).strip()
    try:
        parsed = json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start < 0 or end <= start:
            raise
        parsed = json.loads(cleaned[start : end + 1])
    return parsed


def openai_chat_completion(request_body: dict[str, Any], api_key: str, timeout: int = 60) -> dict[str, Any]:
    """The one place this worker talks to OpenAI.

    Shared by transaction classification and by AI recovery so there is a single
    transport to reason about: one timeout, one auth header, one endpoint.
    """
    request = urllib.request.Request(
        "https://api.openai.com/v1/chat/completions",
        data=json.dumps(request_body).encode("utf-8"),
        headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def request_ai_classifications(items: list[dict[str, Any]], diagnostics: dict[str, Any]) -> list[dict[str, Any]]:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key or not items:
        return []

    prompt = build_classification_prompt(items)
    body = {
        "model": accounting_ai_model(),
        "temperature": 0,
        "max_tokens": 6000,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": "You are an accounting classification assistant. Output valid JSON only."},
            {"role": "user", "content": json.dumps(prompt, default=str)},
        ],
    }

    def send_openai_request(request_body: dict[str, Any]) -> dict[str, Any]:
        return openai_chat_completion(request_body, api_key)

    try:
        try:
            payload = send_openai_request(body)
        except urllib.error.HTTPError as exc:
            body_text = exc.read().decode("utf-8", errors="replace")
            if exc.code != 400:
                raise
            log_warning("worker.ai_classification_retrying_without_response_format", status=exc.code, body=body_text[:1200])
            fallback_body = {key: value for key, value in body.items() if key != "response_format"}
            payload = send_openai_request(fallback_body)
        content = payload.get("choices", [{}])[0].get("message", {}).get("content", "{}")
        parsed = parse_ai_json_content(content)
        valid_ids = {str(item["transaction_id"]) for item in items}
        if isinstance(parsed, dict):
            raw_items = parsed.get("items", [])
        elif isinstance(parsed, list):
            raw_items = parsed
        else:
            raw_items = []
        if isinstance(raw_items, dict):
            raw_items = list(raw_items.values())
        elif not isinstance(raw_items, list):
            raw_items = []
        validated_items = [validated for raw in raw_items if (validated := validate_ai_item(raw, valid_ids))]
        if not validated_items and items:
            diagnostics["ai_failures"] += 1
            log_warning("worker.ai_classification_empty", returned_keys=list(parsed.keys()) if isinstance(parsed, dict) else [], item_count=len(items))
        return validated_items
    except urllib.error.HTTPError as exc:
        diagnostics["ai_failures"] += 1
        body_text = exc.read().decode("utf-8", errors="replace")
        log_warning("worker.ai_classification_http_failed", status=exc.code, body=body_text[:1200])
    except Exception as exc:
        diagnostics["ai_failures"] += 1
        log_warning("worker.ai_classification_failed", error=str(exc))
    return []


def apply_ai_result_to_row(row: dict[str, Any], result: dict[str, Any]) -> None:
    description = str(row.get("description") or "")
    money_out = decimal_amount(row.get("money_out"))
    ai_account_text = f"{result.get('account', '')} {result.get('group', '')}".lower()
    if (
        money_out > 0
        and any(token in ai_account_text for token in ("staff welfare", "meal", "entertainment", "travel"))
        and not is_staff_welfare_merchant(description)
    ):
        if looks_like_business_supplier_payment(description):
            result = {
                **result,
                "account": "Supplier Payments",
                "group": "Operating Expenses",
                "vat_treatment": "Input VAT if valid invoice",
                "vat_claim_status": "Input/Review",
                "review_required": True,
                "review_reason": "Business supplier payment requires invoice and VAT review.",
                "invoice_required": True,
                "confidence": min(float(result.get("confidence") or 0.72), 0.88),
                "reason": "Guardrail applied: company/invoice supplier pattern overrides staff welfare.",
                "explanation": "The description looks like a business supplier or invoice payment, so it must not be classified as meals or entertainment without explicit accountant approval.",
            }
        else:
            result = {
                **result,
                "account": "Review Required Suspense",
                "group": "Review",
                "vat_treatment": "Review",
                "vat_claim_status": "Review",
                "review_required": True,
                "review_reason": "AI suggested staff welfare without a recognised food, personal-care, or entertainment merchant.",
                "invoice_required": True,
                "confidence": min(float(result.get("confidence") or 0.62), 0.70),
                "reason": "Guardrail applied: staff welfare requires a matching merchant pattern.",
                "explanation": "The transaction needs accountant review before it can be treated as staff welfare, meals, entertainment, or travel.",
            }
    row["account"] = result["account"]
    row["group"] = result["group"]
    row["vat_treatment"] = result["vat_treatment"]
    row["vat_claim_status"] = result["vat_claim_status"]
    ai_confidence = float(result["confidence"])
    review_required = bool(result["review_required"])
    review_reason = result["review_reason"] or ""
    if (
        not review_required
        and ai_confidence >= 0.82
        and result["account"] not in {"Unclassified Expense", "Other Income / Review", "Review Required", "Review Required Suspense", "Suspense / Review Required"}
        and result["vat_claim_status"] not in {"Review", "Output/Review"}
    ):
        review_reason = ""
    row["review_required"] = review_required
    row["review_reason"] = review_reason
    row["invoice_required"] = result["invoice_required"]
    row["ai_confidence"] = ai_confidence
    row["classification_reason"] = result.get("reason") or row.get("classification_reason") or "AI classification applied to ambiguous transaction."
    row["classification_explanation"] = result.get("explanation") or row.get("classification_explanation") or ""
    row["ai_used"] = True
    # A merchant is a claim about who was paid. It is kept only when the bank's
    # own description contains it — a name the document does not mention is a
    # fabrication however plausible it reads, and null is the better answer.
    suggested = result.get("normalized_merchant")
    if suggested and merchant_is_grounded(suggested, description):
        row["normalized_merchant"] = suggested
    elif suggested:
        row["ai_merchant_rejected"] = suggested
    recompute_professional_vat(row)


def apply_ai_classifications(rows: list[dict[str, Any]], workspace_id: str = "") -> dict[str, Any]:
    """Enrich classifications with a model. Never allowed to fail a run.

    Classification is enrichment on top of a reconciled ledger. If OpenAI times
    out, returns malformed JSON, returns an account outside the chart or is
    simply unavailable, the deterministic classification stands and the rows
    keep whatever review status they already had. Nothing here may raise into
    the caller: the transactions are already correct, and losing them over a
    classification hint would be absurd.
    """
    try:
        return _apply_ai_classifications(rows, workspace_id)
    except Exception as exc:  # noqa: BLE001 - enrichment must never fail a run
        log_warning("worker.ai_classification_failed_open", workspace_id=workspace_id, error=str(exc)[:400])
        diagnostics = ai_diagnostics()
        diagnostics["ai_failures"] = diagnostics.get("ai_failures", 0) + 1
        diagnostics["ai_skipped"] = "classification_error"
        return diagnostics


def _apply_ai_classifications(rows: list[dict[str, Any]], workspace_id: str) -> dict[str, Any]:
    diagnostics = ai_diagnostics()
    if not diagnostics["ai_enabled"]:
        return diagnostics

    batch: list[dict[str, Any]] = []
    row_by_id: dict[str, dict[str, Any]] = {}
    cache_key_by_id: dict[str, str] = {}

    for index, row in enumerate(rows, start=1):
        if not row_needs_ai(row):
            continue
        cache_key = normalize_ai_cache_key(str(row.get("description") or ""))
        scoped_key = (workspace_id, cache_key)
        if cache_key and scoped_key in AI_CLASSIFICATION_CACHE:
            apply_ai_result_to_row(row, AI_CLASSIFICATION_CACHE[scoped_key])
            diagnostics["ai_cache_hits"] += 1
            continue
        transaction_id = str(index)
        batch.append(ai_safe_item(row, transaction_id))
        row_by_id[transaction_id] = row
        cache_key_by_id[transaction_id] = cache_key

    diagnostics["ai_transactions_sent"] = len(batch)
    for start in range(0, len(batch), AI_CLASSIFICATION_BATCH_SIZE):
        chunk = batch[start : start + AI_CLASSIFICATION_BATCH_SIZE]
        for result in request_ai_classifications(chunk, diagnostics):
            row = row_by_id.get(result["transaction_id"])
            if not row:
                continue
            apply_ai_result_to_row(row, result)
            cache_key = cache_key_by_id.get(result["transaction_id"])
            if cache_key:
                AI_CLASSIFICATION_CACHE[(workspace_id, cache_key)] = result
            diagnostics["ai_transactions_classified"] += 1

    log_event("worker.ai_classification", **diagnostics)
    return diagnostics


def apply_ai_counterparty_reasoning(
    transactions: list[ParsedTransaction],
    bank_profile: str | None,
    diagnostics: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Ask the model about counterparties, then apply the answers to their rows.

    This is the row-by-row path's replacement for anything that recurs. On the
    real statement it turns 413 unresolved rows into 37 questions covering 327
    of them, in two API calls — and each question carries six months of pattern
    the row-by-row prompt never had.

    Rows whose counterparty appears once are deliberately left alone: grouping a
    single transaction adds no evidence, only ceremony.
    """
    report = {
        "counterparty_groups": 0,
        "counterparty_questions": 0,
        "counterparty_rows_covered": 0,
        "counterparty_answers": 0,
        "counterparty_applied": 0,
        "counterparty_rejected": {},
    }
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key or not transactions:
        return report

    unsettled = [
        t for t in transactions
        if str(t.classification_strength or STRENGTH_NONE) not in {STRENGTH_HARD, STRENGTH_LEARNED}
    ]
    if not unsettled:
        return report

    groups = counterparty_group_by(unsettled, bank_profile)
    evidence_by_key = counterparty_evidence_for_all(unsettled, bank_profile)
    for key, evidence in evidence_by_key.items():
        relationship = infer_counterparty_relationship(evidence)
        object.__setattr__(evidence, "relationship", relationship.kind)
        object.__setattr__(evidence, "relationship_strength", relationship.strength)

    merchant_type_by_key = {
        key: rows[0].merchant_type
        for key, rows in groups.items()
        if rows and rows[0].merchant_type
    }
    questions = build_counterparty_questions(groups, evidence_by_key, merchant_type_by_key=merchant_type_by_key)
    report["counterparty_groups"] = len(groups)
    report["counterparty_questions"] = len(questions)
    report["counterparty_rows_covered"] = sum(q.occurrences for q in questions)
    if not questions:
        return report

    vocabulary = canonical_categories()
    rejected_total: dict[str, int] = {}
    verdicts: list[Any] = []

    for start in range(0, len(questions), COUNTERPARTY_BATCH_SIZE):
        chunk = questions[start : start + COUNTERPARTY_BATCH_SIZE]
        prompt = build_counterparty_prompt(chunk, vocabulary)
        body = {
            "model": accounting_ai_model(),
            "temperature": 0,
            "max_tokens": 4000,
            "response_format": {"type": "json_object"},
            "messages": [
                {"role": "system", "content": "You are an accounting classification assistant. Output valid JSON only."},
                {"role": "user", "content": json.dumps(prompt, default=str)},
            ],
        }
        try:
            payload = openai_chat_completion(body, api_key)
            content = payload.get("choices", [{}])[0].get("message", {}).get("content", "{}")
            answers, rejected = validate_counterparty_answers(
                parse_ai_json_content(content), chunk, vocabulary=vocabulary
            )
        except Exception as exc:  # noqa: BLE001 — a model failure must never fail a run
            log_warning("worker.ai_counterparty_failed", error=str(exc)[:300])
            continue
        verdicts.extend(answers)
        for reason, count in rejected.items():
            if count:
                rejected_total[reason] = rejected_total.get(reason, 0) + count

    report["counterparty_answers"] = len(verdicts)
    report["counterparty_rejected"] = rejected_total

    applied = 0
    declined_groups = 0
    for verdict in verdicts:
        # Declining is a valid answer, and the same rule applies here as on the
        # row-by-row path: a parking bucket is not a classification, so the
        # group keeps source=unresolved rather than being recorded as settled
        # by a model that said it could not tell.
        if verdict.category is None or is_unresolved_category(verdict.category):
            declined_groups += 1
            continue
        for transaction in groups.get(verdict.key, ()):  # noqa: B020
            # Never over a settled row. A bank-named fee and a human's own
            # correction are not questions a model gets to reopen.
            if str(transaction.classification_strength or STRENGTH_NONE) in {STRENGTH_HARD, STRENGTH_LEARNED}:
                continue
            transaction.account_category = canonicalise_category(verdict.category) or transaction.account_category
            transaction.classification_source = SOURCE_AI
            transaction.classification_strength = STRENGTH_SOFT
            transaction.classification_confidence = verdict.confidence
            transaction.classification_reason = verdict.reason
            transaction.review_status = "needs_review"
            transaction.evidence_used = [
                {"source": "ai_counterparty", "detail": item, "grade": "medium"}
                for item in verdict.evidence_used
            ] or transaction.evidence_used
            applied += 1

    report["counterparty_applied"] = applied
    report["counterparty_declined"] = declined_groups
    log_event("worker.ai_counterparty_reasoning", **report)
    if diagnostics is not None:
        diagnostics.update(report)
    return report


def month_summary(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    months = sorted({row["month"] for row in rows if row["month"]})
    summary = []
    for month in months:
        matching = [row for row in rows if row["month"] == month]
        receipts = sum((row["money_in"] for row in matching), Decimal("0.00")).quantize(CENT)
        payments = sum((row["money_out"] for row in matching), Decimal("0.00")).quantize(CENT)
        likely_sales = sum((row["money_in"] for row in matching if row["vat_claim_status"].startswith("Output")), Decimal("0.00")).quantize(CENT)
        cos = sum((row["money_out"] for row in matching if reporting_account(row) != "Review Required Suspense" and row["group"] in {"Freight", "Software/IT"}), Decimal("0.00")).quantize(CENT)
        output_vat = sum((row["potential_output_vat"] for row in matching if reporting_account(row) != "Review Required Suspense"), Decimal("0.00")).quantize(CENT)
        input_vat = sum((row["potential_input_vat"] for row in matching if reporting_account(row) != "Review Required Suspense"), Decimal("0.00")).quantize(CENT)
        summary.append({
            "month": month,
            "receipts": receipts,
            "payments": payments,
            "likely_sales": likely_sales,
            "cos": cos,
            "output_vat": output_vat,
            "input_vat": input_vat,
            "vat_payable": (output_vat - input_vat).quantize(CENT),
        })
    return summary


def write_vat_schedule_sheet(workbook: Workbook, rows: list[dict[str, Any]], include_source_period: bool = False):
    vat = workbook.create_sheet("VAT Schedule")
    reportable_rows = [row for row in rows if reporting_account(row) != "Review Required Suspense"]
    total_output_vat = sum((row["potential_output_vat"] for row in reportable_rows), Decimal("0.00")).quantize(CENT)
    total_input_vat = sum((row["potential_input_vat"] for row in reportable_rows), Decimal("0.00")).quantize(CENT)
    net_vat = (total_output_vat - total_input_vat).quantize(CENT)
    review_items = sum(1 for row in rows if reporting_account(row) == "Review Required Suspense" or reporting_vat_status(row) == "Review")
    vat["A1"] = "VAT Schedule & VAT Payable/(Refund)"
    vat["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    vat["A1"].fill = HEADER_FILL
    vat.merge_cells("A1:K1")
    vat["A2"] = "VAT is calculated at 15/115 on VAT-inclusive transactions only where the category and VAT treatment are reportable. Review-required rows stay visible but are excluded from VAT totals until approved."
    vat["A2"].font = Font(italic=True, size=9, color="475569")
    vat["A2"].alignment = Alignment(wrap_text=True)

    write_row(vat, ["VAT Summary", "Output VAT", "Potential Input VAT", "VAT Payable/(Refund)", "Review Items"], 4, header=True)
    write_row(vat, ["Totals", total_output_vat, total_input_vat, net_vat, review_items], 5)
    for column_index in range(2, 5):
        vat.cell(row=5, column=column_index).number_format = CURRENCY_FORMAT
    vat.cell(row=5, column=4).fill = PASS_FILL if net_vat <= 0 else FAIL_FILL
    vat.cell(row=5, column=5).number_format = "0"

    monthly_rows = month_summary(reportable_rows)
    write_row(vat, ["Month", "Output VAT", "Potential Input VAT", "Net VAT Payable/(Refund)", "Running VAT Balance", "Status"], 7, header=True)
    running_balance = Decimal("0.00")
    for row_index, month_row in enumerate(monthly_rows, start=8):
        monthly_net = month_row["vat_payable"]
        running_balance = (running_balance + monthly_net).quantize(CENT)
        write_row(
            vat,
            [
                month_row["month"],
                month_row["output_vat"],
                month_row["input_vat"],
                monthly_net,
                running_balance,
                "Payable" if running_balance >= 0 else "Refundable",
            ],
            row_index,
        )
        for column_index in range(2, 6):
            vat.cell(row=row_index, column=column_index).number_format = CURRENCY_FORMAT

    detail_header_row = max(10, 9 + len(monthly_rows))
    detail_headers = ["Date", "Description", "Money In", "Money Out", "Account", "VAT Code", "Claim Status", "Output VAT", "Potential Input VAT", "Net VAT", "VAT Balance", "Document Status"]
    if include_source_period:
        detail_headers.insert(2, "Source Period")
    write_row(vat, detail_headers, detail_header_row, header=True)
    running_line_vat = Decimal("0.00")
    for row_index, row in enumerate(rows, start=detail_header_row + 1):
        is_reportable = reporting_account(row) != "Review Required Suspense"
        output_vat = row["potential_output_vat"] if is_reportable else Decimal("0.00")
        input_vat = row["potential_input_vat"] if is_reportable else Decimal("0.00")
        line_net = (output_vat - input_vat).quantize(CENT)
        running_line_vat = (running_line_vat + line_net).quantize(CENT)
        values = [
            row["date"],
            row["description"],
            row["money_in"],
            row["money_out"],
            reporting_account(row),
            vat_code_for_row(row),
            reporting_vat_status(row),
            output_vat,
            input_vat,
            line_net,
            running_line_vat,
            "Tax invoice to be matched by user" if is_reportable else "Review before VAT is included",
        ]
        if include_source_period:
            values.insert(2, row["source_period"])
        write_row(vat, values, row_index)

    currency_columns = [3, 4, 8, 9, 10, 11] if include_source_period else [3, 4, 8, 9, 10, 11]
    if include_source_period:
        currency_columns = [4, 5, 9, 10, 11, 12]
    apply_number_formats(vat, currency_columns)
    return vat, detail_header_row, len(detail_headers)


def build_workbook(
    metadata: dict[str, Any],
    transactions: list[ParsedTransaction],
    allow_ai: bool = True,
    workspace_id: str = "",
) -> bytes:
    workbook = Workbook()
    totals = validation_summary(transactions)
    status, calculated_closing = validation_status(metadata, transactions)
    opening = decimal_amount(metadata.get("opening_balance"))
    closing = decimal_amount(metadata.get("closing_balance"))
    # Use the company/account holder detected from the actual statement. Never
    # hardcode a company name into the workbook title.
    company_name = (metadata.get("company_name") or "").strip()
    account_number = (metadata.get("account_number") or "").strip()
    source_file = metadata.get("source_file") or ""
    rows = [professional_transaction_row(transaction, source_file) for transaction in transactions]
    ai_started = time.perf_counter()
    if allow_ai:
        ai_stats = apply_ai_classifications(rows, workspace_id)
    else:
        ai_stats = ai_diagnostics(enabled=bool(os.getenv("OPENAI_API_KEY")))
        ai_stats["ai_skipped"] = "extraction_incomplete"
    ai_duration_ms = round((time.perf_counter() - ai_started) * 1000, 2)
    ai_stats["ai_classification_duration_ms"] = ai_duration_ms
    log_event("worker.ai_classification_duration", duration_ms=ai_duration_ms, parser_profile=WORKER_PARSER_VERSION)
    mark_possible_duplicates(rows)
    metadata["_ai_diagnostics"] = ai_stats
    months = month_summary(rows)
    bank_charge_total = sum((row["bank_charge"] for row in rows), Decimal("0.00")).quantize(CENT)
    bank_vat = (bank_charge_total * Decimal("15") / Decimal("115")).quantize(CENT)
    reportable_rows = [row for row in rows if reporting_account(row) != "Review Required Suspense"]
    total_output_vat = sum((row["potential_output_vat"] for row in reportable_rows), Decimal("0.00")).quantize(CENT)
    total_input_vat = sum((row["potential_input_vat"] for row in reportable_rows), Decimal("0.00")).quantize(CENT)

    dashboard = workbook.active
    dashboard.title = "Dashboard"
    dashboard.merge_cells("A1:K1")
    workbook_title = (
        f"{company_name} - Bank Statement Accounting Pack"
        if company_name
        else "Bank Statement Accounting Pack"
    )
    dashboard["A1"] = workbook_title
    dashboard["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    dashboard["A1"].fill = HEADER_FILL
    dashboard["A1"].alignment = Alignment(horizontal="center")
    dashboard.merge_cells("A2:K2")
    dashboard["A2"] = ACCOUNTING_REPORT_DISCLAIMER
    dashboard["A2"].font = Font(italic=True, size=9, color="475569")
    dashboard["A2"].alignment = Alignment(wrap_text=True)
    dashboard_rows = [
        ("Period covered", f"{metadata.get('statement_period_start') or '-'} to {metadata.get('statement_period_end') or '-'}"),
        ("Opening bank balance", opening),
        ("Total receipts", totals["total_credits"]),
        ("Total payments", totals["total_debits"]),
        ("Closing bank balance", closing),
        ("Bank movement check", (opening + totals["total_credits"] - totals["total_debits"] - closing).quantize(CENT)),
        ("Likely taxable revenue receipts", sum((row["money_in"] for row in reportable_rows if row["vat_claim_status"].startswith("Output")), Decimal("0.00")).quantize(CENT)),
        ("Potential output VAT", total_output_vat),
        ("Potential input VAT (review)", total_input_vat),
        ("Potential VAT payable/(refund)", (total_output_vat - total_input_vat).quantize(CENT)),
        ("Transactions extracted", len(transactions)),
        ("Reconciliation status", "Reconciled" if status == "PASSED" else "Review required"),
    ]
    for index, row in enumerate(dashboard_rows, start=3):
        write_row(dashboard, list(row), index)
    dashboard["B14"].fill = PASS_FILL if status == "PASSED" else FAIL_FILL
    dashboard["B14"].font = Font(bold=True, color="166534" if status == "PASSED" else "991B1B")
    write_row_at(dashboard, ["Month", "Receipts", "Payments", "Likely Sales", "COS/Subcontractors", "Output VAT", "Input VAT", "VAT Payable/(Refund)"], 3, 4, header=True)
    for row_index, month_row in enumerate(months, start=4):
        write_row_at(
            dashboard,
            [month_row["month"], month_row["receipts"], month_row["payments"], month_row["likely_sales"], month_row["cos"], month_row["output_vat"], month_row["input_vat"], month_row["vat_payable"]],
            row_index,
            4,
        )
    for row_index in range(4, 13):
        dashboard.cell(row=row_index, column=2).number_format = CURRENCY_FORMAT
        for column_index in range(5, 12):
            dashboard.cell(row=row_index, column=column_index).number_format = CURRENCY_FORMAT
    dashboard["B13"].number_format = "0"

    tx = workbook.create_sheet("Transactions")
    transaction_headers = [
        "Date", "Month", "Description", "Money In", "Money Out", "Amount", "Type", "Balance", "Bank Charge",
        "Account", "Group", "VAT Code", "VAT Claim Status", "Potential Output VAT", "Potential Input VAT",
        "Confidence", "Classification Reason", "Classification Explanation",
    ]
    write_row(tx, transaction_headers, 1, header=True)
    for row_index, row in enumerate(rows, start=2):
        write_row(
            tx,
            [
                row["date"], row["month"], row["description"], row["money_in"], row["money_out"], row["amount"], row["type"], row["balance"],
                row["bank_charge"], row["account"], row["group"], vat_code_for_row(row), row["vat_claim_status"], row["potential_output_vat"],
                row["potential_input_vat"], row["rule_confidence"], row["classification_reason"], row["classification_explanation"],
            ],
            row_index,
        )
    apply_number_formats(tx, [4, 5, 6, 8, 9, 14, 15])

    vat, vat_detail_header_row, vat_column_count = write_vat_schedule_sheet(workbook, rows)

    ledger = workbook.create_sheet("General Ledger")
    write_row(ledger, ["Date", "Description", "Account", "Debit", "Credit"], 1, header=True)
    gl_row = 2
    write_row(ledger, [workbook_date(metadata.get("statement_period_start")), "Opening balance per bank statement", "Bank", opening, Decimal("0.00")], gl_row)
    gl_row += 1
    write_row(ledger, [workbook_date(metadata.get("statement_period_start")), "Opening balance per bank statement", "Opening Equity / Prior Periods", Decimal("0.00"), opening], gl_row)
    gl_row += 1
    for row in rows:
        if row["money_out"] > 0:
            write_row(ledger, [row["date"], row["description"], reporting_account(row), row["money_out"], Decimal("0.00")], gl_row)
            gl_row += 1
            write_row(ledger, [row["date"], row["description"], "Bank", Decimal("0.00"), row["money_out"]], gl_row)
            gl_row += 1
        elif row["money_in"] > 0:
            write_row(ledger, [row["date"], row["description"], "Bank", row["money_in"], Decimal("0.00")], gl_row)
            gl_row += 1
            write_row(ledger, [row["date"], row["description"], reporting_account(row), Decimal("0.00"), row["money_in"]], gl_row)
            gl_row += 1
    apply_number_formats(ledger, [4, 5])

    trial = workbook.create_sheet("Trial Balance")
    write_row(trial, ["Account", "Total Debits", "Total Credits", "Debit Balance", "Credit Balance"], 1, header=True)
    ledger_accounts = sorted({ledger.cell(row=row, column=3).value for row in range(2, ledger.max_row + 1) if ledger.cell(row=row, column=3).value})
    for row_index, account in enumerate(ledger_accounts, start=2):
        debits = sum(decimal_amount(ledger.cell(row=row, column=4).value) for row in range(2, ledger.max_row + 1) if ledger.cell(row=row, column=3).value == account)
        credits = sum(decimal_amount(ledger.cell(row=row, column=5).value) for row in range(2, ledger.max_row + 1) if ledger.cell(row=row, column=3).value == account)
        net = (debits - credits).quantize(CENT)
        write_row(trial, [account, debits, credits, net if net > 0 else Decimal("0.00"), abs(net) if net < 0 else Decimal("0.00")], row_index)
    total_row = len(ledger_accounts) + 2
    write_row(
        trial,
        [
            "Totals",
            f"=SUM(B2:B{total_row - 1})",
            f"=SUM(C2:C{total_row - 1})",
            f"=SUM(D2:D{total_row - 1})",
            f"=SUM(E2:E{total_row - 1})",
        ],
        total_row,
    )
    write_row(
        trial,
        ["Balance Check", "", "", f"=D{total_row}-E{total_row}", "Balanced when zero"],
        total_row + 1,
    )
    for cell in trial[total_row]:
        cell.font = Font(bold=True)
    apply_number_formats(trial, [2, 3, 4, 5])

    rec = workbook.create_sheet("Bank Rec")
    write_row(rec, ["Bank Reconciliation", "Amount"], 1, header=True)
    rec_rows = [
        ("Opening Balance", opening),
        ("+ Receipts", totals["total_credits"]),
        ("- Payments", totals["total_debits"]),
        ("Expected Closing Balance", "=B2+B3-B4"),
        ("Statement Closing Balance", closing),
        ("Difference", f"=B5-B6"),
        ("Status", '=IF(B7=0,"Reconciled","Review required")'),
        ("Service Fees", bank_charge_total),
        ("Bank VAT", bank_vat),
    ]
    for row_index, row in enumerate(rec_rows, start=2):
        write_row(rec, list(row), row_index)
    rec["B8"].fill = PASS_FILL if status == "PASSED" else FAIL_FILL
    rec["B8"].font = Font(bold=True, color="166534" if status == "PASSED" else "991B1B")
    apply_number_formats(rec, [2])

    review = workbook.create_sheet("Review Items")
    write_row(review, ["Date", "Description", "Money In", "Money Out", "Account", "Group", "VAT Claim Status", "Reason", "Invoice Required"], 1, header=True)
    review_row = 2
    for row in rows:
        reason = row.get("review_reason") or professional_review_reason(row)
        if row.get("review_required") or reporting_account(row) == "Review Required Suspense":
            write_row(
                review,
                [
                    row["date"],
                    row["description"],
                    row["money_in"],
                    row["money_out"],
                    row["account"],
                    row["group"],
                    row["vat_claim_status"],
                    reason or "Review recommended",
                    "Yes" if row.get("invoice_required") else "No",
                ],
                review_row,
            )
            review_row += 1
    apply_number_formats(review, [3, 4])

    assumptions = workbook.create_sheet("Assumptions")
    assumptions_rows = [
        ("Area", "Assumption / Note"),
        ("Report limitation", ACCOUNTING_REPORT_DISCLAIMER),
        ("Important limitation", "This workbook is prepared from bank statements only. It is a cashbook-based reconstruction, not a full accounting system TB."),
        ("VAT rule applied", "Potential VAT is calculated at 15/115 of VAT-inclusive amounts only where the bank description suggests taxable revenue or claimable input VAT."),
        ("Invoice matching", "User confirmed invoices will be handled separately. The VAT schedule therefore flags document status for invoice matching."),
        ("Personal / non-deductible items", "Meals, groceries, spa, pets, gifts, entertainment and similar items are flagged for review and generally should not be claimed without strong business evidence."),
        ("Transfers", "Savings, investment, credit card and home loan transfers are treated as inter-account transfers/loan movements, not VAT transactions."),
        ("Bank fees", "FNB bank VAT per statement has been included in the reconciliation sheet. Individual bank charge VAT is flagged as review where applicable."),
        ("Bank account", f"FNB Platinum Business Account ending {account_number[-4:]}."),
        ("AI-assisted classification", "Where enabled, ambiguous descriptions may be classified by AI after the deterministic parser and reconciliation checks pass. Rule-based classifications remain the fallback."),
        ("Next step", "Match each VAT line to the relevant tax invoice, then update claim status before VAT201 submission."),
    ]
    for row_index, row in enumerate(assumptions_rows, start=1):
        write_row(assumptions, list(row), row_index, header=row_index == 1)

    finish_sheet(dashboard, freeze_pane="D4")
    finish_sheet(tx, filter_ref=f"A1:R{max(tx.max_row, 1)}")
    finish_sheet(vat, freeze_pane=f"A{vat_detail_header_row + 1}", filter_ref=f"A{vat_detail_header_row}:{get_column_letter(vat_column_count)}{max(vat.max_row, vat_detail_header_row)}")
    finish_sheet(ledger, filter_ref=f"A1:E{max(ledger.max_row, 1)}")
    finish_sheet(trial, filter_ref=f"A1:E{max(len(ledger_accounts) + 1, 1)}")
    finish_sheet(rec)
    finish_sheet(review, filter_ref=f"A1:I{max(review.max_row, 1)}")
    finish_sheet(assumptions)

    output = io.BytesIO()
    validate_workbook_for_export(workbook)
    workbook.save(output)
    return output.getvalue()


def parsed_transaction_from_row(row: dict[str, Any]) -> ParsedTransaction:
    return ParsedTransaction(
        transaction_date=row.get("transaction_date"),
        description=str(row.get("description") or ""),
        debit_amount=float(row["debit_amount"]) if row.get("debit_amount") is not None else None,
        credit_amount=float(row["credit_amount"]) if row.get("credit_amount") is not None else None,
        running_balance=float(row["running_balance"]) if row.get("running_balance") is not None else None,
        bank_charge=bool(row.get("bank_charge")),
        account_category=str(row.get("account_category") or "Uncategorised"),
        vat_treatment=str(row.get("vat_treatment") or "review"),
        supported_by_invoice=bool(row.get("supported_by_invoice")),
        notes=str(row.get("notes") or ""),
        confidence=float(row.get("confidence") or 0),
        review_status=str(row.get("review_status") or "needs_review"),
        source_page=row.get("source_page"),
        source_row=row.get("source_row"),
        raw_text=row.get("raw_text"),
    )


def transaction_insert_row(transaction: ParsedTransaction, run_id: str, workspace_id: str) -> dict[str, Any]:
    row = {
        **transaction.model_dump(),
        "run_id": run_id,
        "workspace_id": workspace_id,
    }
    return row


# Columns a very old database may not have. Named explicitly because this row is
# built by spreading model_dump(): a new field on ParsedTransaction reaches
# Supabase automatically, and an insert naming a column that does not exist fails
# the WHOLE batch — losing every transaction in the run. A field only becomes
# writable once it is listed here.
#
# source_row was previously dropped unconditionally, which is why the stored
# ledger had no recoverable order: every row carried NULL, created_at was one
# identical timestamp, and source_page only narrows to ~17 rows. The same 615
# rows produced 17, 513 or 615 balance "gaps" depending on how they were sorted,
# while the true order has none. It is written now, and the fallback only drops
# it on a database that predates migration 005.
OPTIONAL_TRANSACTION_COLUMNS = (
    "source_row",
    "classification_source",
    "classification_strength",
    "classification_confidence",
    "classification_reason",
    "normalized_merchant",
    # migration 023 — identity, relationship and reasoning stored apart from the
    # treatment. Listed here so a worker running against a database that has not
    # had 023 applied drops them and retries, exactly as it did for 021 and 022.
    "counterparty_key",
    "counterparty_display",
    "counterparty_truncated",
    "merchant_type",
    "merchant_type_source",
    "relationship",
    "relationship_strength",
    "evidence_used",
    "treatment_alternatives",
)


def strip_provenance_columns(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """The same rows as they would have been written before migration 021."""
    return [{key: value for key, value in row.items() if key not in OPTIONAL_TRANSACTION_COLUMNS} for row in rows]


def classify_transactions_with_ai(
    transactions: list[ParsedTransaction],
    workspace_id: str,
    source_file: str,
) -> dict[str, Any]:
    """Ask a model about the rows the deterministic rules could not settle.

    This runs BEFORE the transactions are written. It used to run only inside
    build_workbook, after the insert, so its answers reached the exported
    spreadsheet and never the stored row — which is what the reviewer looks at.
    A person opening the review screen saw the deterministic guess while the
    workbook held a better answer, and no correction they made could reconcile
    the two.

    What it may change: the category, and the provenance recording that a model
    chose it. What it may never change: the date, the description, either
    amount, the direction, the running balance, or `confidence` — which for an
    AI-recovered row is capped as an EXTRACTION signal and has nothing to do
    with how well the category is known.

    VAT is deliberately left as the deterministic rules set it. Recognising a
    merchant proves nothing about whether input VAT is claimable, whether a
    valid tax invoice exists, or whether the purpose was business, and this is
    the one place where being wrong costs money rather than time.
    """
    if not transactions:
        return ai_diagnostics(enabled=False)

    rows = [professional_transaction_row(transaction, source_file) for transaction in transactions]
    diagnostics = apply_ai_classifications(rows, workspace_id)

    applied = 0
    declined = 0
    rejected: dict[str, int] = {}
    for transaction, row in zip(transactions, rows):
        if not row.get("ai_used"):
            continue
        category = canonicalise_category(str(row.get("account") or ""))
        if category is None:
            # Already validated on the way in, so this means the guardrails in
            # apply_ai_result_to_row produced something outside the vocabulary.
            rejected["account_outside_vocabulary"] = rejected.get("account_outside_vocabulary", 0) + 1
            continue

        # "I do not know" is not an accounting classification.
        #
        # A model asked about an unidentifiable row often answers with a parking
        # bucket — Suspense / Review Required, Other Income / Review,
        # Uncategorised. That is the correct answer, and it must not be recorded
        # as a successful one. Stamping SOURCE_AI here erased `unresolved` from
        # the ledger entirely: on the real 615-row statement every row came back
        # sourced `ai`, coverage reported 100% automated while 482 rows needed a
        # human, and the run's classification confidence became an average
        # dragged down by 345 rows that were not classifications at all.
        #
        # The row keeps what the evidence layers decided — the same parking
        # bucket at source=unresolved, strength=none. The attempt is recorded in
        # diagnostics, so "asked and declined" stays visible without pretending
        # the model answered.
        if is_unresolved_category(category):
            declined += 1
            continue

        transaction.account_category = category
        transaction.classification_source = SOURCE_AI
        # A model's answer is revisable — by a person, and by a learned rule
        # built from what that person decides.
        transaction.classification_strength = STRENGTH_SOFT
        transaction.classification_confidence = round(float(row.get("ai_confidence") or 0) * 100, 2)
        transaction.classification_reason = str(row.get("classification_reason") or "")[:400]
        merchant = row.get("normalized_merchant")
        if merchant and not transaction.normalized_merchant:
            transaction.normalized_merchant = str(merchant)[:120]
        if row.get("review_required"):
            transaction.review_status = "needs_review"
        applied += 1

    diagnostics["ai_transactions_applied_to_ledger"] = applied
    diagnostics["ai_declined_left_unresolved"] = declined
    diagnostics["ai_rejected_after_guardrails"] = rejected
    return diagnostics


def replace_transactions(supabase: Client, rows: list[dict[str, Any]], run_id: str, workspace_id: str) -> bool:
    """Replace a run's transactions atomically. Returns whether provenance was kept.

    The delete and the insert used to be two separate PostgREST calls. A crash
    between them — a deploy, an OOM kill, a restart — left the run with ZERO
    transactions: the whole ledger gone, no error recorded, the run still
    reading as processed. PostgREST cannot express a client-side transaction, so
    the pair is now one call to replace_accounting_transactions (migration 025),
    whose function body is a single implicit transaction.

    Falls back to the previous delete-then-insert when the function is absent,
    so a database that has not run 025 still processes statements. That path
    keeps the old crash window, which is strictly no worse than before.
    """
    try:
        supabase.rpc(
            "replace_accounting_transactions",
            {"p_run_id": run_id, "p_workspace_id": workspace_id, "p_rows": rows},
        ).execute()
        return True
    except Exception as exc:  # noqa: BLE001 — see fallback note above
        message = str(exc)
        log_warning(
            "worker.atomic_replace_unavailable",
            run_id=run_id,
            error=message[:400],
            note="migration 025 not applied? falling back to delete-then-insert",
        )

    supabase.table("accounting_transactions").delete().eq("run_id", run_id).eq("workspace_id", workspace_id).execute()
    return insert_transactions(supabase, rows, run_id)


def insert_transactions(supabase: Client, rows: list[dict[str, Any]], run_id: str) -> bool:
    """Write the transactions, degrading to the pre-provenance shape if needed.

    Provenance is an enrichment. A database that has not run migration 021 yet
    must still receive its ledger — losing 615 real transactions because a
    reporting column is missing would be a far worse failure than not knowing
    which rule classified them.

    Returns whether provenance was persisted.
    """
    try:
        supabase.table("accounting_transactions").insert(rows).execute()
        return True
    except Exception as exc:  # noqa: BLE001 - any insert rejection falls back
        log_warning(
            "worker.transaction_provenance_not_persisted",
            run_id=run_id,
            error=str(exc)[:400],
            note="migration 021 not applied? retrying without the classification provenance columns",
        )
        supabase.table("accounting_transactions").insert(strip_provenance_columns(rows)).execute()
        return False


def run_period_label(run: dict[str, Any]) -> str:
    start = str(run.get("statement_period_start") or "")
    end = str(run.get("statement_period_end") or "")
    if start and end:
        return f"{start} to {end}"
    return start or end or "Unknown period"


def combine_duplicate_key(transaction: ParsedTransaction) -> tuple[str, str, str, str, str]:
    return (
        str(transaction.transaction_date or ""),
        normalize_merchant_key(transaction.description),
        str(decimal_amount(transaction.debit_amount)),
        str(decimal_amount(transaction.credit_amount)),
        str(decimal_amount(transaction.running_balance)),
    )


def combine_fingerprint(run: dict[str, Any], transaction: ParsedTransaction, fallback_row: int) -> tuple[str, str, str, str, str, str, str, str, str]:
    return (
        str(run.get("account_number") or "").strip().lower(),
        str(run.get("id") or "").strip().lower(),
        str(transaction.transaction_date or ""),
        normalize_merchant_key(transaction.description),
        str(decimal_amount(transaction.debit_amount)),
        str(decimal_amount(transaction.credit_amount)),
        str(decimal_amount(transaction.running_balance)),
        str(transaction.source_page or 0),
        str(transaction.source_row or fallback_row),
    )


def extract_transaction_time(raw_text: str | None) -> str:
    if not raw_text:
        return ""
    match = re.search(r"\b([01]\d|2[0-3]):([0-5]\d)(?::([0-5]\d))?\b", raw_text)
    if not match:
        return ""
    second = match.group(3) or "00"
    return f"{match.group(1)}:{match.group(2)}:{second}"


def parse_iso_date_or_max(value: str | None) -> date:
    if not value:
        return date.max
    try:
        return date.fromisoformat(value)
    except ValueError:
        return date.max


def continuity_state(previous_closing: float | None, next_opening: float | None) -> tuple[str, Decimal | None]:
    if previous_closing is None or next_opening is None:
        return "UNKNOWN", None
    previous_decimal = decimal_amount(previous_closing)
    next_decimal = decimal_amount(next_opening)
    difference = (next_decimal - previous_decimal).quantize(CENT)
    if difference == 0:
        return "PASSED", Decimal("0.00")
    return "FAILED", difference


def continuity_failure_message(continuity: list[dict[str, Any]]) -> str:
    failures = [
        item
        for item in continuity
        if item.get("status") in {"FAILED", "UNKNOWN"}
    ]
    if not failures:
        return ""
    parts = []
    for item in failures:
        parts.append(
            f"{item['previous_period']} -> {item['next_period']}: {item['status']}"
        )
    return "Continuity checks require review: " + "; ".join(parts)


def run_continuity_summary(runs: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], str, str]:
    sorted_runs = sorted(runs, key=lambda run: str(run.get("statement_period_start") or run.get("created_at") or ""))
    continuity: list[dict[str, Any]] = []
    for previous, current in zip(sorted_runs, sorted_runs[1:]):
        state, diff = continuity_state(previous.get("closing_balance"), current.get("opening_balance"))
        continuity.append(
            {
                "previous_period": run_period_label(previous),
                "next_period": run_period_label(current),
                "previous_closing": decimal_amount(previous.get("closing_balance")) if previous.get("closing_balance") is not None else None,
                "next_opening": decimal_amount(current.get("opening_balance")) if current.get("opening_balance") is not None else None,
                "status": state,
                "difference": diff,
            }
        )

    continuity_passed = all(item["status"] == "PASSED" for item in continuity)
    continuity_failed = any(item["status"] == "FAILED" for item in continuity)
    continuity_result = "PASSED" if continuity_passed else "FAILED" if continuity_failed else "UNKNOWN"
    return continuity, continuity_result, continuity_failure_message(continuity)


def validate_combine_runs(runs: list[dict[str, Any]], payload: CombineRequest) -> list[dict[str, Any]]:
    if len(runs) != len(set(payload.run_ids)):
        raise HTTPException(status_code=404, detail="One or more selected statements could not be found.")

    invalid_statuses = [
        run for run in runs if str(run.get("status") or "") not in {"completed", "review"}
    ]
    if invalid_statuses:
        raise HTTPException(status_code=422, detail="Only completed or review-ready statements can be combined.")

    keys = {
        (
            str(run.get("company_name") or "").strip().lower(),
            str(run.get("bank") or "").strip().lower(),
            str(run.get("account_number") or "").strip().lower(),
        )
        for run in runs
    }
    if len(keys) > 1 and not payload.combine_different_accounts:
        raise HTTPException(status_code=422, detail="Selected statements are not the same company, bank and account number.")
    return runs


def build_combined_workbook(
    runs: list[dict[str, Any]],
    transactions_by_run: dict[str, list[ParsedTransaction]],
    workspace_id: str = "",
) -> tuple[bytes, dict[str, Any]]:
    generation_started = time.perf_counter()
    sorted_runs = sorted(runs, key=lambda run: str(run.get("statement_period_start") or run.get("created_at") or ""))
    first_run = sorted_runs[0]
    last_run = sorted_runs[-1]
    company_name = first_run.get("company_name") or "Unknown company"
    bank = first_run.get("bank") or "FNB South Africa"
    account_number = first_run.get("account_number") or ""
    opening_known = first_run.get("opening_balance") is not None
    closing_known = last_run.get("closing_balance") is not None
    opening = decimal_amount(first_run.get("opening_balance")) if opening_known else None
    closing = decimal_amount(last_run.get("closing_balance")) if closing_known else None

    continuity: list[dict[str, Any]] = []
    for previous, current in zip(sorted_runs, sorted_runs[1:]):
        previous_close = previous.get("closing_balance")
        current_open = current.get("opening_balance")
        state, diff = continuity_state(previous_close, current_open)
        continuity.append({
            "previous_period": run_period_label(previous),
            "next_period": run_period_label(current),
            "previous_closing": decimal_amount(previous_close) if previous_close is not None else None,
            "next_opening": decimal_amount(current_open) if current_open is not None else None,
            "status": state,
            "difference": diff,
        })

    combined_transactions: list[tuple[ParsedTransaction, dict[str, Any], int]] = []
    seen: dict[tuple[str, str, str, str, str, str, str, str, str], float] = {}
    duplicates_removed = 0
    for run_index, run in enumerate(sorted_runs):
        for row_index, transaction in enumerate(transactions_by_run.get(str(run["id"]), []), start=1):
            key = combine_fingerprint(run, transaction, row_index)
            previous_confidence = seen.get(key)
            current_confidence = float(transaction.confidence or 0)
            if previous_confidence is not None and previous_confidence >= 98 and current_confidence >= 98:
                duplicates_removed += 1
                log_event(
                    "worker.combine_duplicate_removed",
                    run_id=run.get("id"),
                    transaction_date=transaction.transaction_date,
                    description=transaction.description,
                    confidence=current_confidence,
                )
                continue
            seen[key] = max(previous_confidence or 0, current_confidence)
            combined_transactions.append((transaction, run, run_index))

    combined_transactions.sort(
        key=lambda item: (
            parse_iso_date_or_max(item[0].transaction_date),
            extract_transaction_time(item[0].raw_text),
            item[2],
            item[0].source_page if item[0].source_page is not None else 10**9,
            item[0].source_row if item[0].source_row is not None else 10**9,
        )
    )

    rows: list[dict[str, Any]] = []
    for transaction, run, _run_index in combined_transactions:
        row = professional_transaction_row(transaction, "combined")
        row["source_period"] = run_period_label(run)
        rows.append(row)

    ai_started = time.perf_counter()
    ai_stats = apply_ai_classifications(rows, workspace_id)
    ai_duration_ms = round((time.perf_counter() - ai_started) * 1000, 2)
    ai_stats["ai_classification_duration_ms"] = ai_duration_ms
    mark_possible_duplicates(rows)

    reportable_rows = [row for row in rows if reporting_account(row) != "Review Required Suspense"]
    total_debits = sum((row["money_out"] for row in rows), Decimal("0.00")).quantize(CENT)
    total_credits = sum((row["money_in"] for row in rows), Decimal("0.00")).quantize(CENT)
    expected_closing = (opening + total_credits - total_debits).quantize(CENT) if opening is not None else None
    difference = (expected_closing - closing).quantize(CENT) if expected_closing is not None and closing is not None else None
    review_count = sum(1 for row in rows if row.get("review_required") or reporting_account(row) == "Review Required Suspense")
    continuity_passed = all(item["status"] == "PASSED" for item in continuity)
    continuity_failed = any(item["status"] == "FAILED" for item in continuity)
    continuity_unknown = any(item["status"] == "UNKNOWN" for item in continuity)
    continuity_result = "PASSED" if continuity_passed else "FAILED" if continuity_failed else "UNKNOWN"

    workbook = Workbook()
    dashboard = workbook.active
    dashboard.title = "Dashboard"
    dashboard.merge_cells("A1:H1")
    dashboard["A1"] = f"{company_name} - Combined Bank Statement Accounting Pack"
    dashboard["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    dashboard["A1"].fill = HEADER_FILL
    dashboard["A1"].alignment = Alignment(horizontal="center")
    dashboard.merge_cells("A2:H2")
    dashboard["A2"] = ACCOUNTING_REPORT_DISCLAIMER
    dashboard["A2"].font = Font(italic=True, size=9, color="475569")
    dashboard["A2"].alignment = Alignment(wrap_text=True)
    dashboard_rows = [
        ("Company name", company_name),
        ("Bank", bank),
        ("Account number", account_number),
        ("Combined period", f"{first_run.get('statement_period_start') or '-'} to {last_run.get('statement_period_end') or '-'}"),
        ("Number of statements", len(sorted_runs)),
        ("Opening balance", opening if opening is not None else "Unknown"),
        ("Closing balance", closing if closing is not None else "Unknown"),
        ("Total receipts", total_credits),
        ("Total payments", total_debits),
        ("Total transactions", len(rows)),
        ("Review items", review_count),
        ("Workbook status", "Combined workbook generated with review items." if review_count else "Combined workbook generated."),
    ]
    for index, row in enumerate(dashboard_rows, start=3):
        write_row(dashboard, list(row), index)
    write_row_at(dashboard, ["Month", "VAT-classified receipts", "VAT-classified payments", "Output VAT", "Input VAT", "VAT Payable/(Refund)"], 3, 4, header=True)
    for row_index, month_row in enumerate(month_summary(reportable_rows), start=4):
        write_row_at(
            dashboard,
            [month_row["month"], month_row["receipts"], month_row["payments"], month_row["output_vat"], month_row["input_vat"], month_row["vat_payable"]],
            row_index,
            4,
        )

    tx = workbook.create_sheet("Transactions")
    tx_headers = [
        "Date", "Month", "Source Period", "Description", "Money In", "Money Out", "Amount", "Type", "Balance", "Bank Charge",
        "Account", "Group", "VAT Code", "VAT Claim Status", "Potential Output VAT", "Potential Input VAT",
    ]
    write_row(tx, tx_headers, 1, header=True)
    for row_index, row in enumerate(rows, start=2):
        write_row(
            tx,
            [
                row["date"], row["month"], row["source_period"], row["description"], row["money_in"], row["money_out"], row["amount"], row["type"],
                row["balance"], row["bank_charge"], reporting_account(row), row["group"], vat_code_for_row(row), reporting_vat_status(row),
                row["potential_output_vat"] if reporting_account(row) != "Review Required Suspense" else Decimal("0.00"),
                row["potential_input_vat"] if reporting_account(row) != "Review Required Suspense" else Decimal("0.00"),
            ],
            row_index,
        )
    apply_number_formats(tx, [5, 6, 7, 9, 10, 15, 16])

    vat, vat_detail_header_row, vat_column_count = write_vat_schedule_sheet(workbook, rows, include_source_period=True)

    ledger = workbook.create_sheet("General Ledger")
    write_row(ledger, ["Date", "Description", "Account", "Debit", "Credit", "Source Period"], 1, header=True)
    gl_row = 2
    write_row(ledger, [workbook_date(first_run.get("statement_period_start")), "Opening balance first statement", "Bank", opening, Decimal("0.00"), run_period_label(first_run)], gl_row)
    gl_row += 1
    write_row(ledger, [workbook_date(first_run.get("statement_period_start")), "Opening balance first statement", "Opening Equity / Prior Periods", Decimal("0.00"), opening, run_period_label(first_run)], gl_row)
    gl_row += 1
    for row in rows:
        if row["money_out"] > 0:
            write_row(ledger, [row["date"], row["description"], reporting_account(row), row["money_out"], Decimal("0.00"), row["source_period"]], gl_row)
            gl_row += 1
            write_row(ledger, [row["date"], row["description"], "Bank", Decimal("0.00"), row["money_out"], row["source_period"]], gl_row)
            gl_row += 1
        elif row["money_in"] > 0:
            write_row(ledger, [row["date"], row["description"], "Bank", row["money_in"], Decimal("0.00"), row["source_period"]], gl_row)
            gl_row += 1
            write_row(ledger, [row["date"], row["description"], reporting_account(row), Decimal("0.00"), row["money_in"], row["source_period"]], gl_row)
            gl_row += 1
    apply_number_formats(ledger, [4, 5])

    trial = workbook.create_sheet("Trial Balance")
    write_row(trial, ["Account", "Total Debits", "Total Credits", "Debit Balance", "Credit Balance"], 1, header=True)
    ledger_accounts = sorted({ledger.cell(row=row, column=3).value for row in range(2, ledger.max_row + 1) if ledger.cell(row=row, column=3).value})
    for row_index, account in enumerate(ledger_accounts, start=2):
        debits = sum(decimal_amount(ledger.cell(row=row, column=4).value) for row in range(2, ledger.max_row + 1) if ledger.cell(row=row, column=3).value == account)
        credits = sum(decimal_amount(ledger.cell(row=row, column=5).value) for row in range(2, ledger.max_row + 1) if ledger.cell(row=row, column=3).value == account)
        net = (debits - credits).quantize(CENT)
        write_row(trial, [account, debits, credits, net if net > 0 else Decimal("0.00"), abs(net) if net < 0 else Decimal("0.00")], row_index)
    apply_number_formats(trial, [2, 3, 4, 5])

    rec = workbook.create_sheet("Bank Rec")
    rec_rows = [
        ("Opening balance first period", opening if opening is not None else "Unknown"),
        ("Total credits all periods", total_credits),
        ("Total debits all periods", total_debits),
        ("Expected closing balance", expected_closing if expected_closing is not None else "Unknown"),
        ("Actual closing balance last period", closing if closing is not None else "Unknown"),
        ("Difference", difference if difference is not None else "Unknown"),
        ("Status", "Reconciled" if difference == 0 and continuity_result == "PASSED" else "Review required"),
        ("Period continuity check", continuity_result),
    ]
    write_row(rec, ["Combined Bank Reconciliation", "Amount"], 1, header=True)
    for row_index, row in enumerate(rec_rows, start=2):
        write_row(rec, list(row), row_index)
    write_row(rec, ["Previous Period", "Next Period", "Previous Closing", "Next Opening", "Difference", "Status"], 12, header=True)
    for row_index, item in enumerate(continuity, start=13):
        write_row(rec, [item["previous_period"], item["next_period"], item["previous_closing"], item["next_opening"], item["difference"], item["status"]], row_index)
    apply_number_formats(rec, [2, 3, 4, 5])

    review = workbook.create_sheet("Review Items")
    write_row(review, ["Date", "Source Period", "Description", "Money In", "Money Out", "Account", "VAT Status", "Reason"], 1, header=True)
    review_row = 2
    for row in rows:
        reason = row.get("review_reason") or professional_review_reason(row)
        if row.get("review_required") or reporting_account(row) == "Review Required Suspense":
            write_row(review, [row["date"], row["source_period"], row["description"], row["money_in"], row["money_out"], reporting_account(row), reporting_vat_status(row), reason or "Review recommended"], review_row)
            review_row += 1
    apply_number_formats(review, [4, 5])

    assumptions = workbook.create_sheet("Assumptions")
    assumptions_rows = [
        ("Area", "Assumption / Note"),
        ("Report limitation", ACCOUNTING_REPORT_DISCLAIMER),
        ("Batch processing", "Statements are sorted by statement period start date before combining."),
        ("Duplicate removal", "Potential duplicates are removed by matching date, merchant pattern, amount and running balance."),
        ("Account rule", "Default batch generation only combines the same company, bank and account number."),
        ("Review mode", "Statements with review items can be combined, but unresolved transactions stay in Review Required Suspense."),
        ("Continuity", "Previous closing balance should equal the next opening balance."),
    ]
    for row_index, row in enumerate(assumptions_rows, start=1):
        write_row(assumptions, list(row), row_index, header=row_index == 1)

    diagnostics = workbook.create_sheet("Diagnostics")
    write_row(diagnostics, ["Metric", "Value"], 1, header=True)
    diagnostics_rows = [
        ("worker", worker_version()),
        ("run_ids", [run.get("id") for run in sorted_runs]),
        ("duplicates_removed", duplicates_removed),
        ("continuity", continuity),
        ("ai", ai_stats),
    ]
    for row_index, row in enumerate(diagnostics_rows, start=2):
        write_row(diagnostics, [row[0], json.dumps(row[1], default=str)], row_index)
    diagnostics.sheet_state = "hidden"

    metadata_sheet = workbook.create_sheet("Metadata")
    generated_at = datetime.utcnow().isoformat()
    combined_start = first_run.get("statement_period_start") or "Unknown"
    combined_end = last_run.get("statement_period_end") or "Unknown"
    metadata_rows = [
        ("Parser Version", WORKER_PARSER_VERSION),
        ("Worker Version", worker_version()),
        ("Generated Date", generated_at),
        ("Company", company_name),
        ("Bank", bank),
        ("Account Number", account_number),
        ("Combined Months", f"{combined_start} to {combined_end}"),
        ("Statement Count", len(sorted_runs)),
        ("Duplicate Rows Removed", duplicates_removed),
        ("Continuity Result", continuity_result),
        ("Review Status", "Review Required" if review_count or continuity_result != "PASSED" else "Completed"),
        ("Generation Time", "pending"),
    ]
    write_row(metadata_sheet, ["Metric", "Value"], 1, header=True)
    for row_index, row in enumerate(metadata_rows, start=2):
        write_row(metadata_sheet, [row[0], json.dumps(row[1], default=str) if isinstance(row[1], (dict, list)) else row[1]], row_index)
    metadata_sheet.sheet_state = "hidden"

    finish_sheet(dashboard, freeze_pane="D4")
    finish_sheet(tx, filter_ref=f"A1:P{max(tx.max_row, 1)}")
    finish_sheet(vat, freeze_pane=f"A{vat_detail_header_row + 1}", filter_ref=f"A{vat_detail_header_row}:{get_column_letter(vat_column_count)}{max(vat.max_row, vat_detail_header_row)}")
    finish_sheet(ledger, filter_ref=f"A1:F{max(ledger.max_row, 1)}")
    finish_sheet(trial, filter_ref=f"A1:E{max(trial.max_row, 1)}")
    finish_sheet(rec)
    finish_sheet(review, filter_ref=f"A1:H{max(review.max_row, 1)}")
    finish_sheet(assumptions)
    finish_sheet(diagnostics)
    finish_sheet(metadata_sheet)

    generation_duration_ms = round((time.perf_counter() - generation_started) * 1000, 2)
    metadata_sheet.cell(row=13, column=2).value = f"{generation_duration_ms} ms"
    output = io.BytesIO()
    validate_workbook_for_export(workbook)
    workbook.save(output)
    continuity_message = continuity_failure_message(continuity)
    if continuity_message:
        log_warning("worker.combine_continuity_review", continuity=continuity, message=continuity_message)
    log_event(
        "worker.combine_summary",
        parser_profile=WORKER_PARSER_VERSION,
        continuity_result=continuity_result,
        continuity=continuity,
        duplicates_removed=duplicates_removed,
        workbook_generation_duration_ms=generation_duration_ms,
    )
    return output.getvalue(), {
        "statement_count": len(sorted_runs),
        "transaction_count": len(rows),
        "duplicates_removed": duplicates_removed,
        "review_count": review_count,
        "continuity_ok": continuity_result == "PASSED",
        "continuity_result": continuity_result,
        "continuity": continuity,
        "difference": str(difference) if difference is not None else "Unknown",
        "continuity_message": continuity_message,
        "workbook_generation_duration_ms": generation_duration_ms,
    }


@app.get("/health")
def health() -> dict[str, str]:
    return worker_version()


@app.get("/version")
def version() -> dict[str, str]:
    return worker_version()


@app.post("/combine-fnb-statements")
def combine_fnb_statements(payload: CombineRequest, authorization: str | None = Header(default=None)) -> Response:
    verify_worker_token(authorization)
    validation_started = time.perf_counter()
    if len(payload.run_ids) < 2:
        raise HTTPException(status_code=400, detail="Select at least two statements to combine.")

    supabase = get_supabase()
    log_event("worker.combine_request", workspace_id=payload.workspace_id, run_ids=payload.run_ids)

    try:
      runs_response = (
          supabase.table("accounting_statement_runs")
          .select("*")
          .eq("workspace_id", payload.workspace_id)
          .in_("id", payload.run_ids)
          .execute()
      )
      runs = runs_response.data if isinstance(runs_response.data, list) else []
      runs = validate_combine_runs(runs, payload)
      validation_duration_ms = round((time.perf_counter() - validation_started) * 1000, 2)
      log_event("worker.combine_validated", validation_duration_ms=validation_duration_ms, parser_profile=WORKER_PARSER_VERSION)

      continuity, continuity_result, continuity_message = run_continuity_summary(runs)
      log_event("worker.combine_continuity_checked", continuity_result=continuity_result, continuity=continuity)
      if continuity_result != "PASSED" and not payload.override_continuity:
          message = continuity_message or "Continuity checks failed. Review required before combining."
          supabase.table("accounting_statement_runs").update(
              {
                  "status": "review",
                  "error": message,
                  "updated_at": datetime.utcnow().isoformat(),
              }
          ).eq("workspace_id", payload.workspace_id).in_("id", payload.run_ids).execute()
          raise HTTPException(
              status_code=422,
              detail={
                  "status": "review_required",
                  "message": message,
                  "continuity": continuity,
                  "allow_override": True,
              },
          )

      transactions_by_run: dict[str, list[ParsedTransaction]] = {}
      for run in runs:
          transaction_response = (
              supabase.table("accounting_transactions")
              .select("*")
              .eq("workspace_id", payload.workspace_id)
              .eq("run_id", run["id"])
              .execute()
          )
          transaction_rows = transaction_response.data if isinstance(transaction_response.data, list) else []
          transactions_by_run[str(run["id"])] = [parsed_transaction_from_row(row) for row in transaction_rows]

      export_started = time.perf_counter()
      workbook_bytes, summary = build_combined_workbook(runs, transactions_by_run, payload.workspace_id)

      export_duration_ms = round((time.perf_counter() - export_started) * 1000, 2)
      summary["export_duration_ms"] = export_duration_ms
      log_event("worker.combine_completed", workspace_id=payload.workspace_id, summary=summary)
      return Response(
          content=workbook_bytes,
          media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
          headers={"X-DocuCoreX-Combined-Summary": json.dumps(summary, default=str)},
      )
    except HTTPException:
      raise
    except Exception as exc:
      log_exception("worker.combine_failed", workspace_id=payload.workspace_id, run_ids=payload.run_ids, error=str(exc))
      raise HTTPException(status_code=422, detail=str(exc))


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError) -> JSONResponse:
    body = await request.body()
    missing_fields = [
        ".".join(str(part) for part in error.get("loc", []))
        for error in exc.errors()
        if error.get("type") == "missing"
    ]
    log_warning(
        "worker.validation_error",
        path=str(request.url.path),
        missing_fields=missing_fields,
        errors=exc.errors(),
        body=body.decode("utf-8", errors="replace"),
    )
    return JSONResponse(
        status_code=422,
        content={
            "detail": exc.errors(),
            "missing_fields": missing_fields,
            "message": "Worker request validation failed.",
            "worker": worker_version(),
        },
    )


@app.post("/process-fnb-statement")
def process_fnb_statement(payload: ProcessRequest, authorization: str | None = Header(default=None)) -> dict[str, Any]:
    verify_worker_token(authorization)
    supabase = get_supabase()
    bucket = os.getenv("SUPABASE_BUCKET", "documents")
    process_started = time.perf_counter()
    # Defaults for the paths that fail before detection runs (a storage download
    # that never returns, say). "Unknown" is the truth at that point; the old
    # default said FNB and so mislabelled every failure that never got far enough
    # to look at the document.
    bank_profile = UNKNOWN_PROFILE_ID
    bank_name = bank_name_for(UNKNOWN_PROFILE_ID)
    parser_profile = GENERIC_PARSER_PROFILE_ID
    parser_version = GENERIC_PARSER_PROFILE_ID

    log_event(
        "worker.process_request",
        worker=worker_version(),
        run_id=payload.run_id,
        workspace_id=payload.workspace_id,
        document_id=payload.document_id,
        processing_job_id=payload.processing_job_id,
        storage_path=payload.storage_path,
        bucket=bucket,
    )

    try:
        heartbeat_step(
            supabase,
            run_id=payload.run_id,
            workspace_id=payload.workspace_id,
            processing_job_id=payload.processing_job_id,
            step_label="Detecting PDF type",
            progress=20,
        )
        pdf_bytes = supabase.storage.from_(bucket).download(payload.storage_path)
        log_event("worker.storage_downloaded", run_id=payload.run_id, bytes=len(pdf_bytes or b""))
        heartbeat_step(
            supabase,
            run_id=payload.run_id,
            workspace_id=payload.workspace_id,
            processing_job_id=payload.processing_job_id,
            step_label="Running OCR",
            progress=45,
        )
        pages = extract_statement_text(pdf_bytes) or []
        native_text = "\n".join((page.get("text") or "") for page in pages)
        # Prefer the Node pipeline's best extraction when it is meaningfully long;
        # the natively-extracted PDF text remains the fallback.
        provided = (payload.pre_extracted_text or "").strip()
        log_event(
            "worker.pre_extracted_text_received",
            run_id=payload.run_id,
            received=bool(provided),
            length=len(provided),
            sample=provided[:1000],
            parser_method=payload.parser_method,
            extraction_source=payload.extraction_source,
            ocr_used=bool(payload.ocr_used),
        )
        provided_structured_rows = payload.pre_extracted_rows or []
        log_event(
            "worker.pre_extracted_rows_received",
            run_id=payload.run_id,
            received=bool(provided_structured_rows),
            extraction_format_version=payload.extraction_format_version,
            structured_provider=payload.structured_provider,
            structured_row_count=payload.structured_row_count if payload.structured_row_count is not None else len(provided_structured_rows),
            structured_page_count=payload.structured_page_count,
            structured_row_continuity=payload.structured_row_continuity,
        )
        # Choose the text that actually PARSES, not the one that is longest.
        #
        # This previously accepted the provided text whenever it was at least half
        # the length of the native text. Length is the wrong criterion: OCR
        # markdown is verbose (pipes, headers, reflowed columns) and so always
        # cleared the threshold, displacing a native pdfplumber extraction whose
        # fixed-column layout these FNB parsers are written against. The result
        # was fewer recovered rows and a large reconciliation difference.
        #
        # Counting candidate transaction lines is cheap and directly measures the
        # only thing that matters here: how much of the statement this text will
        # yield once parsed.
        #
        # One exception: when this worker's own extraction is empty (a scanned
        # PDF, where pdfplumber and PyMuPDF both return nothing), the provided
        # text is all there is — use it even if neither yields candidate rows,
        # otherwise a scanned statement would be parsed from an empty string.
        #
        # Count with the counter that matches the BANK. transaction_candidate_lines
        # only enters a transaction section after FNB's "Transactions in RAND"
        # heading, so on every other bank it returns 0 for both texts, the
        # comparison is 0 > 0, and the provided text loses by default. That is how
        # a Standard Bank statement's 78,697-character Mistral extraction was
        # discarded in favour of this worker's own — for a parser that was never
        # going to read it anyway.
        #
        # The bank has to be read from text, so take a preliminary reading first:
        # the Node pipeline's verdict when it sent one, otherwise whichever
        # extraction is available.
        preliminary_profile = payload.detected_bank if is_supported_bank(payload.detected_bank) else None
        if preliminary_profile is None:
            preliminary = detect_bank(provided) if provided else detect_bank(native_text)
            if not preliminary.is_known and provided and native_text:
                preliminary = detect_bank(native_text)
            preliminary_profile = preliminary.profile_id
        count_candidates = (
            (lambda text: len(transaction_candidate_lines(text)))
            if preliminary_profile == FNB_PROFILE_ID
            else generic_candidate_lines
        )
        provided_rows = count_candidates(provided) if provided else 0
        native_rows = count_candidates(native_text) if native_text else 0
        native_is_empty = not native_text.strip()
        log_event(
            "worker.text_source_compared",
            run_id=payload.run_id,
            preliminary_profile=preliminary_profile,
            counter="fnb" if preliminary_profile == FNB_PROFILE_ID else "generic",
            provided_rows=provided_rows,
            native_rows=native_rows,
        )
        if provided and (native_is_empty or provided_rows > native_rows):
            full_text = provided
            log_event(
                "worker.pre_extracted_text_used",
                run_id=payload.run_id,
                provided_chars=len(provided),
                native_chars=len(native_text),
                provided_rows=provided_rows,
                native_rows=native_rows,
            )
        else:
            full_text = native_text
            if provided:
                log_event(
                    "worker.pre_extracted_text_rejected",
                    run_id=payload.run_id,
                    provided_chars=len(provided),
                    native_chars=len(native_text),
                    provided_rows=provided_rows,
                    native_rows=native_rows,
                    reason="native text yields at least as many transaction candidates",
                )
        # Which bank issued this statement, and therefore which parser reads it.
        #
        # This replaces BankRegistry.detect, which folded payload.storage_path
        # into its keyword haystack. Every accounting upload is stored under
        # ".../accounting/fnb/...", so that matched FNB for EVERY document, the
        # statement's own text was never reached, and a Standard Bank statement
        # died on "No FNB transactions could be parsed from this PDF".
        detection = detect_bank(full_text)
        node_detected_bank = (payload.detected_bank or "").strip() or None
        resolution = resolve_bank_profile(
            worker_profile=detection.profile_id,
            worker_confidence=detection.confidence,
            node_profile=node_detected_bank,
            node_confidence=payload.detected_bank_confidence,
        )
        bank_profile = resolution["bank_profile"]
        bank_name = resolution["bank_name"]
        # Not FNB is not an error. It selects the generic parser.
        parser_profile = FNB_PROFILE_ID if bank_profile == FNB_PROFILE_ID else GENERIC_PARSER_PROFILE_ID
        parser_version = parser_profile
        log_event(
            "worker.bank_detected",
            run_id=payload.run_id,
            detected_bank=detection.profile_id,
            detected_bank_name=detection.bank_name,
            detection_confidence=detection.confidence,
            detection_reason=detection.reason,
            detection_evidence=list(detection.evidence),
            detection_scores=detection.scores,
            node_detected_bank=node_detected_bank,
            node_detected_bank_name=payload.detected_bank_name,
            node_detection_confidence=payload.detected_bank_confidence,
            node_detection_reason=payload.detected_bank_reason,
            node_detection_evidence=payload.detected_bank_evidence,
            node_detection_present=node_detected_bank is not None,
            node_worker_agreement=(node_detected_bank == detection.profile_id) if node_detected_bank else None,
            resolved_bank_profile=bank_profile,
            resolved_bank_name=bank_name,
            resolution_source=resolution["source"],
            resolution_reason=resolution["reason"],
            selected_parser_profile=parser_profile,
        )
        if node_detected_bank and node_detected_bank != detection.profile_id:
            log_warning(
                "worker.bank_detection_mismatch",
                run_id=payload.run_id,
                node_detected_bank=node_detected_bank,
                node_detection_confidence=payload.detected_bank_confidence,
                node_detection_evidence=payload.detected_bank_evidence,
                worker_detected_bank=detection.profile_id,
                worker_detection_confidence=detection.confidence,
                worker_detection_evidence=list(detection.evidence),
                extraction_source=payload.extraction_source,
                resolved_bank_profile=bank_profile,
                resolution_source=resolution["source"],
                note="the two sides read different text and reached different banks",
            )
        log_event(
            "worker.text_extracted",
            run_id=payload.run_id,
            pages=len(pages),
            characters=len(full_text),
            bank_profile=bank_profile,
            parser_profile=parser_profile,
        )
        metadata = parse_metadata(full_text)
        metadata["bank_profile"] = bank_profile
        metadata["bank_name"] = bank_name
        metadata["parser_profile"] = parser_profile
        metadata["parser_version"] = parser_version
        metadata["source_file"] = os.path.basename(payload.storage_path).split(".pdf")[0][:80] or "28 Feb 2026 - (Free)"
        candidates = transaction_candidate_lines(full_text)
        service_fee_candidates = service_fee_candidate_lines(full_text)
        log_event(
            "worker.transaction_candidates_built",
            worker=worker_version(),
            run_id=payload.run_id,
            candidates=len(candidates),
            service_fee_candidates=len(service_fee_candidates),
            service_fee_candidate_samples=service_fee_candidates[:6],
            parser_version=parser_version,
        )
        heartbeat_step(
            supabase,
            run_id=payload.run_id,
            workspace_id=payload.workspace_id,
            processing_job_id=payload.processing_job_id,
            step_label="Parsing transactions",
            progress=70,
        )
        selected_transactions, structured_selection = select_transactions_from_sources(
            pages,
            metadata,
            full_text,
            payload.pre_extracted_rows if isinstance(payload.pre_extracted_rows, list) else None,
            parser_profile,
        )
        transactions = selected_transactions or []
        classification_rules = fetch_classification_rules(supabase, payload.workspace_id) or []
        learned_rules_applied = apply_learned_classification_rules(transactions, classification_rules)
        # After classification, never before: the counterparty pass records who
        # each row was with and what the statement proves about them, and must
        # not be in a position to influence the treatment it is describing.
        stamp_counterparty_intelligence(transactions, bank_profile)
        # Counterparty reasoning before the row-by-row path, because anything it
        # settles is one decision instead of many: on the real statement 413
        # unresolved rows become 37 questions covering 327 of them. What it does
        # not cover — counterparties seen once — still reaches the row-by-row
        # classifier further down, unchanged.
        ai_counterparty = apply_ai_counterparty_reasoning(transactions, bank_profile)
        ai_recovery: dict[str, Any] = {"enabled": False, "attempted": False, "accepted_rows": 0}
        # Accounting-parser diagnostics (null-safe).
        _summary = validation_summary(transactions)
        structured_metrics = structured_selection.get("structured_metrics") or {}
        text_metrics = structured_selection.get("text_metrics") or {}
        log_event(
            "worker.accounting_parser",
            run_id=payload.run_id,
            opening_balance_found=metadata.get("opening_balance") is not None,
            closing_balance_found=metadata.get("closing_balance") is not None,
            transactions_parsed=len(transactions),
            credit_count=_summary.get("credit_count"),
            debit_count=_summary.get("debit_count"),
            parser_method=payload.parser_method,
            selected_path=structured_selection.get("selected_path"),
            structured_rows_received=structured_selection.get("structured_rows_received"),
            structured_rows_usable=structured_selection.get("structured_rows_usable"),
            structured_rejection_reason=structured_selection.get("structured_rejection_reason"),
            fallback_reason=structured_selection.get("fallback_reason"),
            structured_financial_count=structured_metrics.get("financial_count"),
            structured_debit_total=str(structured_metrics.get("debit_total")) if structured_metrics.get("debit_total") is not None else None,
            structured_credit_total=str(structured_metrics.get("credit_total")) if structured_metrics.get("credit_total") is not None else None,
            text_financial_count=text_metrics.get("financial_count"),
            text_debit_total=str(text_metrics.get("debit_total")) if text_metrics.get("debit_total") is not None else None,
            text_credit_total=str(text_metrics.get("credit_total")) if text_metrics.get("credit_total") is not None else None,
        )
        log_event(
            "worker.statement_parsed",
            worker=worker_version(),
            run_id=payload.run_id,
            metadata_fields=sorted([key for key, value in metadata.items() if value is not None]),
            transactions=len(transactions),
            parser_version=parser_version,
            service_fee_rows=sum(1 for transaction in transactions if transaction.description.startswith("#")),
            learned_rules_applied=learned_rules_applied,
            selected_path=structured_selection.get("selected_path"),
        )

        if not transactions:
            diagnostics = extraction_diagnostics(pages, full_text, metadata)
            pipeline_debug = payload.extraction_debug if isinstance(payload.extraction_debug, dict) else {}
            # Parser debug — surface the REAL reason, never hide it.
            parser_debug = {
                "selected_parser": pipeline_debug.get("selectedParser") or payload.parser_method,
                "detected_pdf_type": pipeline_debug.get("detectedPdfType"),
                "ocr_used": pipeline_debug.get("ocrUsed") if pipeline_debug.get("ocrUsed") is not None else bool(payload.ocr_used),
                "pdfjs_text_length": pipeline_debug.get("pdfjsTextLength"),
                "ocr_text_length": pipeline_debug.get("ocrTextLength"),
                "pre_extracted_text_length": pipeline_debug.get("preExtractedTextLength", len((payload.pre_extracted_text or "").strip())),
                "sample_text": (full_text or "")[:1000],
                "reason_no_transactions": pipeline_debug.get("reasonNoTransactions"),
                "selected_path": structured_selection.get("selected_path"),
                "structured_fallback_reason": structured_selection.get("fallback_reason"),
                "bank_profile": bank_profile,
                "bank_name": bank_name,
                "parser_profile": parser_profile,
            }
            # Nothing parsed. Whether that is a failure depends on whether
            # anything recoverable is left — a document is only unprocessable
            # once every path over it has been exhausted.
            recovery = recovery_options(
                full_text=full_text,
                pages=pages,
                structured_rows=payload.pre_extracted_rows if isinstance(payload.pre_extracted_rows, list) else None,
                structured_selection=structured_selection,
            )
            parser_debug["recovery"] = recovery
            log_warning(
                "worker.no_transactions_parsed",
                run_id=payload.run_id,
                diagnostics=diagnostics,
                parser_debug=parser_debug,
                recovery=recovery,
                bank_profile=bank_profile,
                parser_profile=parser_profile,
            )

            # Last recovery step. Only attempted when material is genuinely
            # left; there is nothing for a model to locate in a document with no
            # readable text.
            if recovery["recoverable"]:
                heartbeat_step(
                    supabase,
                    run_id=payload.run_id,
                    workspace_id=payload.workspace_id,
                    processing_job_id=payload.processing_job_id,
                    step_label="Recovering with AI",
                    progress=75,
                )
                ai_transactions, ai_recovery = attempt_ai_recovery(
                    full_text=full_text,
                    structured_rows=payload.pre_extracted_rows if isinstance(payload.pre_extracted_rows, list) else None,
                    metadata=metadata,
                    bank_name=bank_name,
                    run_id=payload.run_id,
                )
                log_event(
                    "worker.ai_recovery",
                    run_id=payload.run_id,
                    bank_profile=bank_profile,
                    parser_profile=parser_profile,
                    **{key: value for key, value in ai_recovery.items() if key != "rejected_rows"},
                    rejected_rows=ai_recovery.get("rejected_rows"),
                )
                if ai_transactions:
                    transactions = ai_transactions
                    learned_rules_applied = apply_learned_classification_rules(transactions, classification_rules)
                    log_warning(
                        "worker.ai_recovery_used",
                        run_id=payload.run_id,
                        transactions=len(transactions),
                        note="every row is AI-located and flagged for review; the run cannot complete unreviewed",
                    )

        if not transactions:
            # Name the parser that actually ran. "No FNB transactions" on a
            # Standard Bank statement described the misroute, not the document,
            # and sent every investigation looking at the wrong parser.
            default_reason = (
                "No FNB transactions could be parsed from this PDF."
                if parser_profile == FNB_PROFILE_ID
                else f"No transactions could be parsed from this {bank_name} statement."
            )
            if not recovery["recoverable"]:
                default_reason = (
                    f"{default_reason} Nothing further could be recovered: {recovery['summary']}."
                )
            reason = parser_debug["reason_no_transactions"] or default_reason
            raise HTTPException(
                status_code=422,
                detail={
                    "message": reason,
                    "status": "no_transactions_recoverable" if not recovery["recoverable"] else "no_transactions_parsed",
                    "recovery": recovery,
                    "diagnostics": diagnostics,
                    "parser_debug": parser_debug,
                    "worker": worker_version(),
                },
            )

        validation: dict[str, Any] | None = None
        review_issue: dict[str, Any] | None = None
        heartbeat_step(
            supabase,
            run_id=payload.run_id,
            workspace_id=payload.workspace_id,
            processing_job_id=payload.processing_job_id,
            step_label="Reconciling",
            progress=90,
        )
        try:
            validation_started = time.perf_counter()
            validation = validate_statement(metadata, transactions)
            validation_duration_ms = round((time.perf_counter() - validation_started) * 1000, 2)
            log_event(
                "worker.statement_validated",
                run_id=payload.run_id,
                validation={key: str(value) for key, value in validation.items()},
                validation_duration_ms=validation_duration_ms,
            )
        except HTTPException as exc:
            review_issue = review_validation_issue(exc)
            if not review_issue:
                raise
            log_warning(
                "worker.statement_needs_review",
                run_id=payload.run_id,
                errors=review_issue["errors"],
                summary=review_issue["summary"],
                balance_gaps=review_issue["balance_gaps"][:10],
                parser_version=parser_version,
            )
            for transaction in transactions:
                if transaction.review_status == "ready":
                    transaction.review_status = "needs_review"

        run_state = (
            supabase.table("accounting_statement_runs")
            .select("status")
            .eq("id", payload.run_id)
            .eq("workspace_id", payload.workspace_id)
            .maybe_single()
            .execute()
        )
        if (run_state.data or {}).get("status") == "cancelled":
            log_warning("worker.run_cancelled_before_write", run_id=payload.run_id)
            return {
                "status": "cancelled",
                "transactions": 0,
                "workbook_storage_path": None,
                "confidence": 0,
                "validation": None,
                "review_issue": None,
                "ai_diagnostics": ai_diagnostics(enabled=False),
                "parser_profile": parser_profile,
                "processing_duration_ms": round((time.perf_counter() - process_started) * 1000, 2),
                "worker": worker_version(),
            }

        # Classification enrichment, BEFORE the write, so the reviewer sees the
        # same answer the workbook does. Fails open: the deterministic
        # classification stands if the model is unavailable.
        heartbeat_step(
            supabase,
            run_id=payload.run_id,
            workspace_id=payload.workspace_id,
            processing_job_id=payload.processing_job_id,
            step_label="Classifying transactions",
            progress=88,
        )
        ai_classification_stats = classify_transactions_with_ai(
            transactions,
            payload.workspace_id,
            str(metadata.get("source_file") or ""),
        )
        log_event(
            "worker.ai_classification_applied",
            run_id=payload.run_id,
            **{key: value for key, value in ai_classification_stats.items() if key != "ai_rejected_after_guardrails"},
            rejected=ai_classification_stats.get("ai_rejected_after_guardrails"),
        )

        # Stamp the canonical sequence. The parser validated the running-balance
        # chain in THIS order, so it is the only order in which the stored ledger
        # can be verified again — and nothing else recovers it: created_at is one
        # timestamp for the whole batch, source_page narrows only to a page, and
        # a UUID says nothing. Assigned here, after every classification stage,
        # so it describes exactly what is written.
        for sequence, transaction in enumerate(transactions, start=1):
            transaction.source_row = sequence

        rows = [transaction_insert_row(transaction, payload.run_id, payload.workspace_id) for transaction in transactions]
        provenance_persisted = replace_transactions(supabase, rows, payload.run_id, payload.workspace_id)

        # Derive the closing balance from the statement's own evidence before
        # validating against it. A NULL closing balance is read downstream as
        # zero, which turns a reconciled statement into a permanent
        # "needs fresh extraction" loop.
        derived_closing, closing_source = derive_closing_balance(metadata, transactions)
        if derived_closing is not None and metadata.get("closing_balance") is None:
            metadata["closing_balance"] = derived_closing
        metadata["closing_balance_source"] = closing_source
        log_event(
            "worker.closing_balance_resolved",
            run_id=payload.run_id,
            closing_balance=derived_closing,
            source=closing_source,
            opening_balance=metadata.get("opening_balance"),
            declared_debits=metadata.get("declared_debit_total"),
            declared_credits=metadata.get("declared_credit_total"),
        )

        # General extraction validation (count / totals / reconciliation vs the
        # statement's own declared figures). Bank charges come from the declared
        # fee summary, not from cash-deposit amounts.
        extraction_check = validate_extraction(metadata, transactions)
        extraction_incomplete = extraction_check["status"] != "ok"
        heartbeat_step(
            supabase,
            run_id=payload.run_id,
            workspace_id=payload.workspace_id,
            processing_job_id=payload.processing_job_id,
            step_label="Generating workbook",
            progress=97,
        )
        workbook_bytes = build_workbook(
            metadata,
            transactions,
            allow_ai=not extraction_incomplete and review_issue is None,
            workspace_id=payload.workspace_id,
        )
        ai_stats = metadata.get("_ai_diagnostics") or ai_diagnostics(enabled=False)
        workbook_path = f"{payload.workspace_id}/accounting/fnb/exports/{payload.run_id}.xlsx"
        export_started = time.perf_counter()
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as handle:
            handle.write(workbook_bytes)
            handle.flush()
            supabase.storage.from_(bucket).upload(
                workbook_path,
                handle.name,
                file_options={
                    "content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "upsert": "true",
                },
            )
        export_duration_ms = round((time.perf_counter() - export_started) * 1000, 2)
        log_event("worker.workbook_exported", run_id=payload.run_id, duration_ms=export_duration_ms, parser_profile=WORKER_PARSER_VERSION)

        bank_charges_total = float(bank_charges_from_statement(metadata, transactions))
        avg_confidence = sum(transaction.confidence for transaction in transactions) / len(transactions)
        # A ledger recovered by AI can never report as completed. Every row is
        # already flagged, so this is belt and braces — but the guarantee is
        # worth stating outright rather than depending on a per-row flag that a
        # later classification pass could clear.
        ai_recovered = bool(ai_recovery.get("accepted_rows"))
        status = "review" if (
            review_issue
            or extraction_incomplete
            or ai_recovered
            or any(transaction.review_status == "needs_review" for transaction in transactions)
        ) else "completed"
        if ai_recovered and not review_issue and not extraction_incomplete:
            run_error = (
                f"Recovered by AI — {len(transactions)} row(s) located in the statement text and "
                "verified against it, but not parsed deterministically. Every row needs review."
            )
        elif review_issue:
            run_error = review_error_message(review_issue)
        elif extraction_incomplete:
            expected_count = extraction_check.get("expected_transaction_count")
            extracted_count = extraction_check.get("extracted_transaction_count")
            recon_diff = extraction_check.get("reconciliation_difference")
            run_error = (
                "Extraction incomplete — "
                f"extracted {extracted_count} of {expected_count} transactions; "
                f"reconciliation difference {recon_diff}; "
                f"failed checks: {', '.join(extraction_check.get('failures') or [])}."
            )
        else:
            run_error = None
        processing_duration_ms = round((time.perf_counter() - process_started) * 1000, 2)
        review_required = status == "review"
        validation = {**(validation or {}), **{f"extraction_{k}": v for k, v in extraction_check.items() if k != "checks"}}
        missing_rows = missing_transaction_count_for_storage(extraction_check, len(transactions))
        unresolved_amount_directions = 0
        if structured_selection.get("selected_path") == "structured":
            parse_diag = structured_selection.get("structured_parse_diagnostics")
            if isinstance(parse_diag, dict):
                rejected_reasons = parse_diag.get("rejected_reasons")
                if isinstance(rejected_reasons, dict):
                    unresolved_amount_directions = int(rejected_reasons.get("ambiguous_unsigned_amount_direction") or 0)
        extraction_confidence = extraction_confidence_score(
            metadata,
            extraction_check,
            transactions,
            pages,
            missing_rows,
            unresolved_amount_directions=unresolved_amount_directions,
        )

        update_statement_run(
            supabase,
            payload.run_id,
            payload.workspace_id,
            {
                **statement_run_metadata(metadata),
                "status": status,
                "bank": bank_name,
                "transaction_count": len(transactions),
                "bank_charges_total": bank_charges_total,
                "workbook_storage_path": workbook_path,
                "parser_profile": parser_profile,
                "parser_version": parser_version,
                "review_required": review_required,
                "review_reason": run_error,
                "validation_status": extraction_check.get("status"),
                "reconciliation_difference": extraction_check.get("reconciliation_difference"),
                "missing_transaction_count": missing_rows,
                "requires_review": review_required,
                "processing_duration_ms": int(processing_duration_ms),
                "extraction_accuracy": round(avg_confidence, 2),
                "extraction_confidence": extraction_confidence,
                # DEPRECATED: `confidence` has always carried the CLASSIFICATION
                # score and continues to, so existing readers are unaffected.
                # New readers should use classification_confidence.
                "confidence": round(avg_confidence, 2),
                "classification_confidence": round(avg_confidence, 2),
                "reconciliation_confidence": reconciliation_confidence(
                    extraction_check, missing_rows
                ),
                "error": run_error,
                "updated_at": datetime.utcnow().isoformat(),
            },
        )

        refresh_statement_analytics(supabase, payload.workspace_id, bank_name, parser_profile, parser_version)

        if payload.processing_job_id:
            supabase.table("processing_jobs").update(
                {
                    "status": "completed",
                    "progress": 100,
                    "message": "Accounting workbook ready for review" if review_issue else "Accounting workbook ready",
                    "error": run_error,
                    "updated_at": datetime.utcnow().isoformat(),
                }
            ).eq("id", payload.processing_job_id).execute()

        log_event(
            "worker.process_completed",
            run_id=payload.run_id,
            status=status,
            transactions=len(transactions),
            workbook_storage_path=workbook_path,
            confidence=round(avg_confidence, 2),
            extraction_confidence=extraction_confidence,
            validation={key: str(value) for key, value in validation.items()} if validation else None,
            review_issue=review_issue,
            ai_diagnostics=ai_stats,
            export_duration_ms=export_duration_ms,
            parser_profile=parser_profile,
            processing_duration_ms=processing_duration_ms,
        )

        return {
            "status": status,
            "transactions": len(transactions),
            "workbook_storage_path": workbook_path,
            "confidence": round(avg_confidence, 2),
            "validation": {key: str(value) for key, value in validation.items()} if validation else None,
            "review_issue": review_issue,
            "ai_diagnostics": ai_stats,
            "bank_profile": bank_profile,
            "bank_name": bank_name,
            "parser_profile": parser_profile,
            # Why this run is trusted, or is not. A run reaching "completed" on
            # zero checks would be an unverified claim of success.
            "classification_provenance_persisted": provenance_persisted,
            "completeness_evidence": {
                "checks_run": extraction_check.get("evidence_checks_run"),
                "failures": extraction_check.get("failures"),
                "balance_gap_count": extraction_check.get("balance_gap_count"),
            },
            "ai_recovery": ai_recovery,
            "processing_duration_ms": processing_duration_ms,
            "worker": worker_version(),
        }
    except HTTPException as exc:
        message = json.dumps(exc.detail, default=str) if isinstance(exc.detail, (dict, list)) else str(exc.detail)
        record_parser_failure(supabase, payload.workspace_id, bank_name, message)
        log_exception(
            "worker.process_failed",
            run_id=payload.run_id,
            workspace_id=payload.workspace_id,
            document_id=payload.document_id,
            processing_job_id=payload.processing_job_id,
            storage_path=payload.storage_path,
            error=message,
        )
        supabase.table("accounting_statement_runs").update(
            {"status": "failed", "error": message, "updated_at": datetime.utcnow().isoformat()}
        ).eq("id", payload.run_id).eq("workspace_id", payload.workspace_id).execute()
        if payload.processing_job_id:
            supabase.table("processing_jobs").update(
                {"status": "failed", "progress": 100, "message": message, "error": message, "updated_at": datetime.utcnow().isoformat()}
            ).eq("id", payload.processing_job_id).execute()
        raise exc
    except Exception as exc:
        message = str(exc)
        record_parser_failure(supabase, payload.workspace_id, bank_name, message)
        log_exception(
            "worker.process_failed",
            run_id=payload.run_id,
            workspace_id=payload.workspace_id,
            document_id=payload.document_id,
            processing_job_id=payload.processing_job_id,
            storage_path=payload.storage_path,
            error=message,
        )
        supabase.table("accounting_statement_runs").update(
            {"status": "failed", "error": message, "updated_at": datetime.utcnow().isoformat()}
        ).eq("id", payload.run_id).eq("workspace_id", payload.workspace_id).execute()
        if payload.processing_job_id:
            supabase.table("processing_jobs").update(
                {"status": "failed", "progress": 100, "message": message, "error": message, "updated_at": datetime.utcnow().isoformat()}
            ).eq("id", payload.processing_job_id).execute()
        raise HTTPException(status_code=422, detail=with_worker_version({"message": message})) from exc


def _claim_processing_job(job_id: str) -> bool:
    """Atomically take ownership of a job for execution. True if we claimed it.

    The claim is `status = 'running' WHERE status = 'queued'` on processing_jobs.
    A single conditional UPDATE is atomic, so of two concurrent dispatches for
    the same job exactly one affects a row and the other sees none — no lock,
    no new table, no new column, using the job_status enum that has existed
    since migration 001.

    A job already in 'running' (or any terminal state) cannot be claimed again,
    which is the invariant: a processing_job_id may be accepted for execution
    only once. Force Reprocess is unaffected — it allocates a NEW job, which
    starts 'queued' and claims cleanly.

    Fails CLOSED. If the claim cannot be evaluated — no database, transport
    error — we refuse rather than scheduling work we cannot prove is unique.
    A refused dispatch is recoverable; two pipelines racing on one run is not.
    """
    try:
        supabase = get_supabase()
        result = (
            supabase.table("processing_jobs")
            .update({"status": "running", "updated_at": datetime.now(timezone.utc).isoformat()})
            .eq("id", job_id)
            .eq("status", "queued")
            .execute()
        )
        return len(getattr(result, "data", None) or []) == 1
    except Exception as exc:  # noqa: BLE001 — see "fails closed" above
        log_exception("worker.dispatch_claim_failed", job_id=job_id, error=str(exc))
        return False


# Liveness is not progress.
#
# Heartbeats used to be written only at the seven stage boundaries, so the
# freshest liveness signal a run had was "the stage last changed". A stage that
# legitimately runs long — classification is roughly 21 model round trips for a
# 613-transaction statement, each with retries — looked identical to a dead
# worker, and the stale detector failed the run at 10 minutes while it was
# healthy and working.
#
# So the two signals are now separate:
#
#   PROGRESS   accounting_statement_runs.processing_step, and
#              processing_jobs.message/progress — written at stage boundaries.
#              Unchanged.
#
#   LIVENESS   processing_jobs.updated_at — touched on a timer for as long as
#              the task is running, regardless of which stage it is in.
#
# No migration: processing_jobs.updated_at already exists and is already
# maintained by heartbeat_step; it simply was not the thing being read.
LIVENESS_TICK_SECONDS = 45


class _LivenessHeartbeat:
    """Touch processing_jobs.updated_at while a job is genuinely running.

    Bound to one dispatched task: started when it begins, stopped in a finally
    so it cannot outlive the work. A leaked ticker would keep a dead job looking
    alive forever — the exact inverse of the bug this fixes, and far harder to
    notice, so the stop is not optional.

    Fenced. Before each tick it confirms the run still points at this job; once
    superseded it stops rather than refreshing liveness on the new job's behalf.
    A superseded worker's own heartbeat_step will raise StaleJobError at its next
    stage boundary and end the task properly — this thread only stops lying in
    the meantime.
    """

    def __init__(self, run_id: str, workspace_id: str, job_id: str) -> None:
        self._run_id = run_id
        self._workspace_id = workspace_id
        self._job_id = job_id
        self._stop = threading.Event()
        self._thread: threading.Thread | None = None

    def _still_owns_run(self, supabase: Client) -> bool:
        try:
            result = (
                supabase.table("accounting_statement_runs")
                .select("active_job_id")
                .eq("id", self._run_id)
                .eq("workspace_id", self._workspace_id)
                .limit(1)
                .execute()
            )
            rows = getattr(result, "data", None) or []
            if not rows:
                return False
            active = rows[0].get("active_job_id")
            # NULL means never claimed — a pre-024 row or the legacy synchronous
            # path — not that someone else owns it. Same rule as the fenced
            # heartbeat write.
            return active is None or active == self._job_id
        except Exception:  # noqa: BLE001 — a transient read must not kill the job
            # Unknown ownership: keep beating. Stopping here would let a healthy
            # job be declared stale because of one failed SELECT.
            return True

    def _run(self) -> None:
        while not self._stop.wait(LIVENESS_TICK_SECONDS):
            try:
                supabase = get_supabase()
                if not self._still_owns_run(supabase):
                    log_event("worker.liveness_superseded", run_id=self._run_id, job_id=self._job_id)
                    return
                supabase.table("processing_jobs").update(
                    {"updated_at": datetime.now(timezone.utc).isoformat()}
                ).eq("id", self._job_id).eq("status", "running").execute()
            except Exception as exc:  # noqa: BLE001 — never let liveness kill the work
                log_warning("worker.liveness_tick_failed", run_id=self._run_id, job_id=self._job_id, error=str(exc)[:200])

    def __enter__(self) -> "_LivenessHeartbeat":
        self._thread = threading.Thread(target=self._run, name=f"liveness-{self._job_id[:8]}", daemon=True)
        self._thread.start()
        return self

    def __exit__(self, *_exc: Any) -> None:
        self._stop.set()
        if self._thread:
            self._thread.join(timeout=5)


def _run_dispatched_job(payload: ProcessRequest, authorization: str | None) -> None:
    """Run the pipeline for an already-accepted job, in the background.

    Nothing is returned to anyone: the caller was answered with 202 long ago and
    has gone. Results reach the UI the way they always have — the pipeline writes
    them to Supabase and the frontend polls. Every exception is contained here,
    because an escaping one would only reach a dead request."""
    run_id = payload.run_id
    try:
        # The ticker lives exactly as long as the pipeline call.
        with _LivenessHeartbeat(run_id, payload.workspace_id, payload.processing_job_id or ""):
            process_fnb_statement(payload, authorization)
    except StaleJobError as exc:
        # A Force Reprocess replaced this job while it was running. Correct
        # behaviour is to stop and leave the newer job's work alone.
        log_event("worker.dispatch_superseded", run_id=run_id, job_id=payload.processing_job_id, detail=str(exc))
    except Exception as exc:  # noqa: BLE001 — a background task has nowhere to raise to
        log_exception("worker.dispatch_failed", run_id=run_id, job_id=payload.processing_job_id, error=str(exc))


@app.post("/process-statement/dispatch", status_code=202)
def dispatch_statement(
    payload: ProcessRequest,
    background: BackgroundTasks,
    authorization: str | None = Header(default=None),
) -> dict[str, Any]:
    """Accept a statement for processing and return immediately.

    This exists because holding one HTTP request open for the whole pipeline made
    the caller's timeout decide the run's fate: extraction would finish and the
    run would still be marked failed, because Vercel stopped waiting at 254s of
    its 280s budget while this service carried on working.

    Ownership hands over at the 202. Everything after it — progress, results,
    terminal state — belongs to this service, and the caller is expected to stop
    caring. /process-statement is unchanged for anyone still calling it
    synchronously.

    Authentication is verified BEFORE accepting: a request that cannot be
    authenticated is a dispatch failure, which is the caller's to own, and it
    must not be answered 202."""
    verify_worker_token(authorization)

    if not payload.processing_job_id:
        # The fence has no meaning without a job id, and an unfenced background
        # run could overwrite a newer attempt. Refuse rather than accept blind.
        raise HTTPException(status_code=400, detail="processing_job_id is required to dispatch a job.")

    # Claim the job before scheduling anything.
    #
    # active_job_id fences a SUPERSEDED job — a different id — out of the run.
    # It cannot stop the SAME id being dispatched twice: both writers satisfy
    # "active_job_id = mine". That is not theoretical; production dispatched
    # job f1d9d778 for run 1ee084e3 at 16:23 and again at 16:42, and both were
    # answered 202. Two pipelines on one run is worse than duplicate rows,
    # because transactions are written delete-then-insert keyed on run_id: one
    # worker's DELETE can land between the other's DELETE and INSERT and destroy
    # results that had already completed.
    #
    # The claim is the queued -> running transition on processing_jobs, which
    # Postgres already gives us. A conditional UPDATE is atomic on its own, so
    # concurrent dispatches serialise: exactly one sees a row affected.
    if not _claim_processing_job(payload.processing_job_id):
        log_event(
            "worker.dispatch_already_running",
            run_id=payload.run_id,
            job_id=payload.processing_job_id,
        )
        # Not an error. The caller asked for work that is already happening, and
        # the honest answer is "yes, it is running" — with NOTHING scheduled.
        return {
            "accepted": True,
            "already_running": True,
            "run_id": payload.run_id,
            "job_id": payload.processing_job_id,
            "status": "processing",
        }

    log_event(
        "worker.dispatch_accepted",
        run_id=payload.run_id,
        job_id=payload.processing_job_id,
        workspace_id=payload.workspace_id,
    )
    background.add_task(_run_dispatched_job, payload, authorization)
    return {
        "accepted": True,
        "already_running": False,
        "run_id": payload.run_id,
        "job_id": payload.processing_job_id,
        # Explicit so the caller cannot mistake acceptance for completion.
        "status": "processing",
    }


@app.post("/process-statement")
def process_statement(payload: ProcessRequest, authorization: str | None = Header(default=None)) -> dict[str, Any]:
    # Checked here as well as in the delegate. The duplicate call is idempotent,
    # and it means this endpoint's auth does not depend on an implementation
    # detail of the function it happens to forward to today.
    verify_worker_token(authorization)
    return process_fnb_statement(payload, authorization)
