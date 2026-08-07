from __future__ import annotations

import importlib.util
import io
import json
import os
import sys
import types
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

if importlib.util.find_spec("fitz") is None:
    sys.modules["fitz"] = types.ModuleType("fitz")

if importlib.util.find_spec("pdfplumber") is None:
    sys.modules["pdfplumber"] = types.ModuleType("pdfplumber")

if importlib.util.find_spec("fastapi") is None:
    fastapi = types.ModuleType("fastapi")

    class FastAPI:
        def __init__(self, *args, **kwargs):
            pass

        def get(self, *args, **kwargs):
            return lambda func: func

        def post(self, *args, **kwargs):
            return lambda func: func

        def exception_handler(self, *args, **kwargs):
            return lambda func: func

    class HTTPException(Exception):
        def __init__(self, status_code=500, detail=None):
            super().__init__(detail)
            self.status_code = status_code
            self.detail = detail

    def Header(default=None, *args, **kwargs):
        return default

    fastapi.FastAPI = FastAPI
    fastapi.Header = Header
    fastapi.HTTPException = HTTPException
    fastapi.Request = object
    sys.modules["fastapi"] = fastapi

    fastapi_exceptions = types.ModuleType("fastapi.exceptions")
    fastapi_exceptions.RequestValidationError = Exception
    sys.modules["fastapi.exceptions"] = fastapi_exceptions

    fastapi_responses = types.ModuleType("fastapi.responses")
    fastapi_responses.JSONResponse = dict
    fastapi_responses.Response = dict
    sys.modules["fastapi.responses"] = fastapi_responses

if importlib.util.find_spec("pydantic") is None:
    pydantic = types.ModuleType("pydantic")

    class BaseModel:
        def __init__(self, **kwargs):
            for cls in reversed(self.__class__.mro()):
                for key, value in getattr(cls, "__dict__", {}).items():
                    if not key.startswith("_") and key not in {"model_dump"} and not callable(value):
                        setattr(self, key, value)
            for key, value in kwargs.items():
                setattr(self, key, value)

        def model_dump(self, *args, **kwargs):
            return dict(self.__dict__)

    pydantic.BaseModel = BaseModel
    sys.modules["pydantic"] = pydantic

if importlib.util.find_spec("supabase") is None:
    supabase = types.ModuleType("supabase")
    supabase.Client = object

    def create_client(*args, **kwargs):
        return object()

    supabase.create_client = create_client
    sys.modules["supabase"] = supabase

if importlib.util.find_spec("openpyxl") is None:
    raise RuntimeError(
        "openpyxl is required for regression workbook verification. Install worker deps before running regression suite."
    )

from openpyxl import load_workbook

import main
from main import (
    ParsedTransaction,
    apply_ai_result_to_row,
    apply_learned_classification_rules,
    build_combined_workbook,
    build_workbook,
    parse_metadata,
    parse_fnb_transactions,
    professional_transaction_row,
    validate_statement,
    validation_summary,
)


FNB_PROFILE = "fnb_business_v1"
GENERIC_PROFILE = "generic_bank_statement_v1"


def run_statement_period_case() -> None:
    # ALLIANZ 31 March 2026 statement: the period end and statement date must be
    # read from the PDF so the app names it March 2026 (not the July upload date).
    case_id = "allianz-statement-period"
    text = (
        "ALLIANZ HOLDINGS (PTY) LTD\n"
        "Account Number: 63012589818\n"
        "Statement Period: 28 February 2026 to 31 March 2026\n"
        "Statement Date: 31 March 2026\n"
    )
    meta = parse_metadata(text)
    assert_equal(meta["statement_period_start"], "2026-02-28", f"{case_id} period start")
    assert_equal(meta["statement_period_end"], "2026-03-31", f"{case_id} period end")
    assert_equal(meta["statement_date"], "2026-03-31", f"{case_id} statement date")
    # Naming must come from period end (March), not the upload month.
    assert_equal(meta["statement_period_end"][:7], "2026-03", f"{case_id} names to March 2026")


ROOT = Path(__file__).resolve().parents[3]
MANIFEST_PATH = ROOT / "workers" / "accounting_worker" / "tests" / "fixtures" / "regression_manifest.json"


def assert_equal(actual, expected, label: str) -> None:
    if actual != expected:
        raise AssertionError(f"{label}: expected {expected!r}, got {actual!r}")


def run_synthetic_case(case_id: str, fixture_path: Path) -> None:
    fixture = json.loads(fixture_path.read_text())
    metadata = fixture["metadata"]
    transactions = [ParsedTransaction(**row) for row in fixture["transactions"]]
    expected = fixture["expected"]

    validation = validate_statement(metadata, transactions)
    summary = validation_summary(transactions)

    assert_equal(str(validation["opening_balance"]), expected["opening_balance"], f"{case_id} opening balance")
    assert_equal(str(validation["closing_balance"]), expected["closing_balance"], f"{case_id} closing balance")
    assert_equal(str(summary["total_credits"]), expected["total_credits"], f"{case_id} total credits")
    assert_equal(str(summary["total_debits"]), expected["total_debits"], f"{case_id} total debits")
    assert_equal(validation["transaction_count"], expected["transaction_count"], f"{case_id} transaction count")
    assert_equal(
        str(validation["calculated_closing"]),
        str(validation["closing_balance"]),
        f"{case_id} reconciliation",
    )

    workbook_bytes = build_workbook(metadata, transactions)
    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    for sheet_name in expected["workbook_required_sheets"]:
        if sheet_name not in workbook.sheetnames:
            raise AssertionError(f"{case_id} missing required sheet: {sheet_name}")

    vat_sheet = workbook["VAT Schedule"]
    if vat_sheet.max_row < 2:
        raise AssertionError(f"{case_id} VAT extraction failed: expected transaction rows in VAT Schedule")

    tx_sheet = workbook["Transactions"]
    for row_index in range(2, tx_sheet.max_row + 1):
        account = tx_sheet.cell(row=row_index, column=11).value
        if not account:
            raise AssertionError(f"{case_id} AI/account categorisation missing at transaction row {row_index}")

    ai_diagnostics = metadata.get("_ai_diagnostics")
    if not isinstance(ai_diagnostics, dict):
        raise AssertionError(f"{case_id} AI diagnostics missing from workbook metadata")
    required_ai_keys = {
        "ai_enabled",
        "ai_model",
        "ai_transactions_sent",
        "ai_transactions_classified",
        "ai_failures",
        "ai_cache_hits",
        "ai_classification_duration_ms",
    }
    if not required_ai_keys.issubset(ai_diagnostics.keys()):
        missing = sorted(required_ai_keys.difference(ai_diagnostics.keys()))
        raise AssertionError(f"{case_id} AI diagnostics missing keys: {missing}")



# ── FNB date-grouped extraction regression (ACAPOLITE class of failure) ──────
#
# FNB prints a transaction date only once per date group; later rows in the
# group (debit orders, app / RTC payments, fee lines) print WITHOUT a leading
# date. Those rows were being merged into the previous line and dropped, so the
# statement lost 6 debits (R29,912.54) and failed to reconcile. This fixture
# reproduces that exact layout at full scale (143 transactions) and fails unless
# every row is extracted and the reconciliation difference is R0.00.
FNB_EXPECTED = {
    "transaction_count": 143,
    "credit_count": 15,
    "debit_count": 128,
    "total_credits": "419700.00",
    "total_debits": "422747.72",
    "opening_balance": "3390.09",
    "closing_balance": "342.37",
}


def _build_acapolite_style_statement() -> tuple[str, dict]:
    from decimal import Decimal as D

    lines = ["Transactions in Rand (ZAR)"]
    bal = D("3390.09")  # true running balance for the balance-bearing rows

    def money(value: D) -> str:
        return f"{value:,.2f}"

    # 15 credits summing 419,700.00.
    for i in range(15):
        bal += D("27980.00")
        lines.append(f"0{(i % 9) + 1} Mar Eft Deposit Customer {i:03d} 27,980.00Cr {money(bal)} Cr")

    # 121 ordinary debits of 3,200.00 (each date-led with a running balance).
    for i in range(121):
        bal -= D("3200.00")
        lines.append(f"1{(i % 9)} Mar Card Purchase Merchant {i:03d} 3,200.00 {money(bal)} Cr")
    # One more ordinary debit (5,635.18) immediately followed by the date-LESS
    # debit-order row (the ACAPOLITE 02 Mar case) that used to be swallowed.
    bal -= D("5635.18")
    lines.append(f"02 Mar Card Purchase Fuel Filling Station 5,635.18 {money(bal)} Cr")
    lines.append("Internal Debit Order Fnbfuneral Fi11941792 J62730 696.30")

    # 18 Mar group: date printed once, second payment has no leading date.
    lines.append("18 Mar Fnb App Payment To 819035690 3,000.00")
    lines.append("Fnb App Rtc Pmt To Patric 25,000.00")

    # 24 Mar fee group: three fee rows, date printed once.
    lines.append("24 Mar Monthly Account Fee 93.00")
    lines.append("Service Fees 523.80")
    lines.append("Cash Deposit Fee 599.44")

    lines.append("Closing Balance 342.37")

    metadata = {
        "statement_period_start": "2026-03-01",
        "statement_period_end": "2026-03-31",
        "opening_balance": 3390.09,
        "closing_balance": 342.37,
    }
    return "\n".join(lines), metadata


def run_fnb_extraction_case() -> None:
    case_id = "fnb-acapolite-grouped-rows"
    text, metadata = _build_acapolite_style_statement()
    transactions = parse_fnb_transactions([], metadata, text)

    # All 143 rows must be extracted — including the 6 date-grouped debits.
    assert_equal(len(transactions), FNB_EXPECTED["transaction_count"], f"{case_id} transaction count")

    # The six previously-missing rows must be present with the exact amounts.
    missing_rows = {
        ("Internal Debit Order Fnbfuneral Fi11941792 J62730", "696.30"),
        ("Fnb App Payment To 819035690", "3000.00"),
        ("Fnb App Rtc Pmt To Patric", "25000.00"),
        ("Monthly Account Fee", "93.00"),
        ("Service Fees", "523.80"),
        ("Cash Deposit Fee", "599.44"),
    }
    extracted = {(t.description, f"{(t.debit_amount or 0):.2f}") for t in transactions}
    for desc, amount in missing_rows:
        if (desc, amount) not in extracted:
            raise AssertionError(f"{case_id}: missing row not extracted: {desc} R{amount}")

    summary = validation_summary(transactions)
    assert_equal(str(summary["total_credits"]), FNB_EXPECTED["total_credits"], f"{case_id} total credits")
    assert_equal(str(summary["total_debits"]), FNB_EXPECTED["total_debits"], f"{case_id} total debits")
    assert_equal(summary["credit_count"], FNB_EXPECTED["credit_count"], f"{case_id} credit count")
    assert_equal(summary["debit_count"], FNB_EXPECTED["debit_count"], f"{case_id} debit count")

    # Reconciliation difference must be exactly R0.00 (validate_statement raises otherwise).
    validation = validate_statement(metadata, transactions)
    if validation["calculated_closing"] != validation["closing_balance"]:
        raise AssertionError(
            f"{case_id}: reconciliation difference not zero — calculated {validation['calculated_closing']} vs closing {validation['closing_balance']}"
        )
    assert_equal(str(validation["closing_balance"]), FNB_EXPECTED["closing_balance"], f"{case_id} closing balance")


def run_missing_column_fallback_case() -> None:
    # If the DB schema lacks an optional column (e.g. statement_date before its
    # migration is applied), the run update must drop it and still save — never
    # fail the whole processing job with HTTP 422.
    from main import update_statement_run

    case_id = "missing-statement-date-column"
    calls: list[dict] = []

    class Query:
        def __init__(self, fields, fail_on):
            self.fields = fields
            self.fail_on = fail_on

        def eq(self, *args, **kwargs):
            return self

        def execute(self):
            calls.append(self.fields)
            if self.fail_on and self.fail_on(self.fields):
                raise RuntimeError(self.fail_on(self.fields))
            return None

    def make_supabase(fail_on):
        class Table:
            def update(self, fields):
                return Query(fields, fail_on)

        class Supabase:
            def table(self, _name):
                return Table()

        return Supabase()

    schema_error = "Could not find the 'statement_date' column of 'accounting_statement_runs' in the schema cache"
    supabase = make_supabase(lambda fields: schema_error if "statement_date" in fields else None)
    update_statement_run(supabase, "run-1", "ws-1", {"statement_date": "2026-03-31", "status": "completed", "confidence": 90})
    assert_equal(len(calls), 2, f"{case_id} retried once without the optional column")
    assert_equal("statement_date" in calls[1], False, f"{case_id} dropped the missing column")
    assert_equal(calls[1]["status"], "completed", f"{case_id} core fields preserved")

    # Non-schema errors must still propagate (not silently swallowed).
    raised = False
    try:
        update_statement_run(make_supabase(lambda _fields: "permission denied"), "run-1", "ws-1", {"statement_date": "2026-03-31", "status": "completed"})
    except RuntimeError:
        raised = True
    assert_equal(raised, True, f"{case_id} non-schema errors propagate")


def run_validation_diagnostics_case() -> None:
    # When extraction does not match the statement's declared figures, validation
    # must fail with the SPECIFIC rules and extracted-vs-declared values (not a
    # generic "layout needs review"), and must never pass silently.
    from main import HTTPException, ParsedTransaction, validate_statement

    case_id = "validation-diagnostics"

    def txn(debit=None, credit=None):
        return ParsedTransaction(
            transaction_date="2026-03-10", description="x", debit_amount=debit, credit_amount=credit,
            running_balance=None, bank_charge=False, account_category="X", vat_treatment="review",
            supported_by_invoice=False, confidence=90, review_status="ready", source_page=1, raw_text="r",
        )

    # Declares 143 (15 credits / 128 debits) but only 137 extracted (6 debits missing).
    meta = {
        "opening_balance": 3390.09, "closing_balance": 342.37,
        "expected_transaction_count": 143, "expected_credit_count": 15, "expected_debit_count": 128,
        "declared_credit_total": 419700.00, "declared_debit_total": 422747.72,
    }
    txns = [txn(credit=419700.00 / 15) for _ in range(15)] + [txn(debit=392835.18 / 122) for _ in range(122)]

    raised = False
    try:
        validate_statement(meta, txns)
    except HTTPException as exc:
        raised = True
        detail = exc.detail
        for rule in ("reconciliation", "transaction_count", "debit_count", "debit_total"):
            if rule not in detail["failed_rules"]:
                raise AssertionError(f"{case_id}: expected failed rule {rule}, got {detail['failed_rules']}")
        assert_equal(detail["suspected_missing_rows"], 6, f"{case_id} suspected missing rows")
        assert_equal(detail["extracted_transaction_count"], 137, f"{case_id} extracted count")
        assert_equal(detail["expected_transaction_count"], 143, f"{case_id} expected count")
        joined = " ".join(detail["errors"])
        if "extracted 137 vs declared 143" not in joined:
            raise AssertionError(f"{case_id}: error must show extracted vs declared, got {detail['errors']}")
        # Must not hardcode a specific statement's expected figures.
        import inspect
        import main as worker_main
        if "111600.56" in inspect.getsource(worker_main.validate_statement):
            raise AssertionError(f"{case_id}: validate_statement must not hardcode a per-statement expectation")
    if not raised:
        raise AssertionError(f"{case_id}: validation must fail (not pass silently) when extraction is short")


def run_april_missing_rows_case() -> None:
    # April 2026 (ACAPOLITE) statement: three rows print an amount + running
    # balance WITHOUT a Cr/Dr suffix (two Internal Debit Order / FnbFuneral rows
    # and one "#Excess Item Fee") and were dropped, breaking reconciliation by
    # R1,682.32. The parser must now capture them and the fee must be Bank Charges.
    from decimal import Decimal as D

    from main import parse_fnb_transactions, validate_statement, validation_summary

    case_id = "april-missing-rows"
    opening, closing = D("342.37"), D("368.96")
    lines = ["Transactions in Rand (ZAR)"]
    bal = opening

    def money(value: D) -> str:
        return f"{value:,.2f}"

    def balance_cell(value: D) -> str:
        # FNB prints Cr for a positive balance; overdrawn balances print without
        # "Cr" (magnitude only) — exactly the case the parser must handle.
        return f"{money(value)} Cr" if value >= 0 else money(value.copy_abs())

    # 3 rows FIRST, pushing the account overdrawn, with amount + balance and NO
    # Cr/Dr suffix on the (negative) balance — the previously-dropped rows.
    specials = [
        ("01 Apr Internal Debit Order Fnbfuneral Fi11941792A Ex6460", D("676.02")),
        ("01 Apr Internal Debit Order Fnbfuneral Fi11941792 Ex6462", D("696.30")),
        ("02 Apr #Excess Item Fee 2 Items On 26/04/01", D("310.00")),
    ]
    for desc, amt in specials:
        bal -= amt  # now negative
        lines.append(f"{desc} {money(amt)} {money(bal.copy_abs())}")
    # 7 credits summing 226,361.00 (Cr-suffixed amount) — recover the balance.
    for i, amt in enumerate([D("37000.00")] * 6 + [D("4361.00")]):
        bal += amt
        lines.append(f"0{(i % 9) + 1} Apr Eft Credit Customer {i:03d} {money(amt)}Cr {balance_cell(bal)}")
    # 57 ordinary debits summing 224,652.09.
    for i, amt in enumerate([D("3900.00")] * 56 + [D("6252.09")]):
        bal -= amt
        lines.append(f"1{i % 9} Apr Card Purchase Merchant {i:03d} {money(amt)} {balance_cell(bal)}")
    lines.append(f"Closing Balance {money(closing)}")

    metadata = {
        "opening_balance": 342.37, "closing_balance": 368.96,
        "expected_transaction_count": 67, "expected_credit_count": 7, "expected_debit_count": 60,
        "declared_credit_total": 226361.00, "declared_debit_total": 226334.41,
    }
    txns = parse_fnb_transactions([], metadata, "\n".join(lines))

    assert_equal(len(txns), 67, f"{case_id} transaction count")
    summary = validation_summary(txns)
    assert_equal(summary["credit_count"], 7, f"{case_id} credit count")
    assert_equal(summary["debit_count"], 60, f"{case_id} debit count")
    assert_equal(str(summary["total_credits"]), "226361.00", f"{case_id} credit total")
    assert_equal(str(summary["total_debits"]), "226334.41", f"{case_id} debit total")

    # The three named rows must be present with the exact amounts.
    extracted = {(t.description, f"{(t.debit_amount or 0):.2f}") for t in txns}
    for desc, amount in [("Internal Debit Order Fnbfuneral Fi11941792A Ex6460", "676.02"),
                         ("Internal Debit Order Fnbfuneral Fi11941792 Ex6462", "696.30")]:
        if (desc, amount) not in extracted:
            raise AssertionError(f"{case_id}: missing row {desc} R{amount}")
    fee = next((t for t in txns if "excess item fee" in t.description.lower()), None)
    if fee is None or not fee.bank_charge or fee.account_category != "Bank Charges":
        raise AssertionError(f"{case_id}: #Excess Item Fee must be captured as Bank Charges, got {fee}")

    # Reconciliation must be exactly R0.00 (validate_statement raises otherwise).
    validation = validate_statement(metadata, txns)
    if validation["calculated_closing"] != validation["closing_balance"]:
        raise AssertionError(f"{case_id}: reconciliation not zero ({validation['calculated_closing']} vs {validation['closing_balance']})")
    assert_equal(str(validation["closing_balance"]), "368.96", f"{case_id} closing balance")


def run_freight_aces_case() -> None:
    # FNBBSJAN2026 (FREIGHT ACES) Gold Business statement: header
    # "Transactions in RAND (ZAR) : 62905786151", transactions across several
    # pages, "#" fee rows and rows whose description is lost (date + amount +
    # balance). Previously returned "No FNB transactions could be parsed".
    from decimal import Decimal as D

    from main import extraction_diagnostics, parse_fnb_transactions, validate_statement, validation_summary

    case_id = "freight-aces-jan"
    opening, closing = D("1869.10"), D("295242.68")

    def money(v: D) -> str:
        return f"{v:,.2f}"

    def balance_cell(v: D) -> str:
        return f"{money(v)} Cr" if v >= 0 else money(v.copy_abs())

    rows: list[tuple[str, D, bool]] = []
    # 24 credits summing 909,530.63.
    for i, amt in enumerate([D("39000.00")] * 23 + [D("12530.63")]):
        rows.append((f"Eft Credit Customer {i:03d}", amt, True))
    # 116 debits summing 616,157.05: two "#" fee rows (= declared service fees
    # 1,168.52) plus 114 others, ~11 of which lose their description.
    rows.append(("#Monthly Account Fee", D("349.00"), False))
    rows.append(("#Service Fees", D("819.52"), False))
    for i, amt in enumerate([D("5400.00")] * 113 + [D("4788.53")]):
        desc = "" if i % 10 == 0 else f"Card Purchase Merchant {i:03d}"
        rows.append((desc, amt, False))

    lines = [
        "FREIGHT ACES (PTY)LTD",
        "Account Number : 62905786151",
        "Opening Balance 1,869.10 Cr",
        "Closing Balance 295,242.68 Cr",
        "Transactions in RAND (ZAR) : 62905786151",
    ]
    bal = opening
    for idx, (desc, amt, is_credit) in enumerate(rows):
        day = f"{(idx % 28) + 1:02d} Jan"
        if is_credit:
            bal += amt
            lines.append(f"{day} {desc} {money(amt)}Cr {balance_cell(bal)}")
        else:
            bal -= amt
            lines.append(f"{day} {desc} {money(amt)} {balance_cell(bal)}".replace("  ", " "))
        if idx in (39, 79, 119):  # page breaks: carried-forward line + repeated header
            lines.append(f"Balance Brought Forward {balance_cell(bal)}")
            lines.append("Transactions in RAND (ZAR) : 62905786151")
    lines.append(f"Closing Balance {money(closing)} Cr")
    lines.append("Turnover for Statement Period")
    text = "\n".join(lines)

    metadata = {
        "opening_balance": 1869.10, "closing_balance": 295242.68,
        "expected_transaction_count": 140, "expected_credit_count": 24, "expected_debit_count": 116,
        "declared_credit_total": 909530.63, "declared_debit_total": 616157.05,
    }
    txns = parse_fnb_transactions([], metadata, text)

    # Must NOT be rejected as unparseable; diagnostics must see the section.
    diagnostics = extraction_diagnostics([], text, metadata)
    if not diagnostics["transaction_section_found"]:
        raise AssertionError(f"{case_id}: transaction section not detected")

    assert_equal(len(txns), 140, f"{case_id} transaction count")
    summary = validation_summary(txns)
    assert_equal(summary["credit_count"], 24, f"{case_id} credit count")
    assert_equal(summary["debit_count"], 116, f"{case_id} debit count")
    assert_equal(str(summary["total_credits"]), "909530.63", f"{case_id} credit total")
    assert_equal(str(summary["total_debits"]), "616157.05", f"{case_id} debit total")

    fees = [t for t in txns if t.description.startswith("#") and t.bank_charge]
    if len(fees) < 2:
        raise AssertionError(f"{case_id}: fee rows not captured as bank charges ({len(fees)})")

    validation = validate_statement(metadata, txns)
    if validation["calculated_closing"] != validation["closing_balance"]:
        raise AssertionError(f"{case_id}: reconciliation not zero ({validation['calculated_closing']} vs {validation['closing_balance']})")
    assert_equal(str(validation["closing_balance"]), "295242.68", f"{case_id} closing balance")


def run_december_multi_page_closing_balance_case() -> None:
    # Regression: multi-page OCR text can include "Closing Balance" between repeated
    # page headers. The parser must NOT stop at that intermediate line.
    from decimal import Decimal as D

    from main import parse_fnb_transactions, validate_statement, validation_summary

    case_id = "freight-aces-dec-multipage"
    opening, closing = D("4378.76"), D("97489.87")

    def money(v: D) -> str:
        return f"{v:,.2f}"

    def balance_cell(v: D) -> str:
        return f"{money(v)} Cr" if v >= 0 else money(v.copy_abs())

    credits = [D("50000.00")] * 11 + [D("12345.67")]
    debits = [D("4500.00")] * 104 + [D("1234.56")]
    rows: list[tuple[str, D, bool]] = []
    for idx, amount in enumerate(credits):
        rows.append((f"Eft Credit Customer {idx:03d}", amount, True))
    for idx, amount in enumerate(debits):
        desc = "" if idx % 11 == 0 else f"Card Purchase Merchant {idx:03d}"
        rows.append((desc, amount, False))

    lines = [
        "FREIGHT ACES (PTY)LTD",
        "Account Number : 62905786151",
        "Opening Balance 4,378.76 Cr",
        "Closing Balance 97,489.87 Cr",
        "Transactions in RAND (ZAR) : 62905786151",
    ]
    bal = opening
    for idx, (desc, amount, is_credit) in enumerate(rows):
        day = f"{(idx % 28) + 1:02d} Dec"
        if is_credit:
            bal += amount
            lines.append(f"{day} {desc} {money(amount)}Cr {balance_cell(bal)}")
        else:
            bal -= amount
            lines.append(f"{day} {desc} {money(amount)} {balance_cell(bal)}".replace("  ", " "))
        if idx in (38, 76):  # pseudo page breaks
            lines.append(f"Closing Balance {balance_cell(bal)}")
            lines.append(f"Balance Brought Forward {balance_cell(bal)}")
            lines.append("Transactions in RAND (ZAR) : 62905786151")
    lines.append("Closing Balance 97,489.87 Cr")
    lines.append("Turnover for Statement Period")
    text = "\n".join(lines)

    metadata = {
        "opening_balance": 4378.76,
        "closing_balance": 97489.87,
        "expected_transaction_count": 117,
        "expected_credit_count": 12,
        "expected_debit_count": 105,
        "declared_credit_total": 562345.67,
        "declared_debit_total": 469234.56,
    }
    txns = parse_fnb_transactions([], metadata, text)
    assert_equal(len(txns), 117, f"{case_id} transaction count")
    summary = validation_summary(txns)
    assert_equal(summary["credit_count"], 12, f"{case_id} credit count")
    assert_equal(summary["debit_count"], 105, f"{case_id} debit count")
    assert_equal(str(summary["total_credits"]), "562345.67", f"{case_id} credit total")
    assert_equal(str(summary["total_debits"]), "469234.56", f"{case_id} debit total")
    validation = validate_statement(metadata, txns)
    if validation["calculated_closing"] != validation["closing_balance"]:
        raise AssertionError(f"{case_id}: reconciliation not zero ({validation['calculated_closing']} vs {validation['closing_balance']})")


def run_compound_ocr_line_case() -> None:
    # OCR occasionally merges adjacent transaction rows onto one physical line.
    # The parser must split those compound lines back into separate movements.
    case_id = "compound-ocr-line-split"
    text = (
        "Transactions in RAND (ZAR)\n"
        "01 Dec Diesel Depot 1,200.00 8,800.00 Cr 02 Dec Eft Credit Customer Alpha 9,500.00Cr 18,300.00 Cr\n"
        "03 Dec Sanral Toll 450.00 17,850.00 Cr\n"
        "Closing Balance 17,850.00 Cr\n"
        "Turnover for Statement Period\n"
    )
    metadata = {
        "statement_period_start": "2025-12-01",
        "statement_period_end": "2025-12-31",
        "opening_balance": 10000.00,
        "closing_balance": 17850.00,
        "expected_transaction_count": 3,
        "expected_credit_count": 1,
        "expected_debit_count": 2,
        "declared_credit_total": 9500.00,
        "declared_debit_total": 1650.00,
    }
    txns = parse_fnb_transactions([], metadata, text)
    assert_equal(len(txns), 3, f"{case_id} transaction count")
    extracted = {(t.description, f"{(t.debit_amount or t.credit_amount or 0):.2f}") for t in txns}
    for expected in {
        ("Diesel Depot", "1200.00"),
        ("Eft Credit Customer Alpha", "9500.00"),
        ("Sanral Toll", "450.00"),
    }:
        if expected not in extracted:
            raise AssertionError(f"{case_id}: missing split transaction {expected}")
    validate_statement(metadata, txns)


def run_professional_classification_case() -> None:
    case_id = "professional-classification"
    fuel = ParsedTransaction(
        transaction_date="2026-01-05",
        description="Shell Diesel Depot",
        debit_amount=1500.0,
        credit_amount=None,
        running_balance=8500.0,
        bank_charge=False,
        account_category="Motor Vehicle Expenses",
        vat_treatment="standard",
        supported_by_invoice=False,
        confidence=92,
        review_status="ready",
        source_page=1,
        raw_text="05 Jan Shell Diesel Depot 1,500.00 8,500.00 Cr",
    )
    receipt = ParsedTransaction(
        transaction_date="2026-01-06",
        description="Eft Credit Customer Freight Aces",
        debit_amount=None,
        credit_amount=12500.0,
        running_balance=21000.0,
        bank_charge=False,
        account_category="Sales / Revenue",
        vat_treatment="standard",
        supported_by_invoice=False,
        confidence=92,
        review_status="ready",
        source_page=1,
        raw_text="06 Jan Eft Credit Customer Freight Aces 12,500.00Cr 21,000.00 Cr",
    )
    gov_receipt = ParsedTransaction(
        transaction_date="2026-05-02",
        description="Magtape Credit 047-Gp Hea-000052034",
        debit_amount=None,
        credit_amount=1234021.00,
        running_balance=2700250.85,
        bank_charge=False,
        account_category="Sales / Revenue",
        vat_treatment="standard",
        supported_by_invoice=False,
        confidence=94,
        review_status="ready",
        source_page=1,
        raw_text="02 May Magtape Credit 047-Gp Hea-000052034 1,234,021.00Cr 2,700,250.85Cr",
    )
    supplier_payment = ParsedTransaction(
        transaction_date="2026-05-03",
        description="FNB App Payment To Rmsp Trading Allianz Holdings",
        debit_amount=2770250.85,
        credit_amount=None,
        running_balance=0.00,
        bank_charge=False,
        account_category="Supplier Payments",
        vat_treatment="review",
        supported_by_invoice=False,
        confidence=86,
        review_status="needs_review",
        source_page=1,
        raw_text="03 May FNB App Payment To Rmsp Trading Allianz Holdings 2,770,250.85 0.00Cr",
    )
    msi_payment = ParsedTransaction(
        transaction_date="2026-05-15",
        description="FNB App Payment To Msi Industries Inv109034",
        debit_amount=1012000.00,
        credit_amount=None,
        running_balance=328320.91,
        bank_charge=False,
        account_category="Supplier Payments",
        vat_treatment="review",
        supported_by_invoice=False,
        confidence=88,
        review_status="needs_review",
        source_page=1,
        raw_text="15 May FNB App Payment To Msi Industries Inv109034 1,012,000.00 328,320.91Cr",
    )
    fuel_row = professional_transaction_row(fuel, "fixture")
    receipt_row = professional_transaction_row(receipt, "fixture")
    gov_receipt_row = professional_transaction_row(gov_receipt, "fixture")
    supplier_payment_row = professional_transaction_row(supplier_payment, "fixture")
    msi_payment_row = professional_transaction_row(msi_payment, "fixture")
    apply_ai_result_to_row(
        msi_payment_row,
        {
            "transaction_id": "1",
            "account": "Travel / Meals / Entertainment",
            "group": "Expense",
            "vat_treatment": "Staff Welfare / Meals / Entertainment",
            "vat_claim_status": "Review",
            "review_required": True,
            "review_reason": "AI guessed meals",
            "invoice_required": False,
            "confidence": 0.91,
            "reason": "bad model answer",
            "explanation": "bad model answer",
        },
    )
    assert_equal(fuel_row["review_required"], False, f"{case_id} fuel review")
    assert_equal(receipt_row["review_required"], False, f"{case_id} receipt review")
    assert_equal(fuel_row["account"], "Motor Vehicle Expenses", f"{case_id} fuel account")
    assert_equal(receipt_row["account"], "Sales / Revenue", f"{case_id} receipt account")
    assert_equal(receipt_row["vat_claim_status"], "Output", f"{case_id} receipt vat")
    assert_equal(gov_receipt_row["account"], "Sales / Revenue", f"{case_id} government health receipt account")
    assert_equal(gov_receipt_row["group"], "Income", f"{case_id} government health receipt group")
    assert_equal(supplier_payment_row["account"], "Supplier Payments", f"{case_id} supplier payment account")
    assert_equal(supplier_payment_row["group"], "Operating Expenses", f"{case_id} supplier payment group")
    assert_equal(supplier_payment_row["invoice_required"], True, f"{case_id} supplier invoice review")
    assert_equal(str(supplier_payment_row["potential_input_vat"]), "361337.07", f"{case_id} supplier potential input VAT")
    assert_equal(msi_payment_row["account"], "Supplier Payments", f"{case_id} AI guardrail supplier account")
    assert_equal(msi_payment_row["group"], "Operating Expenses", f"{case_id} AI guardrail supplier group")
    assert_equal(msi_payment_row["invoice_required"], True, f"{case_id} AI guardrail invoice support")
    assert_equal(str(msi_payment_row["potential_input_vat"]), "132000.00", f"{case_id} AI guardrail potential input VAT")


def run_learned_supplier_rules_case() -> None:
    case_id = "learned-supplier-rules"
    transactions = [
        ParsedTransaction(
            transaction_date="2026-04-02",
            description="POS Purchase New Uber Eats 400568*7629 01 Apr",
            debit_amount=94.0,
            credit_amount=None,
            running_balance=1598939.08,
            bank_charge=False,
            account_category="Suspense / Review Required",
            vat_treatment="review",
            supported_by_invoice=False,
            confidence=55,
            review_status="needs_review",
            source_page=1,
            raw_text="02 Apr POS Purchase New Uber Eats 400568*7629 01 Apr 94.00 1,598,939.08Cr",
        ),
        ParsedTransaction(
            transaction_date="2026-04-04",
            description="POS Purchase Google Chatgpt 400568*7629 03 Apr",
            debit_amount=424.99,
            credit_amount=None,
            running_balance=1598514.09,
            bank_charge=False,
            account_category="Suspense / Review Required",
            vat_treatment="review",
            supported_by_invoice=False,
            confidence=55,
            review_status="needs_review",
            source_page=1,
            raw_text="04 Apr POS Purchase Google Chatgpt 400568*7629 03 Apr 424.99 1,598,514.09Cr",
        ),
        ParsedTransaction(
            transaction_date="2026-04-25",
            description="25 Apr Byc Debit 63012593504",
            debit_amount=8.51,
            credit_amount=None,
            running_balance=1450166.60,
            bank_charge=False,
            account_category="Suspense / Review Required",
            vat_treatment="review",
            supported_by_invoice=False,
            confidence=55,
            review_status="needs_review",
            source_page=2,
            raw_text="25 Apr Byc Debit 63012593504 8.51 1,450,166.60Cr",
        ),
        ParsedTransaction(
            transaction_date="2026-04-07",
            description="POS Purchase Sage SA 400568*7629 06 Apr",
            debit_amount=599.0,
            credit_amount=None,
            running_balance=1449567.60,
            bank_charge=False,
            account_category="Suspense / Review Required",
            vat_treatment="review",
            supported_by_invoice=False,
            confidence=55,
            review_status="needs_review",
            source_page=2,
            raw_text="07 Apr POS Purchase Sage SA 400568*7629 06 Apr 599.00 1,449,567.60Cr",
        ),
        ParsedTransaction(
            transaction_date="2026-04-08",
            description="Scheduled Payment To Home Loan Emporers Home Loan Payment",
            debit_amount=10000.0,
            credit_amount=None,
            running_balance=1439567.60,
            bank_charge=False,
            account_category="Suspense / Review Required",
            vat_treatment="review",
            supported_by_invoice=False,
            confidence=55,
            review_status="needs_review",
            source_page=2,
            raw_text="08 Apr Scheduled Payment To Home Loan Emporers Home Loan Payment 10,000.00 1,439,567.60Cr",
        ),
    ]
    rules = [
        {
            "merchant_key": "google",
            "account_category": "Software / IT",
            "vat_treatment": "review",
            "review_status": "needs_review",
            "confidence": 84,
        },
        {
            "merchant_key": "google chatgpt",
            "account_category": "Software Subscriptions",
            "vat_treatment": "standard",
            "review_status": "needs_review",
            "confidence": 90,
        },
        {
            "merchant_key": "uber eats",
            "account_category": "Staff Welfare / Meals / Entertainment",
            "vat_treatment": "review",
            "review_status": "needs_review",
            "confidence": 88,
        },
        {
            "merchant_key": "byc debit",
            "account_category": "Bank Charges",
            "vat_treatment": "standard",
            "review_status": "approved",
            "confidence": 98,
        },
        {
            "merchant_key": "sage sa",
            "account_category": "Software Subscriptions",
            "vat_treatment": "standard",
            "review_status": "needs_review",
            "confidence": 90,
        },
        {
            "merchant_key": "home loan payment",
            "account_category": "Loan / Liability",
            "vat_treatment": "out_of_scope",
            "review_status": "approved",
            "confidence": 92,
        },
    ]
    applied = apply_learned_classification_rules(transactions, rules)
    assert_equal(applied, 5, f"{case_id} applied count")
    assert_equal(transactions[0].account_category, "Staff Welfare / Meals / Entertainment", f"{case_id} uber")
    assert_equal(transactions[1].account_category, "Software Subscriptions", f"{case_id} specific google rule")
    assert_equal(transactions[2].account_category, "Bank Charges", f"{case_id} bank fee")
    assert_equal(transactions[2].review_status, "approved", f"{case_id} bank fee review status")
    assert_equal(transactions[2].confidence, 98.0, f"{case_id} bank fee confidence")
    assert_equal(transactions[3].account_category, "Software Subscriptions", f"{case_id} sage")
    assert_equal(transactions[4].account_category, "Loan / Liability", f"{case_id} home loan")
    assert_equal(transactions[4].review_status, "approved", f"{case_id} home loan review status")


def run_combined_workbook_case() -> None:
    case_id = "combined-workbook-months"
    december_run = {
        "id": "run-dec",
        "company_name": "Freight Aces (Pty) Ltd",
        "bank": "FNB South Africa",
        "account_number": "62905786151",
        "statement_period_start": "2025-12-01",
        "statement_period_end": "2025-12-31",
        "opening_balance": 1000.0,
        "closing_balance": 3200.0,
        "created_at": "2026-01-01T00:00:00",
    }
    january_run = {
        "id": "run-jan",
        "company_name": "Freight Aces (Pty) Ltd",
        "bank": "FNB South Africa",
        "account_number": "62905786151",
        "statement_period_start": "2026-01-01",
        "statement_period_end": "2026-01-31",
        "opening_balance": 3200.0,
        "closing_balance": 6400.0,
        "created_at": "2026-02-01T00:00:00",
    }
    december_txns = [
        ParsedTransaction(
            transaction_date="2025-12-05",
            description="Eft Credit Customer Afrigreen",
            debit_amount=None,
            credit_amount=4000.0,
            running_balance=5000.0,
            bank_charge=False,
            account_category="Sales / Revenue",
            vat_treatment="standard",
            supported_by_invoice=False,
            confidence=92,
            review_status="ready",
            source_page=1,
            raw_text="05 Dec Eft Credit Customer Afrigreen 4,000.00Cr 5,000.00 Cr",
        ),
        ParsedTransaction(
            transaction_date="2025-12-09",
            description="Diesel Depot",
            debit_amount=1800.0,
            credit_amount=None,
            running_balance=3200.0,
            bank_charge=False,
            account_category="Motor Vehicle Expenses",
            vat_treatment="standard",
            supported_by_invoice=False,
            confidence=92,
            review_status="ready",
            source_page=1,
            raw_text="09 Dec Diesel Depot 1,800.00 3,200.00 Cr",
        ),
    ]
    january_txns = [
        ParsedTransaction(
            transaction_date="2026-01-03",
            description="Eft Credit Customer Freight Aces",
            debit_amount=None,
            credit_amount=5000.0,
            running_balance=8200.0,
            bank_charge=False,
            account_category="Sales / Revenue",
            vat_treatment="standard",
            supported_by_invoice=False,
            confidence=92,
            review_status="ready",
            source_page=1,
            raw_text="03 Jan Eft Credit Customer Freight Aces 5,000.00Cr 8,200.00 Cr",
        ),
        ParsedTransaction(
            transaction_date="2026-01-10",
            description="Sanral Toll",
            debit_amount=1800.0,
            credit_amount=None,
            running_balance=6400.0,
            bank_charge=False,
            account_category="Road Tolls",
            vat_treatment="standard",
            supported_by_invoice=False,
            confidence=90,
            review_status="ready",
            source_page=1,
            raw_text="10 Jan Sanral Toll 1,800.00 6,400.00 Cr",
        ),
    ]
    workbook_bytes, summary = build_combined_workbook(
        [january_run, december_run],
        {"run-dec": december_txns, "run-jan": january_txns},
    )
    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    tx_sheet = workbook["Transactions"]
    headers = [tx_sheet.cell(row=1, column=column).value for column in range(1, tx_sheet.max_column + 1)]
    if "VAT Code" not in headers or "VAT Treatment" in headers:
        raise AssertionError(f"{case_id}: Transactions sheet must expose VAT Code and hide ambiguous VAT Treatment")
    source_periods = {tx_sheet.cell(row=row, column=3).value for row in range(2, tx_sheet.max_row + 1)}
    if {"2025-12-01 to 2025-12-31", "2026-01-01 to 2026-01-31"} - source_periods:
        raise AssertionError(f"{case_id}: combined workbook must preserve both statement periods")
    assert_equal(summary["transaction_count"], 4, f"{case_id} transaction count")
    assert_equal(summary["review_count"], 0, f"{case_id} review count")
    diagnostics = workbook["Diagnostics"]
    ai_row = next((row for row in range(2, diagnostics.max_row + 1) if diagnostics.cell(row=row, column=1).value == "ai"), None)
    if ai_row is None:
        raise AssertionError(f"{case_id}: combined diagnostics missing AI row")
    vat = workbook["VAT Schedule"]
    assert_equal(vat["A1"].value, "VAT Schedule & VAT Payable/(Refund)", f"{case_id} VAT title")
    assert_equal(vat["C7"].value, "Potential Input VAT", f"{case_id} VAT input header")
    assert_equal(vat["D7"].value, "Net VAT Payable/(Refund)", f"{case_id} VAT monthly net header")
    assert_equal(vat["E7"].value, "Running VAT Balance", f"{case_id} VAT running balance header")
    assert_equal(vat["A11"].value, "Date", f"{case_id} VAT detail starts after monthly summary")
    assert_equal(vat["G11"].value, "VAT Code", f"{case_id} VAT detail must include VAT Code")


REAL_STATEMENT_RESULTS: dict[str, object] = {}


def run_local_real_statement_files_case() -> None:
    """Optional guard for the two real FNB statements supplied during support.
    These files live outside the repository, so CI/deploys skip this test. On the
    affected Mac it verifies that the current parser reconciles the exact March
    and May PDFs whose older production runs displayed incorrect Money In/Out."""
    # NEVER silently return. This case used to swallow a missing pdfplumber and
    # report success while covering nothing — the runner used the system python3,
    # which has no pdfplumber, so it skipped on every run and hid a live parser
    # defect. A missing dependency is now a hard failure.
    try:
        import pdfplumber
    except Exception as exc:  # noqa: BLE001
        raise AssertionError(
            "real-statement regression requires pdfplumber; run via "
            "workers/accounting_worker/tests/run_regression.sh so the worker venv is used"
        ) from exc
    if not hasattr(pdfplumber, "open"):
        raise AssertionError("pdfplumber is present but unusable (no .open)")

    cases = [
        {
            "id": "real-march-2026",
            "path": Path("/Users/patric/Library/Mobile Documents/com~apple~CloudDocs/Desktop Mac Downloads/31 Mar 2026 - (Free)..A1N1WRAFCAUDGU1_EVNXHAEGdVVAU1RUVBgaIEYBVhkGAHlSEVdUVA8cGHIRB1UYUFV0AkAEBldXHxAkSgAETw.XhdUVD1zfHZqdiBmUQIBAAIFDQFtUVRWUgUCVgRUAwAFBwcDUldRXgAJAQw8UwU.pdf"),
            "credits": "7043521.68",
            "debits": "5388160.19",
            "closing": "1666557.95",
        },
        {
            "id": "real-may-2026",
            "path": Path("/Users/patric/Library/Mobile Documents/com~apple~CloudDocs/Desktop Mac Downloads/30 May 2026 - (Free)..CgYpVEALUQoGSxsjSlRWGQFSeAVJBwQCBE5KJEdRBB8DUS4FTAIEBwYbGSAQAQcYAAUpVExUUgoCSB1yEV0ASg.V0MIWWZ9JXlvJHY6CgUABQJQAFFvBQQAAlNSUgUEUQYAV1BUBAIBDQkOAF49VVc.pdf"),
            "credits": "12214591.85",
            "debits": "6758364.90",
            "closing": "6957593.75",
        },
    ]

    for case in cases:
        pdf_path = case["path"]
        if not pdf_path.exists():
            REAL_STATEMENT_RESULTS[case["id"]] = "missing-file"
            continue
        pages = []
        full_text_parts = []
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page_number, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ""
                pages.append({"page": page_number, "text": text, "tables": []})
                full_text_parts.append(text)
        full_text = "\n".join(full_text_parts)
        metadata = parse_metadata(full_text)
        transactions = parse_fnb_transactions(pages, metadata, full_text)
        validation = validate_statement(metadata, transactions)
        summary = validation_summary(transactions)
        assert_equal(str(summary["total_credits"]), case["credits"], f"{case['id']} credits")
        assert_equal(str(summary["total_debits"]), case["debits"], f"{case['id']} debits")
        assert_equal(str(validation["closing_balance"]), case["closing"], f"{case['id']} closing")
        extraction_check = main.validate_extraction(metadata, transactions)
        missing_rows = main.missing_transaction_count_for_storage(extraction_check, len(transactions))
        extraction_confidence = main.extraction_confidence_score(
            metadata,
            extraction_check,
            transactions,
            pages,
            missing_rows,
            unresolved_amount_directions=0,
        )
        if extraction_confidence is None or extraction_confidence < 98:
            raise AssertionError(
                f"{case['id']}: fully reconciled real statement should score at least 98 extraction confidence, got {extraction_confidence}"
            )
        summary_counts = validation_summary(transactions)
        REAL_STATEMENT_RESULTS[case["id"]] = {
            "ledger_rows": summary_counts["ledger_row_count"],
            "financial": summary_counts["transaction_count"],
            "informational": summary_counts["informational_row_count"],
            "credits": summary_counts["credit_count"],
            "debits": summary_counts["debit_count"],
            "extraction_confidence": extraction_confidence,
        }


def test_informational_rows_are_kept_but_not_counted() -> None:
    """Zero-value status rows stay in the ledger but are excluded from the count
    compared against the bank's declared transaction total.

    March 2026 prints 103 rows; FNB declares 101. The two extra are
    "Express Pmt Pending" status lines carrying no money and leaving the balance
    untouched. They must be preserved (they are printed) and not counted.
    """
    import main

    def row(desc, debit, credit, balance):
        t = main.build_transaction("27 Mar", desc, debit, credit, balance, {}, None, "raw", 96)
        assert t is not None
        return t

    # Realistic ordering, as printed: a real movement lands the balance, then the
    # two status rows leave it exactly where it was.
    real_debit = row("FNB App Rtc Pmt To Someone", 96300.0, None, 3394030.08)
    pending = row("Express Pmt Pending", 0.0, None, 3394030.08)
    complete = row("Express Pmt Pending Express Pmt Complete", 0.0, None, 3394030.08)

    rows = [real_debit, pending, complete]
    financial, informational = main.split_ledger_rows(rows)

    assert len(informational) == 2, informational
    assert len(financial) == 1, financial
    assert main.financial_transaction_count(rows) == 1

    # Detection needs ALL THREE conditions — a bare zero amount is not enough.
    zero_but_named = row("Interest Adjustment", 0.0, None, 3394030.08)
    assert main.is_non_financial_informational_row(zero_but_named, main.decimal_amount(3394030.08)) is False

    moved_balance = row("Express Pmt Pending", 0.0, None, 999.99)
    assert main.is_non_financial_informational_row(moved_balance, main.decimal_amount(3394030.08)) is False

    with_money = row("Express Pmt Pending", 25.0, None, 3394005.08)
    assert main.is_non_financial_informational_row(with_money, main.decimal_amount(3394030.08)) is False

    # The rows are never removed from the ledger.
    summary = main.validation_summary(rows)
    assert summary["ledger_row_count"] == 3, summary
    assert summary["transaction_count"] == 1, summary
    assert summary["informational_row_count"] == 2, summary


def test_real_statement_cases_actually_execute() -> None:
    """Guard against the silent-skip that hid a live parser defect.

    run_local_real_statement_files_case used to swallow a missing pdfplumber and
    return, so it reported success while never opening a PDF. This asserts the
    cases genuinely ran and produced counts.
    """
    if not REAL_STATEMENT_RESULTS:
        raise AssertionError(
            "real-statement cases did not execute — run_local_real_statement_files_case "
            "must run before this test and must never skip silently"
        )
    executed = {k: v for k, v in REAL_STATEMENT_RESULTS.items() if isinstance(v, dict)}
    if not executed:
        raise AssertionError(
            "no real statement was opened; all cases reported missing-file: "
            f"{REAL_STATEMENT_RESULTS}"
        )
    for case_id, counts in executed.items():
        assert counts["ledger_rows"] >= counts["financial"], case_id
        assert counts["financial"] > 0, case_id


def test_worker_auth_fails_closed() -> None:
    """An unconfigured secret must reject every request, not admit every request.

    The old verify_worker_token returned early when ACCOUNTING_WORKER_TOKEN was
    unset, and it was never set on Render — so /process-statement was callable by
    anyone who knew the hostname (verified live, HTTP 422 on an unauthenticated
    POST). The unconfigured case is the one that actually shipped, so it is the
    one asserted first.
    """
    import auth

    secret = "s3cret-token"

    # The regression that mattered: no secret configured -> refuse, and refuse
    # with 503 (server misconfiguration) rather than 401 (caller's fault).
    for missing in (None, "", "   "):
        assert auth.check_bearer(f"Bearer {secret}", missing) == auth.UNCONFIGURED, repr(missing)
        assert auth.check_bearer(None, missing) == auth.UNCONFIGURED, repr(missing)
    assert auth.STATUS_FOR_VERDICT[auth.UNCONFIGURED][0] == 503

    # Configured: only the exact credential passes.
    assert auth.check_bearer(f"Bearer {secret}", secret) == auth.OK
    assert auth.check_bearer(f"bearer {secret}", secret) == auth.OK, "scheme is case-insensitive"
    assert auth.check_bearer(f"Bearer   {secret}", secret) == auth.OK, "whitespace runs collapse"

    assert auth.check_bearer(None, secret) == auth.MISSING
    assert auth.check_bearer("", secret) == auth.MISSING
    assert auth.check_bearer("Bearer", secret) == auth.MALFORMED
    assert auth.check_bearer("Bearer ", secret) == auth.MALFORMED, "scheme with no token"
    assert auth.check_bearer("   ", secret) == auth.MISSING, "whitespace-only header is no header"
    assert auth.check_bearer(secret, secret) == auth.MALFORMED, "raw token without scheme"
    assert auth.check_bearer(f"Basic {secret}", secret) == auth.MALFORMED
    assert auth.check_bearer("Bearer wrong", secret) == auth.INVALID
    assert auth.check_bearer(f"Bearer {secret}x", secret) == auth.INVALID
    assert auth.check_bearer(f"Bearer {secret[:-1]}", secret) == auth.INVALID

    # Everything that is not OK maps to a status, and none of them is a 2xx.
    for verdict, (status, detail) in auth.STATUS_FOR_VERDICT.items():
        assert status >= 400, f"{verdict} -> {status}"
        assert secret not in detail, f"{verdict} message leaks the secret"
    assert auth.OK not in auth.STATUS_FOR_VERDICT, "OK must have no error mapping"


def test_worker_auth_matches_pdf_plumber_contract() -> None:
    """The two services duplicate this logic because they have separate rootDirs
    and cannot import across the repo. Pin them to one truth table so they cannot
    drift apart silently. Skipped when the pdfplumber module is not present.
    """
    import importlib.util

    peer = ROOT / "services" / "pdf-plumber" / "auth.py"
    if not peer.exists():
        return

    spec = importlib.util.spec_from_file_location("pdf_plumber_auth", peer)
    assert spec and spec.loader
    peer_auth = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(peer_auth)

    import auth

    cases = [
        (None, None),
        ("Bearer tok", None),
        ("Bearer tok", ""),
        (None, "tok"),
        ("", "tok"),
        ("Bearer", "tok"),
        ("Bearer tok", "tok"),
        ("bearer tok", "tok"),
        ("Basic tok", "tok"),
        ("Bearer nope", "tok"),
        ("tok", "tok"),
    ]
    for header, expected in cases:
        assert auth.check_bearer(header, expected) == peer_auth.check_bearer(header, expected), (
            f"accounting and pdfplumber disagree on {header!r}/{expected!r}"
        )


def test_worker_token_check_raises_on_unconfigured_secret() -> None:
    """End-to-end through main.verify_worker_token, which is what the endpoints
    call. Asserts the HTTP status the caller actually receives.
    """
    from fastapi import HTTPException

    previous = os.environ.pop("ACCOUNTING_WORKER_TOKEN", None)
    try:
        for header in (None, "Bearer anything"):
            try:
                main.verify_worker_token(header)
            except HTTPException as exc:
                assert exc.status_code == 503, f"{header!r} -> {exc.status_code}"
            else:
                raise AssertionError(
                    f"verify_worker_token({header!r}) allowed the request with no secret configured"
                )

        os.environ["ACCOUNTING_WORKER_TOKEN"] = "correct-horse"
        main.verify_worker_token("Bearer correct-horse")  # must not raise

        for header, status in ((None, 401), ("Bearer wrong", 401), ("correct-horse", 401)):
            try:
                main.verify_worker_token(header)
            except HTTPException as exc:
                assert exc.status_code == status, f"{header!r} -> {exc.status_code}"
            else:
                raise AssertionError(f"verify_worker_token({header!r}) should have been rejected")
    finally:
        os.environ.pop("ACCOUNTING_WORKER_TOKEN", None)
        if previous is not None:
            os.environ["ACCOUNTING_WORKER_TOKEN"] = previous


def test_mutating_endpoints_are_all_authenticated() -> None:
    """Every @app.post handler must call verify_worker_token in its own body.

    /health and /version stay open by design — Render's health check must reach
    them and they return only version metadata.

    This parses main.py rather than inspecting main.app.routes, so it works even
    when the suite runs against the FastAPI stub at the top of this file. Each
    handler must check for itself: a handler that merely delegates to another
    protected handler is rejected, because that guarantee is one refactor away
    from disappearing.
    """
    import ast

    tree = ast.parse((ROOT / "workers" / "accounting_worker" / "main.py").read_text())
    posts = []
    for node in ast.walk(tree):
        if not isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
            continue
        for dec in node.decorator_list:
            func = dec.func if isinstance(dec, ast.Call) else dec
            if (
                isinstance(func, ast.Attribute)
                and func.attr == "post"
                and isinstance(func.value, ast.Name)
                and func.value.id == "app"
            ):
                posts.append(node)

    assert posts, "found no @app.post handlers — the parser is wrong, not the code"

    unprotected = []
    for node in posts:
        calls = {
            c.func.id
            for c in ast.walk(node)
            if isinstance(c, ast.Call) and isinstance(c.func, ast.Name)
        }
        if "verify_worker_token" not in calls:
            unprotected.append(node.name)
    if unprotected:
        raise AssertionError(
            "POST handlers with no verify_worker_token call of their own: " + ", ".join(unprotected)
        )


def test_non_ascii_credential_is_invalid_not_a_500() -> None:
    """A credential containing a non-ASCII character must be INVALID, not a crash.

    hmac.compare_digest REFUSES str arguments containing non-ASCII characters: it
    raises TypeError rather than returning False. check_bearer called it directly
    on a token taken from a request header, so a credential carrying e.g. a
    zero-width space became an unhandled exception and an HTTP 500 instead of the
    401 the truth table promises — and let a caller distinguish "non-ASCII" from
    "wrong" by status code alone.

    This also retires a hypothesis about the 2026-08-06 token mismatch: because
    the pre-fix code raised on non-ASCII on EITHER side, and the live worker
    answers 401 rather than 500, both the configured secret and the token sent to
    it are already proven pure ASCII.
    """
    import auth

    secret = "s3cret-token"
    for bad in (secret + "\u200b", secret + "\u00e9", "\u200b" + secret):
        assert auth.check_bearer(f"Bearer {bad}", secret) == auth.INVALID, repr(bad)
    # And a non-ASCII secret on the SERVER side must not crash either.
    assert auth.check_bearer(f"Bearer {secret}", secret + "\u200b") == auth.INVALID
    # Equality still holds when both sides carry the same non-ASCII bytes.
    assert auth.check_bearer(f"Bearer {secret}\u200b", secret + "\u200b") == auth.OK
    assert auth.constant_time_equal("a", "a") is True
    assert auth.constant_time_equal("a", "b") is False


def test_auth_diagnostics_never_leak_the_secret() -> None:
    """The diagnostics must reveal nothing about the token beyond digest+length.

    A diagnostic that leaks the value it measures is worse than no diagnostic.
    Asserted as a NEGATIVE over every substring of length >= 4.
    """
    import auth

    secret = "s3cret-token-value-abcdefghijklmnop"
    record = auth.auth_compare_diagnostics(f"Bearer {secret}", secret)
    blob = repr(record)
    for n in range(4, len(secret) + 1):
        assert secret[:n] not in blob, f"leaked a {n}-char prefix"
        assert secret[-n:] not in blob, f"leaked a {n}-char suffix"

    import hashlib

    expected = hashlib.sha256(secret.encode()).hexdigest()
    assert record["configured_sha256"] == expected
    assert record["received_sha256"] == expected
    assert len(record["configured_sha256"]) == 64, "full digest, not truncated"
    assert record["compare_digest_result"] is True
    assert record["digests_match"] is True


def test_auth_diagnostics_hash_exactly_what_is_compared() -> None:
    """The digest must cover the same bytes the comparison uses.

    parse_bearer was split out of check_bearer for this reason. If the two ever
    drift, the diagnostic misreports the very comparison it exists to measure —
    which is the failure mode that would send an investigation the wrong way.
    """
    import auth

    secret = "s3cret-token"
    for header in (
        f"Bearer {secret}",
        f"bearer {secret}",
        f"Bearer   {secret}",
        f"  Bearer {secret}  ",
    ):
        record = auth.auth_compare_diagnostics(header, secret)
        agrees = auth.check_bearer(header, secret) == auth.OK
        assert record["compare_digest_result"] is agrees, header
        assert record["digests_match"] is agrees, header

    # A missing header is unambiguous — this is the case the caller's silent
    # header omission previously made indistinguishable from a wrong value.
    missing = auth.auth_compare_diagnostics(None, secret)
    assert missing["received_present"] is False
    assert missing["received_sha256"] is None
    assert missing["bearer_prefix_valid"] is False
    assert missing["configured_present"] is True

    # A stored secret with surrounding whitespace is REPORTED but still matches,
    # so a trailing newline in a dashboard is visible without breaking auth.
    padded = auth.auth_compare_diagnostics(f"Bearer {secret}", secret + "\n")
    assert padded["configured_had_surrounding_whitespace"] is True
    assert padded["configured_raw_length"] == len(secret) + 1
    assert padded["configured_length"] == len(secret)
    assert padded["compare_digest_result"] is True


# ── Bank detection ────────────────────────────────────────────────────────────
#
# A real 37-page Standard Bank statement extracted cleanly (66k-78k characters
# from four extractors, account number and opening balance both found) and then
# failed with "No FNB transactions could be parsed from this PDF", because it
# was routed to parser_profile fnb_business_v1. These cases pin the detection
# that stops that happening.

STANDARD_BANK_SAMPLE = """
STANDARD BANK 6 month statement
Website:
www.standardbank.co.za
Customer Care Line 0860 123 000
Account Number 123 456 789
Date Description Payments Deposits Balance
STATEMENT OPENING BALANCE -992,452.57
30 Apr 25 ADT JHB 1,204.55 -993,657.12
30 Apr 25 SBSARETAIL 340.00 -993,997.12
02 May 25 SALARY DEPOSIT 45,000.00 -948,997.12
"""

FNB_SAMPLE = """
FIRST NATIONAL BANK
A division of FirstRand Bank Limited
www.fnb.co.za
Platinum Business Account
Statement Number 118
Transactions in RAND (ZAR)
01 Mar EFT Deposit Client 1,000.00Cr 1,000.00 Cr
01 Mar Card Purchase Fuel 300.00 700.00 Cr
"""

BANK_SAMPLES = {
    "fnb_business_v1": FNB_SAMPLE,
    "standard_bank_business_v1": STANDARD_BANK_SAMPLE,
    "absa_business_v1": "ABSA BANK LIMITED\nwww.absa.co.za\nCheque Account Statement\nDate Description Debit Credit Balance\n",
    "nedbank_business_v1": "NEDBANK LIMITED\nwww.nedbank.co.za\nBusiness Account Statement\nDate Description Debit Credit Balance\n",
    "capitec_business_v1": "CAPITEC BANK LIMITED\nwww.capitecbank.co.za\nBusiness Account Statement\nDate Description Money In Money Out Balance\n",
    "investec_business_v1": "INVESTEC BANK LIMITED\nwww.investec.co.za\nPrivate Bank Account Statement\nDate Description Debit Credit Balance\n",
}


def test_bank_detection_identifies_standard_bank_from_text() -> None:
    from engine.detection import detect_bank

    detection = detect_bank(STANDARD_BANK_SAMPLE)
    assert_equal(detection.profile_id, "standard_bank_business_v1", "Standard Bank detected from its own text")
    assert_equal(detection.bank_name, "Standard Bank", "Standard Bank name reported")
    assert_equal(detection.reason, "matched_bank_markers", "detection reason is evidence, not a default")
    if detection.confidence < 90:
        raise AssertionError(f"unopposed Standard Bank evidence should be high confidence, got {detection.confidence}")
    if not detection.evidence:
        raise AssertionError("a positive detection must name the evidence it used")


def test_bank_detection_covers_every_supported_bank() -> None:
    from engine.detection import detect_bank

    for expected_profile, sample in BANK_SAMPLES.items():
        detection = detect_bank(sample)
        assert_equal(detection.profile_id, expected_profile, f"{expected_profile} detected from statement text")


def test_bank_detection_keeps_fnb_unchanged() -> None:
    """The FNB path must not regress: this is the one bank with a real parser."""
    from engine.detection import detect_bank

    detection = detect_bank(FNB_SAMPLE)
    assert_equal(detection.profile_id, "fnb_business_v1", "FNB still detected from its own text")
    assert_equal(detection.bank_name, "FNB South Africa", "FNB name unchanged")


def test_bank_detection_never_defaults_to_a_bank() -> None:
    """No text, unrecognised text and a bankless ledger must all be `unknown`.

    BankRegistry.detect returns `_parsers[0]` — FNB — when nothing matches, so
    every unsupported bank became an FNB statement. There is no default here.
    """
    from engine.detection import UNKNOWN_PROFILE_ID, detect_bank

    for label, sample in (
        ("empty text", ""),
        ("whitespace only", "   \n\t  "),
        ("no bank markers", "Ledger Export\nDate Description Debit Credit Balance\n01 Jan Opening 0.00 0.00 100.00"),
    ):
        detection = detect_bank(sample)
        assert_equal(detection.profile_id, UNKNOWN_PROFILE_ID, f"{label} is unknown, not a default bank")
        assert_equal(detection.confidence, 0.0, f"{label} carries no confidence")


def test_bank_detection_ignores_a_counterparty_named_in_a_transaction() -> None:
    """A bank named in a description is not the issuer.

    Both directions are checked: an FNB statement paying Standard Bank stays
    FNB, and a Standard Bank statement paying FNB stays Standard Bank. The
    second case is also the storage-path guard — the literal token "fnb" in the
    body must not pull the statement back to the FNB parser.
    """
    from engine.detection import detect_bank

    fnb_paying_standard_bank = FNB_SAMPLE + "\n02 Mar EFT STANDARD BANK TRANSFER 5,000.00 -4,300.00\n"
    detection = detect_bank(fnb_paying_standard_bank)
    assert_equal(detection.profile_id, "fnb_business_v1", "FNB letterhead outweighs a Standard Bank payee")

    standard_bank_paying_fnb = STANDARD_BANK_SAMPLE + "\n05 May 25 EFT FNB TRANSFER 2,500.00 -951,497.12\n"
    detection = detect_bank(standard_bank_paying_fnb)
    assert_equal(detection.profile_id, "standard_bank_business_v1", "Standard Bank letterhead outweighs an FNB payee")


def test_bank_detection_reads_no_file_path() -> None:
    """The reported defect, pinned.

    Every accounting upload is stored at "{workspace}/accounting/fnb/{uuid}-{name}"
    (accountingStoragePath, lib/accounting/server.ts). BankRegistry.detect folds
    that path into its keyword haystack, so the literal "fnb" in it matched the
    FNB parser for every document and the statement text was never reached.

    detect_bank takes text only — there is no parameter through which a path
    could reach it, and the path-aware detector is now gone rather than merely
    unused. Runs uploaded before the storage path went neutral still carry the
    old ".../accounting/fnb/..." prefix, so this must keep holding for them.
    """
    import inspect

    from engine.detection import detect_bank
    from engine.registry import BankRegistry

    parameters = list(inspect.signature(detect_bank).parameters)
    if "text" not in parameters:
        raise AssertionError(f"detect_bank must take statement text, got {parameters}")
    for parameter in parameters:
        if any(token in parameter for token in ("path", "file", "name")):
            raise AssertionError(f"detect_bank must not accept a path/file input, found {parameter!r}")

    legacy_storage_path = "ws-1/accounting/fnb/2f6c-Standard_Bank_Statement.pdf"
    assert "fnb" in legacy_storage_path
    assert_equal(detect_bank(STANDARD_BANK_SAMPLE).profile_id, "standard_bank_business_v1", "text-only detection is correct")

    if hasattr(BankRegistry, "detect"):
        raise AssertionError("BankRegistry.detect matched bank keywords against the storage path; it must stay deleted")


def test_bank_detection_is_ambiguous_rather_than_wrong() -> None:
    """Balanced evidence for two banks is `unknown`, not a coin flip."""
    from engine.detection import UNKNOWN_PROFILE_ID, detect_bank

    detection = detect_bank("STANDARD BANK\nNEDBANK\nDate Description Debit Credit Balance\n")
    assert_equal(detection.profile_id, UNKNOWN_PROFILE_ID, "two equally-evidenced banks resolve to unknown")
    if not detection.reason.startswith("ambiguous"):
        raise AssertionError(f"expected an ambiguity reason, got {detection.reason!r}")


def test_bank_detection_survives_broken_letterhead_whitespace() -> None:
    """OCR breaks a letterhead across lines and pads it with non-breaking spaces."""
    from engine.detection import detect_bank

    broken = "STANDARD\n  BANK\n  6 month   statement\nwww.standardbank.co.za\n"
    assert_equal(detect_bank(broken).profile_id, "standard_bank_business_v1", "wrapped letterhead still detected")


# ── Routing ───────────────────────────────────────────────────────────────────


def _standard_bank_pages() -> list[dict]:
    """A Standard Bank Payments/Deposits layout, as pdfplumber returns it."""
    rows = [
        ["Date", "Description", "Payments", "Deposits", "Balance"],
        ["30 Apr 25", "ADT JHB", "1,204.55", "", "-993,657.12"],
        ["02 May 25", "SALARY DEPOSIT", "", "45,000.00", "-948,657.12"],
        ["03 May 25", "SBSARETAIL", "340.00", "", "-948,997.12"],
    ]
    text = "\n".join(" ".join(cell for cell in row if cell) for row in rows)
    return [{"page": 1, "text": f"STANDARD BANK 6 month statement\nwww.standardbank.co.za\n{text}", "tables": [rows]}]


def test_standard_bank_is_not_routed_to_the_fnb_parser() -> None:
    """The reported failure, end to end at the routing layer."""
    import main

    resolution = main.resolve_bank_profile(
        worker_profile="standard_bank_business_v1",
        worker_confidence=99.0,
        node_profile="standard_bank_business_v1",
        node_confidence=99.0,
    )
    assert_equal(resolution["bank_profile"], "standard_bank_business_v1", "Standard Bank resolved")
    assert_equal(resolution["bank_name"], "Standard Bank", "Standard Bank named")

    parser_profile = FNB_PROFILE if resolution["bank_profile"] == FNB_PROFILE else GENERIC_PROFILE
    assert_equal(parser_profile, GENERIC_PROFILE, "Standard Bank selects the generic parser, not FNB")


def test_standard_bank_layout_parses_payments_and_deposits() -> None:
    """Both money columns are read. Reading only the first dropped every deposit."""
    import main

    pages = _standard_bank_pages()
    metadata = main.parse_metadata(pages[0]["text"])
    transactions = main.parse_transactions(pages, metadata, pages[0]["text"], GENERIC_PROFILE)

    assert_equal(len(transactions), 3, "all three Standard Bank rows parsed")
    debits = [t for t in transactions if t.debit_amount]
    credits = [t for t in transactions if t.credit_amount]
    assert_equal(len(debits), 2, "both payments read as debits")
    assert_equal(len(credits), 1, "the deposit read as a credit")
    assert_equal(credits[0].credit_amount, 45000.0, "deposit amount")
    assert_equal(debits[0].debit_amount, 1204.55, "payment amount")


def test_generic_parser_never_invents_fnb_fee_rows() -> None:
    """FNB fee reconstruction infers rows from FNB's own fee summary.

    Running it on another bank's statement would add transactions that are not
    printed anywhere on the document.
    """
    import main

    pages = _standard_bank_pages()
    metadata = main.parse_metadata(pages[0]["text"])
    # A fee summary of the shape the FNB inference reads, on a non-FNB statement.
    metadata["total_service_fees"] = 250.00
    metadata["opening_balance"] = -992452.57

    generic = main.parse_transactions(pages, metadata, pages[0]["text"], GENERIC_PROFILE)
    assert_equal(len(generic), 3, "generic parser returns only the printed rows")
    for transaction in generic:
        if "#" in transaction.description:
            raise AssertionError(f"generic parser produced an FNB fee row: {transaction.description}")


def test_structured_rows_skip_fnb_reconstruction_for_other_banks() -> None:
    """Structured rows stay the preferred source for every bank.

    Only the three FNB-specific reconstruction steps are gated by profile.
    """
    import main

    metadata = {"statement_period_end": "2025-05-31"}
    rows = [
        {"pageNumber": 1, "cells": {"date": "30 Apr 25", "description": "ADT JHB", "debit": "1,204.55", "balance": "-993,657.12"}},
        {"pageNumber": 1, "cells": {"date": "02 May 25", "description": "SALARY DEPOSIT", "credit": "45,000.00", "balance": "-948,657.12"}},
    ]

    generic_txns, generic_diag = main.parse_structured_rows(rows, dict(metadata), GENERIC_PROFILE)
    assert_equal(len(generic_txns), 2, "structured rows parse for a non-FNB bank")
    assert_equal(generic_diag["fnb_reconstruction_applied"], False, "FNB reconstruction stays off")
    assert_equal(generic_diag["profile"], GENERIC_PROFILE, "diagnostics name the profile used")

    fnb_txns, fnb_diag = main.parse_structured_rows(rows, dict(metadata), FNB_PROFILE)
    assert_equal(fnb_diag["fnb_reconstruction_applied"], True, "FNB reconstruction still runs for FNB")
    assert_equal(len(fnb_txns), 2, "FNB path unchanged on rows that need no reconstruction")


def test_unknown_bank_routes_to_generic_not_fnb() -> None:
    """An unidentified statement is parsed generically, never as FNB."""
    import main
    from engine.detection import UNKNOWN_PROFILE_ID

    resolution = main.resolve_bank_profile(
        worker_profile=UNKNOWN_PROFILE_ID,
        worker_confidence=0.0,
        node_profile=None,
        node_confidence=None,
    )
    assert_equal(resolution["bank_profile"], UNKNOWN_PROFILE_ID, "unknown stays unknown")
    assert_equal(resolution["bank_name"], "Unknown", "unknown is named honestly")
    assert_equal(resolution["source"], "none", "neither side identified a bank")

    parser_profile = FNB_PROFILE if resolution["bank_profile"] == FNB_PROFILE else GENERIC_PROFILE
    assert_equal(parser_profile, GENERIC_PROFILE, "unknown selects the generic parser")


def test_fnb_still_routes_to_the_fnb_parser() -> None:
    """The one bank with a real parser must keep reaching it."""
    import main

    resolution = main.resolve_bank_profile(
        worker_profile=FNB_PROFILE,
        worker_confidence=99.0,
        node_profile=FNB_PROFILE,
        node_confidence=99.0,
    )
    assert_equal(resolution["bank_profile"], FNB_PROFILE, "FNB resolved")
    parser_profile = FNB_PROFILE if resolution["bank_profile"] == FNB_PROFILE else GENERIC_PROFILE
    assert_equal(parser_profile, FNB_PROFILE, "FNB selects the FNB parser")


def test_bank_resolution_precedence_between_the_two_detections() -> None:
    """Each side can be the better witness; the rules say which one wins."""
    import main
    from engine.detection import UNKNOWN_PROFILE_ID

    only_node = main.resolve_bank_profile(UNKNOWN_PROFILE_ID, 0.0, "standard_bank_business_v1", 99.0)
    assert_equal(only_node["bank_profile"], "standard_bank_business_v1", "node identifies what the worker could not")
    assert_equal(only_node["source"], "node", "attributed to the node pipeline")

    only_worker = main.resolve_bank_profile("nedbank_business_v1", 88.0, "unknown", 0.0)
    assert_equal(only_worker["bank_profile"], "nedbank_business_v1", "worker identifies what node could not")
    assert_equal(only_worker["source"], "worker", "attributed to the worker")

    node_more_confident = main.resolve_bank_profile("absa_business_v1", 60.0, "capitec_business_v1", 95.0)
    assert_equal(node_more_confident["bank_profile"], "capitec_business_v1", "conflict goes to the more confident read")

    worker_more_confident = main.resolve_bank_profile("absa_business_v1", 95.0, "capitec_business_v1", 60.0)
    assert_equal(worker_more_confident["bank_profile"], "absa_business_v1", "and the other way round")

    tie = main.resolve_bank_profile("absa_business_v1", 90.0, "capitec_business_v1", 90.0)
    assert_equal(tie["bank_profile"], "absa_business_v1", "an exact tie goes to the text about to be parsed")

    # A newer frontend naming a bank this worker does not know is not a bank.
    unknown_id = main.resolve_bank_profile(UNKNOWN_PROFILE_ID, 0.0, "tymebank_business_v1", 99.0)
    assert_equal(unknown_id["bank_profile"], UNKNOWN_PROFILE_ID, "unrecognised ids count as no identification")


def test_parser_profile_not_implemented_rejection_is_gone() -> None:
    """Being a bank without a dedicated parser must not be a 422.

    The worker used to raise "parser_profile_not_implemented" for anything that
    was not FNB. That guard was unreachable in production (everything detected
    as FNB via the storage path) and is wrong now that a generic parser exists.
    """
    source = (ROOT / "workers" / "accounting_worker" / "main.py").read_text()
    if "parser_profile_not_implemented" in source:
        raise AssertionError("the not-implemented rejection must be gone; unsupported banks parse generically")
    if "only fnb_business_v1 is implemented" in source:
        raise AssertionError("the not-implemented message must be gone")


# ── Generic parser ────────────────────────────────────────────────────────────

STANDARD_BANK_TEXT_STATEMENT = """STANDARD BANK 6 month statement
www.standardbank.co.za
Page 1 of 37
Date Description Payments Deposits Balance
STATEMENT OPENING BALANCE -992,452.57
30 Apr 25 ADT JHB SECURITY SERVICES
MONTHLY CONTRACT 1,204.55 -993,657.12
30 Apr 25 SBSARETAIL 340.00 -993,997.12
02 May 25 SALARY DEPOSIT 45,000.00 -948,997.12
02 May 25 DEBIT ORDER INSURANCE 890.00 -949,887.12
FUNERAL COVER 220.00 -950,107.12
09 May 25 CLIENT EFT PAYMENT
INVOICE 20551 12,500.00 -937,607.12
Page 2 of 37
Date Description Payments Deposits Balance
TOTAL PAYMENTS 2,654.55
"""


def _generic_text_pages(text: str) -> list[dict]:
    return [{"page": 1, "text": text, "tables": []}]


def test_generic_parser_reads_a_text_only_statement_end_to_end() -> None:
    """No tables at all — the realistic shape of a provider extraction."""
    import main

    pages = _generic_text_pages(STANDARD_BANK_TEXT_STATEMENT)
    metadata = main.parse_metadata(STANDARD_BANK_TEXT_STATEMENT)
    transactions = main.parse_transactions(pages, metadata, STANDARD_BANK_TEXT_STATEMENT, GENERIC_PROFILE)
    summary = main.validation_summary(transactions)

    assert_equal(len(transactions), 6, "every printed movement recovered")
    assert_equal(summary["debit_count"], 4, "four payments")
    assert_equal(summary["credit_count"], 2, "two deposits")
    assert_equal(str(summary["total_debits"]), "2654.55", "debit total")
    assert_equal(str(summary["total_credits"]), "57500.00", "credit total")

    # The whole point of a ledger: opening plus movements must land on the
    # printed closing balance.
    opening = main.decimal_amount(metadata["opening_balance"])
    closing = opening - summary["total_debits"] + summary["total_credits"]
    assert_equal(str(closing), "-937607.12", "reconciles to the printed closing balance")


def test_generic_parser_merges_wrapped_descriptions() -> None:
    """A description that wraps must not become a row of its own."""
    import main

    pages = _generic_text_pages(STANDARD_BANK_TEXT_STATEMENT)
    metadata = main.parse_metadata(STANDARD_BANK_TEXT_STATEMENT)
    transactions = main.parse_transactions(pages, metadata, STANDARD_BANK_TEXT_STATEMENT, GENERIC_PROFILE)

    descriptions = [t.description for t in transactions]
    if "ADT JHB SECURITY SERVICES MONTHLY CONTRACT" not in descriptions:
        raise AssertionError(f"wrapped description not rejoined: {descriptions}")
    if "CLIENT EFT PAYMENT INVOICE 20551" not in descriptions:
        raise AssertionError(f"wrapped description not rejoined: {descriptions}")
    for description in descriptions:
        if description in {"MONTHLY CONTRACT", "INVOICE 20551"}:
            raise AssertionError(f"description fragment became its own row: {description}")


def test_generic_parser_inherits_the_date_of_a_grouped_movement() -> None:
    """Banks print the date once per date group; the rows under it are separate."""
    import main

    pages = _generic_text_pages(STANDARD_BANK_TEXT_STATEMENT)
    metadata = main.parse_metadata(STANDARD_BANK_TEXT_STATEMENT)
    transactions = main.parse_transactions(pages, metadata, STANDARD_BANK_TEXT_STATEMENT, GENERIC_PROFILE)

    funeral = next((t for t in transactions if t.description == "FUNERAL COVER"), None)
    if funeral is None:
        raise AssertionError("the dateless grouped movement was dropped")
    assert_equal(funeral.transaction_date, "2025-05-02", "inherited the date of its group")
    assert_equal(funeral.debit_amount, 220.0, "and kept its own amount")


def test_generic_parser_infers_direction_from_balance_continuity() -> None:
    """Text loses the column a figure was printed in; the arithmetic recovers it.

    "45,000.00" on a Payments/Deposits statement carries no sign and no column
    once flattened to text. The balance moving UP by exactly that amount is what
    makes it a deposit.
    """
    import main

    pages = _generic_text_pages(STANDARD_BANK_TEXT_STATEMENT)
    metadata = main.parse_metadata(STANDARD_BANK_TEXT_STATEMENT)
    transactions = main.parse_transactions(pages, metadata, STANDARD_BANK_TEXT_STATEMENT, GENERIC_PROFILE)

    salary = next(t for t in transactions if t.description == "SALARY DEPOSIT")
    assert_equal(salary.credit_amount, 45000.0, "balance rose, so it is a credit")
    assert_equal(main.decimal_amount(salary.debit_amount), main.decimal_amount(0), "and not a debit")

    fee = next(t for t in transactions if t.description == "SBSARETAIL")
    assert_equal(fee.debit_amount, 340.0, "balance fell, so it is a debit")
    assert_equal(main.decimal_amount(fee.credit_amount), main.decimal_amount(0), "and not a credit")


def test_generic_parser_drops_page_furniture_and_totals() -> None:
    """Repeated headers, page numbers and summary totals are not transactions."""
    import main

    pages = _generic_text_pages(STANDARD_BANK_TEXT_STATEMENT)
    metadata = main.parse_metadata(STANDARD_BANK_TEXT_STATEMENT)
    transactions = main.parse_transactions(pages, metadata, STANDARD_BANK_TEXT_STATEMENT, GENERIC_PROFILE)

    for transaction in transactions:
        lowered = (transaction.description or "").lower()
        for forbidden in ("page 1 of", "page 2 of", "total payments", "date description", "standard bank"):
            if forbidden in lowered:
                raise AssertionError(f"furniture parsed as a transaction: {transaction.description!r}")
    if str(main.decimal_amount(2654.55)) in [str(main.decimal_amount(t.credit_amount)) for t in transactions]:
        raise AssertionError("the totals line was read as a credit")


def test_generic_parser_reads_a_debit_credit_layout() -> None:
    """The other common column pair, with Dr/Cr suffixes carrying the sign."""
    import main

    text = """NEDBANK LIMITED
www.nedbank.co.za
Date Description Debit Credit Balance
Opening Balance 5,000.00
03 Jun 2025 EFT RECEIVED 2,500.00Cr 7,500.00
04 Jun 2025 SERVICE FEE 125.00Dr 7,375.00
"""
    pages = _generic_text_pages(text)
    metadata = main.parse_metadata(text)
    transactions = main.parse_transactions(pages, metadata, text, GENERIC_PROFILE)

    assert_equal(len(transactions), 2, "both rows read")
    received = next(t for t in transactions if "EFT RECEIVED" in t.description)
    fee = next(t for t in transactions if "SERVICE FEE" in t.description)
    assert_equal(received.credit_amount, 2500.0, "Cr suffix read as a credit")
    assert_equal(fee.debit_amount, 125.0, "Dr suffix read as a debit")


def test_negative_opening_balance_keeps_its_sign() -> None:
    """An overdrawn statement's opening balance is negative.

    The separator pattern used to swallow the minus, so -992,452.57 was recorded
    as +992,452.57 — and on a statement whose amounts carry no sign of their own,
    that made every row unresolvable.
    """
    import main

    assert_equal(
        main.parse_metadata("STATEMENT OPENING BALANCE -992,452.57")["opening_balance"],
        -992452.57,
        "a minus against the figure is a sign",
    )
    assert_equal(
        main.parse_metadata("Opening Balance - 1,000.00")["opening_balance"],
        1000.0,
        "a dash with a space after it is a separator",
    )
    assert_equal(main.parse_metadata("Opening Balance: -50.25")["opening_balance"], -50.25, "colon then sign")
    assert_equal(main.parse_metadata("Balance Brought Forward 342.37")["opening_balance"], 342.37, "unsigned is positive")


def test_candidate_counting_is_bank_independent() -> None:
    """The text-source comparison must not be blind on non-FNB statements.

    transaction_candidate_lines only enters a section after FNB's "Transactions
    in RAND" heading, so it returns 0 for every other bank. Comparing 0 with 0
    is how a Standard Bank statement's provider extraction lost to this worker's
    own text.
    """
    import main
    from engine.generic_parser import count_candidate_lines

    assert_equal(len(main.transaction_candidate_lines(STANDARD_BANK_TEXT_STATEMENT)), 0, "the FNB counter is blind here")
    if count_candidate_lines(STANDARD_BANK_TEXT_STATEMENT) < 6:
        raise AssertionError("the generic counter must see the Standard Bank rows")
    # And it must still count an FNB statement, so the comparison stays sane if
    # a preliminary detection is wrong.
    if count_candidate_lines(FNB_SAMPLE) < 2:
        raise AssertionError("the generic counter must also see FNB rows")


def test_generic_rows_and_provider_rows_take_the_same_path() -> None:
    """Text rows and Azure/Mistral rows converge on one transformer.

    Two transformers would mean two sets of Dr/Cr, date-inheritance and
    informational-row rules, drifting apart.
    """
    import main
    from engine.generic_parser import extract_generic_rows

    rows = extract_generic_rows(_generic_text_pages(STANDARD_BANK_TEXT_STATEMENT))
    if not rows:
        raise AssertionError("the generic parser produced no rows")
    for row in rows:
        if not isinstance(row.get("cells"), dict):
            raise AssertionError(f"row is not in StructuredRow shape: {row}")
        if "raw" not in row or "pageNumber" not in row:
            raise AssertionError(f"row is missing StructuredRow fields: {row}")

    # The rows go through the same function the providers' rows do.
    transactions, diagnostics = main.parse_structured_rows(rows, main.parse_metadata(STANDARD_BANK_TEXT_STATEMENT), GENERIC_PROFILE)
    assert_equal(len(transactions), 6, "the shared transformer reads the generic rows")
    assert_equal(diagnostics["fnb_reconstruction_applied"], False, "and applies no FNB reconstruction")


# ── Outcome semantics ─────────────────────────────────────────────────────────


def _txn(main, date, description, debit=None, credit=None, balance=None):
    return main.ParsedTransaction(
        transaction_date=date,
        description=description,
        debit_amount=debit,
        credit_amount=credit,
        running_balance=balance,
    )


def test_a_statement_that_ties_out_completes() -> None:
    """Evidence that everything was recovered means the run is finished."""
    import main

    metadata = {"opening_balance": 1000.00, "closing_balance": 700.00}
    transactions = [
        _txn(main, "2025-06-01", "PAYMENT A", debit=100.00, balance=900.00),
        _txn(main, "2025-06-02", "PAYMENT B", debit=200.00, balance=700.00),
    ]
    check = main.validate_extraction(metadata, transactions)
    assert_equal(check["status"], "ok", "a reconciling statement is complete")
    assert_equal(check["failures"], [], "no failures")


def test_a_broken_balance_chain_is_partial_not_complete() -> None:
    """Rows missing from the middle break continuity even when nothing declares totals.

    This is the only completeness evidence an unsupported bank is guaranteed to
    give us, and without it a generic parse that dropped half a statement was
    reported as a clean success.
    """
    import main

    metadata = {"opening_balance": 1000.00}
    transactions = [
        _txn(main, "2025-06-01", "PAYMENT A", debit=100.00, balance=900.00),
        # 900 - 200 is 700, but the statement prints 400: a row is missing.
        _txn(main, "2025-06-02", "PAYMENT B", debit=200.00, balance=400.00),
    ]
    check = main.validate_extraction(metadata, transactions)
    assert_equal(check["status"], "review_required", "a broken chain is not a completed run")
    if "balance_continuity" not in check["failures"]:
        raise AssertionError(f"continuity must be the failing check, got {check['failures']}")
    assert_equal(check["balance_gap_count"], 1, "the gap is counted")


def test_continuity_does_not_second_guess_a_statement_whose_money_ties_out() -> None:
    """Continuity substitutes for declared evidence; it does not add to it.

    FNB prints an overdrawn balance as a magnitude with a Dr marker, so the row
    chain shows gaps of twice the balance on statements whose declared totals
    reconcile to the cent. The stronger signal wins.
    """
    import main

    metadata = {"opening_balance": 1000.00, "closing_balance": 700.00}
    transactions = [
        _txn(main, "2025-06-01", "PAYMENT A", debit=100.00, balance=500.00),
        _txn(main, "2025-06-02", "PAYMENT B", debit=200.00, balance=700.00),
    ]
    check = main.validate_extraction(metadata, transactions)
    ran = [item["name"] for item in check["checks"]]
    if "balance_continuity" in ran:
        raise AssertionError("continuity must not run when the money already ties out")
    assert_equal(check["status"], "ok", "the reconciling statement completes")
    assert_equal(check["balance_gap_count"], 2, "the gaps are still reported as diagnostics")


def test_no_evidence_at_all_goes_to_review_never_to_completed() -> None:
    """"Nothing failed" is not "it is right".

    A statement that declares no totals, prints no closing balance and gives no
    opening balance to chain from leaves every check unrun. Reporting that as a
    clean success would be an unverified claim.
    """
    import main

    metadata: dict = {}
    transactions = [_txn(main, "2025-06-01", "PAYMENT A", debit=100.00)]
    check = main.validate_extraction(metadata, transactions)

    assert_equal(check["evidence_checks_run"], 0, "nothing could be checked")
    assert_equal(check["status"], "review_required", "so it cannot be called complete")
    if "no_completeness_evidence" not in check["failures"]:
        raise AssertionError(f"the reason must be named, got {check['failures']}")


def test_recovery_options_separates_unreadable_from_unparsed() -> None:
    """"We could not read it" and "there is nothing to read" are different runs."""
    import main

    empty_selection: dict = {}

    nothing = main.recovery_options(full_text="", pages=[], structured_rows=None, structured_selection=empty_selection)
    assert_equal(nothing["recoverable"], False, "a document with no text, tables or rows is exhausted")
    if "no usable text" not in nothing["summary"]:
        raise AssertionError(f"the summary must say what is missing: {nothing['summary']}")

    text_remains = main.recovery_options(
        full_text=STANDARD_BANK_TEXT_STATEMENT,
        pages=_generic_text_pages(STANDARD_BANK_TEXT_STATEMENT),
        structured_rows=None,
        structured_selection=empty_selection,
    )
    assert_equal(text_remains["recoverable"], True, "legible dated money rows are recoverable material")
    if text_remains["generic_candidate_lines"] < 6:
        raise AssertionError("the remaining rows must be counted")

    rows_remain = main.recovery_options(
        full_text="",
        pages=[],
        structured_rows=[{"pageNumber": 1, "cells": {"date": "01 Jun", "description": "X", "amount": "1.00"}}],
        structured_selection={"structured_rows_usable": False, "structured_rejection_reason": "too_many_duplicates:9"},
    )
    assert_equal(rows_remain["recoverable"], True, "rows that arrived are material even when rejected")
    assert_equal(rows_remain["structured_rejection_reason"], "too_many_duplicates:9", "why they were rejected is kept")


def test_failure_messages_name_the_parser_that_ran() -> None:
    """A Standard Bank statement reported as an FNB failure describes the routing."""
    source = (ROOT / "workers" / "accounting_worker" / "main.py").read_text()
    if '"message": "FNB parser validation failed."' in source:
        raise AssertionError("validation failure must not be attributed to FNB on every bank")
    if "no_transactions_recoverable" not in source:
        raise AssertionError("an exhausted run must be reported apart from an unparsed one")


# ── AI recovery ───────────────────────────────────────────────────────────────
#
# The instruction "do not invent transactions" is not a control. These cases
# check the control: nothing survives that cannot be traced back to a line we
# actually sent.

AI_SOURCE_LINES = [
    "30 Apr 25 ADT JHB SECURITY 1,204.55 -993,657.12",
    "02 May 25 SALARY DEPOSIT 45,000.00 -948,657.12",
    "05 May 25 MONTHLY SERVICE FEE 150.00 -948,807.12",
]


def _ai_row(**over):
    row = {
        "source_line": AI_SOURCE_LINES[0],
        "date": "30 Apr 25",
        "description": "ADT JHB SECURITY",
        "amount": "1,204.55",
        "balance": "-993,657.12",
        "direction": "debit",
        "confidence": 0.9,
    }
    row.update(over)
    return row


def test_ai_rows_must_come_from_a_line_we_sent() -> None:
    """A row whose source line is not in the document is discarded outright."""
    from engine.ai_recovery import ground_rows

    invented = _ai_row(
        source_line="07 May 25 CONSULTING FEE 9,500.00 -958,307.12",
        description="CONSULTING FEE",
        amount="9,500.00",
        balance="-958,307.12",
    )
    report = ground_rows([_ai_row(), invented], AI_SOURCE_LINES)
    assert_equal(len(report.accepted), 1, "only the row that exists survives")
    assert_equal(report.rejected.get("source_line_not_in_document"), 1, "the invented row is counted, not silently dropped")


def test_ai_amounts_must_appear_in_their_source_line() -> None:
    """A real line with an altered figure is still a fabrication."""
    from engine.ai_recovery import ground_rows

    report = ground_rows([_ai_row(amount="1,240.55")], AI_SOURCE_LINES)
    assert_equal(len(report.accepted), 0, "a figure not printed on the line is rejected")
    assert_equal(report.rejected.get("amount_not_in_source_line"), 1, "and the reason is recorded")


def test_ai_balances_that_are_not_printed_are_dropped_but_the_row_is_kept() -> None:
    """A computed balance is discarded; the row it belongs to is not."""
    from engine.ai_recovery import ground_rows

    report = ground_rows([_ai_row(balance="-993,000.00")], AI_SOURCE_LINES)
    assert_equal(len(report.accepted), 1, "the row survives")
    assert_equal(report.accepted[0]["balance"], "", "the unprinted balance does not")
    assert_equal(report.accepted[0]["balance_dropped"], True, "and that is flagged")


def test_ai_descriptions_may_not_introduce_words() -> None:
    """A narrative the document does not contain is a fabrication too."""
    from engine.ai_recovery import ground_rows

    report = ground_rows([_ai_row(description="ADT JHB SECURITY monthly alarm monitoring contract")], AI_SOURCE_LINES)
    assert_equal(len(report.accepted), 0, "invented narrative is rejected")
    assert_equal(report.rejected.get("description_not_in_source_line"), 1, "and counted")

    kept = ground_rows([_ai_row(description="ADT SECURITY")], AI_SOURCE_LINES)
    assert_equal(len(kept.accepted), 1, "a description drawn from the line is fine")


def test_ai_never_receives_or_returns_a_guessed_direction() -> None:
    """An unreadable direction stays unknown; it is not filled in."""
    from engine.ai_recovery import ground_rows

    report = ground_rows([_ai_row(direction="probably debit")], AI_SOURCE_LINES)
    assert_equal(report.accepted[0]["direction"], "unknown", "an unrecognised direction becomes unknown, not debit")


def test_ai_candidate_lines_only_carry_figures() -> None:
    """Prose and page chrome cannot be ledger rows and are not worth the context."""
    from engine.ai_recovery import candidate_lines

    text = (
        "STANDARD BANK 6 month statement\n"
        "Please retain this statement for your records\n"
        "30 Apr 25 ADT JHB SECURITY 1,204.55 -993,657.12\n"
        "Customer Care 0860 123 000\n"
        "02 May 25 SALARY DEPOSIT 45,000.00 -948,657.12\n"
    )
    lines = candidate_lines(text)
    assert_equal(len(lines), 2, "only the two lines carrying figures are sent")
    for line in lines:
        if "Please retain" in line or "Customer Care" in line:
            raise AssertionError(f"non-ledger line sent to the model: {line}")


def test_ai_line_cap_is_reported_never_silent() -> None:
    """A capped read must not look like a complete one."""
    from engine.ai_recovery import LINES_PER_BATCH, MAX_BATCHES, batches, dropped_line_count

    lines = [f"0{index % 9 + 1} May 25 ROW {index} 1{index}.00 -1,000.00" for index in range(LINES_PER_BATCH * MAX_BATCHES + 37)]
    assert_equal(len(batches(lines)), MAX_BATCHES, "batches are capped")
    assert_equal(dropped_line_count(lines), 37, "and the remainder is counted so it can be logged")
    assert_equal(dropped_line_count(lines[:10]), 0, "nothing dropped when everything fits")


def test_ai_recovery_is_skipped_without_a_key_and_never_invents_a_ledger() -> None:
    """No provider configured is a clean no-op, not a failure and not a guess."""
    import os

    import main

    previous = os.environ.pop("OPENAI_API_KEY", None)
    try:
        transactions, diagnostics = main.attempt_ai_recovery(
            full_text=STANDARD_BANK_TEXT_STATEMENT,
            structured_rows=None,
            metadata={},
            bank_name="Standard Bank",
            run_id="run-1",
        )
        assert_equal(transactions, [], "no key means no transactions")
        assert_equal(diagnostics["enabled"], False, "and it says so")
        assert_equal(diagnostics["reason"], "no_api_key", "with the reason named")
    finally:
        if previous is not None:
            os.environ["OPENAI_API_KEY"] = previous


def test_ai_recovered_rows_are_all_flagged_for_review() -> None:
    """Located is not understood. A person has to see every AI-derived row."""
    import os

    import main
    from engine import ai_recovery as ai_module

    rows = [
        {"source_line": AI_SOURCE_LINES[0], "date": "30 Apr 25", "description": "ADT JHB SECURITY",
         "amount": "1,204.55", "balance": "-993,657.12", "direction": "debit", "confidence": 0.95},
        {"source_line": AI_SOURCE_LINES[1], "date": "02 May 25", "description": "SALARY DEPOSIT",
         "amount": "45,000.00", "balance": "-948,657.12", "direction": "credit", "confidence": 0.9},
        # An invented row, returned alongside the real ones.
        {"source_line": "11 May 25 CONSULTING 9,500.00 -939,157.12", "date": "11 May 25",
         "description": "CONSULTING", "amount": "9,500.00", "balance": "-939,157.12", "direction": "debit", "confidence": 0.99},
    ]

    text = "\n".join(AI_SOURCE_LINES)
    previous_key = os.environ.get("OPENAI_API_KEY")
    original_transport = main.openai_chat_completion
    os.environ["OPENAI_API_KEY"] = "test-key"
    main.openai_chat_completion = lambda body, api_key, timeout=60: {
        "choices": [{"message": {"content": json.dumps({"rows": rows})}}]
    }
    try:
        transactions, diagnostics = main.attempt_ai_recovery(
            full_text=text,
            structured_rows=None,
            metadata={"opening_balance": -992452.57, "statement_period_end": "2025-05-31"},
            bank_name="Standard Bank",
            run_id="run-1",
        )
    finally:
        main.openai_chat_completion = original_transport
        if previous_key is None:
            os.environ.pop("OPENAI_API_KEY", None)
        else:
            os.environ["OPENAI_API_KEY"] = previous_key

    assert_equal(len(transactions), 2, "the two grounded rows are recovered")
    assert_equal(diagnostics["returned_rows"], 3, "the model returned three")
    assert_equal(diagnostics["rejected_rows"].get("source_line_not_in_document"), 1, "the third was invented and rejected")

    for transaction in transactions:
        assert_equal(transaction.review_status, "needs_review", f"{transaction.description} must be flagged")
        if "recovered_by: ai" not in transaction.notes:
            raise AssertionError(f"{transaction.description} must record where it came from: {transaction.notes!r}")
        if transaction.confidence > 60:
            raise AssertionError(f"AI rows must not carry parser-grade confidence, got {transaction.confidence}")

    credit = next(t for t in transactions if "SALARY" in t.description)
    assert_equal(credit.credit_amount, 45000.0, "amounts are the printed ones")


def test_an_ai_recovered_run_can_never_report_completed() -> None:
    """AI succeeding is a review outcome, by construction, not by luck."""
    source = (ROOT / "workers" / "accounting_worker" / "main.py").read_text()
    if "or ai_recovered" not in source:
        raise AssertionError("the run status must force review when AI produced the ledger")
    if "append_note(transaction, AI_RECOVERY_NOTE)" not in source:
        raise AssertionError("every AI row must be traceable in the ledger itself")
    if 'AI_RECOVERY_NOTE = "recovered_by: ai"' not in source:
        raise AssertionError("the marker the ledger and the classification guard share must not drift")


# ── Ledger section boundaries ─────────────────────────────────────────────────
#
# Modelled on the real 37-page Standard Bank statement (SBSA_Statement_2025-10-27,
# run 27142721-ab81-4981-ad66-89fb8f299ed1), which cannot be committed: it is a
# client bank statement. The structure is reproduced exactly — a letterhead
# repeated on every page carrying both a date and an available-balance figure,
# the column header, transactions whose narrative wraps onto the next line, and
# a footer whose summary prints figures in transaction shape.
#
# On the real document this shape produced 4 rows that appear nowhere on the
# statement and 15 running-balance gaps.

def _sbsa_shaped_page(page_number: int, body: list[str], footer: bool = False) -> dict:
    letterhead = [
        "Customer Care: 0860 123 000",
        "Website: www.standardbank.co.za",
        "STANDARD BANK 6 month statement",
        "From: 30 Apr 25",
        "To: 27 Oct 25",
        # A bare date line. It opens a row unless the section is closed.
        "27 Oct 2025",
        "Account number: 30 198 148 5 Address:",
        "Account holder: TEST HOLDER PTY LTD 3 UITKYK ST",
        "Product name: CURRENT ACC FLAMINGO PARK",
        # Carries a figure, and completes the row the bare date opened.
        "Transaction details Available Balance: R96,313.45",
        "Date Description Payments Deposits Balance",
    ]
    tail = [
        "Please verify all transactions reflected on this statement and notify any discrepancies to the bank.",
        "Statement Summary",
        "Payments -R2,000.00",
        "Deposits R5,000.00",
        "Today's debits have not yet been paid",
        "The Standard Bank of South Africa Limited (Reg. No. 1962/000738/06).",
    ] if footer else []
    return {"page": page_number, "text": "\n".join(letterhead + body + tail), "tables": []}


SBSA_SHAPED_PAGES = [
    _sbsa_shaped_page(1, [
        "STATEMENT OPENING BALANCE -1,000.00",
        "30 Apr 25 ADT JHB 2117556751ADT5087498 -380.00 -1,380.00",
        "ACCOUNT PAYMENT",
        "30 Apr 25 SBSARETAIL895F 00040202771 -620.00 -2,000.00",
        "LOAN REPAYMENT",
    ]),
    _sbsa_shaped_page(2, [
        "02 May 25 CLIENT SETTLEMENT 20250502 5,000.00 3,000.00",
        "ELECTRONIC TRANSFER CREDIT",
        "05 May 25 CARTRACK CART25D5S58NYRV -1,000.00 2,000.00",
        "ACCOUNT PAYMENT",
    ], footer=True),
]


def test_the_repeated_letterhead_is_not_read_as_a_transaction() -> None:
    """The defect the real Standard Bank statement exposed.

    The letterhead carries a bare statement date and an available-balance
    figure. Read as ledger lines, the date opened a row, the address block wrapped
    into it, and the available balance completed it — one invented transaction
    per page, each breaking the running-balance chain twice.
    """
    import main

    full_text = "\n".join(page["text"] for page in SBSA_SHAPED_PAGES)
    metadata = main.parse_metadata(full_text)
    transactions = main.parse_transactions(SBSA_SHAPED_PAGES, metadata, full_text, GENERIC_PROFILE)

    for transaction in transactions:
        for forbidden in ("Account number", "Available Balance", "Account holder", "Product name"):
            if forbidden in transaction.description:
                raise AssertionError(f"letterhead read as a transaction: {transaction.description!r}")
    if any(main.decimal_amount(t.running_balance) == main.decimal_amount(96313.45) for t in transactions if t.running_balance is not None):
        raise AssertionError("the available-balance figure became a running balance")


def test_the_footer_summary_is_not_read_as_transactions() -> None:
    """"Payments -R2,000.00" is a summary of the statement, not a movement."""
    import main

    full_text = "\n".join(page["text"] for page in SBSA_SHAPED_PAGES)
    metadata = main.parse_metadata(full_text)
    transactions = main.parse_transactions(SBSA_SHAPED_PAGES, metadata, full_text, GENERIC_PROFILE)

    descriptions = [t.description for t in transactions]
    for forbidden in ("Payments", "Deposits", "Today's debits"):
        if any(description.strip().startswith(forbidden) for description in descriptions):
            raise AssertionError(f"footer summary read as a transaction: {forbidden}")
    # The closing disclaimer must not be absorbed into the last transaction either.
    for description in descriptions:
        if "Please verify all transactions" in description:
            raise AssertionError(f"footer prose absorbed as a wrapped description: {description!r}")


def test_the_section_bounded_ledger_reconciles_exactly() -> None:
    """With both boundaries in place the chain is continuous and ties out.

    This is the shape that, on the real 37-page statement, produced 615
    transactions whose debit and credit totals match the bank's own declared
    Payments and Deposits to the cent, with zero running-balance gaps.
    """
    import main

    full_text = "\n".join(page["text"] for page in SBSA_SHAPED_PAGES)
    metadata = main.parse_metadata(full_text)
    transactions = main.parse_transactions(SBSA_SHAPED_PAGES, metadata, full_text, GENERIC_PROFILE)
    summary = main.validation_summary(transactions)

    assert_equal(len(transactions), 4, "the four printed movements, and nothing else")
    assert_equal(str(summary["total_debits"]), "2000.00", "debits match the statement's declared Payments")
    assert_equal(str(summary["total_credits"]), "5000.00", "credits match the statement's declared Deposits")
    assert_equal(len(main.balance_gap_diagnostics(metadata, transactions)), 0, "the running balance is continuous")

    check = main.validate_extraction(metadata, transactions)
    assert_equal(check["status"], "ok", "a continuous, complete ledger")


def test_a_statement_with_no_column_header_is_still_read() -> None:
    """A layout we cannot recognise is better parsed than skipped."""
    import main

    text = "\n".join([
        "SOME BANK LIMITED",
        "01 Jun 2025 EFT RECEIVED 2,500.00Cr 7,500.00",
        "02 Jun 2025 SERVICE FEE 125.00Dr 7,375.00",
    ])
    pages = [{"page": 1, "text": text, "tables": []}]
    transactions = main.parse_transactions(pages, main.parse_metadata(text), text, GENERIC_PROFILE)
    assert_equal(len(transactions), 2, "no header means read the whole page, not none of it")


# ── AI recovery confidence cap ────────────────────────────────────────────────


def _ai_recovered_row(main, description="CARTRACK CART25D5S58NYRV ACCOUNT PAYMENT"):
    """A row exactly as attempt_ai_recovery leaves it."""
    return main.ParsedTransaction(
        transaction_date="2025-05-02",
        description=description,
        debit_amount=1710.15,
        running_balance=-1001484.52,
        confidence=54.0,
        review_status="needs_review",
        notes=main.AI_RECOVERY_NOTE,
    )


def _learned_rule(main, description, confidence=96):
    return {
        "merchant_key": main.normalize_merchant_key(description),
        "account_category": "Motor Vehicle Expenses",
        "vat_treatment": "standard",
        "review_status": "ready",
        "confidence": confidence,
    }


def test_a_learned_rule_cannot_promote_an_ai_recovered_row() -> None:
    """Classification certainty is not extraction certainty.

    A learned rule knows what a merchant is. It says nothing about whether the
    row was read off the document reliably, and an AI-recovered row was located
    by a model rather than parsed. A 96% rule was lifting such a row to 96 and
    "ready", so a row nobody had checked read as a confident extraction.
    """
    import main

    transaction = _ai_recovered_row(main)
    applied = main.apply_learned_classification_rules([transaction], [_learned_rule(main, transaction.description)])

    assert_equal(applied, 1, "the rule still matches and is applied")
    if transaction.confidence > main.AI_RECOVERED_MAX_CONFIDENCE:
        raise AssertionError(f"AI-recovered row promoted to {transaction.confidence}")
    assert_equal(transaction.confidence, 54.0, "its own confidence is untouched, not raised")
    assert_equal(transaction.review_status, "needs_review", "and it stays flagged for review")


def test_the_learned_rule_still_classifies_the_ai_recovered_row() -> None:
    """Only the certainty is withheld; the classification is still useful."""
    import main

    transaction = _ai_recovered_row(main)
    main.apply_learned_classification_rules([transaction], [_learned_rule(main, transaction.description)])

    assert_equal(transaction.account_category, "Motor Vehicle Expenses", "category comes from the rule")
    assert_equal(transaction.vat_treatment, "standard", "VAT treatment comes from the rule")


def test_the_ai_marker_survives_classification() -> None:
    """A promoted row that lost its marker would be untraceable."""
    import main

    transaction = _ai_recovered_row(main)
    main.apply_learned_classification_rules([transaction], [_learned_rule(main, transaction.description)])

    if main.AI_RECOVERY_NOTE not in transaction.notes:
        raise AssertionError(f"the AI marker was lost: {transaction.notes!r}")
    assert_equal(main.is_ai_recovered(transaction), True, "and the row still reports as AI-recovered")


def test_a_parsed_row_keeps_normal_learned_rule_behaviour() -> None:
    """Nothing changes for rows a parser actually read."""
    import main

    transaction = main.ParsedTransaction(
        transaction_date="2025-05-02",
        description="CARTRACK CART25D5S58NYRV ACCOUNT PAYMENT",
        debit_amount=1710.15,
        running_balance=-1001484.52,
        confidence=70.0,
        review_status="needs_review",
        notes="",
    )
    main.apply_learned_classification_rules([transaction], [_learned_rule(main, transaction.description)])

    assert_equal(transaction.confidence, 96.0, "a parsed row is still promoted by the rule")
    assert_equal(transaction.review_status, "ready", "and still takes the rule's review status")
    assert_equal(main.is_ai_recovered(transaction), False, "it was never AI-recovered")


# ── Deterministic fee and drawings classification ─────────────────────────────
#
# Terminology taken from the banks themselves: a real 37-page Standard Bank
# statement and the FNB fixtures in this suite. Not from a general fee
# vocabulary — a "consulting fee" is not a bank charge.


def test_bank_fee_terminology_classifies_as_bank_charges() -> None:
    """The wording the banks actually print for their own charges."""
    import main

    observed = [
        # Standard Bank, from the real statement
        "301981485 OVERDRAFT SERVICE FEE",
        "ACC 301981485 SERVICE FEE",
        "ACC 301981485 MONTHLY MANAGEMENT FEE",
        "ACC 301981485 0051 NOTIFICATION FEE: MYUPDATES FOR BUSINESS",
        "301981485 3004 HONOURING FEE",
        "301981485 - 0526-0624 - 13DAYS FEE: UNUSED FACILITY",
        "PRELLER TRUST FEE: PAYMENT CONFIRM - EMAIL",
        "301981485 0000H00 FEE IMMEDIATE PAYMENT",
        # FNB, from this suite's fixtures
        "#Service Fees",
        "#Monthly Account Fee",
        "#Excess Item Fee 2 Items On 26/04/01",
        # Plain terms that were falling through to the debit fallback entirely
        "BANK CHARGE",
        "BANK FEE",
        "TRANSACTION FEE",
        "ACCOUNT FEE",
        "CASH DEPOSIT FEE",
        "CARD FEE",
    ]
    for description in observed:
        category, vat, bank_charge, confidence = main.classify_transaction(description, 100.0, None)
        assert_equal(category, "Bank Charges", f"{description!r} is a bank charge")
        assert_equal(vat, "standard", f"{description!r} carries claimable input VAT")
        assert_equal(bank_charge, True, f"{description!r} is flagged as a bank charge")
        if confidence < 90:
            raise AssertionError(f"{description!r} is high-certainty, got {confidence}")


def test_an_explicit_bank_fee_is_never_owner_drawings() -> None:
    """The reported production defect, exactly as it appeared.

    "301981485 10H00 FEE - INSTANT MONEY" matched no fee keyword, fell through
    every rule, and was then caught by the "instant money" payment-channel
    marker, which returned Related Party / Drawings at out-of-scope VAT. A bank
    charge became an owner withdrawal and its input VAT was discarded.
    """
    import main

    category, vat, bank_charge, _ = main.classify_transaction("301981485 10H00 FEE - INSTANT MONEY", 27.50, None)
    assert_equal(category, "Bank Charges", "a bank fee is a bank charge")
    if "Drawings" in category:
        raise AssertionError("a bank fee must never be booked to drawings")
    assert_equal(vat, "standard", "and its input VAT is not discarded as out of scope")
    assert_equal(bank_charge, True, "and it counts as a bank charge")


def test_non_bank_fees_are_not_swept_into_bank_charges() -> None:
    """A fee someone else charges is not a bank charge.

    The point of reading the banks' own terminology rather than matching the
    word "fee": this must not become a keyword net that captures every invoice.
    """
    import main

    for description in ("CONSULTING FEE", "SCHOOL FEES TERM 2", "ARCHITECT FEE INV0093"):
        category, _, bank_charge, _ = main.classify_transaction(description, 5000.0, None)
        if category == "Bank Charges" or bank_charge:
            raise AssertionError(f"{description!r} was wrongly classified as a bank charge")


def test_an_unknown_payee_does_not_default_to_owner_drawings() -> None:
    """Drawings by exhaustion was the defect. Unresolved is the honest answer."""
    import main

    for description in (
        "Payshap Payment To Joe Bloggs",
        "Fnb App Transfer To John",
        "Send Money To Wallet 0821234567",
        "Fnb App Rtc Pmt To Sunfield",
    ):
        category, vat, _, confidence = main.classify_transaction(description, 1500.0, None)
        if "Drawings" in category:
            raise AssertionError(f"{description!r} became drawings without evidence")
        assert_equal(category, "Suspense / Review Required", f"{description!r} is unresolved, not drawings")
        assert_equal(vat, "review", "and its VAT stays open rather than out of scope")
        if confidence >= 70:
            raise AssertionError(f"an unresolved row must not read as confident, got {confidence}")


def test_owner_drawings_requires_positive_evidence() -> None:
    """It remains reachable — but only when the statement says so."""
    import main
    from engine.classification import owner_drawings_evidence

    for description in (
        "Fnb App Payment To Drawings",
        "DIRECTOR LOAN REPAYMENT",
        "OWNER WITHDRAWAL",
        "MEMBERS DRAWINGS",
    ):
        if not owner_drawings_evidence(description):
            raise AssertionError(f"{description!r} is explicit drawings evidence and must be recognised")
        category, _, _, _ = main.classify_transaction(description, 5000.0, None)
        if "Drawings" not in category and "Loan" not in category:
            raise AssertionError(f"{description!r} should reach a drawings/loan account, got {category}")

    for description in ("PAYMENT TO ABC TRADING", "POS PURCHASE WOOLWORTHS MENLYN 004829"):
        if owner_drawings_evidence(description):
            raise AssertionError(f"{description!r} is not drawings evidence")


def test_a_debit_alone_is_not_evidence_of_anything() -> None:
    """Direction carries no accounting meaning on its own."""
    import main

    debit_category, _, _, _ = main.classify_transaction("ZZZ UNRECOGNISED 12345", 900.0, None)
    credit_category, _, _, _ = main.classify_transaction("ZZZ UNRECOGNISED 12345", None, 900.0)
    for category in (debit_category, credit_category):
        if "Drawings" in category:
            raise AssertionError(f"direction alone produced drawings: {category}")
    assert_equal(debit_category, "Suspense / Review Required", "an unknown debit is unresolved")


def test_classification_changes_do_not_touch_the_ledger() -> None:
    """Classification must not be able to alter extracted evidence."""
    import main

    transactions = main.parse_transactions(
        SBSA_SHAPED_PAGES,
        main.parse_metadata("\n".join(p["text"] for p in SBSA_SHAPED_PAGES)),
        "\n".join(p["text"] for p in SBSA_SHAPED_PAGES),
        GENERIC_PROFILE,
    )
    before = [(t.transaction_date, t.description, t.debit_amount, t.credit_amount, t.running_balance) for t in transactions]
    for transaction in transactions:
        main.classify_transaction(transaction.description, transaction.debit_amount, transaction.credit_amount)
    after = [(t.transaction_date, t.description, t.debit_amount, t.credit_amount, t.running_balance) for t in transactions]
    assert_equal(after, before, "classification left every extracted field untouched")


# ── Rule strength ─────────────────────────────────────────────────────────────


def test_every_classification_reports_the_standing_of_its_rule() -> None:
    """A settled classification and a guess must be distinguishable."""
    import main
    from engine.classification import STRENGTH_HARD, STRENGTH_NONE, STRENGTH_SOFT

    cases = [
        ("ACC 301981485 SERVICE FEE", STRENGTH_HARD, "Bank Charges"),
        ("OWNER WITHDRAWAL", STRENGTH_HARD, "Director Loan / Drawings"),
        ("ENGEN GARAGE WELKOM", STRENGTH_SOFT, "Motor Vehicle Expenses"),
        ("ZZZ UNRECOGNISED 12345", STRENGTH_NONE, "Suspense / Review Required"),
        ("Payshap Payment To Joe Bloggs", STRENGTH_NONE, "Suspense / Review Required"),
    ]
    for description, expected_strength, expected_category in cases:
        result = main.classify_transaction_detailed(description, 100.0, None)
        assert_equal(result.strength, expected_strength, f"{description!r} standing")
        assert_equal(result.category, expected_category, f"{description!r} category")
        if not result.reason:
            raise AssertionError(f"{description!r} must say why it was classified that way")


def test_the_compatibility_wrapper_cannot_drift_from_the_detailed_form() -> None:
    """Two views of one decision, never two decisions."""
    import main

    for description in ("ACC 301981485 SERVICE FEE", "ENGEN GARAGE", "ZZZ UNKNOWN", "OWNER DRAWINGS"):
        detailed = main.classify_transaction_detailed(description, 100.0, None)
        legacy = main.classify_transaction(description, 100.0, None)
        assert_equal(
            legacy,
            (detailed.category, detailed.vat_treatment, detailed.bank_charge, detailed.confidence),
            f"{description!r} agrees between both views",
        )


def test_settled_classifications_are_not_sent_to_ai() -> None:
    """Standing decides, not the confidence number.

    A merchant-keyword guess and a fee the bank named itself both scored in the
    eighties, so a confidence threshold could not tell them apart — and sent 98%
    of a real 615-row statement to the model.
    """
    import main
    from engine.classification import STRENGTH_HARD, STRENGTH_LEARNED, STRENGTH_NONE, STRENGTH_SOFT

    def row(strength, review=False, vat_claim="Standard"):
        return {"rule_strength": strength, "review_required": review, "vat_claim_status": vat_claim}

    assert_equal(main.row_needs_ai(row(STRENGTH_HARD)), False, "a bank-named fee is settled")
    assert_equal(main.row_needs_ai(row(STRENGTH_LEARNED)), False, "a workspace correction is settled")
    assert_equal(main.row_needs_ai(row(STRENGTH_SOFT)), True, "a keyword guess is revisable")
    assert_equal(main.row_needs_ai(row(STRENGTH_NONE)), True, "an unresolved row is revisable")
    assert_equal(main.row_needs_ai(row(STRENGTH_HARD, review=True)), False, "settled stays settled")


def test_a_workspace_correction_upgrades_the_standing() -> None:
    """A person's decision outranks a keyword guess."""
    import main
    from engine.classification import STRENGTH_LEARNED, STRENGTH_SOFT

    transaction = main.ParsedTransaction(
        transaction_date="2025-05-02", description="ENGEN GARAGE WELKOM",
        debit_amount=900.0, confidence=84.0, classification_strength=STRENGTH_SOFT,
    )
    rule = {"merchant_key": main.normalize_merchant_key(transaction.description),
            "account_category": "Motor Vehicle Expenses", "vat_treatment": "standard",
            "review_status": "ready", "confidence": 96, "reason": "Learned from accountant correction."}
    main.apply_learned_classification_rules([transaction], [rule])
    assert_equal(transaction.classification_strength, STRENGTH_LEARNED, "standing upgraded")
    assert_equal(main.row_needs_ai({"rule_strength": transaction.classification_strength}), False, "and no longer needs AI")


def test_standing_does_not_reopen_the_ai_recovery_cap() -> None:
    """PR #36 protections are unaffected by the new standing field."""
    import main
    from engine.classification import STRENGTH_LEARNED

    transaction = main.ParsedTransaction(
        transaction_date="2025-05-02", description="CARTRACK CART25D5S58NYRV ACCOUNT PAYMENT",
        debit_amount=1710.15, confidence=54.0, review_status="needs_review", notes=main.AI_RECOVERY_NOTE,
    )
    rule = {"merchant_key": main.normalize_merchant_key(transaction.description),
            "account_category": "Motor Vehicle Expenses", "vat_treatment": "standard",
            "review_status": "ready", "confidence": 96}
    main.apply_learned_classification_rules([transaction], [rule])

    # The CLASSIFICATION is now a person's decision; the EXTRACTION is still a
    # model's guess, and those are different certainties.
    assert_equal(transaction.classification_strength, STRENGTH_LEARNED, "classification standing upgraded")
    assert_equal(transaction.confidence, 54.0, "extraction confidence still capped")
    assert_equal(transaction.review_status, "needs_review", "and the row stays flagged")


def test_provenance_is_written_but_never_at_the_ledgers_cost() -> None:
    """Provenance is enrichment. A ledger is not.

    transaction_insert_row is built by spreading model_dump(), so a new field
    reaches Supabase automatically — and an insert naming a column that does not
    exist fails the WHOLE batch, losing every transaction in the run. Migration
    021 adds these columns; a database that has not run it yet must still get
    its ledger.
    """
    import main

    transaction = main.ParsedTransaction(
        transaction_date="2025-05-02", description="ACC 1 SERVICE FEE", debit_amount=10.0,
    )
    row = main.transaction_insert_row(transaction, "run-1", "ws-1")
    for column in main.PROVENANCE_COLUMNS:
        if column not in row:
            raise AssertionError(f"{column} has a column as of migration 021 and should be written")

    stripped = main.strip_provenance_columns([row])[0]
    for column in main.PROVENANCE_COLUMNS:
        if column in stripped:
            raise AssertionError(f"{column} must be droppable for a database without migration 021")
    for essential in ("transaction_date", "description", "debit_amount", "running_balance", "run_id", "workspace_id"):
        if essential not in stripped:
            raise AssertionError(f"the fallback dropped {essential}, which is the ledger itself")


# ── Classification provenance ─────────────────────────────────────────────────


def test_unresolved_is_recorded_as_unresolved_not_as_a_decision() -> None:
    """"We do not know" must not be recorded as a deterministic classification.

    On the real 615-row statement 434 rows are unresolved. Recording those as
    `deterministic` would present a fallback as an answer, which is how the
    review UI came to show them as confidently classified.
    """
    from engine.classification import (
        SOURCE_DETERMINISTIC,
        SOURCE_LEARNED_RULE,
        SOURCE_UNRESOLVED,
        STRENGTH_HARD,
        STRENGTH_LEARNED,
        STRENGTH_NONE,
        STRENGTH_SOFT,
        source_for_strength,
    )

    assert_equal(source_for_strength(STRENGTH_HARD), SOURCE_DETERMINISTIC, "a hard rule is deterministic")
    assert_equal(source_for_strength(STRENGTH_SOFT), SOURCE_DETERMINISTIC, "a soft rule is still deterministic")
    assert_equal(source_for_strength(STRENGTH_LEARNED), SOURCE_LEARNED_RULE, "a workspace rule is a learned rule")
    assert_equal(source_for_strength(STRENGTH_NONE), SOURCE_UNRESOLVED, "no rule fired means unresolved")
    assert_equal(source_for_strength("something-new"), SOURCE_UNRESOLVED, "an unknown standing is not a decision")


def test_a_transaction_records_who_classified_it() -> None:
    """Provenance is recorded, never inferred from the confidence number."""
    import main
    from engine.classification import SOURCE_DETERMINISTIC, SOURCE_UNRESOLVED

    fee = main.build_transaction("01 May 2025", "ACC 301981485 SERVICE FEE", 574.30, None, -1000.0, {}, 1, "raw", 90)
    assert_equal(fee.classification_source, SOURCE_DETERMINISTIC, "a bank fee is a deterministic classification")
    assert_equal(fee.classification_confidence, 97, "and carries its own classification confidence")
    if not fee.classification_reason:
        raise AssertionError("the evidence must be recorded for the review screen")

    unknown = main.build_transaction("01 May 2025", "ZZZ UNRECOGNISED 12345", 100.0, None, -1100.0, {}, 1, "raw", 90)
    assert_equal(unknown.classification_source, SOURCE_UNRESOLVED, "an unclassified row says so")


def test_a_learned_rule_records_itself_as_the_source() -> None:
    import main
    from engine.classification import SOURCE_LEARNED_RULE

    transaction = main.ParsedTransaction(
        transaction_date="2025-05-02", description="ENGEN GARAGE WELKOM", debit_amount=900.0, confidence=84.0,
    )
    main.apply_learned_classification_rules([transaction], [{
        "merchant_key": main.normalize_merchant_key(transaction.description),
        "account_category": "Motor Vehicle Expenses", "vat_treatment": "standard",
        "review_status": "ready", "confidence": 96, "reason": "Learned from accountant correction.",
    }])
    assert_equal(transaction.classification_source, SOURCE_LEARNED_RULE, "the workspace rule is the source")
    assert_equal(transaction.classification_confidence, 96.0, "with the rule's own confidence")


def test_classification_confidence_is_not_extraction_confidence() -> None:
    """The two numbers answer different questions and must not merge.

    An AI-recovered row is capped at 60 as an EXTRACTION signal (PR #36). Its
    CATEGORY can still be certain — a workspace correction is a person's
    decision — and recording that must not raise the extraction cap.
    """
    import main

    transaction = main.ParsedTransaction(
        transaction_date="2025-05-02", description="CARTRACK CART25D5S58NYRV ACCOUNT PAYMENT",
        debit_amount=1710.15, confidence=54.0, review_status="needs_review", notes=main.AI_RECOVERY_NOTE,
    )
    main.apply_learned_classification_rules([transaction], [{
        "merchant_key": main.normalize_merchant_key(transaction.description),
        "account_category": "Motor Vehicle Expenses", "vat_treatment": "standard",
        "review_status": "ready", "confidence": 96,
    }])
    assert_equal(transaction.classification_confidence, 96.0, "the category is certain")
    assert_equal(transaction.confidence, 54.0, "the extraction cap is untouched")
    assert_equal(transaction.review_status, "needs_review", "and the row stays flagged")


def test_the_migration_adds_every_column_the_worker_writes() -> None:
    """A column the worker writes but the migration omits fails the whole insert."""
    import main

    migration = (ROOT / "supabase" / "migrations" / "021_classification_provenance.sql").read_text()
    for column in main.PROVENANCE_COLUMNS:
        if f"add column if not exists {column}" not in migration:
            raise AssertionError(f"migration 021 must add {column}")


def run() -> None:
    test_unresolved_is_recorded_as_unresolved_not_as_a_decision()
    test_a_transaction_records_who_classified_it()
    test_a_learned_rule_records_itself_as_the_source()
    test_classification_confidence_is_not_extraction_confidence()
    test_the_migration_adds_every_column_the_worker_writes()
    test_every_classification_reports_the_standing_of_its_rule()
    test_the_compatibility_wrapper_cannot_drift_from_the_detailed_form()
    test_settled_classifications_are_not_sent_to_ai()
    test_a_workspace_correction_upgrades_the_standing()
    test_standing_does_not_reopen_the_ai_recovery_cap()
    test_provenance_is_written_but_never_at_the_ledgers_cost()
    test_bank_fee_terminology_classifies_as_bank_charges()
    test_an_explicit_bank_fee_is_never_owner_drawings()
    test_non_bank_fees_are_not_swept_into_bank_charges()
    test_an_unknown_payee_does_not_default_to_owner_drawings()
    test_owner_drawings_requires_positive_evidence()
    test_a_debit_alone_is_not_evidence_of_anything()
    test_classification_changes_do_not_touch_the_ledger()
    test_a_learned_rule_cannot_promote_an_ai_recovered_row()
    test_the_learned_rule_still_classifies_the_ai_recovered_row()
    test_the_ai_marker_survives_classification()
    test_a_parsed_row_keeps_normal_learned_rule_behaviour()
    test_the_repeated_letterhead_is_not_read_as_a_transaction()
    test_the_footer_summary_is_not_read_as_transactions()
    test_the_section_bounded_ledger_reconciles_exactly()
    test_a_statement_with_no_column_header_is_still_read()
    test_ai_rows_must_come_from_a_line_we_sent()
    test_ai_amounts_must_appear_in_their_source_line()
    test_ai_balances_that_are_not_printed_are_dropped_but_the_row_is_kept()
    test_ai_descriptions_may_not_introduce_words()
    test_ai_never_receives_or_returns_a_guessed_direction()
    test_ai_candidate_lines_only_carry_figures()
    test_ai_line_cap_is_reported_never_silent()
    test_ai_recovery_is_skipped_without_a_key_and_never_invents_a_ledger()
    test_ai_recovered_rows_are_all_flagged_for_review()
    test_an_ai_recovered_run_can_never_report_completed()
    test_a_statement_that_ties_out_completes()
    test_a_broken_balance_chain_is_partial_not_complete()
    test_continuity_does_not_second_guess_a_statement_whose_money_ties_out()
    test_no_evidence_at_all_goes_to_review_never_to_completed()
    test_recovery_options_separates_unreadable_from_unparsed()
    test_failure_messages_name_the_parser_that_ran()
    test_generic_parser_reads_a_text_only_statement_end_to_end()
    test_generic_parser_merges_wrapped_descriptions()
    test_generic_parser_inherits_the_date_of_a_grouped_movement()
    test_generic_parser_infers_direction_from_balance_continuity()
    test_generic_parser_drops_page_furniture_and_totals()
    test_generic_parser_reads_a_debit_credit_layout()
    test_negative_opening_balance_keeps_its_sign()
    test_candidate_counting_is_bank_independent()
    test_generic_rows_and_provider_rows_take_the_same_path()
    test_standard_bank_is_not_routed_to_the_fnb_parser()
    test_standard_bank_layout_parses_payments_and_deposits()
    test_generic_parser_never_invents_fnb_fee_rows()
    test_structured_rows_skip_fnb_reconstruction_for_other_banks()
    test_unknown_bank_routes_to_generic_not_fnb()
    test_fnb_still_routes_to_the_fnb_parser()
    test_bank_resolution_precedence_between_the_two_detections()
    test_parser_profile_not_implemented_rejection_is_gone()
    test_bank_detection_identifies_standard_bank_from_text()
    test_bank_detection_covers_every_supported_bank()
    test_bank_detection_keeps_fnb_unchanged()
    test_bank_detection_never_defaults_to_a_bank()
    test_bank_detection_ignores_a_counterparty_named_in_a_transaction()
    test_bank_detection_reads_no_file_path()
    test_bank_detection_is_ambiguous_rather_than_wrong()
    test_bank_detection_survives_broken_letterhead_whitespace()
    test_worker_auth_fails_closed()
    test_worker_auth_matches_pdf_plumber_contract()
    test_worker_token_check_raises_on_unconfigured_secret()
    test_mutating_endpoints_are_all_authenticated()
    test_non_ascii_credential_is_invalid_not_a_500()
    test_auth_diagnostics_never_leak_the_secret()
    test_auth_diagnostics_hash_exactly_what_is_compared()
    test_descriptionless_fee_rows_are_preserved()
    test_unnamed_fee_rows_are_labelled_from_statement_figures_only()
    test_informational_rows_are_kept_but_not_counted()
    test_structured_rows_are_selected_when_equal_or_better()
    test_structured_rows_fallback_to_text_when_weaker_or_unusable()
    test_structured_rows_preserve_empty_date_fee_info_and_overdrawn_balance()
    test_no_structured_rows_keeps_text_path_unchanged()
    test_structured_unsigned_amount_debit_proven_by_balance_continuity()
    test_structured_unsigned_amount_credit_proven_by_balance_continuity()
    test_structured_unsigned_amount_unresolved_falls_back_to_text()
    test_structured_unsigned_amount_unresolved_never_persisted_as_debit()
    test_extraction_confidence_scores_financially_complete_statement_near_100()
    test_extraction_confidence_large_fully_reconciled_statement_near_100()
    test_extraction_confidence_is_independent_from_classification_confidence()
    test_extraction_confidence_drops_when_one_transaction_is_missing()
    test_extraction_confidence_drops_on_debit_total_mismatch()
    test_extraction_confidence_drops_on_running_balance_gap()
    test_extraction_confidence_normalises_when_optional_summary_fields_absent()
    test_extraction_confidence_ignores_informational_rows()
    test_extraction_confidence_equivalent_for_structured_and_text_financially_identical_results()
    run_fnb_extraction_case()
    run_statement_period_case()
    run_missing_column_fallback_case()
    run_validation_diagnostics_case()
    run_april_missing_rows_case()
    run_freight_aces_case()
    run_december_multi_page_closing_balance_case()
    run_compound_ocr_line_case()
    run_professional_classification_case()
    run_learned_supplier_rules_case()
    run_combined_workbook_case()
    run_local_real_statement_files_case()
    test_real_statement_cases_actually_execute()

    manifest = json.loads(MANIFEST_PATH.read_text())
    cases = manifest.get("cases") if isinstance(manifest, dict) else None
    if not isinstance(cases, list) or not cases:
        raise AssertionError("Regression manifest has no cases. Add at least one statement fixture.")

    for case in cases:
        case_id = str(case.get("id") or "unnamed")
        source = str(case.get("source") or "")
        fixture_rel_path = case.get("fixture")
        if source != "synthetic":
            raise AssertionError(f"{case_id}: unsupported source {source!r}")
        if not isinstance(fixture_rel_path, str) or not fixture_rel_path:
            raise AssertionError(f"{case_id}: fixture path is required")

        fixture_path = ROOT / fixture_rel_path
        if not fixture_path.exists():
            raise AssertionError(f"{case_id}: fixture file not found: {fixture_path}")

        run_synthetic_case(case_id, fixture_path)


def test_descriptionless_fee_rows_are_preserved() -> None:
    """FNB prints fee rows with no narrative: "DD Mon <amount> <balance>".

    LOOSE_DATE's optional year group used to swallow the amount's integer part
    ("26 Apr 550.00" parsed its date as "26 Apr 550"), which build_transaction
    could not parse, so the row was dropped. Two such rows vanished per
    statement and the debit count came up one short of the bank's own summary.
    """
    import main

    for line, expected_date in (
        ("26 Apr 550.00 148,157.78Cr", "26 Apr"),
        ("26 Apr 15.00 148,142.78Cr", "26 Apr"),
        ("26 May 30.00 106,128.78Cr", "26 May"),
        ("02 Jun 1,234.56 10,000.00Cr", "02 Jun"),
    ):
        assert main.LOOSE_DATE.match(line).group("date") == expected_date, line

    # A genuine year must still be recognised.
    for line, expected_date in (
        ("01 Apr 2025 Something 10.00 5.00Cr", "01 Apr 2025"),
        ("01/04/2025 Foo 10.00 5.00Cr", "01/04/2025"),
        ("01 Apr Magtape 330.00 123,839.78Cr", "01 Apr"),
    ):
        assert main.LOOSE_DATE.match(line).group("date") == expected_date, line

    # The row must now parse into a real debit, not be discarded.
    parsed = main.parse_fnb_transaction_line("26 Apr 550.00 148,157.78Cr", {})
    assert parsed is not None, "descriptionless fee row must not be dropped"
    assert parsed.debit_amount == 550.0
    assert parsed.running_balance == 148157.78
    assert parsed.description == main.UNNAMED_FEE_DESCRIPTION


def test_unnamed_fee_rows_are_labelled_from_statement_figures_only() -> None:
    """Naming must come from the statement's own figures, never a guess."""
    import main

    charged = main.build_transaction(
        "26 Apr", "FNB App Rtc Pmt To Someone", 2500.0, None, 148707.78, {}, None, "raw", 96
    )
    assert charged is not None
    charged.notes = "Accrued bank charges: 15.00"
    fee = main.build_transaction("26 Apr", "", 15.0, None, 148692.78, {}, None, "raw", 96)
    unknown = main.build_transaction("26 Apr", "", 30.0, None, 148662.78, {}, None, "raw", 96)
    assert fee is not None and unknown is not None

    rows = [charged, fee, unknown]
    main.label_unnamed_fee_rows(rows, {})

    # Matches a preceding accrued charge -> named.
    assert fee.description == "Transaction Fee"
    assert fee.bank_charge is True
    # Nothing proves what the 30.00 is, so it keeps the neutral placeholder
    # rather than being guessed at.
    assert unknown.description == main.UNNAMED_FEE_DESCRIPTION
    # Amounts are never altered by labelling.
    assert fee.debit_amount == 15.0 and unknown.debit_amount == 30.0


def test_structured_rows_are_selected_when_equal_or_better() -> None:
    import main

    metadata = {"statement_period_end": "2026-03-31"}
    full_text = "\n".join(
        [
            "Transactions in Rand (ZAR)",
            "01 Mar EFT Deposit Client 1,000.00Cr 1,000.00 Cr",
            "01 Mar Card Purchase Fuel 300.00 700.00 Cr",
        ]
    )
    structured_rows = [
        {
            "pageNumber": 1,
            "confidence": 0.95,
            "raw": "01 Mar EFT Deposit Client 1,000.00Cr 1,000.00 Cr",
            "cells": {
                "date": "01 Mar",
                "description": "EFT Deposit Client",
                "credit": "1,000.00",
                "balance": "1,000.00 Cr",
            },
        },
        {
            "pageNumber": 1,
            "confidence": 0.94,
            "raw": "01 Mar Card Purchase Fuel 300.00 700.00 Cr",
            "cells": {
                "date": "01 Mar",
                "description": "Card Purchase Fuel",
                "debit": "300.00",
                "balance": "700.00 Cr",
            },
        },
    ]

    selected, diagnostics = main.select_transactions_from_sources([], metadata, full_text, structured_rows, FNB_PROFILE)
    assert diagnostics["selected_path"] == "structured", diagnostics
    summary = main.validation_summary(selected)
    assert_equal(summary["transaction_count"], 2, "structured-selected financial count")
    assert_equal(str(summary["total_credits"]), "1000.00", "structured-selected credit total")
    assert_equal(str(summary["total_debits"]), "300.00", "structured-selected debit total")

    # Better-than-text case: no extractable text transactions, but structured rows
    # still produce a valid ledger.
    selected_better, diagnostics_better = main.select_transactions_from_sources(
        [],
        metadata,
        "Statement header only",
        structured_rows,
        FNB_PROFILE,
    )
    assert diagnostics_better["selected_path"] == "structured", diagnostics_better
    assert_equal(main.financial_transaction_count(selected_better), 2, "structured better-than-text selection")


def test_structured_rows_fallback_to_text_when_weaker_or_unusable() -> None:
    import main

    text, metadata = _build_acapolite_style_statement()
    weak_rows = [
        {
            "pageNumber": 1,
            "raw": "01 Mar Card Purchase Example 100.00 3,290.09 Cr",
            "cells": {"date": "01 Mar", "description": "Card Purchase Example", "debit": "100.00", "balance": "3,290.09 Cr"},
        }
    ]
    selected_weak, weak_diag = main.select_transactions_from_sources([], metadata, text, weak_rows, FNB_PROFILE)
    assert weak_diag["selected_path"] == "text", weak_diag
    assert str(weak_diag.get("fallback_reason") or "").startswith("structured_weaker_than_text"), weak_diag
    assert_equal(len(selected_weak), FNB_EXPECTED["transaction_count"], "weak structured falls back to full text parse")

    unusable_rows = [{"pageNumber": 1, "cells": {"date": "", "description": "", "amount": ""}}]
    selected_unusable, unusable_diag = main.select_transactions_from_sources([], metadata, text, unusable_rows, FNB_PROFILE)
    assert unusable_diag["selected_path"] == "text", unusable_diag
    assert str(unusable_diag.get("fallback_reason") or "").startswith("structured_unusable"), unusable_diag
    assert_equal(len(selected_unusable), FNB_EXPECTED["transaction_count"], "unusable structured falls back to text")


def test_structured_rows_preserve_empty_date_fee_info_and_overdrawn_balance() -> None:
    import main

    metadata = {"statement_period_end": "2026-03-31"}
    rows = [
        {
            "pageNumber": 1,
            "confidence": 0.91,
            "raw": "01 Mar Card Purchase Supplier 75,496.08 72,814.46 Dr",
            "cells": {
                "date": "01 Mar",
                "description": "Card Purchase Supplier",
                "reference": "ABCD123",
                "debit": "75,496.08",
                "balance": "72,814.46 Dr",
            },
        },
        {
            "pageNumber": 1,
            "confidence": 0.90,
            "raw": "Service Fees 523.80 73,338.26 Dr",
            "cells": {
                "date": "",
                "description": "Service Fees",
                "debit": "523.80",
                "balance": "73,338.26 Dr",
            },
        },
        {
            "pageNumber": 1,
            "confidence": 0.88,
            "raw": "Express Pmt Pending",
            "cells": {
                "date": "",
                "description": "Express Pmt Pending",
                "amount": "0.00",
                "balance": "73,338.26 Dr",
            },
        },
        {
            "pageNumber": 1,
            "confidence": 0.86,
            "raw": "26 Apr 550.00 148,157.78Cr",
            "cells": {
                "date": "",
                "description": "",
                "debit": "550.00",
                "balance": "148,157.78Cr",
            },
        },
    ]

    txns, parse_diag = main.parse_structured_rows(rows, metadata, FNB_PROFILE)
    assert_equal(len(txns), 4, "structured rows parsed")
    assert parse_diag["date_inferred_rows"] >= 2, parse_diag

    first = txns[0]
    assert_equal(first.running_balance, -72814.46, "Dr balance parsed as negative")
    assert "reference: ABCD123" in (first.notes or ""), first.notes

    second = txns[1]
    assert second.transaction_date == first.transaction_date, "empty-date row must inherit group date"

    fee = next((txn for txn in txns if txn.debit_amount == 550.0), None)
    assert fee is not None, "descriptionless fee row must be preserved"
    assert fee.description in {"Transaction Fee", main.UNNAMED_FEE_DESCRIPTION}, fee.description

    financial, informational = main.split_ledger_rows(txns)
    assert_equal(len(financial), 3, "structured financial row count")
    assert_equal(len(informational), 1, "structured informational row count")


def test_no_structured_rows_keeps_text_path_unchanged() -> None:
    import main

    text, metadata = _build_acapolite_style_statement()
    expected = main.parse_fnb_transactions([], metadata, text)
    selected, diagnostics = main.select_transactions_from_sources([], metadata, text, None, FNB_PROFILE)

    assert diagnostics["selected_path"] == "text", diagnostics
    assert diagnostics["fallback_reason"] == "structured_rows_absent", diagnostics
    assert_equal(len(selected), len(expected), "no structured rows should keep text output count")
    expected_summary = main.validation_summary(expected)
    selected_summary = main.validation_summary(selected)
    assert_equal(expected_summary["transaction_count"], selected_summary["transaction_count"], "text transaction count unchanged")
    assert_equal(str(expected_summary["total_debits"]), str(selected_summary["total_debits"]), "text debit total unchanged")
    assert_equal(str(expected_summary["total_credits"]), str(selected_summary["total_credits"]), "text credit total unchanged")


def test_structured_unsigned_amount_debit_proven_by_balance_continuity() -> None:
    import main

    metadata = {"statement_period_end": "2026-03-31"}
    rows = [
        {
            "pageNumber": 1,
            "cells": {
                "date": "01 Mar",
                "description": "Seed credit",
                "credit": "1,000.00",
                "balance": "1,000.00 Cr",
            },
        },
        {
            "pageNumber": 1,
            "cells": {
                "date": "01 Mar",
                "description": "Unsigned amount row",
                "amount": "200.00",
                "balance": "800.00 Cr",
            },
        },
    ]
    txns, diag = main.parse_structured_rows(rows, metadata, FNB_PROFILE)
    assert_equal(diag["rejected_reasons"].get("ambiguous_unsigned_amount_direction", 0), 0, "no ambiguity rejection")
    assert_equal(len(txns), 2, "both rows parsed")
    assert_equal(txns[1].debit_amount, 200.0, "unsigned amount resolved to debit")
    assert_equal(txns[1].credit_amount, None, "unsigned amount not credit")


def test_structured_unsigned_amount_credit_proven_by_balance_continuity() -> None:
    import main

    metadata = {"statement_period_end": "2026-03-31"}
    rows = [
        {
            "pageNumber": 1,
            "cells": {
                "date": "01 Mar",
                "description": "Seed credit",
                "credit": "1,000.00",
                "balance": "1,000.00 Cr",
            },
        },
        {
            "pageNumber": 1,
            "cells": {
                "date": "01 Mar",
                "description": "Unsigned amount row",
                "amount": "200.00",
                "balance": "1,200.00 Cr",
            },
        },
    ]
    txns, diag = main.parse_structured_rows(rows, metadata, FNB_PROFILE)
    assert_equal(diag["rejected_reasons"].get("ambiguous_unsigned_amount_direction", 0), 0, "no ambiguity rejection")
    assert_equal(len(txns), 2, "both rows parsed")
    assert_equal(txns[1].credit_amount, 200.0, "unsigned amount resolved to credit")
    assert_equal(txns[1].debit_amount, None, "unsigned amount not debit")


def test_structured_unsigned_amount_unresolved_falls_back_to_text() -> None:
    import main

    metadata = {"statement_period_end": "2026-03-31"}
    text = "\n".join(
        [
            "Transactions in Rand (ZAR)",
            "01 Mar EFT Deposit Seed 1,000.00Cr 1,000.00 Cr",
            "01 Mar Card Purchase Proven Debit 200.00 800.00 Cr",
        ]
    )
    rows = [
        {
            "pageNumber": 1,
            "cells": {
                "date": "01 Mar",
                "description": "Unsigned amount row",
                "amount": "200.00",
            },
        }
    ]
    selected, diagnostics = main.select_transactions_from_sources([], metadata, text, rows, FNB_PROFILE)
    assert diagnostics["selected_path"] == "text", diagnostics
    assert str(diagnostics.get("fallback_reason") or "").startswith("structured_unusable"), diagnostics
    assert "ambiguous_unsigned_amount_direction" in str(diagnostics.get("structured_parse_diagnostics", {}).get("rejected_reasons", {}))
    assert_equal(main.financial_transaction_count(selected), 2, "text fallback keeps financial rows")


def test_structured_unsigned_amount_unresolved_never_persisted_as_debit() -> None:
    import main

    metadata = {"statement_period_end": "2026-03-31"}
    rows = [
        {
            "pageNumber": 1,
            "cells": {
                "date": "01 Mar",
                "description": "Unsigned amount row",
                "amount": "200.00",
            },
        }
    ]
    txns, diag = main.parse_structured_rows(rows, metadata, FNB_PROFILE)
    assert_equal(len(txns), 0, "ambiguous unsigned row is rejected")
    assert_equal(diag["rejected_reasons"].get("ambiguous_unsigned_amount_direction"), 1, "ambiguity reason counted")
    debits = [txn for txn in txns if txn.debit_amount == 200.0]
    assert_equal(len(debits), 0, "ambiguous unsigned row was not stored as debit")


def test_extraction_confidence_scores_financially_complete_statement_near_100() -> None:
    import main

    metadata = {
        "opening_balance": 1000.00,
        "closing_balance": 900.00,
        "expected_transaction_count": 2,
        "expected_credit_count": 1,
        "expected_debit_count": 1,
        "declared_credit_total": 250.00,
        "declared_debit_total": 350.00,
    }
    credit = main.build_transaction("01 Mar", "Deposit", None, 250.0, 1250.0, metadata, 1, "01 Mar Deposit 250.00Cr 1,250.00 Cr", 92)
    debit = main.build_transaction("01 Mar", "Card Purchase", 350.0, None, 900.0, metadata, 1, "01 Mar Card Purchase 350.00 900.00 Cr", 92)
    assert credit is not None and debit is not None
    txns = [credit, debit]
    extraction_check = main.validate_extraction(metadata, txns)
    missing_rows = main.missing_transaction_count_for_storage(extraction_check, len(txns))
    score = main.extraction_confidence_score(
        metadata,
        extraction_check,
        txns,
        [{"page": 1, "text": "seed", "tables": []}],
        missing_rows,
        unresolved_amount_directions=0,
    )
    assert score is not None
    if score < 95:
        raise AssertionError(f"clean reconciled statement should score near 100, got {score}")


def test_extraction_confidence_is_independent_from_classification_confidence() -> None:
    import main

    metadata = {
        "opening_balance": 1000.00,
        "closing_balance": 900.00,
        "expected_transaction_count": 2,
        "expected_credit_count": 1,
        "expected_debit_count": 1,
        "declared_credit_total": 250.00,
        "declared_debit_total": 350.00,
    }
    credit = main.build_transaction("01 Mar", "Deposit", None, 250.0, 1250.0, metadata, 1, "01 Mar Deposit 250.00Cr 1,250.00 Cr", 92)
    debit = main.build_transaction("01 Mar", "Card Purchase", 350.0, None, 900.0, metadata, 1, "01 Mar Card Purchase 350.00 900.00 Cr", 92)
    assert credit is not None and debit is not None
    txns = [credit, debit]
    extraction_check = main.validate_extraction(metadata, txns)
    missing_rows = main.missing_transaction_count_for_storage(extraction_check, len(txns))
    baseline = main.extraction_confidence_score(
        metadata,
        extraction_check,
        txns,
        [{"page": 1, "text": "seed", "tables": []}],
        missing_rows,
        unresolved_amount_directions=0,
    )
    for txn in txns:
        txn.confidence = 12
        txn.review_status = "needs_review"
    after_low_classification = main.extraction_confidence_score(
        metadata,
        extraction_check,
        txns,
        [{"page": 1, "text": "seed", "tables": []}],
        missing_rows,
        unresolved_amount_directions=0,
    )
    assert_equal(after_low_classification, baseline, "extraction confidence must not depend on classification confidence")


def _confidence_score(main, metadata, txns, pages=None, unresolved=0):
    extraction_check = main.validate_extraction(metadata, txns)
    missing_rows = main.missing_transaction_count_for_storage(extraction_check, len(txns))
    return main.extraction_confidence_score(
        metadata,
        extraction_check,
        txns,
        pages or [{"page": 1, "text": "seed", "tables": []}],
        missing_rows,
        unresolved_amount_directions=unresolved,
    )


def test_extraction_confidence_large_fully_reconciled_statement_near_100() -> None:
    import main

    text, metadata = _build_acapolite_style_statement()
    txns = main.parse_fnb_transactions([], metadata, text)
    score = _confidence_score(main, metadata, txns, pages=[{"page": 1, "text": text, "tables": []}])
    assert score is not None
    if score < 95:
        raise AssertionError(f"large fully reconciled statement should score near 100, got {score}")


def test_extraction_confidence_drops_when_one_transaction_is_missing() -> None:
    import main

    text, metadata = _build_acapolite_style_statement()
    txns = main.parse_fnb_transactions([], metadata, text)
    baseline = _confidence_score(main, metadata, txns, pages=[{"page": 1, "text": text, "tables": []}])
    reduced = txns[:-1]
    lower = _confidence_score(main, metadata, reduced, pages=[{"page": 1, "text": text, "tables": []}])
    assert baseline is not None and lower is not None
    if not lower < baseline:
        raise AssertionError(f"score should drop when one row is missing: baseline={baseline}, missing={lower}")


def test_extraction_confidence_drops_on_debit_total_mismatch() -> None:
    import main

    metadata = {
        "opening_balance": 1000.00,
        "closing_balance": 900.00,
        "expected_transaction_count": 2,
        "expected_credit_count": 1,
        "expected_debit_count": 1,
        "declared_credit_total": 250.00,
        "declared_debit_total": 350.00,
    }
    credit = main.build_transaction("01 Mar", "Deposit", None, 250.0, 1250.0, metadata, 1, "01 Mar Deposit 250.00Cr 1,250.00 Cr", 92)
    debit = main.build_transaction("01 Mar", "Card Purchase", 350.0, None, 900.0, metadata, 1, "01 Mar Card Purchase 350.00 900.00 Cr", 92)
    assert credit is not None and debit is not None
    txns = [credit, debit]
    baseline = _confidence_score(main, metadata, txns)
    mismatch_meta = dict(metadata)
    mismatch_meta["declared_debit_total"] = 999.00
    lowered = _confidence_score(main, mismatch_meta, txns)
    assert baseline is not None and lowered is not None
    if not lowered < baseline:
        raise AssertionError(f"score should drop on debit-total mismatch: baseline={baseline}, mismatch={lowered}")


def test_extraction_confidence_drops_on_running_balance_gap() -> None:
    import main

    metadata = {
        "opening_balance": 1000.00,
        "closing_balance": 900.00,
        "expected_transaction_count": 2,
        "expected_credit_count": 1,
        "expected_debit_count": 1,
        "declared_credit_total": 250.00,
        "declared_debit_total": 350.00,
    }
    credit = main.build_transaction("01 Mar", "Deposit", None, 250.0, 1250.0, metadata, 1, "01 Mar Deposit 250.00Cr 1,250.00 Cr", 92)
    debit_ok = main.build_transaction("01 Mar", "Card Purchase", 350.0, None, 900.0, metadata, 1, "01 Mar Card Purchase 350.00 900.00 Cr", 92)
    debit_gap = main.build_transaction("01 Mar", "Card Purchase", 350.0, None, 910.0, metadata, 1, "01 Mar Card Purchase 350.00 910.00 Cr", 92)
    assert credit is not None and debit_ok is not None and debit_gap is not None
    baseline = _confidence_score(main, metadata, [credit, debit_ok])
    gapped = _confidence_score(main, metadata, [credit, debit_gap])
    assert baseline is not None and gapped is not None
    if not gapped < baseline:
        raise AssertionError(f"score should drop on running-balance gaps: baseline={baseline}, gap={gapped}")


def test_extraction_confidence_normalises_when_optional_summary_fields_absent() -> None:
    import main

    metadata = {
        "opening_balance": 1000.00,
        "closing_balance": 900.00,
        # Intentionally no expected/declared count/totals fields.
    }
    credit = main.build_transaction("01 Mar", "Deposit", None, 250.0, 1250.0, metadata, 1, "01 Mar Deposit 250.00Cr 1,250.00 Cr", 92)
    debit = main.build_transaction("01 Mar", "Card Purchase", 350.0, None, 900.0, metadata, 1, "01 Mar Card Purchase 350.00 900.00 Cr", 92)
    assert credit is not None and debit is not None
    score = _confidence_score(main, metadata, [credit, debit])
    assert score is not None
    if score < 90:
        raise AssertionError(f"missing optional summary fields should not force a low score, got {score}")


def test_extraction_confidence_ignores_informational_rows() -> None:
    import main

    metadata = {
        "opening_balance": 1000.00,
        "closing_balance": 900.00,
        "expected_transaction_count": 2,
        "expected_credit_count": 1,
        "expected_debit_count": 1,
        "declared_credit_total": 250.00,
        "declared_debit_total": 350.00,
    }
    credit = main.build_transaction("01 Mar", "Deposit", None, 250.0, 1250.0, metadata, 1, "01 Mar Deposit 250.00Cr 1,250.00 Cr", 92)
    debit = main.build_transaction("01 Mar", "Card Purchase", 350.0, None, 900.0, metadata, 1, "01 Mar Card Purchase 350.00 900.00 Cr", 92)
    info = main.build_transaction("01 Mar", "Express Pmt Pending", 0.0, None, 900.0, metadata, 1, "01 Mar Express Pmt Pending 0.00 900.00 Cr", 92)
    assert credit is not None and debit is not None and info is not None
    without_info = _confidence_score(main, metadata, [credit, debit])
    with_info = _confidence_score(main, metadata, [credit, debit, info])
    assert_equal(with_info, without_info, "informational rows must not lower extraction confidence")


def test_extraction_confidence_equivalent_for_structured_and_text_financially_identical_results() -> None:
    import main

    metadata = {"statement_period_end": "2026-03-31"}
    text = "\n".join(
        [
            "Transactions in Rand (ZAR)",
            "01 Mar EFT Deposit Client 1,000.00Cr 1,000.00 Cr",
            "01 Mar Card Purchase Fuel 300.00 700.00 Cr",
        ]
    )
    text_txns = main.parse_fnb_transactions([], metadata, text)
    for txn in text_txns:
        if txn.source_page is None:
            txn.source_page = 1
    structured_rows = [
        {
            "pageNumber": 1,
            "cells": {
                "date": "01 Mar",
                "description": "EFT Deposit Client",
                "credit": "1,000.00",
                "balance": "1,000.00 Cr",
            },
        },
        {
            "pageNumber": 1,
            "cells": {
                "date": "01 Mar",
                "description": "Card Purchase Fuel",
                "debit": "300.00",
                "balance": "700.00 Cr",
            },
        },
    ]
    structured_txns, diag = main.parse_structured_rows(structured_rows, metadata, FNB_PROFILE)
    assert_equal(diag["rejected_reasons"].get("ambiguous_unsigned_amount_direction", 0), 0, "structured rows parsed cleanly")

    # Use extraction checks generated from each source's own rows; if the financial
    # facts are identical, extraction confidence must be equivalent.
    text_meta = main.parse_metadata(text)
    structured_meta = main.parse_metadata(text)
    text_check = main.validate_extraction(text_meta, text_txns)
    structured_check = main.validate_extraction(structured_meta, structured_txns)
    text_missing = main.missing_transaction_count_for_storage(text_check, len(text_txns))
    structured_missing = main.missing_transaction_count_for_storage(structured_check, len(structured_txns))
    text_score = main.extraction_confidence_score(text_meta, text_check, text_txns, [{"page": 1, "text": text, "tables": []}], text_missing, unresolved_amount_directions=0)
    structured_score = main.extraction_confidence_score(
        structured_meta,
        structured_check,
        structured_txns,
        [{"page": 1, "text": text, "tables": []}],
        structured_missing,
        unresolved_amount_directions=0,
    )
    assert_equal(structured_score, text_score, "equivalent financial outputs must score equally")


if __name__ == "__main__":
    run()
    print("Accounting regression suite passed.")
